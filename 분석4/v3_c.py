# -*- coding: utf-8 -*-
"""3단계 : 분석 시트"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
warnings.filterwarnings('ignore')
from data3 import MONTHS, TOT, HS, HSNAME, OLDX, OLDM, NEW, OLDX_ITEMS, OLDM_ITEMS, NEW_ITEMS, Y, M17
from hs_meta import SECTIONS, CH2SEC, DESC

FONT='맑은 고딕'; NAVY='1F3864'; BLUE='0000FF'; GREEN='008000'; RED='C00000'
HDR=PatternFill('solid',fgColor='1F3864'); SUB=PatternFill('solid',fgColor='D9E2F3')
IN=PatternFill('solid',fgColor='FFFF00'); WARN=PatternFill('solid',fgColor='FCE4D6')
G24=PatternFill('solid',fgColor='EDEDED'); G26=PatternFill('solid',fgColor='E2EFDA')
THIN=Side(style='thin',color='BFBFBF'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
NUM='#,##0'; NUM1='#,##0.0'; PCT='0.0%;[RED]-0.0%;-'; PCT2='0.00%'; UNIT='#,##0.000'
PP='+0.00"%p";[RED]-0.00"%p";-'
wb=openpyxl.load_workbook('s2b.xlsx')

DT="'데이터_총괄'";      T0,T1=3,33
DH="'데이터_품목별'";    H0,H1=3,2984
DX="'데이터_성질별수출'"; X0,X1=3,653
DM="'데이터_성질별수입'"; M0,M1=3,627
DN="'데이터_신성질별'";   N0,N1=3,529
NM=len(MONTHS)
YRS=[2024,2025,2026]
def rg(sh,c,a,b): return f'{sh}!${c}${a}:${c}${b}'
tR=lambda c: rg(DT,c,T0,T1); hR=lambda c: rg(DH,c,H0,H1)
xR=lambda c: rg(DX,c,X0,X1); mR=lambda c: rg(DM,c,M0,M1); nR=lambda c: rg(DN,c,N0,N1)

def title(ws,row,text,span):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=14,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); ws.row_dimensions[row].height=22
def note(ws,row,col,text,span=10,color='595959'):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,i=True,color=color)
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    c.alignment=Alignment(vertical='center',wrap_text=True); return c
def sec(ws,row,text,span=10):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=11,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); c.fill=SUB; return c
def head(ws,row,col,text,fill=None):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,b=True,color='FFFFFF'); c.fill=fill or HDR
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BOX; return c
def put(ws,row,col,v,fmt=None,bold=False,color='000000',align=None,sz=10,border=True,fill=None,wrap=False):
    c=ws.cell(row,col,v); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    if fmt: c.number_format=fmt
    if align or wrap: c.alignment=Alignment(horizontal=align or 'general',vertical='center',wrap_text=wrap)
    if border: c.border=BOX
    if fill: c.fill=fill
    return c
def widths(ws,ws_w):
    for i,w_ in enumerate(ws_w): ws.column_dimensions[L(i+1)].width=w_
def SRCTOT(col): return f'IFERROR(VALUE(SUBSTITUTE(TRIM(\'수출입 총괄\'!{col}6),",","")),"")'

# ══════════════════════════════════════════════════════════════════
# 분석1_정합성검증
# ══════════════════════════════════════════════════════════════════
ws=wb.create_sheet('분석1_정합성검증'); ws.sheet_view.showGridLines=False
widths(ws,[24,18,18,15,20,18,18,15,20,12])
title(ws,1,'① 합계 정합성 검증 — 개별항목의 합 vs 총합 (5개 원본 시트 교차)',10)
note(ws,2,1,'원본은 숫자가 텍스트로 저장되어 있어 「데이터_○○」 정제 시트(수식)를 거쳐 검증합니다. 원본 시트는 수정하지 않았습니다. 기간 : 2024.01~2026.07(31개월).',10)
put(ws,3,1,'허용오차(중량, 톤)',bold=True); put(ws,3,2,1.0,NUM1,color=BLUE,fill=IN)
put(ws,3,3,'허용오차(금액, 천달러)',bold=True); put(ws,3,4,50,NUM,color=BLUE,fill=IN)
put(ws,3,5,'품목별 허용오차',bold=True); put(ws,3,6,500,NUM,color=BLUE,fill=IN)
TW,TA,TB='$B$3','$D$3','$F$3'
r=5
sec(ws,r,'[검증 1] 「수출입 총괄」 총계행 vs 31개월 합계'); r+=1
for i,h in enumerate(['항목','총계행(원본)','월별 합계','차이','판정']): head(ws,r,i+1,h)
r+=1
for nm,sc,dc,fmt,tol in [('조업일수','B','E',NUM1,TW),('수출 건수','C','F',NUM,TA),('수출 중량(톤)','D','G',NUM1,TW),
        ('수출 금액(천$)','E','H',NUM,TA),('수입 건수','F','I',NUM,TA),('수입 중량(톤)','G','J',NUM1,TW),
        ('수입 금액(천$)','H','K',NUM,TA),('무역수지(천$)','I','L',NUM,TA)]:
    put(ws,r,1,nm,bold=True); put(ws,r,2,'='+SRCTOT(sc),fmt)
    put(ws,r,3,f'=SUM({tR(dc)})',fmt); put(ws,r,4,f'=B{r}-C{r}',NUM1)
    put(ws,r,5,f'=IF(ABS(D{r})<={tol},"일치(반올림오차 이내)","불일치")',align='center'); r+=1
r+=1
sec(ws,r,'[검증 2] 무역수지 항등식 및 시트 간 교차검증 (31개월 총계)'); r+=1
for i,h in enumerate(['구분','수출 금액','수입 금액','무역수지','수출 중량','비고']): head(ws,r,i+1,h)
r+=1
X=r
put(ws,r,1,'수출입 총괄',bold=True)
put(ws,r,2,f'=SUM({tR("H")})',NUM); put(ws,r,3,f'=SUM({tR("K")})',NUM)
put(ws,r,4,f'=B{r}-C{r}',NUM); put(ws,r,5,f'=SUM({tR("G")})',NUM1); put(ws,r,6,'기준값'); r+=1
_xm=['1. 식료 및 직접소비재','2. 원료 및 연료','3. 경공업품','4. 중화학 공업품']
_mm=['1. 소비재','2. 원자재','3. 자본재']
_nm=['1.소비재','2.원자재','3.자본재']
put(ws,r,1,'성질별(수출)',bold=True)
put(ws,r,2,'='+'+'.join(f'SUMIFS({xR("G")},{xR("E")},"{k}")' for k in _xm),NUM)
put(ws,r,3,'-',align='center'); put(ws,r,4,'-',align='center')
put(ws,r,5,'='+'+'.join(f'SUMIFS({xR("F")},{xR("E")},"{k}")' for k in _xm),NUM1)
put(ws,r,6,'대분류 4개 합계'); r+=1
put(ws,r,1,'성질별(수입)',bold=True)
put(ws,r,2,'-',align='center')
put(ws,r,3,'='+'+'.join(f'SUMIFS({mR("G")},{mR("E")},"{k}")' for k in _mm),NUM)
put(ws,r,4,'-',align='center'); put(ws,r,5,'-',align='center')
put(ws,r,6,'대분류 3개 합계'); r+=1
put(ws,r,1,'신성질별',bold=True)
put(ws,r,2,'='+'+'.join(f'SUMIFS({nR("G")},{nR("E")},"{k}")' for k in _nm),NUM)
put(ws,r,3,'='+'+'.join(f'SUMIFS({nR("I")},{nR("E")},"{k}")' for k in _nm),NUM)
put(ws,r,4,f'=B{r}-C{r}',NUM)
put(ws,r,5,'='+'+'.join(f'SUMIFS({nR("F")},{nR("E")},"{k}")' for k in _nm),NUM1)
put(ws,r,6,'대분류 3개 합계'); r+=1
put(ws,r,1,'품목별(HS)',bold=True)
put(ws,r,2,f'=SUM({hR("H")})',NUM,color=RED); put(ws,r,3,f'=SUM({hR("J")})',NUM,color=RED)
put(ws,r,4,f'=B{r}-C{r}',NUM); put(ws,r,5,f'=SUM({hR("G")})',NUM1,color=RED)
put(ws,r,6,'HS 01~97·99류만 → 과소',fill=WARN); r+=1
put(ws,r,1,'총괄 대비 최대 차이',bold=True)
put(ws,r,2,f'=MAX(ABS(B{X+1}-B{X}),ABS(B{X+3}-B{X}))',NUM,bold=True)
put(ws,r,3,f'=MAX(ABS(C{X+2}-C{X}),ABS(C{X+3}-C{X}))',NUM,bold=True)
put(ws,r,4,f'=ABS(D{X+3}-D{X})',NUM,bold=True)
put(ws,r,5,f'=MAX(ABS(E{X+1}-E{X}),ABS(E{X+3}-E{X}))',NUM1,bold=True)
put(ws,r,6,'품목별 제외 기준'); r+=1
put(ws,r,1,'판정',bold=True)
put(ws,r,2,f'=IF(B{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,3,f'=IF(C{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,4,f'=IF(D{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,5,f'=IF(E{r-1}<={TW},"일치","불일치")',align='center',bold=True,color=GREEN)
r+=2
sec(ws,r,'[검증 3] 「품목별(HS)」 합 vs 「총괄」 — 연도별 차이'); r+=1
for i,h in enumerate(['연도','품목별 수출 합','총괄 수출금액','차이','차이 비율','품목별 수입 합','총괄 수입금액','차이','차이 비율']): head(ws,r,i+1,h)
r+=1
for y in YRS:
    put(ws,r,1,y,align='center',bold=True)
    put(ws,r,2,f'=SUMIFS({hR("H")},{hR("B")},$A{r})',NUM)
    put(ws,r,3,f'=SUMIFS({tR("H")},{tR("B")},$A{r})',NUM)
    put(ws,r,4,f'=B{r}-C{r}',NUM,color=RED); put(ws,r,5,f'=D{r}/C{r}',PCT2)
    put(ws,r,6,f'=SUMIFS({hR("J")},{hR("B")},$A{r})',NUM)
    put(ws,r,7,f'=SUMIFS({tR("K")},{tR("B")},$A{r})',NUM)
    put(ws,r,8,f'=F{r}-G{r}',NUM,color=RED); put(ws,r,9,f'=H{r}/G{r}',PCT2)
    r+=1
note(ws,r,1,'★ 품목별 시트는 총괄보다 매년 0.03~0.05% 작습니다. 이 시트에 실제로 존재하는 HS류는 01~97류 중 96개(77류는 HS 결번)와 값이 대부분 0인 99류뿐이고, '
            '98류(특수분류 — 여행자 휴대품·별송품·소액물품 등)가 통째로 빠져 있습니다. 품목 간 비교·순위 분석에는 영향이 없으나 총액 인용 시에는 총괄 시트를 쓰십시오.',9,color=RED)
r+=2
sec(ws,r,'[검증 4] 월별 항등식 (수출−수입=수지) 및 성질별 대분류 합 검증'); r+=1
for i,h in enumerate(['기간','수출−수입','원본 수지','차이','성질별(수출) 대분류합','총괄 수출','차이','신성질별 수입 합','총괄 수입','차이']): head(ws,r,i+1,h)
r+=1
V4=r
for k in range(NM):
    dr=T0+k; m=MONTHS[k]
    put(ws,r,1,f'={DT}!A{dr}',align='center')
    put(ws,r,2,f'={DT}!H{dr}-{DT}!K{dr}',NUM); put(ws,r,3,f'={DT}!L{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM)
    put(ws,r,5,'='+'+'.join(f'SUMIFS({xR("G")},{xR("E")},"{k2}",{xR("A")},$A{r})' for k2 in _xm),NUM)
    put(ws,r,6,f'={DT}!H{dr}',NUM); put(ws,r,7,f'=E{r}-F{r}',NUM1)
    put(ws,r,8,'='+'+'.join(f'SUMIFS({nR("I")},{nR("E")},"{k2}",{nR("A")},$A{r})' for k2 in _nm),NUM)
    put(ws,r,9,f'={DT}!K{dr}',NUM); put(ws,r,10,f'=H{r}-I{r}',NUM1)
    r+=1
put(ws,r,1,'최대 절대차이',bold=True)
for col,src in [(4,'D'),(7,'G'),(10,'J')]:
    put(ws,r,col,f'=MAX(MAX({src}{V4}:{src}{r-1}),-MIN({src}{V4}:{src}{r-1}))',NUM1,bold=True)
r+=1
put(ws,r,1,'판정',bold=True)
put(ws,r,4,f'=IF(D{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
put(ws,r,7,f'=IF(G{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
put(ws,r,10,f'=IF(J{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
ws.freeze_panes='A5'
print('분석1 완료', r)

# ══════════════════════════════════════════════════════════════════
# 분석2_월별추이 (31개월)
# ══════════════════════════════════════════════════════════════════
w2=wb.create_sheet('분석2_월별추이'); w2.sheet_view.showGridLines=False
widths(w2,[12,9,12,14,15,15,15,10,10,10,10,13,14,10,12,12,12,15,15,14])
title(w2,1,'② 월별 추이 (2024.01 ~ 2026.07, 31개월)',20)
note(w2,2,1,'「데이터_총괄」 참조. 일평균 = 금액÷조업일수, 단가 = 금액÷중량, 순상품교역조건 = 수출단가÷수입단가, 12개월 이동합 = 최근 12개월 누계(계절성 제거 추세).',20)
H2=['기간','조업일수','수출 건수','수출 중량(톤)','수출 금액(천$)','수입 금액(천$)','무역수지(천$)',
    '수출 전월비','수입 전월비','수출 전년동월비','수입 전년동월비','수지 전년동월증감','일평균 수출(천$)',
    '일평균 전년동월비','수출단가(천$/톤)','수입단가(천$/톤)','순상품교역조건','수출 12개월 이동합','수지 12개월 이동합','연내 누계 수출']
HR2=4
for i,h in enumerate(H2): head(w2,HR2,i+1,h)
w2.row_dimensions[HR2].height=36
D2=HR2+1
for k in range(NM):
    r=D2+k; dr=T0+k
    put(w2,r,1,f'={DT}!A{dr}',align='center',bold=True)
    put(w2,r,2,f'={DT}!E{dr}',NUM1); put(w2,r,3,f'={DT}!F{dr}',NUM)
    put(w2,r,4,f'={DT}!G{dr}',NUM1); put(w2,r,5,f'={DT}!H{dr}',NUM)
    put(w2,r,6,f'={DT}!K{dr}',NUM); put(w2,r,7,f'={DT}!L{dr}',NUM)
    put(w2,r,8,'' if k==0 else f'=E{r}/E{r-1}-1',PCT)
    put(w2,r,9,'' if k==0 else f'=F{r}/F{r-1}-1',PCT)
    put(w2,r,10,'' if k<12 else f'=E{r}/E{r-12}-1',PCT)
    put(w2,r,11,'' if k<12 else f'=F{r}/F{r-12}-1',PCT)
    put(w2,r,12,'' if k<12 else f'=G{r}-G{r-12}',NUM)
    put(w2,r,13,f'=E{r}/B{r}',NUM)
    put(w2,r,14,'' if k<12 else f'=M{r}/M{r-12}-1',PCT)
    put(w2,r,15,f'=E{r}/D{r}',UNIT); put(w2,r,16,f'=F{r}/{DT}!J{dr}',UNIT)
    put(w2,r,17,f'=O{r}/P{r}','0.000')
    put(w2,r,18,'' if k<11 else f'=SUM(E{r-11}:E{r})',NUM)
    put(w2,r,19,'' if k<11 else f'=SUM(G{r-11}:G{r})',NUM)
    put(w2,r,20,f'=SUMIFS({tR("H")},{tR("B")},VALUE(LEFT($A{r},4)),{tR("C")},"<="&VALUE(RIGHT($A{r},2)))',NUM)
E2=D2+NM-1
r=E2+2
SR={}
for lbl,cond in [('2024년 연간',2024),('2025년 연간',2025),('2026년 1~7월',2026),
                 ('2024년 1~7월',(2024,7)),('2025년 1~7월',(2025,7))]:
    put(w2,r,1,lbl,bold=True,align='center'); SR[lbl]=r
    if isinstance(cond,tuple):
        y,mm=cond
        f=lambda c: f'=SUMIFS({tR(c)},{tR("B")},{y},{tR("C")},"<={mm}")'
    else:
        f=lambda c: f'=SUMIFS({tR(c)},{tR("B")},{cond})'
    for col,src,fmt in [(2,'E',NUM1),(3,'F',NUM),(4,'G',NUM1),(5,'H',NUM),(6,'K',NUM),(7,'L',NUM)]:
        put(w2,r,col,f(src),fmt,bold=True)
    put(w2,r,13,f'=E{r}/B{r}',NUM,bold=True)
    put(w2,r,15,f'=E{r}/D{r}',UNIT,bold=True)
    put(w2,r,17,f'=O{r}/P{r}','0.000',bold=True)
    r+=1
put(w2,r,1,'25→26 동기비(1~7월)',bold=True,align='center')
for col in [2,3,4,5,6,13,15]:
    put(w2,r,col,f'={L(col)}{SR["2026년 1~7월"]}/{L(col)}{SR["2025년 1~7월"]}-1',PCT,bold=True)
put(w2,r,7,f'=G{SR["2026년 1~7월"]}-G{SR["2025년 1~7월"]}',NUM,bold=True)
r+=1
put(w2,r,1,'24→25 동기비(1~7월)',bold=True,align='center')
for col in [2,3,4,5,6,13,15]:
    put(w2,r,col,f'={L(col)}{SR["2025년 1~7월"]}/{L(col)}{SR["2024년 1~7월"]}-1',PCT,bold=True)
put(w2,r,7,f'=G{SR["2025년 1~7월"]}-G{SR["2024년 1~7월"]}',NUM,bold=True)
# 수입단가 요약행 보정
for lbl in SR:
    rr=SR[lbl]
    if lbl.endswith('연간'):
        y=int(lbl[:4]); put(w2,rr,16,f'=F{rr}/SUMIFS({tR("J")},{tR("B")},{y})',UNIT,bold=True)
    else:
        y=int(lbl[:4]); put(w2,rr,16,f'=F{rr}/SUMIFS({tR("J")},{tR("B")},{y},{tR("C")},"<=7")',UNIT,bold=True)
w2.freeze_panes=f'B{D2}'
print('분석2 완료')

# ══════════════════════════════════════════════════════════════════
# 분석3_MoM·QoQ·YoY
# ══════════════════════════════════════════════════════════════════
w3=wb.create_sheet('분석3_MoM·QoQ·YoY'); w3.sheet_view.showGridLines=False
widths(w3,[13,15,11,11,11,11,15,11,11,15,11,11,13,13])
title(w3,1,'③ MoM · QoQ · YoY 분석',14)
note(w3,2,1,'MoM(전월비)은 조업일수 차이에 취약하므로 일평균 기준 MoM을 함께 제시했습니다. QoQ는 분기 대비, YoY는 전년 동기 대비입니다. 2026년은 7월까지만 있어 3분기는 7월 1개월만 반영된 부분분기입니다.',14)
r=4
sec(w3,r,'[A] 월별 MoM · YoY (2024.01~2026.07)',14); r+=1
for i,h in enumerate(['기간','수출 금액','MoM','일평균 MoM','YoY','YoY 3개월이평','수입 금액','MoM','YoY','무역수지','수지 MoM증감','수지 YoY증감']): head(w3,r,i+1,h)
w3.row_dimensions[r].height=30
r+=1
A3=r
for k in range(NM):
    rr=A3+k; d2=D2+k
    put(w3,rr,1,f"='분석2_월별추이'!A{d2}",align='center',bold=True)
    put(w3,rr,2,f"='분석2_월별추이'!E{d2}",NUM)
    put(w3,rr,3,'' if k==0 else f'=B{rr}/B{rr-1}-1',PCT)
    put(w3,rr,4,'' if k==0 else f"='분석2_월별추이'!M{d2}/'분석2_월별추이'!M{d2-1}-1",PCT)
    put(w3,rr,5,'' if k<12 else f'=B{rr}/B{rr-12}-1',PCT)
    put(w3,rr,6,'' if k<14 else f'=SUM(B{rr-2}:B{rr})/SUM(B{rr-14}:B{rr-12})-1',PCT)
    put(w3,rr,7,f"='분석2_월별추이'!F{d2}",NUM)
    put(w3,rr,8,'' if k==0 else f'=G{rr}/G{rr-1}-1',PCT)
    put(w3,rr,9,'' if k<12 else f'=G{rr}/G{rr-12}-1',PCT)
    put(w3,rr,10,f"='분석2_월별추이'!G{d2}",NUM)
    put(w3,rr,11,'' if k==0 else f'=J{rr}-J{rr-1}',NUM)
    put(w3,rr,12,'' if k<12 else f'=J{rr}-J{rr-12}',NUM)
E3=A3+NM-1
r=E3+2
sec(w3,r,'[B] 분기별 QoQ · YoY',14); r+=1
for i,h in enumerate(['분기','수출 금액','QoQ','YoY','수입 금액','QoQ','YoY','무역수지','QoQ 증감','YoY 증감','조업일수','일평균 수출','비고']): head(w3,r,i+1,h)
r+=1
Q3=r
QL=[(y,q) for y in YRS for q in range(1,5)]
QL=[(y,q) for (y,q) in QL if not (y==2026 and q==4)]
for i,(y,q) in enumerate(QL):
    rr=Q3+i
    part = (y==2026 and q==3)
    put(w3,rr,1,f'{y} Q{q}',align='center',bold=True,fill=G26 if y==2026 else (G24 if y==2024 else None))
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y},{tR("D")},{q})',NUM)
    put(w3,rr,5,f'=SUMIFS({tR("K")},{tR("B")},{y},{tR("D")},{q})',NUM)
    put(w3,rr,8,f'=B{rr}-E{rr}',NUM)
    put(w3,rr,11,f'=SUMIFS({tR("E")},{tR("B")},{y},{tR("D")},{q})',NUM1)
    put(w3,rr,12,f'=IFERROR(B{rr}/K{rr},"-")',NUM)
    if i>0 and not part and not (y==2026 and q==3):
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,6,f'=E{rr}/E{rr-1}-1',PCT)
        put(w3,rr,9,f'=H{rr}-H{rr-1}',NUM)
    if i>=4 and not part:
        put(w3,rr,4,f'=B{rr}/B{rr-4}-1',PCT); put(w3,rr,7,f'=E{rr}/E{rr-4}-1',PCT)
        put(w3,rr,10,f'=H{rr}-H{rr-4}',NUM)
    put(w3,rr,13,'7월만 포함된 부분분기 — 증감률 계산 제외' if part else '',sz=9)
E3Q=Q3+len(QL)-1
r=E3Q+2
sec(w3,r,'[C] 연도별 비교 (연간 및 1~7월 동기)',14); r+=1
for i,h in enumerate(['구분','수출 금액','YoY','수입 금액','YoY','무역수지','YoY 증감','수출 중량','YoY','수출단가','YoY']): head(w3,r,i+1,h)
r+=1
C3=r
for i,y in enumerate(YRS):
    rr=C3+i
    put(w3,rr,1,f'{y}년 연간'+(' (1~7월만)' if y==2026 else ''),bold=True,fill=G26 if y==2026 else None)
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y})',NUM)
    put(w3,rr,4,f'=SUMIFS({tR("K")},{tR("B")},{y})',NUM)
    put(w3,rr,6,f'=B{rr}-D{rr}',NUM)
    put(w3,rr,8,f'=SUMIFS({tR("G")},{tR("B")},{y})',NUM1)
    put(w3,rr,10,f'=B{rr}/H{rr}',UNIT)
    if i>0 and y!=2026:
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,5,f'=D{rr}/D{rr-1}-1',PCT)
        put(w3,rr,7,f'=F{rr}-F{rr-1}',NUM); put(w3,rr,9,f'=H{rr}/H{rr-1}-1',PCT)
        put(w3,rr,11,f'=J{rr}/J{rr-1}-1',PCT)
r=C3+3
for i,y in enumerate(YRS):
    rr=r+i
    put(w3,rr,1,f'{y}년 1~7월',bold=True,fill=G26 if y==2026 else None)
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},"<=7")',NUM)
    put(w3,rr,4,f'=SUMIFS({tR("K")},{tR("B")},{y},{tR("C")},"<=7")',NUM)
    put(w3,rr,6,f'=B{rr}-D{rr}',NUM)
    put(w3,rr,8,f'=SUMIFS({tR("G")},{tR("B")},{y},{tR("C")},"<=7")',NUM1)
    put(w3,rr,10,f'=B{rr}/H{rr}',UNIT)
    if i>0:
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,5,f'=D{rr}/D{rr-1}-1',PCT)
        put(w3,rr,7,f'=F{rr}-F{rr-1}',NUM); put(w3,rr,9,f'=H{rr}/H{rr-1}-1',PCT)
        put(w3,rr,11,f'=J{rr}/J{rr-1}-1',PCT)
Y17=r
w3.freeze_panes='B6'
print('분석3 완료')

# ══════════════════════════════════════════════════════════════════
# 분석4_연도별계절비교 (1~12월 × 2024/2025/2026)
# ══════════════════════════════════════════════════════════════════
w4=wb.create_sheet('분석4_연도별계절비교'); w4.sheet_view.showGridLines=False
widths(w4,[10,15,15,15,11,11,13,11,11])
title(w4,1,'④ 연도별 계절 비교 — 1~12월을 한 축에 놓고 2024·2025·2026을 겹쳐 본다',9)
note(w4,2,1,'각 블록은 행 = 1~12월, 열 = 2024·2025·2026 구조입니다. 이 표가 분석10의 오버레이 그래프 원본이며, 같은 달끼리 3개 연도를 비교하므로 계절성이 자동으로 통제됩니다.',9)
note(w4,3,1,'2026년은 7월까지만 자료가 있어 8~12월은 수식 없이 비워 두었습니다(그래프에서 선이 끊깁니다). 새 달 자료가 추가되면 바로 윗 셀의 수식을 복사해 내려 채우십시오. YoY 열은 해당 연도와 직전 연도 같은 달의 증감률입니다.',9)
SEASON={}
def season_block(ws,r,name,unit,fx,fmt=NUM,ratio=False):
    sec(ws,r,f'{name} ({unit})',9); r+=1
    hr=r
    head(ws,r,1,'월')
    for i,y in enumerate(YRS): head(ws,r,2+i,f'{y}년',fill=PatternFill('solid',fgColor='7F7F7F') if y==2024 else (PatternFill('solid',fgColor='2E75B6') if y==2025 else PatternFill('solid',fgColor='548235')))
    head(ws,r,5,'2025 YoY'); head(ws,r,6,'2026 YoY'); head(ws,r,7,'26 vs 24'); head(ws,r,8,'25 비중'); head(ws,r,9,'26 비중')
    r+=1; r0=r
    for m in range(1,13):
        put(ws,r,1,f'{m}월',align='center',bold=True)
        for i,y in enumerate(YRS):
            if f'{y}.{m:02d}' in MONTHS: put(ws,r,2+i,fx(y,m),fmt)
            else: put(ws,r,2+i,None,fmt)
        put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT)
        put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"")',PCT)
        put(ws,r,7,f'=IFERROR(D{r}/B{r}-1,"")',PCT)
        put(ws,r,8,f'=IFERROR(C{r}/SUM(C${r0}:C${r0+11}),"")',PCT2)
        put(ws,r,9,f'=IFERROR(D{r}/SUM(D${r0}:D${r0+11}),"")',PCT2)
        r+=1
    put(ws,r,1,'연간 합',bold=True)
    for i in range(3): put(ws,r,2+i,f'=SUM({L(2+i)}{r0}:{L(2+i)}{r0+11})',fmt,bold=True)
    put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT,bold=True); put(ws,r,6,'',PCT)
    r+=1
    put(ws,r,1,'1~7월 합',bold=True)
    for i in range(3): put(ws,r,2+i,f'=SUM({L(2+i)}{r0}:{L(2+i)}{r0+6})',fmt,bold=True)
    put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT,bold=True)
    put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"")',PCT,bold=True)
    put(ws,r,7,f'=IFERROR(D{r}/B{r}-1,"")',PCT,bold=True)
    SEASON[name]=(hr,r0,r0+11,r-1,r)
    return r+2

def F_TOT(col): return lambda y,m: f'=SUMIFS({tR(col)},{tR("B")},{y},{tR("C")},{m})'
def F_NEW(item,col='G'): return lambda y,m: f'=SUMIFS({nR(col)},{nR("E")},"{item}",{nR("B")},{y},{nR("C")},{m})'
def F_HS(code,col='H'): return lambda y,m: f'=SUMIFS({hR(col)},{hR("E")},"{code}",{hR("B")},{y},{hR("C")},{m})'
def F_X(item,col='G'): return lambda y,m: f'=SUMIFS({xR(col)},{xR("E")},"{item}",{xR("B")},{y},{xR("C")},{m})'
def F_M(item,col='G'): return lambda y,m: f'=SUMIFS({mR(col)},{mR("E")},"{item}",{mR("B")},{y},{mR("C")},{m})'
def F_NONIT(y,m):
    tot=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'
    it='+'.join(f'SUMIFS({nR("G")},{nR("E")},"{i}",{nR("B")},{y},{nR("C")},{m})' for i in ['라.IT부품','다.IT제품'])
    return f'=IF({tot}=0,"",{tot}-({it}))'
def F_UNIT(y,m):
    a=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'; b=f'SUMIFS({tR("G")},{tR("B")},{y},{tR("C")},{m})'
    return f'=IFERROR({a}/{b},"")'
def F_DPA(y,m):
    a=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'; b=f'SUMIFS({tR("E")},{tR("B")},{y},{tR("C")},{m})'
    return f'=IFERROR({a}/{b},"")'
r=5
r=season_block(w4,r,'수출 금액','천달러',F_TOT('H'))
r=season_block(w4,r,'수입 금액','천달러',F_TOT('K'))
r=season_block(w4,r,'무역수지','천달러',F_TOT('L'))
r=season_block(w4,r,'수출 중량','톤',F_TOT('G'),NUM1)
r=season_block(w4,r,'수출 단가','천달러/톤',F_UNIT,UNIT)
r=season_block(w4,r,'일평균 수출','천달러/일',F_DPA)
r=season_block(w4,r,'IT부품 수출','천달러',F_NEW('라.IT부품'))
r=season_block(w4,r,'IT제품 수출','천달러',F_NEW('다.IT제품'))
r=season_block(w4,r,'IT 제외 수출','천달러',F_NONIT)
r=season_block(w4,r,'HS85 전기기기 수출','천달러',F_HS('85'))
r=season_block(w4,r,'HS84 기계류 수출','천달러',F_HS('84'))
r=season_block(w4,r,'HS87 차량 수출','천달러',F_HS('87'))
r=season_block(w4,r,'HS27 광물성연료 수출','천달러',F_HS('27'))
r=season_block(w4,r,'HS89 선박 수출','천달러',F_HS('89'))
r=season_block(w4,r,'중화학공업품 수출','천달러',F_X('4. 중화학 공업품'))
r=season_block(w4,r,'수입 원자재','천달러',F_M('2. 원자재'))
r=season_block(w4,r,'수입 자본재','천달러',F_M('3. 자본재'))
w4.freeze_panes='B6'
print('분석4 완료 블록수',len(SEASON))
exec(open('v3_m.py').read())
exec(open('v3_d.py').read())
wb.save('final3.xlsx'); print('saved final3')
