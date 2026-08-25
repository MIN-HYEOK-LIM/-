# -*- coding: utf-8 -*-
# 분석8·9·11·12·13 + 안내 (월별 분해 반영)

# ══════════════════════════════════════════════════════════════════
# 분석8_성질별심층 (월별 분해 중심)
# ══════════════════════════════════════════════════════════════════
w8=wb.create_sheet('분석8_성질별심층'); w8.sheet_view.showGridLines=False
widths(w8,[8,26]+[11]*7+[13]*3+[10,10,10,13,13,13,10])
title(w8,1,'⑧ 성질별·신성질별 — 월별로 쪼개서 본다',23)
note(w8,2,1,'이번 자료는 성질별이 수출·수입 두 파일로 나뉘어 있어 양방향 분석이 가능합니다. 수출과 수입의 분류 체계가 서로 다르므로(수출 : 식료/원료/경공업/중화학, 수입 : 소비재/원자재/자본재) 항목을 1:1로 대응시키지 마십시오.',23)
note(w8,3,1,'★ C~I열이 2026년 각 월의 전년동월비입니다. 누계가 아니라 월별로 나눠 계산했습니다. J열부터는 참고용 누계입니다.',23,color=RED)
k26=[MONTHS.index(m) for m in M26L]
def blk_month(ws,r,title_,items,majors,block,rng,valcol,label):
    sec(ws,r,title_,23); r+=1
    hd=['구분','항목']+[f'{m} YoY' for m in M26L]+[f'{label} 24(1~7)',f'{label} 25(1~7)',f'{label} 26(1~7)',
        '25 YoY','26 YoY','기여도','중량 26','단가 25','단가 26','단가 증감']
    for i,h in enumerate(hd): head(ws,r,i+1,h)
    ws.row_dimensions[r].height=28
    r+=1; r0=r
    for nm in items:
        isM = nm in majors
        idx=items.index(nm)
        put(ws,r,1,'대분류' if isM else '중분류',align='center',bold=isM,fill=SUB if isM else None)
        put(ws,r,2,nm,bold=isM,fill=SUB if isM else None)
        for j,k in enumerate(k26):
            put(ws,r,3+j,f'=IFERROR({mref(block,idx,k)}/{mref(block,idx,k-12)}-1,"-")',PCT,sz=9)
        for i,y in enumerate(YRS):
            put(ws,r,10+i,f'=SUMIFS({rng(valcol)},{rng("E")},$B{r},{rng("B")},{y},{rng("C")},"<=7")',NUM,bold=isM)
        put(ws,r,13,f'=IFERROR(K{r}/J{r}-1,"-")',PCT)
        put(ws,r,14,f'=IFERROR(L{r}/K{r}-1,"-")',PCT,bold=isM)
        put(ws,r,15,f'=IFERROR((L{r}-K{r})/({_e26}-{_e25}),"-")',PCT2)
        put(ws,r,16,f'=SUMIFS({rng("F")},{rng("E")},$B{r},{rng("B")},2026,{rng("C")},"<=7")',NUM1)
        put(ws,r,17,f'=IFERROR(K{r}/SUMIFS({rng("F")},{rng("E")},$B{r},{rng("B")},2025,{rng("C")},"<=7"),"-")',UNIT)
        put(ws,r,18,f'=IFERROR(L{r}/P{r},"-")',UNIT)
        put(ws,r,19,f'=IFERROR(R{r}/Q{r}-1,"-")',PCT)
        r+=1
    return r+1, r0
r=5
r,X8=blk_month(w8,r,'[A] 성질별 — 수출 (월별 전년동월비 + 누계)',OLDX_ITEMS,
    ['1. 식료 및 직접소비재','2. 원료 및 연료','3. 경공업품','4. 중화학 공업품'],MX_X,xR,'G','수출')
r,M8=blk_month(w8,r,'[B] 성질별 — 수입 (월별 전년동월비 + 누계)',OLDM_ITEMS,
    ['1. 소비재','2. 원자재','3. 자본재'],MX_M,mR,'G','수입')
sec(w8,r,'[C] 신성질별 — 수출 월별 YoY · 수입 월별 YoY · 무역수지',23); r+=1
hd=['구분','항목']+[f'{m} 수출YoY' for m in M26L]+['수출 26(1~7)','수입 26(1~7)','수지 25','수지 26','수지 증감']+[f'{m} 수입YoY' for m in M26L]
for i,h in enumerate(hd): head(w8,r,i+1,h)
w8.row_dimensions[r].height=30
r+=1; N8=r
for nm in NEW_ITEMS:
    isM = nm in ['1.소비재','2.원자재','3.자본재']; idx=NEW_ITEMS.index(nm)
    put(w8,r,1,'대분류' if isM else '중분류',align='center',bold=isM,fill=SUB if isM else None)
    put(w8,r,2,nm,bold=isM,fill=SUB if isM else None)
    for j,k in enumerate(k26):
        put(w8,r,3+j,f'=IFERROR({mref(MX_N,idx,k)}/{mref(MX_N,idx,k-12)}-1,"-")',PCT,sz=9)
        put(w8,r,15+j,f'=IFERROR({mref(MX_NI,idx,k)}/{mref(MX_NI,idx,k-12)}-1,"-")',PCT,sz=9)
    put(w8,r,10,f'=SUMIFS({nR("G")},{nR("E")},$B{r},{nR("B")},2026,{nR("C")},"<=7")',NUM,bold=isM)
    put(w8,r,11,f'=SUMIFS({nR("I")},{nR("E")},$B{r},{nR("B")},2026,{nR("C")},"<=7")',NUM)
    put(w8,r,12,f'=SUMIFS({nR("G")},{nR("E")},$B{r},{nR("B")},2025,{nR("C")},"<=7")-SUMIFS({nR("I")},{nR("E")},$B{r},{nR("B")},2025,{nR("C")},"<=7")',NUM)
    put(w8,r,13,f'=J{r}-K{r}',NUM,bold=isM); put(w8,r,14,f'=M{r}-L{r}',NUM)
    r+=1
E8N=r-1
w8.freeze_panes='C6'
print('분석8(월별) 완료')

# ══════════════════════════════════════════════════════════════════
# 분석9_물량·가격분해 (월별)
# ══════════════════════════════════════════════════════════════════
w9=wb.create_sheet('분석9_물량·가격분해'); w9.sheet_view.showGridLines=False
widths(w9,[13,15,15,15,15,11,11,11,11,12,26])
title(w9,1,'⑨ 물량·가격 분해 — 매달 따로 계산한다',11)
note(w9,2,1,'금액 증감 = 물량효과 (Q₁−Q₀)×P₀ + 가격효과 (P₁−P₀)×Q₀ + 교차효과. 세 값의 합은 실제 증감액과 일치합니다(검증 열).',11)
note(w9,3,1,'★ 아래 [A]는 2025.01~2026.07 각 달을 전년 동월과 비교해 따로 분해한 것입니다. 누계 분해는 [B]에 참고로 두었습니다.',11,color=RED)
r=5
sec(w9,r,'[A] 월별 3효과 분해 (전년 동월 대비)',11); r+=1
for i,h in enumerate(['기간','수출 금액','금액 증감','물량효과','가격효과','교차효과','물량 기여','가격 기여','물량 증감률','단가 증감률','검증']): head(w9,r,i+1,h)
r+=1; A9=r
for k in range(12,NM):
    d1=D2+k; d0=D2+k-12
    put(w9,r,1,f"='분석2_월별추이'!A{d1}",align='center',bold=True)
    put(w9,r,2,f"='분석2_월별추이'!E{d1}",NUM)
    put(w9,r,3,f"='분석2_월별추이'!E{d1}-'분석2_월별추이'!E{d0}",NUM)
    q1=f"'분석2_월별추이'!D{d1}"; q0=f"'분석2_월별추이'!D{d0}"
    p1=f"'분석2_월별추이'!O{d1}"; p0=f"'분석2_월별추이'!O{d0}"
    put(w9,r,4,f'=({q1}-{q0})*{p0}',NUM)
    put(w9,r,5,f'=({p1}-{p0})*{q0}',NUM)
    put(w9,r,6,f'=({q1}-{q0})*({p1}-{p0})',NUM)
    put(w9,r,7,f'=IFERROR(D{r}/$C{r},"-")',PCT); put(w9,r,8,f'=IFERROR(E{r}/$C{r},"-")',PCT)
    put(w9,r,9,f'={q1}/{q0}-1',PCT); put(w9,r,10,f'={p1}/{p0}-1',PCT)
    put(w9,r,11,f'=IF(ABS(D{r}+E{r}+F{r}-C{r})<1,"OK","불일치")',align='center')
    r+=1
E9M=r-1
put(w9,r,1,'2026.01~07 합',bold=True)
for col in [3,4,5,6]: put(w9,r,col,f'=SUM({L(col)}{A9+12}:{L(col)}{E9M})',NUM,bold=True)
put(w9,r,7,f'=D{r}/C{r}',PCT,bold=True); put(w9,r,8,f'=E{r}/C{r}',PCT,bold=True)
r+=1
put(w9,r,1,'2025.01~07 합',bold=True)
for col in [3,4,5,6]: put(w9,r,col,f'=SUM({L(col)}{A9}:{L(col)}{A9+6})',NUM,bold=True)
put(w9,r,7,f'=D{r}/C{r}',PCT,bold=True); put(w9,r,8,f'=E{r}/C{r}',PCT,bold=True)
r+=2
sec(w9,r,'[B] 참고 : 누계 구간 분해 (1~7월 동기)',11); r+=1
for i,h in enumerate(['구간','금액 증감','물량효과','가격효과','교차효과','물량 기여','가격 기여','물량 증감률','단가 증감률','검증']): head(w9,r,i+1,h)
r+=1; B9C=r
for lbl,y0,y1 in [('2024 → 2025',2024,2025),('2025 → 2026',2025,2026)]:
    v0=tot_sum(y0,7); v1=tot_sum(y1,7); q0=tot_sum(y0,7,'G'); q1=tot_sum(y1,7,'G')
    put(w9,r,1,lbl,bold=True)
    put(w9,r,2,f'={v1}-{v0}',NUM)
    put(w9,r,3,f'=({q1}-{q0})*({v0}/{q0})',NUM)
    put(w9,r,4,f'=(({v1}/{q1})-({v0}/{q0}))*{q0}',NUM)
    put(w9,r,5,f'=({q1}-{q0})*(({v1}/{q1})-({v0}/{q0}))',NUM)
    put(w9,r,6,f'=C{r}/B{r}',PCT); put(w9,r,7,f'=D{r}/B{r}',PCT)
    put(w9,r,8,f'={q1}/{q0}-1',PCT); put(w9,r,9,f'=({v1}/{q1})/({v0}/{q0})-1',PCT)
    put(w9,r,10,f'=IF(ABS(C{r}+D{r}+E{r}-B{r})<1,"OK","불일치")',align='center')
    r+=1
r+=1
sec(w9,r,'[C] 신성질별 중분류 분해 (2025→2026, 1~7월 누계 기준)',11); r+=1
for i,h in enumerate(['품목','중량 2025','금액 2025','중량 2026','금액 2026','물량 증감률','단가 증감률','물량효과','가격효과','Δ금액']): head(w9,r,i+1,h)
r+=1; B9=r
SUBS=[n for n in NEW_ITEMS if n not in ['1.소비재','2.원자재','3.자본재']]
for nm in SUBS:
    put(w9,r,1,nm,bold=True)
    put(w9,r,2,f'=SUMIFS({nR("F")},{nR("E")},"{nm}",{nR("B")},2025,{nR("C")},"<=7")',NUM1)
    put(w9,r,3,f'=SUMIFS({nR("G")},{nR("E")},"{nm}",{nR("B")},2025,{nR("C")},"<=7")',NUM)
    put(w9,r,4,f'=SUMIFS({nR("F")},{nR("E")},"{nm}",{nR("B")},2026,{nR("C")},"<=7")',NUM1)
    put(w9,r,5,f'=SUMIFS({nR("G")},{nR("E")},"{nm}",{nR("B")},2026,{nR("C")},"<=7")',NUM)
    put(w9,r,6,f'=IFERROR(D{r}/B{r}-1,"-")',PCT)
    put(w9,r,7,f'=IFERROR((E{r}/D{r})/(C{r}/B{r})-1,"-")',PCT)
    put(w9,r,8,f'=IFERROR((D{r}-B{r})*(C{r}/B{r}),"-")',NUM)
    put(w9,r,9,f'=IFERROR(((E{r}/D{r})-(C{r}/B{r}))*B{r},"-")',NUM)
    put(w9,r,10,f'=E{r}-C{r}',NUM)
    r+=1
E9=r-1
put(w9,r,1,'합계',bold=True)
for col in [2,3,4,5,8,9,10]: put(w9,r,col,f'=SUM({L(col)}{B9}:{L(col)}{E9})',NUM1 if col in (2,4) else NUM,bold=True)
put(w9,r,6,f'=D{r}/B{r}-1',PCT,bold=True); put(w9,r,7,f'=(E{r}/D{r})/(C{r}/B{r})-1',PCT,bold=True)
S9=r; r+=2
sec(w9,r,'[D] 지수 분해 (라스파이레스·파셰·피셔, 2025→2026 누계)',11); r+=1
for i,h in enumerate(['지수','값','증감률','의미']): head(w9,r,i+1,h)
r+=1; I9=r
q0p0=f'SUMPRODUCT($B${B9}:$B${E9},$C${B9}:$C${E9}/$B${B9}:$B${E9})'
q1p0=f'SUMPRODUCT($D${B9}:$D${E9},$C${B9}:$C${E9}/$B${B9}:$B${E9})'
q1p1=f'SUM($E${B9}:$E${E9})'
q0p1=f'SUMPRODUCT($B${B9}:$B${E9},$E${B9}:$E${E9}/$D${B9}:$D${E9})'
for nm,f_,mean in [('라스파이레스 물량지수',f'={q1p0}/{q0p0}','기준연도 가격 고정, 물량만 변화'),
                   ('파셰 물량지수',f'={q1p1}/{q0p1}','비교연도 가격 고정, 물량만 변화'),
                   ('피셔 물량지수',None,'두 지수의 기하평균 (권장)'),
                   ('라스파이레스 가격지수',f'={q0p1}/{q0p0}','기준연도 물량 고정, 가격만 변화'),
                   ('파셰 가격지수',f'={q1p1}/{q1p0}','비교연도 물량 고정, 가격만 변화'),
                   ('피셔 가격지수',None,'두 지수의 기하평균 (권장)'),
                   ('금액지수',f'={q1p1}/{q0p0}','피셔 물량 × 피셔 가격과 일치해야 함')]:
    put(w9,r,1,nm,bold=True)
    if nm=='피셔 물량지수': put(w9,r,2,f'=SQRT(B{I9}*B{I9+1})','0.0000',bold=True)
    elif nm=='피셔 가격지수': put(w9,r,2,f'=SQRT(B{I9+3}*B{I9+4})','0.0000',bold=True)
    else: put(w9,r,2,f_,'0.0000')
    put(w9,r,3,f'=B{r}-1',PCT); put(w9,r,4,mean)
    r+=1
put(w9,r,1,'구성(믹스)효과',bold=True)
put(w9,r,2,f'=((E{S9}/D{S9})/(C{S9}/B{S9}))/B{I9+5}','0.0000')
put(w9,r,3,f'=B{r}-1',PCT,color=RED); put(w9,r,4,'단순 단가 상승분 중 품목 구성 변화가 만든 몫')
print('분석9(월별) 완료')

# ══════════════════════════════════════════════════════════════════
# 분석13_월별변화요약  (그 달에 정확히 무슨 일이 있었나)
# ══════════════════════════════════════════════════════════════════
exec(open('movers.py').read())
w13=wb.create_sheet('분석13_월별변화요약'); w13.sheet_view.showGridLines=False
widths(w13,[11,14,9,9,14,10,10,16,13,16,13,16,13,16,13,16,13,16,13,52])
title(w13,1,'⑬ 월별 변화 요약 — 그 달에 정확히 무엇이 달라졌나 (2025.01~2026.07)',20)
note(w13,2,1,'각 행이 한 달입니다. 그 달의 전년 동월 대비 증감액을 누가 만들었는지(증가·감소 상위 3개 HS류)와 반도체/비반도체의 몫을 함께 보여줍니다.',20)
note(w13,3,1,'상위 3개 품목의 선정은 작성 시점의 값 기준으로 고정되어 있습니다(수식 자동 재정렬 아님). 증감액은 모두 수식이며 원본이 바뀌면 갱신됩니다.',20)
r=5
hd=['기간','수출 금액','YoY','MoM','전체 증감액','반도체 몫','비반도체 몫',
    '증가 1위','Δ','증가 2위','Δ','증가 3위','Δ','감소 1위','Δ','감소 2위','Δ','감소 3위','Δ','한 줄 해석']
for i,h in enumerate(hd): head(w13,r,i+1,h)
w13.row_dimensions[r].height=30
r+=1; R13=r
for k in range(12,NM):
    m=MONTHS[k]; d1=D2+k; d0=D2+k-12
    put(w13,r,1,m,align='center',bold=True)
    put(w13,r,2,f"='분석2_월별추이'!E{d1}",NUM)
    put(w13,r,3,f"='분석2_월별추이'!J{d1}",PCT)
    put(w13,r,4,f"='분석2_월별추이'!H{d1}",PCT)
    put(w13,r,5,f"='분석2_월별추이'!E{d1}-'분석2_월별추이'!E{d0}",NUM)
    c1=L(2+k); c0=L(2+k-12)
    put(w13,r,6,f'=IFERROR((\'분석5_반도체vs비반도체\'!{c1}{R_IT2}-\'분석5_반도체vs비반도체\'!{c0}{R_IT2})/E{r},"-")',PCT2)
    put(w13,r,7,f'=IFERROR((\'분석5_반도체vs비반도체\'!{c1}{R_NON}-\'분석5_반도체vs비반도체\'!{c0}{R_NON})/E{r},"-")',PCT2)
    up,dn=MOVERS[m]
    for j,code in enumerate(up):
        idx=HSI[code]
        put(w13,r,8+j*2,f'{code}. {HSNAME[code][:12]}',sz=9)
        put(w13,r,9+j*2,f'={mref(MX_HS_EX,idx,k)}-{mref(MX_HS_EX,idx,k-12)}',NUM,sz=9)
    for j,code in enumerate(dn):
        idx=HSI[code]
        put(w13,r,14+j*2,f'{code}. {HSNAME[code][:12]}',sz=9)
        put(w13,r,15+j*2,f'={mref(MX_HS_EX,idx,k)}-{mref(MX_HS_EX,idx,k-12)}',NUM,sz=9)
    put(w13,r,20,
        f'=IF(C{r}>0.5,"전년 대비 큰 폭 증가. ",IF(C{r}>0.1,"전년 대비 뚜렷한 증가. ",IF(C{r}>0,"전년 대비 소폭 증가. ",'
        f'IF(C{r}>-0.1,"전년 대비 소폭 감소. ","전년 대비 뚜렷한 감소. "))))'
        f'&IF(D{r}>0,"전월보다도 늘었다. ","전월보다는 줄었다. ")'
        f'&IF(AND(E{r}>0,F{r}>0.7),"증가의 대부분(70% 이상)을 반도체가 만들었다. ",'
        f'IF(AND(E{r}>0,F{r}>0.4),"반도체가 증가의 절반 안팎을 만들었다. ",'
        f'IF(AND(E{r}>0,F{r}<=0.4),"반도체 외 품목의 기여가 더 컸다. ","")))'
        f'&"주도 품목은 "&H{r}&"("&TEXT(I{r},"#,##0")&"). "'
        f'&IF(O{r}<0,"반대로 "&N{r}&"("&TEXT(O{r},"#,##0")&")가 가장 크게 줄었다.","")',sz=9,wrap=True)
    w13.row_dimensions[r].height=32
    r+=1
E13=r-1
note(w13,r,1,'단위 : 금액·증감액은 천달러. 반도체는 광의(IT부품+IT제품) 기준입니다. 전체가 감소한 달에는 기여도 부호가 뒤집혀 보일 수 있습니다.',20)
w13.freeze_panes='B6'
print('분석13 완료')

# ══════════════════════════════════════════════════════════════════
# 분석10_그래프_연도오버레이
# ══════════════════════════════════════════════════════════════════
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import ChartLines
PALETTE=['A6A6A6','2E75B6','548235','ED7D31','C00000','7030A0','1F3864','BF8F00',
         '4472C4','70AD47','FFC000','9E480E','636363','255E91','43682B','7CAFDD']
def flat(ch,off=0):
    isbar=isinstance(ch,BarChart)
    for i,s_ in enumerate(ch.series):
        col=PALETTE[(i+off)%len(PALETTE)]
        if isbar:
            s_.graphicalProperties.solidFill=col; s_.graphicalProperties.line.noFill=True
        else:
            s_.smooth=False; s_.graphicalProperties.line.solidFill=col; s_.graphicalProperties.line.width=24000
def subax(ch,off=0):
    ch.x_axis.delete=True; flat(ch,off)
w10=wb.create_sheet('분석10_그래프_연도오버레이'); w10.sheet_view.showGridLines=False
widths(w10,[16]+[12]*9)
title(w10,1,'⑩ 연도 오버레이 그래프 — 1~12월을 한 축에 놓고 2024·2025·2026을 겹쳐 비교',10)
note(w10,2,1,'회색 = 2024년, 파랑 = 2025년, 초록 = 2026년. 2026년은 7월에서 선이 끊깁니다. 같은 달끼리 비교하므로 계절성이 통제되며, 세 선의 간격이 곧 그 달의 전년 대비 증감입니다.',10)
note(w10,3,1,'원본 데이터는 「분석4_연도별계절비교」의 각 블록입니다.',10)
CT=5
def overlay(name,anchor,ylab):
    hr,r0,r1,ytot,y7=SEASON[name]
    ch=LineChart()
    ch.add_data(Reference(w4,min_col=2,max_col=4,min_row=hr,max_row=r1),titles_from_data=True)
    ch.set_categories(Reference(w4,min_col=1,min_row=r0,max_row=r1))
    flat(ch); ch.dispBlanksAs='gap'; ch.title=f'{name} — 연도 비교 (2024·2025·2026)'
    ch.width=24; ch.height=11; ch.style=2
    ch.y_axis.title=ylab; ch.y_axis.majorGridlines=ChartLines()
    w10.add_chart(ch,anchor)
ORDER10=[('수출 금액','천달러'),('수입 금액','천달러'),('무역수지','천달러'),('수출 단가','천달러/톤'),
         ('일평균 수출','천달러/일'),('수출 중량','톤'),('IT부품 수출','천달러'),('IT제품 수출','천달러'),
         ('IT 제외 수출','천달러'),('HS85 전기기기 수출','천달러'),('HS84 기계류 수출','천달러'),
         ('HS87 차량 수출','천달러'),('HS27 광물성연료 수출','천달러'),('HS89 선박 수출','천달러'),
         ('중화학공업품 수출','천달러'),('수입 원자재','천달러'),('수입 자본재','천달러')]
for i,(nm,ylab) in enumerate(ORDER10): overlay(nm,f'A{CT+i*22}',ylab)
print('분석10 완료 :',len(ORDER10),'개')

# ══════════════════════════════════════════════════════════════════
# 분석11_그래프_구조분석 (월별 분해 반영)
# ══════════════════════════════════════════════════════════════════
w11=wb.create_sheet('분석11_그래프_구조분석'); w11.sheet_view.showGridLines=False
widths(w11,[26]+[12]*12)
title(w11,1,'⑪ 구조 분석 그래프 — 월별 분해 · 반도체 · 부문 · 분기',13)
note(w11,2,1,'보조표는 분석3·5·6·7·9·13을 참조하는 수식입니다.',13)
r=4
sec(w11,r,'보조표 A. HS 부문별 1~7월 수출 증감액 (2026 vs 2025, 천달러)',13); r+=1
A11H=r; head(w11,r,1,'부문'); head(w11,r,2,'증감액'); head(w11,r,3,'26 YoY')
r+=1; A11=r
for i,(s,n,ch) in enumerate(SEC_ORDER):
    put(w11,r,1,f'{s}. {n}',sz=9)
    put(w11,r,2,f"='분석6_HS부문별'!E{D6+i}-'분석6_HS부문별'!D{D6+i}",NUM)
    put(w11,r,3,f"='분석6_HS부문별'!G{D6+i}",PCT)
    r+=1
A11E=r-1; r+=1
sec(w11,r,'보조표 B. 성장 유형별 Δ금액 (HS 97류, 천달러)',13); r+=1
B11H=r; head(w11,r,1,'유형'); head(w11,r,2,'Δ금액'); head(w11,r,3,'류 수')
r+=1; B11=r
for i,t in enumerate(['가격 주도','물량 주도','혼합','감소','단가 불안정']):
    put(w11,r,1,t); put(w11,r,2,f"='분석7_HS류별상세'!D{T7+i}",NUM); put(w11,r,3,f"='분석7_HS류별상세'!B{T7+i}",NUM)
    r+=1
B11E=r-1; r+=1
sec(w11,r,'보조표 C. HS 류별 1~7월 수출 증감액 상위 15 / 하위 10 (천달러)',13); r+=1
C11H=r; head(w11,r,1,'HS류'); head(w11,r,2,'증감액')
r+=1; C11=r
for i,code in enumerate(order7[:15]):
    put(w11,r,1,f'{code}. {HSNAME[code][:18]}',sz=9)
    put(w11,r,2,f"='분석7_HS류별상세'!T{R7+i}-'분석7_HS류별상세'!S{R7+i}",NUM)
    r+=1
C11E=r-1; r+=1
D11H=r; head(w11,r,1,'HS류(감소)'); head(w11,r,2,'증감액')
r+=1; D11=r
for code in order7[-10:]:
    idx=order7.index(code)
    put(w11,r,1,f'{code}. {HSNAME[code][:18]}',sz=9)
    put(w11,r,2,f"='분석7_HS류별상세'!T{R7+idx}-'분석7_HS류별상세'!S{R7+idx}",NUM)
    r+=1
D11E=r-1; r+=1
sec(w11,r,'보조표 D. 월별 증감액 분해 — 반도체 vs 비반도체 (천달러)',13); r+=1
E11H=r; head(w11,r,1,'기간'); head(w11,r,2,'반도체(광의)'); head(w11,r,3,'비반도체'); head(w11,r,4,'전체 증감액')
r+=1; E11=r
for i,k in enumerate(range(12,NM)):
    rr=R13+i
    put(w11,r,1,MONTHS[k],align='center',sz=9)
    put(w11,r,2,f"='분석13_월별변화요약'!F{rr}*'분석13_월별변화요약'!E{rr}",NUM,sz=9)
    put(w11,r,3,f"='분석13_월별변화요약'!G{rr}*'분석13_월별변화요약'!E{rr}",NUM,sz=9)
    put(w11,r,4,f"='분석13_월별변화요약'!E{rr}",NUM,sz=9)
    r+=1
E11E=r-1
CT11=E11E+3
def add11(ch,anchor,t,w=26,h=12,ylab=None,xlab=None,legend=True,off=0):
    flat(ch,off); ch.title=t; ch.width=w; ch.height=h; ch.style=2
    if ylab: ch.y_axis.title=ylab
    if xlab: ch.x_axis.title=xlab
    if not legend: ch.legend=None
    ch.y_axis.majorGridlines=ChartLines()
    w11.add_chart(ch,anchor)
# 1. 반도체 vs 비반도체 월별 YoY
l1=LineChart()
l1.add_data(Reference(w5,min_col=1,max_col=1+(NM-12),min_row=Y_IT2,max_row=Y_IT2),from_rows=True,titles_from_data=True)
l1.add_data(Reference(w5,min_col=1,max_col=1+(NM-12),min_row=Y_NON,max_row=Y_NON),from_rows=True,titles_from_data=True)
l1.add_data(Reference(w5,min_col=1,max_col=1+(NM-12),min_row=Y_TOT,max_row=Y_TOT),from_rows=True,titles_from_data=True)
l1.set_categories(Reference(w5,min_col=2,max_col=1+(NM-12),min_row=B5H))
add11(l1,f'A{CT11}','1. 월별 전년동월비 — 반도체(광의) vs 비반도체 vs 전체',ylab='전년동월비',off=3)
# 2. 반도체 비중 추이
l2=LineChart()
l2.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=R_SH,max_row=R_SH),from_rows=True,titles_from_data=True)
l2.set_categories(Reference(w5,min_col=2,max_col=1+NM,min_row=A5H))
add11(l2,f'A{CT11+24}','2. 반도체(광의) 비중 추이 (2024.01~2026.07)',ylab='전체 수출 대비 비중',off=4)
# 3. 월별 증감액 분해 (반도체/비반도체 누적 막대)
b3=BarChart(); b3.type='col'; b3.grouping='stacked'; b3.overlap=100
b3.add_data(Reference(w11,min_col=2,max_col=3,min_row=E11H,max_row=E11E),titles_from_data=True)
b3.set_categories(Reference(w11,min_col=1,min_row=E11,max_row=E11E))
add11(b3,f'A{CT11+48}','3. 월별 수출 증감액 분해 — 반도체와 비반도체가 각각 만든 몫',ylab='증감액(천달러)',off=1)
# 4. HS 부문별 증감액
b4=BarChart(); b4.type='bar'
b4.add_data(Reference(w11,min_col=2,min_row=A11H,max_row=A11E),titles_from_data=True)
b4.set_categories(Reference(w11,min_col=1,min_row=A11,max_row=A11E))
add11(b4,f'A{CT11+72}','4. HS 부문별 1~7월 수출 증감액 (2026 vs 2025)',h=15,ylab='증감액(천달러)',legend=False,off=1)
# 5. HS 상위 15
b5=BarChart(); b5.type='bar'
b5.add_data(Reference(w11,min_col=2,min_row=C11H,max_row=C11E),titles_from_data=True)
b5.set_categories(Reference(w11,min_col=1,min_row=C11,max_row=C11E))
add11(b5,f'A{CT11+102}','5. HS 류별 수출 증감액 상위 15',h=14,ylab='증감액(천달러)',legend=False,off=1)
# 6. HS 하위 10
b6=BarChart(); b6.type='bar'
b6.add_data(Reference(w11,min_col=2,min_row=D11H,max_row=D11E),titles_from_data=True)
b6.set_categories(Reference(w11,min_col=1,min_row=D11,max_row=D11E))
add11(b6,f'A{CT11+130}','6. HS 류별 수출 감소액 하위 10',h=12,ylab='증감액(천달러)',legend=False,off=4)
# 7. 성장 유형별
b7=BarChart(); b7.type='col'
b7.add_data(Reference(w11,min_col=2,min_row=B11H,max_row=B11E),titles_from_data=True)
b7.set_categories(Reference(w11,min_col=1,min_row=B11,max_row=B11E))
add11(b7,f'A{CT11+154}','7. 성장 유형별 수출 증감액 (HS 97류 분류)',h=11,ylab='증감액(천달러)',legend=False,off=1)
# 8. 분기별
b8=BarChart(); b8.type='col'
b8.add_data(Reference(w3,min_col=2,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
b8.add_data(Reference(w3,min_col=5,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
b8.set_categories(Reference(w3,min_col=1,min_row=Q3,max_row=E3Q))
l8=LineChart()
l8.add_data(Reference(w3,min_col=8,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l8.y_axis.axId=200; l8.y_axis.title='무역수지(천$)'; l8.y_axis.crosses='max'; subax(l8,4); b8+=l8
add11(b8,f'A{CT11+178}','8. 분기별 수출·수입·무역수지 (2024 Q1~2026 Q3)',ylab='금액(천달러)',off=1)
# 9. 분기 YoY / QoQ
l9=LineChart()
l9.add_data(Reference(w3,min_col=4,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l9.add_data(Reference(w3,min_col=3,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l9.set_categories(Reference(w3,min_col=1,min_row=Q3,max_row=E3Q))
add11(l9,f'A{CT11+202}','9. 분기별 수출 YoY와 QoQ',ylab='증감률',off=1)
# 10. 월별 물량·가격효과
b10=BarChart(); b10.type='col'; b10.grouping='stacked'; b10.overlap=100
b10.add_data(Reference(w9,min_col=4,max_col=6,min_row=A9-1,max_row=E9M),titles_from_data=True)
b10.set_categories(Reference(w9,min_col=1,min_row=A9,max_row=E9M))
add11(b10,f'A{CT11+226}','10. 월별 물량효과·가격효과·교차효과 (전년 동월 대비)',ylab='금액(천달러)',off=1)
print('분석11 완료')

# ══════════════════════════════════════════════════════════════════
# 분석12_해설 (월별 분해 중심)
# ══════════════════════════════════════════════════════════════════
w12=wb.create_sheet('분석12_해설'); w12.sheet_view.showGridLines=False
widths(w12,[30,15,15,15,13,13,13,13,13])
title(w12,1,'⑫ 해설 — 월별로 쪼개서 본 2024~2026',9)
note(w12,2,1,'이 파일의 모든 증감 분석은 1~7월 누계가 아니라 각 달을 전년 동월과 비교하는 방식으로 재구성했습니다. 누계 수치는 참고용으로만 남겨 두었습니다.',9)
r=4
sec(w12,r,'[핵심 수치] 월별 전년동월비 (2026년, %)',9); r+=1
for i,h in enumerate(['구분']+M26L): head(w12,r,i+1,h)
r+=1
K12=r
for lab,src in [('전체 수출',Y_TOT),('반도체(광의)',Y_IT2),('비반도체',Y_NON),('HS85',Y_HS85)]:
    put(w12,r,1,lab,bold=True)
    for j,m in enumerate(M26L):
        k=MONTHS.index(m)
        put(w12,r,2+j,f"='분석5_반도체vs비반도체'!{L(2+k-12)}{src}",PCT)
    r+=1
put(w12,r,1,'반도체가 만든 몫',bold=True)
for j,m in enumerate(M26L):
    k=MONTHS.index(m)
    put(w12,r,2+j,f"='분석5_반도체vs비반도체'!{L(2+k-12)}{R_CONT}",PCT2)
r+=2
def para(ws,row,text,span=9,sz=10,color='000000',bold=False):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    c.alignment=Alignment(vertical='top',wrap_text=True)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    ws.row_dimensions[row].height=max(16,15*(len(text)//64+1))
SECT=[
 ('1. 이번 개편 — 누계를 걷어내고 월별로 다시 짰습니다',[
  '· 이전 버전은 1~7월을 합산해 전년 동기와 비교했습니다. 그렇게 하면 "7개월 평균 +50.5%"처럼 뭉뚱그려져, 어느 달에 무엇이 달라졌는지가 사라집니다.',
  '· 이번에는 분석4B에 모든 항목의 31개월 월별 값을 한 번에 계산해 두고, 분석5~9·13이 그 표를 참조하도록 바꿨습니다. 품목·부문·성질별 모든 항목이 매달 따로 전년 동월과 비교됩니다.',
  '· 새로 넣은 분석13은 한 행이 한 달입니다. 그 달의 증감액을 누가 만들었는지(증가·감소 상위 3개 HS류)와 반도체/비반도체의 몫을 한 줄로 읽을 수 있습니다.',
 ]),
 ('2. 2025년 — 월별로 보면 "정체"가 아니라 "1월 급락 후 회복"',[
  '· 1~7월 누계로는 +0.8%라 정체처럼 보이지만, 월별로 쪼개면 완전히 다른 그림입니다. 2025년 1월은 −10.1%로 31개월 중 최악이었고, 자동차(−16억$)·광물성연료(−14억$)·전기기기(−6억$)가 동시에 빠졌습니다. 이 달 하나가 상반기 누계를 통째로 깎아먹었습니다.',
  '· 2월부터 회복이 시작됩니다. 2월 +0.4%, 3월 +2.7%, 4월 +3.4%로 완만하게 올라오다가 5월에 −1.3%로 한 번 더 주춤합니다. 5월에는 반도체가 +13.0%였는데도 전체가 마이너스였습니다. 자동차·플라스틱·광물성연료가 동시에 줄었기 때문입니다.',
  '· 하반기에 확실히 방향이 바뀝니다. 9월 +12.6%(전기기기 +23억, 기계 +12억, 자동차 +11억$), 11월 +7.9%, 12월 +13.3%(전기기기 +61억$). 특히 9월은 비반도체도 +11.6%로 이 기간 중 유일하게 두 자릿수를 기록한 달입니다.',
  '· 10월은 선박(+27억$)이 홀로 끌어올린 달입니다. 선박은 인도 시점에 금액이 한꺼번에 잡히므로 이런 달은 추세로 읽으면 안 됩니다. 같은 달 자동차는 −10억$였습니다.',
 ]),
 ('3. 2026년 — 1월에 점프하고 6월에 정점',[
  '· 2026년 1월 전년동월비 +33.9%로 계단을 한 칸 올라섭니다. 전기기기 +105억$, 기계 +26억$가 주도했고, 이 두 품목만으로 그 달 증가액의 78%입니다.',
  '· 2월은 +29.4%로 다소 낮아 보이지만 내용은 더 극단적입니다. 반도체가 만든 몫이 113%로 100%를 넘었습니다. 반도체 외 품목이 순감소(비반도체 −5.2%)했는데 반도체가 그것까지 메우고 전체를 끌어올렸다는 뜻입니다. 이 달 자동차는 −16억$였습니다.',
  '· 3~5월은 +50.4% → +47.8% → +52.9%로 높은 수준을 유지합니다. 이 구간부터 광물성연료가 매달 +13~29억$씩 더해지며 두 번째 축으로 올라옵니다. 2025년 내내 마이너스였던 품목이 방향을 바꾼 시점입니다.',
  '· 6월이 정점입니다. +70.4%, 증가액 +421억$로 31개월 중 최대이며 전기기기 +267억$, 기계 +89억$입니다. 7월은 +63.0%로 여전히 높지만 전월비로는 −2.9%, 조업일수를 보정한 일평균 기준으로는 −9.0%입니다. 월 합계만 보면 실제 둔화 폭을 3분의 1로 과소평가하게 됩니다.',
 ]),
 ('4. 반도체 vs 비반도체 — 19개월 연속 반도체 우위, 그러나 최근 격차는 축소',[
  '· 2025년 1월부터 2026년 7월까지 19개월 내내 반도체(광의)의 전년동월비가 비반도체를 앞섰습니다. 단 한 달의 예외도 없습니다(분석5-B의 「그 달의 우위」 행).',
  '· 2025년에는 반도체 +9.3%, 비반도체 −2.6%였습니다. 그해 수출이 플러스를 지킨 것은 전적으로 반도체 덕분이며, 비반도체만 놓고 보면 역성장한 해였습니다.',
  '· 2026년에는 비반도체도 살아났습니다. 1월 +11.8% → 2월 −5.2%(일시 후퇴) → 3월 +16.7% → 6월 +17.7% → 7월 +17.8%로, 2월을 제외하면 꾸준히 올라옵니다. 반도체가 만든 몫도 2월 113%에서 7월 81%로 낮아졌습니다. 여전히 반도체가 압도적이지만, 온기가 밖으로 번지기 시작한 것은 분명합니다.',
  '· 반도체 비중은 2024년 1월 25.4%에서 2026년 6월 53.7%까지 올라갔다가 7월 51.3%로 소폭 내려왔습니다. 월별 비중 추이는 분석5-A 마지막 행에서 볼 수 있습니다.',
 ]),
 ('5. HS 부문·류 — 어느 달에 무엇이 달라졌나',[
  '· 분석6-A는 21개 부문의 월별 전년동월비를, 분석6-B는 월별 증감액을 담고 있습니다. XVI부(기계·전기전자)는 2025년 3월부터 단 한 달도 빠짐없이 플러스이고, 2026년 들어 매달 +100억$ 이상을 더합니다.',
  '· V부(광물성 생산품)의 전환 시점이 뚜렷합니다. 2025년에는 1·3·4·5·7월 모두 두 자릿수 마이너스였는데, 2026년 3월부터 +29억$, +13억$, +15억$, +16억$, +15억$로 다섯 달 연속 플러스입니다. 2026년 3월이 이 부문의 변곡점입니다.',
  '· XVII부(수송기기)는 반대입니다. 2026년 2월 −16억$, 4월 −6억$, 5월 −5억$로 감소한 달이 더 많습니다. 자동차(HS87)가 2년 연속 감소한 결과이며, 수출 전체가 50% 늘어난 국면에서 최대 소비재 품목이 뒤처졌다는 사실이 이번 호황의 성격을 압축해 보여줍니다.',
  '· 분석7은 97개 류마다 2026년 7개월의 전년동월비와 증감액을 각각 열로 나눠 놓았습니다. 해석 열에는 "월별로는 몇 개월 증가·몇 개월 감소, 최대 몇 %(어느 달), 최소 몇 %(어느 달)"까지 자동으로 계산해 넣었으므로, 같은 증가율이라도 꾸준한 증가인지 한 달에 몰린 증가인지 구분할 수 있습니다.',
  '· 분석7 [요약 2]는 매달 몇 개 류가 늘고 줄었는지를 셉니다. 2026년 2월은 증가 류가 가장 적었고(비반도체 후퇴와 일치), 6월은 가장 많았습니다. 개수 기준 확산과 금액 기준 확산을 함께 보면 "큰 품목만 늘었는지" 여부를 판별할 수 있습니다.',
 ]),
 ('6. 물량인가 가격인가 — 월별로도 답은 같습니다',[
  '· 분석9-A는 19개 달을 각각 3효과로 분해합니다. 2026년 모든 달에서 가격효과가 물량효과를 압도합니다. 2026년 1~7월 합계로는 가격효과 +2,260.5억$, 물량효과 −167.8억$, 교차효과 −95.9억$입니다.',
  '· 2025년은 성격이 다릅니다. 물량효과와 가격효과가 달마다 부호를 바꿔가며 서로 상쇄해 순증가가 거의 나오지 않았습니다. 즉 2025년의 정체는 "물량도 가격도 방향을 못 잡은 해"였고, 2026년의 급증은 "가격이 한 방향으로 몰린 해"입니다.',
  '· 신성질별 중분류로 계산한 피셔 가격지수는 1.514(+51.4%), 물량지수는 0.994(−0.6%)이며 구성(믹스)효과는 +3.8%p입니다. 품목 구성이 바뀌어 평균 단가가 올라 보이는 것이 아니라, 팔던 품목의 값 자체가 오른 것입니다.',
 ]),
 ('7. 이 파일을 쓰는 방법과 한계',[
  '· 매월 갱신 : 원본 5개 시트에 새 달 데이터를 같은 형식으로 붙이고 「데이터_○○」 정제 시트의 참조 범위와 분석4B의 월 열만 늘리면 나머지 분석과 그래프가 자동으로 다시 계산됩니다.',
  '· 반도체 정의 : HS 2단위 자료라 반도체(HS 8541·8542)를 단독으로 뽑을 수 없어 세 가지 대용 정의를 병기했습니다. 정밀한 반도체 분석에는 HS 4단위 이상 자료가 필요합니다.',
  '· 품목별 시트는 98류(특수분류)가 없어 총괄보다 매년 0.03~0.05% 작습니다. 품목 간 비교에는 영향이 없지만 총액 인용 시에는 총괄 시트를 쓰십시오.',
  '· 2026년은 7개월치입니다. 분기 분석에서 2026 Q3는 7월만 포함된 부분분기라 증감률 계산에서 제외했고, 연도 오버레이 그래프에서는 8~12월을 비워 선이 끊기게 했습니다.',
  '· 분석13의 상위 3개 품목은 작성 시점 값 기준으로 고정되어 있습니다(수식 자동 재정렬 아님). 데이터가 크게 바뀌면 순위가 달라질 수 있으니 분석7의 월별 열로 교차 확인하십시오.',
  '· 국가별 자료가 없어 지역·상대국 분석은 포함되지 않았습니다.',
 ]),
]
for t,lines in SECT:
    sec(w12,r,t,9); r+=1
    for ln in lines:
        para(w12,r,ln); r+=1
    r+=1
note(w12,r,1,'작성 : 업로드된 5개 원본 파일만을 근거로 산출했으며 외부 자료를 참조하지 않았습니다.',9)
print('분석12 완료')

# ══════════════════════════════════════════════════════════════════
# 0_안내
# ══════════════════════════════════════════════════════════════════
w0=wb.create_sheet('0_안내',0); w0.sheet_view.showGridLines=False
widths(w0,[26,70,24])
title(w0,1,'수출입 실적 통합 분석 (2024.01~2026.07) — HS 4단위 · 섹터 · 투자 판단판',3)
note(w0,2,1,'원본 9개 파일(2단위 5 + 4단위 4)을 한 파일로 모으고, 원본은 수정하지 않은 채 분석 시트를 추가했습니다. 모든 증감은 월별(전년 동월) 비교이며, 4단위 321개 코드를 40개 투자 섹터로 재편해 스코어보드·스크리너·투자 아이디어까지 제공합니다.',3)
r=4
sec(w0,r,'시트 구성',3); r+=1
for i,h in enumerate(['시트','내용','비고']): head(w0,r,i+1,h)
r+=1
ROWS0=[('수출입 총괄 외 4개','원본 5개 시트 (총괄·품목별·성질별 수출·성질별 수입·신성질별)','원본 · 수정 없음'),
 ('데이터_○○ 5개','텍스트 숫자를 VALUE 수식으로 숫자화 + 연도·월·분기 파생','전부 수식 · 직접 입력 금지'),
 ('분석1_정합성검증','총계 vs 합계, 시트 간 교차, 월별 항등식 검증',''),
 ('분석2_월별추이','31개월 추이 · MoM · YoY · 일평균 · 단가 · 12개월 이동합',''),
 ('분석3_MoM·QoQ·YoY','월별 MoM/YoY, 분기별 QoQ/YoY, 연도별 비교',''),
 ('분석4_연도별계절비교','1~12월 × 2024/2025/2026 매트릭스 17종','오버레이 그래프 원본'),
 ('분석4B_월별매트릭스','HS 97류·부문·성질별 전 항목의 31개월 월별 값','★ 월별 분석의 원천'),
 ('분석5_반도체vs비반도체','월별 금액·YoY·기여도·우위 판정 (정의 3종)','★ 월별 분해'),
 ('분석6_HS부문별','21개 부문의 월별 YoY·월별 증감액 + 구성 류','★ 월별 분해'),
 ('분석7_HS류별상세','97개 류의 월별 YoY 7열·월별 증감액 7열 + 자동 해석','★ 월별 분해'),
 ('분석8_성질별심층','성질별 수출·수입, 신성질별의 월별 YoY','★ 월별 분해'),
 ('분석9_물량·가격분해','19개 달을 각각 3효과로 분해 + 피셔 지수','★ 월별 분해'),
 ('분석10_그래프_연도오버레이','1~12월 축에 3개 연도를 겹친 그래프 17종',''),
 ('분석11_그래프_구조분석','월별 증감액 분해·반도체·부문·분기 그래프 10종',''),
 ('분석12_해설','월별로 본 해석 · 사용법 · 한계','먼저 읽으시면 좋습니다'),
 ('분석13_월별변화요약','한 행이 한 달 — 그 달의 증감 주도 품목 상위 3개와 몫',''),
 ('데이터_HS4수출/수입/중량/수입중량','HS 4단위 321개 × 31개월 피벗 4종 — 수출금액·수입금액·수출중량·수입중량 (원본 셀 직접 참조)','★ 4단위 신규'),
 ('분석14_섹터정의','4단위 321개를 밸류체인 40개 섹터로 재편','★ 투자'),
 ('분석15_섹터월별','섹터별 월별 수출·수입·전년동월비','★ 투자'),
 ('분석16_섹터스코어보드','3M YoY·가속도·국면·성장동력·종합점수·등급','★ 투자 핵심'),
 ('분석17_HS4상세','321개 4단위 전수 + 월별 YoY + 자동 해석','★ 4단위'),
 ('분석18_반도체밸류체인','소자·모듈·저장·장비·검사 단계별 비교와 격차','★ 투자'),
 ('분석19_스크리너','턴어라운드·물량주도·피크아웃·가격의존·회피','★ 투자'),
 ('분석20_마진프록시','정제마진·롤마진·ASP 스프레드 추이','★ 투자'),
 ('분석21_투자아이디어','섹터별 등급·논거·확인지표·리스크','★ 투자'),
 ('분석22_그래프_투자','국면 지도·성장의 질 산점도 등 10종','★ 투자')]
for a,b,c in ROWS0:
    put(w0,r,1,a,bold=True); put(w0,r,2,b); put(w0,r,3,c,align='center'); r+=1
r+=1
sec(w0,r,'이번 판에서 달라진 점',3); r+=1
for t in ['① HS 4단위 321개 코드 반영 — 10개 류(27·29·39·71·72·84·85·87·89·90)를 4단위로 분해',
          '② 밸류체인 관점 40개 섹터로 재편 — 「반도체 vs 비반도체」를 소자·모듈·저장·장비·검사·소재로 세분',
          '③ 투자 판단용 스코어보드 신설 — 3M YoY·가속도·국면 판정·성장동력(가격/물량)·종합점수·등급',
          '④ 스크리너 5종(턴어라운드·물량주도·피크아웃·가격의존·회피)과 마진 프록시, 투자 아이디어 시트',
          '⑤ 마진 프록시에 수입중량 피벗을 신설해 「수출단가 ÷ 수입단가」를 올바르게 계산 (정제마진 2710/2709 등)',
          '',
          '[직전 판에서 반영된 사항 — 월별 분해]',
          '① 모든 증감 분석을 1~7월 누계에서 월별 비교로 전환 (분석5·6·7·8·9)',
          '② 분석4B_월별매트릭스 신설 — HS 97류·21개 부문·성질별 전 항목의 31개월 값을 한 곳에 계산',
          '③ 분석13_월별변화요약 신설 — 한 행이 한 달, 그 달의 증가·감소 상위 3개 품목과 반도체/비반도체 몫',
          '④ 분석7에 2026년 7개월의 월별 YoY·월별 증감액 열 추가, 해석 문장에 "몇 개월 증가·최대/최소는 몇 월" 자동 반영',
          '⑤ 분석9를 19개 달 각각의 3효과 분해로 전환, 분석11에 월별 증감액 분해 그래프 추가',
          '⑥ 누계 수치는 삭제하지 않고 각 시트 뒤쪽에 참고용으로 유지']:
    put(w0,r,1,t,border=False); w0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
r+=1
note(w0,r,1,'※ 모든 분석 셀은 원본을 참조하는 수식입니다. 원본에 다음 달 데이터를 추가하고 정제 시트·분석4B의 범위를 늘리면 분석 전체가 갱신됩니다.',3)
print('0_안내 완료')

exec(open('v4_e.py').read())
