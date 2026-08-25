# -*- coding: utf-8 -*-
# 투자 분석 시트 (분석14~22)
from sectors import SECTORS, SEC_OF
from data4 import H4, M17_4
import openpyxl as _op
_pv=_op.load_workbook('s2b.xlsx',read_only=True)['데이터_HS4수출']
CODES=[]; CNAME={}
for r in range(3,10000):
    v=_pv.cell(r,1).value
    if not v: break
    CODES.append(str(v)); CNAME[str(v)]=_pv.cell(r,2).value
_pv.parent.close()
CROW={c:3+i for i,c in enumerate(CODES)}
M4=MONTHS            # 31개월 (기존과 동일)
PEX="'데이터_HS4수출'"; PIM="'데이터_HS4수입'"; PWT="'데이터_HS4중량'"; PWI="'데이터_HS4수입중량'"
def pcol(k): return L(7+k)
def pref(sheet,code,k): return f'{sheet}!{pcol(k)}{CROW[code]}'
SEC_CODES={sc:[c for c in cd if c in CROW] for sc,nm,g,cd,ds in SECTORS}
SEC_LIVE=[(sc,nm,g,ds) for sc,nm,g,cd,ds in SECTORS if SEC_CODES[sc]]
def ssum(sheet,sc,k): return '+'.join(pref(sheet,c,k) for c in SEC_CODES[sc])
K26=[M4.index(m) for m in M17['2026']]
K25=[M4.index(m) for m in M17['2025']]
K24=[M4.index(m) for m in M17['2024']]
KL3=[len(M4)-3,len(M4)-2,len(M4)-1]          # 최근 3개월 2026.05~07
KL3P=[k-12 for k in KL3]
KL3B=[len(M4)-6,len(M4)-5,len(M4)-4]         # 직전 3개월 2026.02~04
KL3BP=[k-12 for k in KL3B]

# ══════════════════════════════════════════════════════════════════
# 분석14_섹터정의
# ══════════════════════════════════════════════════════════════════
w14=wb.create_sheet('분석14_섹터정의'); w14.sheet_view.showGridLines=False
widths(w14,[9,24,14,46,40,12,12])
title(w14,1,'⑭ 투자 분석용 섹터 정의 — HS 4단위 321개를 40개 섹터로 재편',7)
note(w14,2,1,'2단위(류)는 성격이 다른 품목이 섞여 있어 투자 판단에 쓰기 어렵습니다. 4단위 코드를 밸류체인 관점으로 다시 묶었습니다. 예 : 84류를 「반도체 모듈·부품(8473)」과 「일반기계」로 분리하면 정반대 신호가 드러납니다.',7)
r=4
for i,h in enumerate(['섹터코드','섹터명','그룹','포함 HS 4단위','설명','2026 1~7 수출','전체 비중']): head(w14,r,i+1,h)
r+=1; S14=r
_e26all=tot_sum(2026,7)
for sc,nm,g,ds in SEC_LIVE:
    put(w14,r,1,sc,align='center',bold=True); put(w14,r,2,nm,bold=True); put(w14,r,3,g,align='center')
    put(w14,r,4,', '.join(SEC_CODES[sc]),sz=8,wrap=True)
    put(w14,r,5,ds,sz=9,wrap=True)
    put(w14,r,6,'='+'+'.join(f'({ssum(PEX,sc,k)})' for k in K26),NUM)
    put(w14,r,7,f'=F{r}/{_e26all}',PCT2)
    w14.row_dimensions[r].height=26
    r+=1
E14=r-1
put(w14,r,2,'합계(4단위 수집분)',bold=True)
put(w14,r,6,f'=SUM(F{S14}:F{E14})',NUM,bold=True); put(w14,r,7,f'=F{r}/{_e26all}',PCT2,bold=True)
r+=2
note(w14,r,1,'※ 이번에 4단위를 받은 10개 류(27·29·39·71·72·84·85·87·89·90)만 섹터로 분해됩니다. 나머지 류는 기존 2단위 분석(분석6·7)에서 확인하십시오. '
             '「반도체 소재(화학)」·「철강 강관」 섹터는 해당 4단위(28류·73류)가 아직 수집되지 않아 비어 있습니다.',7)
print('분석14 완료')

# ══════════════════════════════════════════════════════════════════
# 분석15_섹터월별
# ══════════════════════════════════════════════════════════════════
w15=wb.create_sheet('분석15_섹터월별'); w15.sheet_view.showGridLines=False
widths(w15,[9,24]+[12]*len(M4))
title(w15,1,'⑮ 섹터별 월별 수출·수입 및 전년동월비 (2024.01~2026.07)',2+len(M4))
note(w15,2,1,'분석16 이후 모든 투자 지표가 이 표를 참조합니다. [A] 월별 수출금액, [B] 월별 전년동월비, [C] 월별 수입금액입니다.',2+len(M4))
r=4
sec(w15,r,'[A] 섹터별 월별 수출금액 (천달러)',2+len(M4)); r+=1
A15H=r; head(w15,r,1,'코드'); head(w15,r,2,'섹터')
for k in range(len(M4)): head(w15,r,3+k,M4[k])
r+=1; A15=r
for sc,nm,g,ds in SEC_LIVE:
    put(w15,r,1,sc,align='center',sz=9); put(w15,r,2,nm,sz=9,bold=True)
    for k in range(len(M4)): put(w15,r,3+k,'='+ssum(PEX,sc,k),NUM,sz=8)
    r+=1
E15=r-1; r+=1
sec(w15,r,'[B] 섹터별 월별 전년동월비',2+len(M4)); r+=1
B15H=r; head(w15,r,1,'코드'); head(w15,r,2,'섹터')
for k in range(12,len(M4)): head(w15,r,3+k-12,M4[k])
r+=1; B15=r
for i,(sc,nm,g,ds) in enumerate(SEC_LIVE):
    put(w15,r,1,sc,align='center',sz=9); put(w15,r,2,nm,sz=9,bold=True)
    for k in range(12,len(M4)):
        put(w15,r,3+k-12,f'=IFERROR({L(3+k)}{A15+i}/{L(3+k-12)}{A15+i}-1,"-")',PCT,sz=8)
    r+=1
E15B=r-1; r+=1
sec(w15,r,'[C] 섹터별 월별 수입금액 (천달러)',2+len(M4)); r+=1
C15H=r; head(w15,r,1,'코드'); head(w15,r,2,'섹터')
for k in range(len(M4)): head(w15,r,3+k,M4[k])
r+=1; C15=r
for sc,nm,g,ds in SEC_LIVE:
    put(w15,r,1,sc,align='center',sz=9); put(w15,r,2,nm,sz=9,bold=True)
    for k in range(len(M4)): put(w15,r,3+k,'='+ssum(PIM,sc,k),NUM,sz=8)
    r+=1
E15C=r-1
w15.freeze_panes='C6'
print('분석15 완료')

# ══════════════════════════════════════════════════════════════════
# 분석16_섹터스코어보드  ★ 투자 판단 핵심
# ══════════════════════════════════════════════════════════════════
w16=wb.create_sheet('분석16_섹터스코어보드'); w16.sheet_view.showGridLines=False
widths(w16,[8,22,12,13,8,9,9,10,10,9,10,10,10,10,9,13,13,13,10,9,10,44])
title(w16,1,'⑯ 섹터 스코어보드 — 지금 어디에 자금을 넣을 것인가',22)
note(w16,2,1,'최근 3개월(2026.05~07)을 전년 동월과 비교한 값이 「3M YoY」이고, 그 직전 3개월(2026.02~04)과의 차이가 「가속도」입니다. 국면은 이 둘의 부호 조합으로 자동 판정합니다.',22)
note(w16,3,1,'성장동력 : 단가 상승이 주도하면 사이클(되돌림 위험), 물량 증가가 주도하면 실수요(지속성 높음). 같은 성장률이라도 성격이 다르므로 등급 산정에 물량 기여를 가중했습니다.',22)
note(w16,4,1,'※ 이 시트는 무역통계에 근거한 정량 스크리닝 도구입니다. 개별 종목의 실적·밸류에이션·수급은 반영되어 있지 않으므로 종목 선택은 별도 검증이 필요합니다.',22,color=RED)
r=6
H16=['코드','섹터','그룹','2026 수출','비중','2024','2025','2026','25 YoY','26 YoY','3M YoY','직전 3M','가속도','국면',
     '단가Δ','물량Δ','성장동력','증가 월수','변동성','수입 26','수지 Δ','종합','등급']
for i,h in enumerate(H16): head(w16,r,i+1,h)
w16.row_dimensions[r].height=32
r+=1; S16=r
A15S=A15
for i,(sc,nm,g,ds) in enumerate(SEC_LIVE):
    rr=A15+i
    m=lambda ks: '+'.join(f"'분석15_섹터월별'!{L(3+k)}{rr}" for k in ks)
    put(w16,r,1,sc,align='center',bold=True); put(w16,r,2,nm,bold=True); put(w16,r,3,g,align='center',sz=9)
    put(w16,r,4,'='+m(K26),NUM)
    put(w16,r,5,f'=D{r}/{_e26all}',PCT2)
    put(w16,r,6,'='+m(K24),NUM); put(w16,r,7,'='+m(K25),NUM); put(w16,r,8,f'=D{r}',NUM)
    put(w16,r,9,f'=IFERROR(G{r}/F{r}-1,"-")',PCT)
    put(w16,r,10,f'=IFERROR(H{r}/G{r}-1,"-")',PCT)
    put(w16,r,11,f'=IFERROR(({m(KL3)})/({m(KL3P)})-1,"-")',PCT,bold=True)
    put(w16,r,12,f'=IFERROR(({m(KL3B)})/({m(KL3BP)})-1,"-")',PCT)
    put(w16,r,13,f'=IFERROR((K{r}-L{r})*100,"-")',PP,bold=True)
    put(w16,r,14,f'=IF(K{r}="","-",IF(AND(K{r}>0,M{r}>0),"가속 성장",IF(K{r}>0,"성장 둔화",IF(M{r}>0,"바닥 통과","침체 심화"))))',align='center',bold=True)
    wt26='+'.join(f"({ssum(PWT,sc,k)})" for k in K26); wt25='+'.join(f"({ssum(PWT,sc,k)})" for k in K25)
    put(w16,r,15,f'=IFERROR((H{r}/({wt26}))/(G{r}/({wt25}))-1,"-")',PCT)
    put(w16,r,16,f'=IFERROR(({wt26})/({wt25})-1,"-")',PCT)
    put(w16,r,17,f'=IF(OR(O{r}="-",P{r}="-"),"-",IF(O{r}>ABS(P{r})*2,"가격 주도",IF(P{r}>ABS(O{r})*2,"물량 주도","혼합")))',align='center',sz=9)
    put(w16,r,18,f'=COUNTIF(\'분석15_섹터월별\'!{L(3+K26[0]-12)}{B15+i}:{L(3+K26[-1]-12)}{B15+i},">0")',NUM)
    put(w16,r,19,f'=IFERROR(STDEV(\'분석15_섹터월별\'!{L(3+K26[0]-12)}{B15+i}:{L(3+K26[-1]-12)}{B15+i}),"-")',PCT)
    imp='+'.join(f"'분석15_섹터월별'!{L(3+k)}{C15+i}" for k in K26)
    imp25='+'.join(f"'분석15_섹터월별'!{L(3+k)}{C15+i}" for k in K25)
    put(w16,r,20,'='+imp,NUM)
    put(w16,r,21,f'=(H{r}-({imp}))-(G{r}-({imp25}))',NUM)
    put(w16,r,22,
        f'=IFERROR(ROUND(MEDIAN(0,K{r},1)*30'                      # 성장 (0~100% → 0~30)
        f'+MEDIAN(0,M{r}/100+0.3,0.6)/0.6*25'                      # 가속 (-30~+30%p → 0~25)
        f'+IF(Q{r}="물량 주도",20,IF(Q{r}="혼합",12,5))*R{r}/7'      # 지속성 (성장동력 × 증가월수)
        f'+MEDIAN(0,E{r}*100/10,1)*15'                             # 규모 (비중 10% 만점)
        f'+IF(S{r}="-",5,MEDIAN(0,1-S{r}/1.5,1)*10),1),"-")',      # 안정성
        '0.0',bold=True)
    put(w16,r,23,
        f'=IF(V{r}="-","평가 불가",'
        f'IF(AND(N{r}="가속 성장",V{r}>=55),"비중확대",'
        f'IF(AND(N{r}="바닥 통과",M{r}>10),"매집 검토",'
        f'IF(AND(N{r}="가속 성장",V{r}>=40),"중립+",'
        f'IF(N{r}="성장 둔화","비중축소 검토",'
        f'IF(N{r}="침체 심화","회피","중립"))))))',align='center',bold=True)
    r+=1
E16=r-1
w16.freeze_panes='D7'
print('분석16 완료', S16, E16)

# ══════════════════════════════════════════════════════════════════
# 분석17_HS4상세 (321개 전수)
# ══════════════════════════════════════════════════════════════════
w17=wb.create_sheet('분석17_HS4상세'); w17.sheet_view.showGridLines=False
widths(w17,[8,34,9,20,12,12,12,9,9,10,10,11,10,10,10,12,12]+[10]*7+[46])
title(w17,1,'⑰ HS 4단위 전수 분석 — 321개 코드',24+7)
note(w17,2,1,'수집된 10개 류(27·29·39·71·72·84·85·87·89·90)의 4단위 코드를 하나도 빼지 않고 담았습니다. 정렬은 2026년 1~7월 수출 증감액 순(작성 시점 고정).',31)
r=4
H17=['HS4','품목명','2단위','섹터','2024','2025','2026','25 YoY','26 YoY','기여도','비중','3M YoY','가속도','단가Δ','물량Δ','수입 26','수지 26']+[f'{m} YoY' for m in M17['2026']]+['해석']
for i,h in enumerate(H17): head(w17,r,i+1,h)
w17.row_dimensions[r].height=32
r+=1; R17=r
def c26(c): return sum(H4[c][m][1] for m in M17['2026'] if m in H4[c]) if c in H4 else 0
def c25(c): return sum(H4[c][m][1] for m in M17['2025'] if m in H4[c]) if c in H4 else 0
ORD17=sorted(CODES,key=lambda c:-(c26(c)-c25(c)))
for code in ORD17:
    mm=lambda ks,sh=PEX: '+'.join(f'{pref(sh,code,k)}' for k in ks)
    sc=SEC_OF.get(code,'-')
    put(w17,r,1,code,align='center',bold=True).number_format='@'
    put(w17,r,2,CNAME[code],sz=8,wrap=True)
    put(w17,r,3,code[:2],align='center',sz=9).number_format='@'
    put(w17,r,4,dict((s,n) for s,n,g,d in SEC_LIVE).get(sc,'-'),sz=9)
    put(w17,r,5,'='+mm(K24),NUM); put(w17,r,6,'='+mm(K25),NUM); put(w17,r,7,'='+mm(K26),NUM)
    put(w17,r,8,f'=IFERROR(F{r}/E{r}-1,"-")',PCT)
    put(w17,r,9,f'=IFERROR(G{r}/F{r}-1,"-")',PCT)
    put(w17,r,10,f'=(G{r}-F{r})/({_e26all}-{tot_sum(2025,7)})',PCT2)
    put(w17,r,11,f'=G{r}/{_e26all}',PCT2)
    put(w17,r,12,f'=IFERROR(({mm(KL3)})/({mm(KL3P)})-1,"-")',PCT)
    put(w17,r,13,f'=IFERROR((L{r}-(({mm(KL3B)})/({mm(KL3BP)})-1))*100,"-")',PP)
    put(w17,r,14,f'=IFERROR((G{r}/({mm(K26,PWT)}))/(F{r}/({mm(K25,PWT)}))-1,"-")',PCT)
    put(w17,r,15,f'=IFERROR(({mm(K26,PWT)})/({mm(K25,PWT)})-1,"-")',PCT)
    put(w17,r,16,'='+mm(K26,PIM),NUM)
    put(w17,r,17,f'=G{r}-P{r}',NUM)
    for j,k in enumerate(K26):
        put(w17,r,18+j,f'=IFERROR({pref(PEX,code,k)}/{pref(PEX,code,k-12)}-1,"-")',PCT,sz=9)
    put(w17,r,25,
        f'=IF(G{r}=0,"자료 없음",'
        f'IF(I{r}>1,"2배 이상 급증. ",IF(I{r}>0.5,"+50% 초과 급증. ",IF(I{r}>0.15,"뚜렷한 증가. ",'
        f'IF(I{r}>0.02,"소폭 증가. ",IF(I{r}>-0.02,"보합. ",IF(I{r}>-0.15,"소폭 감소. ","뚜렷한 감소. "))))))'
        f'&IF(N{r}="-","",IF(AND(N{r}>ABS(O{r})*2,N{r}>0.05),"단가 상승이 주도(사이클성). ",'
        f'IF(AND(O{r}>ABS(N{r})*2,O{r}>0.05),"물량 확대가 주도(실수요). ","물량·단가 혼재. ")))'
        f'&IF(M{r}="-","",IF(M{r}>5,"최근 3개월 가속. ",IF(M{r}<-5,"최근 3개월 감속. ","최근 3개월 속도 유지. ")))'
        f'&"7개월 중 "&COUNTIF(R{r}:X{r},">0")&"개월 증가. "'
        f'&IF(J{r}>0.03,"전체 증가의 3% 이상을 설명하는 핵심 품목.",IF(J{r}>0.003,"전체 증가에 의미 있게 기여.","총량 영향은 제한적."))',
        sz=8,wrap=True)
    w17.row_dimensions[r].height=30
    r+=1
E17=r-1
put(w17,r,2,'합계',bold=True)
for col in [5,6,7,16,17]: put(w17,r,col,f'=SUM({L(col)}{R17}:{L(col)}{E17})',NUM,bold=True)
put(w17,r,8,f'=F{r}/E{r}-1',PCT,bold=True); put(w17,r,9,f'=G{r}/F{r}-1',PCT,bold=True)
put(w17,r,10,f'=(G{r}-F{r})/({_e26all}-{tot_sum(2025,7)})',PCT2,bold=True)
put(w17,r,11,f'=G{r}/{_e26all}',PCT2,bold=True)
SUM17=r
w17.freeze_panes='E5'
print('분석17 완료', R17, E17)

# ══════════════════════════════════════════════════════════════════
# 분석18_반도체밸류체인
# ══════════════════════════════════════════════════════════════════
w18=wb.create_sheet('분석18_반도체밸류체인'); w18.sheet_view.showGridLines=False
widths(w18,[10,24,13]+[12]*len(M4))
title(w18,1,'⑱ 반도체 밸류체인 — 소자는 폭발했는데 장비는 왜 조용한가',3+len(M4))
note(w18,2,1,'같은 반도체 사이클이라도 밸류체인 단계마다 반응 시점이 다릅니다. 소자(8542)·모듈(8473)·저장장치(8523)가 먼저 뛰고, 통상 장비(8486)와 검사장비(9030·9031)가 뒤따릅니다. 두 그룹의 격차가 곧 투자 타이밍 신호입니다.',3+len(M4))
CHAIN=[('전공정 소자','8542','집적회로 — 메모리·시스템반도체 본체'),
       ('모듈·부품','8473','메모리 모듈 등 자료처리기계 부분품'),
       ('저장장치','8523','SSD·메모리카드'),
       ('개별소자','8541','다이오드·전력반도체'),
       ('제조장비','8486','반도체·FPD 제조장비'),
       ('검사·계측(9030)','9030','오실로스코프·스펙트럼 분석기'),
       ('검사·계측(9031)','9031','그 밖의 측정·검사기기'),
       ('디스플레이 모듈','8524','평판디스플레이 모듈')]
r=4
sec(w18,r,'[A] 단계별 월별 수출금액 (천달러)',3+len(M4)); r+=1
A18H=r; head(w18,r,1,'단계'); head(w18,r,2,'설명'); head(w18,r,3,'HS4')
for k in range(len(M4)): head(w18,r,4+k,M4[k])
r+=1; A18=r
for lab,code,desc in CHAIN:
    put(w18,r,1,lab,bold=True,sz=9); put(w18,r,2,desc,sz=8); put(w18,r,3,code,align='center',sz=9).number_format='@'
    for k in range(len(M4)): put(w18,r,4+k,f'={pref(PEX,code,k)}',NUM,sz=8)
    r+=1
E18=r-1; r+=1
sec(w18,r,'[B] 단계별 월별 전년동월비 — 선행·후행 확인',3+len(M4)); r+=1
B18H=r; head(w18,r,1,'단계'); head(w18,r,2,'설명'); head(w18,r,3,'HS4')
for k in range(12,len(M4)): head(w18,r,4+k-12,M4[k])
r+=1; B18=r
for i,(lab,code,desc) in enumerate(CHAIN):
    put(w18,r,1,lab,bold=True,sz=9); put(w18,r,2,desc,sz=8); put(w18,r,3,code,align='center',sz=9).number_format='@'
    for k in range(12,len(M4)):
        put(w18,r,4+k-12,f'=IFERROR({L(4+k)}{A18+i}/{L(4+k-12)}{A18+i}-1,"-")',PCT,sz=8)
    r+=1
E18B=r-1
put(w18,r,1,'소자−장비 격차(%p)',bold=True); put(w18,r,2,'8542 YoY − 8486 YoY',sz=8)
for k in range(12,len(M4)):
    put(w18,r,4+k-12,f'=IFERROR(({L(4+k-12)}{B18}-{L(4+k-12)}{B18+4})*100,"-")',PP,sz=8)
GAP18=r; r+=1
put(w18,r,1,'소자−검사 격차(%p)',bold=True); put(w18,r,2,'8542 YoY − 9030 YoY',sz=8)
for k in range(12,len(M4)):
    put(w18,r,4+k-12,f'=IFERROR(({L(4+k-12)}{B18}-{L(4+k-12)}{B18+5})*100,"-")',PP,sz=8)
r+=2
sec(w18,r,'[C] 단계별 요약과 투자 시사점',3+len(M4)); r+=1
for i,h in enumerate(['단계','HS4','2025 1~7','2026 1~7','26 YoY','3M YoY','가속도','단가Δ','물량Δ','성장동력','시사점']): head(w18,r,i+1,h)
r+=1; C18=r
for i,(lab,code,desc) in enumerate(CHAIN):
    mm=lambda ks,sh=PEX: '+'.join(pref(sh,code,k) for k in ks)
    put(w18,r,1,lab,bold=True); put(w18,r,2,code,align='center').number_format='@'
    put(w18,r,3,'='+mm(K25),NUM); put(w18,r,4,'='+mm(K26),NUM)
    put(w18,r,5,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(w18,r,6,f'=IFERROR(({mm(KL3)})/({mm(KL3P)})-1,"-")',PCT)
    put(w18,r,7,f'=IFERROR((F{r}-(({mm(KL3B)})/({mm(KL3BP)})-1))*100,"-")',PP)
    put(w18,r,8,f'=IFERROR((D{r}/({mm(K26,PWT)}))/(C{r}/({mm(K25,PWT)}))-1,"-")',PCT)
    put(w18,r,9,f'=IFERROR(({mm(K26,PWT)})/({mm(K25,PWT)})-1,"-")',PCT)
    put(w18,r,10,f'=IF(OR(H{r}="-",I{r}="-"),"-",IF(H{r}>ABS(I{r})*2,"가격 주도",IF(I{r}>ABS(H{r})*2,"물량 주도","혼합")))',align='center',sz=9)
    put(w18,r,11,
        f'=IF(E{r}>1,"이미 폭발 구간 — 신규 진입 시 되돌림 위험 점검",'
        f'IF(AND(E{r}<0.3,G{r}>5,J{r}="물량 주도"),"물량 기반 회복 초기 — 밸류체인 후행 수혜 후보",'
        f'IF(AND(E{r}<0.3,G{r}>0),"완만한 회복 — 소자 사이클 확산 대기",'
        f'IF(E{r}<0,"미반응 — 사이클 수혜 아직 없음","중립"))))',sz=9)
    r+=1
E18C=r-1; r+=1
note(w18,r,1,'읽는 법 : 전공정 소자·모듈·저장장치는 이미 세 자릿수 증가입니다. 반면 제조장비와 검사장비는 10~20%대이고, 특히 검사장비는 단가가 아니라 물량이 늘고 있습니다. '
             '설비 발주가 이제 막 살아나는 국면으로 해석할 수 있으며, 소자−장비 격차([B] 아래 두 행)가 좁혀지기 시작하는 달이 장비 사이클 진입 신호입니다.',3+len(M4))
w18.freeze_panes='D6'
print('분석18 완료')

# ══════════════════════════════════════════════════════════════════
# 분석19_스크리너
# ══════════════════════════════════════════════════════════════════
w19=wb.create_sheet('분석19_스크리너'); w19.sheet_view.showGridLines=False
widths(w19,[8,24,12,12,11,11,11,11,11,11,12,40])
title(w19,1,'⑲ 스크리너 — 지금 사야 할 것, 팔아야 할 것, 지켜볼 것',12)
note(w19,2,1,'분석16 스코어보드를 조건별로 걸러낸 목록입니다. 각 표는 조건에 맞는 섹터만 자동으로 표시되며(조건 불충족 시 공란), 원본이 갱신되면 목록도 바뀝니다.',12)
def screen(ws,r,title_,cond,desc):
    sec(ws,r,title_,12); r+=1
    note(ws,r,1,desc,12); r+=1
    for i,h in enumerate(['코드','섹터','2026 수출','비중','25 YoY','26 YoY','3M YoY','가속도','단가Δ','물량Δ','국면','판정 근거']): head(ws,r,i+1,h)
    r+=1
    for i,(sc,nm,g,ds) in enumerate(SEC_LIVE):
        sr=S16+i
        c=cond.format(r=sr)
        put(ws,r,1,f'=IF({c},\'분석16_섹터스코어보드\'!A{sr},"")',align='center')
        put(ws,r,2,f'=IF({c},\'분석16_섹터스코어보드\'!B{sr},"")')
        put(ws,r,3,f'=IF({c},\'분석16_섹터스코어보드\'!D{sr},"")',NUM)
        put(ws,r,4,f'=IF({c},\'분석16_섹터스코어보드\'!E{sr},"")',PCT2)
        put(ws,r,5,f'=IF({c},\'분석16_섹터스코어보드\'!I{sr},"")',PCT)
        put(ws,r,6,f'=IF({c},\'분석16_섹터스코어보드\'!J{sr},"")',PCT)
        put(ws,r,7,f'=IF({c},\'분석16_섹터스코어보드\'!K{sr},"")',PCT)
        put(ws,r,8,f'=IF({c},\'분석16_섹터스코어보드\'!M{sr},"")',PP)
        put(ws,r,9,f'=IF({c},\'분석16_섹터스코어보드\'!O{sr},"")',PCT)
        put(ws,r,10,f'=IF({c},\'분석16_섹터스코어보드\'!P{sr},"")',PCT)
        put(ws,r,11,f'=IF({c},\'분석16_섹터스코어보드\'!N{sr},"")',align='center')
        put(ws,r,12,f'=IF({c},\'분석16_섹터스코어보드\'!Q{sr}&" · 증가 "&\'분석16_섹터스코어보드\'!R{sr}&"/7개월 · 종합 "&TEXT(\'분석16_섹터스코어보드\'!V{sr},"0.0"),"")',sz=9)
        r+=1
    return r+1
S16C=lambda col,r: f"'분석16_섹터스코어보드'!{col}{r}"
r=4
r=screen(w19,r,'[A] 턴어라운드 — 작년 역성장에서 올해 회복, 최근 가속',
    "AND('분석16_섹터스코어보드'!I{r}<0,'분석16_섹터스코어보드'!K{r}>0,'분석16_섹터스코어보드'!M{r}>0)",
    '2025년에 마이너스였다가 2026년 최근 3개월이 플러스로 돌아섰고 속도까지 붙은 섹터입니다. 기저가 낮아 실적 개선폭이 크게 나타나는 구간입니다.')
r=screen(w19,r,'[B] 물량 주도 성장 — 가격이 아니라 실수요가 끌고 있는 섹터',
    "AND('분석16_섹터스코어보드'!Q{r}=\"물량 주도\",'분석16_섹터스코어보드'!K{r}>0)",
    '단가가 아니라 물량이 늘어 성장한 섹터입니다. 가격 되돌림 위험이 낮아 이익의 지속성이 상대적으로 높습니다.')
r=screen(w19,r,'[C] 피크아웃 경고 — 여전히 플러스지만 속도가 꺾인 섹터',
    "AND('분석16_섹터스코어보드'!K{r}>0,'분석16_섹터스코어보드'!M{r}<-10)",
    '3개월 증가율은 아직 플러스지만 직전 3개월 대비 10%p 이상 감속했습니다. 고점 통과 가능성을 점검해야 합니다.')
r=screen(w19,r,'[D] 가격 의존 경고 — 단가로만 버티는 섹터',
    "AND('분석16_섹터스코어보드'!Q{r}=\"가격 주도\",'분석16_섹터스코어보드'!P{r}<0)",
    '금액은 늘었지만 물량은 오히려 줄어든 섹터입니다. 단가가 되돌아설 때 실적이 빠르게 반락할 수 있습니다.')
r=screen(w19,r,'[E] 회피 — 2년 연속 부진하고 반등 신호도 없는 섹터',
    "AND('분석16_섹터스코어보드'!I{r}<0,'분석16_섹터스코어보드'!K{r}<0,'분석16_섹터스코어보드'!M{r}<0)",
    '2025년에도 2026년에도 마이너스이고 최근 3개월에도 감속 중입니다. 구조적 부진 가능성을 우선 검토해야 합니다.')
w19.freeze_panes='C5'
print('분석19 완료')

# ══════════════════════════════════════════════════════════════════
# 분석20_마진프록시
# ══════════════════════════════════════════════════════════════════
w20=wb.create_sheet('분석20_마진프록시'); w20.sheet_view.showGridLines=False
widths(w20,[26,14]+[12]*len(M4))
title(w20,1,'⑳ 마진 프록시 — 수출단가와 투입 원가의 스프레드',2+len(M4))
note(w20,2,1,'같은 산업 안에서 「팔 때 단가」와 「사올 때 단가」의 비율을 보면 마진 방향을 가늠할 수 있습니다. 무역통계의 톤당 단가를 쓴 근사치이므로 절대 수준이 아니라 방향과 변화폭을 보십시오.',2+len(M4))
MARGIN=[('정유 : 석유제품 수출단가 ÷ 원유 수입단가','2710','2709','수출 2710(석유제품) / 수입 2709(원유) — 정제마진 프록시'),
        ('석유화학 : 합성수지 수출단가 ÷ 석유제품 수입단가','3901','2710','수출 3901(PE) / 수입 2710(나프타 포함 석유제품) — 스프레드 프록시'),
        ('철강 : 도금강판 수출단가 ÷ 스크랩 수입단가','7210','7204','수출 7210(도금강판) / 수입 7204(철스크랩) — 롤마진 프록시'),
        ('반도체 : 집적회로 수출단가','8542',None,'수출 8542 톤당 단가 — 메모리 ASP 프록시'),
        ('반도체 모듈 : 8473 수출단가','8473',None,'수출 8473 톤당 단가 — 모듈 ASP 프록시'),
        ('디스플레이 : 8524 수출단가','8524',None,'수출 8524 톤당 단가'),
        ('자동차 : 승용차 수출단가','8703',None,'수출 8703 톤당 단가 — 대당 ASP 방향 프록시'),
        ('조선 : 상선 수출단가','8901',None,'수출 8901 톤당 단가 — 선가 프록시')]
r=4
sec(w20,r,'[A] 월별 스프레드·단가 추이',2+len(M4)); r+=1
A20H=r; head(w20,r,1,'지표'); head(w20,r,2,'구성')
for k in range(len(M4)): head(w20,r,3+k,M4[k])
r+=1; A20=r
for lab,ex,im,desc in MARGIN:
    put(w20,r,1,lab,bold=True,sz=9); put(w20,r,2,(f'{ex} / {im}' if im else f'{ex}'),align='center',sz=9)
    for k in range(len(M4)):
        if im:
            f=(f'=IFERROR(({pref(PEX,ex,k)}/{pref(PWT,ex,k)})/'
               f'({pref(PIM,im,k)}/{pref(PWI,im,k)}),"-")')
            put(w20,r,3+k,f,'0.000',sz=8)
        else:
            put(w20,r,3+k,f'=IFERROR({pref(PEX,ex,k)}/{pref(PWT,ex,k)},"-")',UNIT,sz=8)
    r+=1
E20=r-1; r+=1
sec(w20,r,'[B] 구간 비교와 방향',2+len(M4)); r+=1
for i,h in enumerate(['지표','구성','2024 1~7 평균','2025 1~7 평균','2026 1~7 평균','25 변화','26 변화','최근 3개월','직전 3개월','방향','해석']): head(w20,r,i+1,h)
r+=1; B20=r
for i,(lab,ex,im,desc) in enumerate(MARGIN):
    a=A20+i
    avg=lambda ks: 'AVERAGE('+','.join(f'{L(3+k)}{a}' for k in ks)+')'
    put(w20,r,1,lab,bold=True,sz=9); put(w20,r,2,desc,sz=8,wrap=True)
    put(w20,r,3,f'=IFERROR({avg(K24)},"-")','0.000')
    put(w20,r,4,f'=IFERROR({avg(K25)},"-")','0.000')
    put(w20,r,5,f'=IFERROR({avg(K26)},"-")','0.000')
    put(w20,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(w20,r,7,f'=IFERROR(E{r}/D{r}-1,"-")',PCT)
    put(w20,r,8,f'=IFERROR({avg(KL3)},"-")','0.000')
    put(w20,r,9,f'=IFERROR({avg(KL3B)},"-")','0.000')
    put(w20,r,10,f'=IF(OR(H{r}="-",I{r}="-"),"-",IF(H{r}>I{r}*1.02,"개선",IF(H{r}<I{r}*0.98,"악화","보합")))',align='center',bold=True)
    put(w20,r,11,f'=IF(J{r}="개선","최근 3개월 스프레드가 직전 3개월보다 확대 — 마진 방향 우호적",'
                 f'IF(J{r}="악화","최근 3개월 스프레드 축소 — 마진 압박 신호","뚜렷한 방향성 없음"))',sz=9)
    w20.row_dimensions[r].height=26
    r+=1
E20B=r-1
w20.freeze_panes='C6'
print('분석20 완료')

# ══════════════════════════════════════════════════════════════════
# 분석21_투자아이디어
# ══════════════════════════════════════════════════════════════════
w21=wb.create_sheet('분석21_투자아이디어'); w21.sheet_view.showGridLines=False
widths(w21,[8,22,11,11,11,10,44,40,34,12])
title(w21,1,'㉑ 투자 아이디어 — 섹터별 등급·논거·확인 지표·리스크',10)
note(w21,2,1,'등급과 수치는 분석16 스코어보드에서 자동으로 가져옵니다. 논거·확인 지표·리스크는 데이터에서 읽히는 사실을 근거로 작성했으며, 개별 종목 판단이 아니라 섹터 배분 우선순위를 정하기 위한 것입니다.',10)
note(w21,3,1,'※ 무역통계 기반 정량 스크리닝입니다. 실제 집행 전에는 개별 기업의 실적·수주·밸류에이션·환율 노출을 반드시 별도 확인하십시오.',10,color=RED)
r=5
for i,h in enumerate(['코드','섹터','2026 수출','3M YoY','가속도','등급','투자 논거 (데이터 근거)','확인해야 할 지표','핵심 리스크','종합점수']): head(w21,r,i+1,h)
w21.row_dimensions[r].height=30
r+=1; R21=r
IDEA={
 'S01':('메모리 사이클의 본체. 3M YoY 세 자릿수에 가속도까지 플러스로, 아직 피크 신호가 없다. 다만 물량은 거의 늘지 않고 단가가 전부를 만들었다는 점이 이 섹터의 성격이자 위험이다.',
        '월별 단가(분석20 8542 ASP 프록시)가 꺾이는 달. 단가가 먼저 꺾이고 금액이 뒤따른다.',
        '단가 되돌림. 물량 방어력이 없어 ASP 하락이 곧바로 실적에 반영된다.'),
 'S02':('84류 급증의 진짜 정체. 모듈·부품이 세 자릿수 후반으로 가장 빠르게 늘고 있으며 가속도도 최상위다. 완제품 컴퓨터(8471)는 미미해 서버·데이터센터향 모듈 수요로 해석된다.',
        '8473 월별 물량. 지금은 단가가 주도하므로 물량이 따라 붙는지가 지속성의 관건.',
        '고객사 재고 조정. 모듈은 소자보다 재고 사이클에 민감하다.'),
 'S03':('SSD·저장매체. 증가율은 밸류체인 전체에서 가장 높고 물량도 함께 늘어 질이 좋다. 규모는 아직 작지만 성장 기여도는 상위권.',
        '물량 증가율이 유지되는지. 물량이 꺾이면 단가만 남는다.',
        '낸드 가격 변동성. 규모가 작아 개별 계약 변동에 크게 흔들린다.'),
 'S05':('제조장비. 소자가 세 자릿수인데 장비는 10%대로, 설비 사이클이 아직 본격화되지 않았다는 신호다. 가속도는 플러스로 돌아섰다.',
        '분석18 [B]의 소자−장비 격차. 격차가 좁혀지기 시작하는 달이 진입 신호.',
        '증설 지연. 소자 호황이 가격 주도면 증설로 이어지지 않을 수 있다.'),
 'S06':('검사·계측장비. 이 섹터의 성장은 단가가 아니라 물량이 만들었다. 밸류체인에서 물량 주도로 늘고 있는 몇 안 되는 구간.',
        '물량 증가율과 9030·9031의 월별 흐름. 후공정 투자 선행 지표로 활용 가능.',
        '규모가 작아 소수 고객 발주에 좌우된다.'),
 'S22':('정유. 2025년 두 자릿수 역성장에서 2026년 큰 폭 반등. 다만 단가 주도이고 물량은 줄어, 유가·정제마진 국면에 전적으로 의존한다.',
        '분석20의 정제마진 프록시(2710/2709). 최근 3개월 방향이 마진의 선행 신호.',
        '유가 반락 시 단가와 마진이 동시에 축소된다.'),
 'S18':('조선 상선. 2년 연속 20%대 증가로 드물게 꾸준하다. 단가 상승이 주도하는데, 이는 과거 고선가 수주분이 인도되는 구조라 향후 몇 년의 실적 가시성이 상대적으로 높다.',
        '월별 인도 편차가 크므로 3개월 이동평균으로 볼 것. 특수선(8905)도 함께 확인.',
        '인도 지연·환율. 단일 월 데이터로 추세를 판단하면 오독한다.'),
 'S15':('완성차. 2년 연속 감소하다 최근 3개월 플러스로 전환했고 가속도가 붙었다. 바닥 통과 국면으로 보이나 아직 증가폭은 미미하다.',
        '8703(승용차) 단가와 물량의 분리. 지금은 단가로 버티는 국면인지 확인 필요.',
        '관세·현지생산 이전이라는 구조적 요인이 남아 있어 회복이 제한될 수 있다.'),
 'S16':('자동차 부품. 완성차보다 부진이 깊고 회복도 느리다. 최근 3개월도 여전히 마이너스이나 감속은 멈췄다.',
        '완성차(S15)와의 시차. 통상 부품이 완성차를 후행한다.',
        '완성차 회복이 현지생산으로 흡수되면 국내 부품 수출은 회복되지 않는다.'),
 'S28':('철강 판재. 2년 연속 감소에 7개월 중 1개월만 증가로 가장 부진한 대형 섹터. 단가·물량 모두 약하다.',
        '분석20 롤마진 프록시(7210/7204)와 봉형강(S29)의 대비.',
        '공급 과잉과 통상 규제. 반등 근거가 데이터에 아직 없다.'),
 'S33':('귀금속 소재. 증가율은 최상위지만 가속도가 크게 꺾였고 단가가 전부를 만들었다. 금·은 가격의 파생 지표로 보는 편이 맞다.',
        '단가 지수. 물량은 오히려 줄고 있어 가격이 유일한 변수.',
        '가격 반락 시 곧바로 역성장. 이미 피크아웃 경고 구간.'),
 'S09':('이차전지. 2025년 역성장에서 반등해 최근 3개월 가속. 규모 대비 회복 각도가 가파르다.',
        '물량 증가율이 유지되는지. 지금은 단가 기여가 더 크다.',
        '전방 수요(전기차) 둔화가 재발하면 회복이 짧게 끝날 수 있다.'),
 'S08':('디스플레이. 2025년 두 자릿수 역성장에서 최근 3개월 플러스로 전환. 다만 물량은 여전히 감소해 회복의 질이 낮다.',
        '8524 물량. 물량이 돌아서지 않으면 단가만의 반등이다.',
        '패널 가격 사이클과 중국 경쟁 심화.'),
 'S10':('통신기기. 최근 3개월 40%대 증가에 가속도도 높다. 단가 주도이나 물량도 플러스로 균형이 나쁘지 않다.',
        '8517 월별 흐름. 규모가 작아 특정 월 수출에 좌우된다.',
        '변동성이 커 추세 확정에 시간이 필요하다.'),
 'S12':('가전. 7개월 내내 감소한 유일한 섹터. 단가·물량 모두 마이너스로 구조적 후퇴로 보인다.',
        '반등 신호가 나타나는지만 확인. 현 시점 데이터에는 없다.',
        '해외 생산 이전이 진행 중이면 수출 지표로는 회복이 잡히지 않는다.'),
}
def idea_row(sc,nm,i):
    global r
    sr=S16+i
    put(w21,r,1,f"='분석16_섹터스코어보드'!A{sr}",align='center',bold=True)
    put(w21,r,2,f"='분석16_섹터스코어보드'!B{sr}",bold=True)
    put(w21,r,3,f"='분석16_섹터스코어보드'!D{sr}",NUM)
    put(w21,r,4,f"='분석16_섹터스코어보드'!K{sr}",PCT)
    put(w21,r,5,f"='분석16_섹터스코어보드'!M{sr}",PP)
    put(w21,r,6,f"='분석16_섹터스코어보드'!W{sr}",align='center',bold=True)
    a,b,c=IDEA[sc]
    put(w21,r,7,a,sz=9,wrap=True); put(w21,r,8,b,sz=9,wrap=True); put(w21,r,9,c,sz=9,wrap=True)
    put(w21,r,10,f"='분석16_섹터스코어보드'!V{sr}",'0.0',bold=True)
    w21.row_dimensions[r].height=58
    r+=1
IDX_OF={sc:i for i,(sc,nm,g,ds) in enumerate(SEC_LIVE)}
for sc in ['S01','S02','S03','S05','S06','S22','S18','S15','S16','S28','S33','S09','S08','S10','S12']:
    if sc in IDX_OF: idea_row(sc,'',IDX_OF[sc])
E21=r-1
r+=1
sec(w21,r,'포트폴리오 관점 정리',10); r+=1
for t in ['· 코어(비중 유지) : 반도체 소자·모듈·저장장치. 증가율과 가속도가 모두 최상위이나 셋 다 단가 주도라 ASP 하나에 연동된다. 세 섹터를 나눠 담아도 분산 효과는 거의 없다는 점을 전제해야 한다.',
          '· 다음 순번 후보 : 반도체 장비·검사장비. 소자 대비 증가율이 현저히 낮은데 가속도는 플러스로 돌아섰고, 검사장비는 물량이 늘고 있다. 소자−장비 격차(분석18)가 좁혀지는 달을 진입 트리거로 삼을 수 있다.',
          '· 턴어라운드 바스켓 : 정유·합성수지·석유화학·이차전지·디스플레이·일반기계. 2025년 역성장에서 최근 3개월 플러스 전환에 가속까지 붙은 조합이다. 다만 대부분 단가 주도라 유가·원자재 가격 국면과 함께 봐야 한다.',
          '· 질이 좋은 성장 : 계측·검사장비, 정밀·기능성 화학, 철강 봉형강·원료, 의료기기. 단가가 아니라 물량이 늘어난 섹터로, 가격 되돌림 위험이 상대적으로 낮다.',
          '· 회피·축소 : 가전(7개월 내내 감소), 철강 판재(7개월 중 1개월만 증가), 귀금속 소재(가속도 −110%p로 피크아웃), 자동차 부품(회복 지연).',
          '· 리스크 관리 : 수출의 절반이 반도체 3개 섹터에서 나온다. 이 파일의 어떤 조합으로 포트폴리오를 짜도 ASP 리스크가 남으므로, 헤지는 섹터 분산이 아니라 물량 주도 섹터 편입으로만 가능하다.']:
    c=put(w21,r,1,t,border=False); c.alignment=Alignment(vertical='top',wrap_text=True)
    w21.merge_cells(start_row=r,start_column=1,end_row=r,end_column=10)
    w21.row_dimensions[r].height=max(16,14*(len(t)//90+1)); r+=1
w21.freeze_panes='C6'
print('분석21 완료')

# ══════════════════════════════════════════════════════════════════
# 분석22_그래프_투자
# ══════════════════════════════════════════════════════════════════
from openpyxl.chart import ScatterChart, Series
w22=wb.create_sheet('분석22_그래프_투자'); w22.sheet_view.showGridLines=False
widths(w22,[26]+[12]*21)
title(w22,1,'㉒ 투자 판단용 그래프',15)
note(w22,2,1,'보조표는 분석16·18·20을 참조하는 수식입니다. 산점도 두 개(3·9번)가 이 시트의 핵심으로, 사분면 위치가 곧 섹터의 국면입니다.',15)
r=4
sec(w22,r,'보조표 A. 섹터 좌표 (3M YoY · 가속도 · 단가Δ · 물량Δ · 규모)',15); r+=1
A22H=r; head(w22,r,1,'섹터'); head(w22,r,2,'3M YoY'); head(w22,r,3,'가속도'); head(w22,r,4,'단가Δ'); head(w22,r,5,'물량Δ'); head(w22,r,6,'2026 수출'); head(w22,r,7,'종합점수')
QUAD=[('가속 성장',8),('바닥 통과',10),('성장 둔화',12),('침체 심화',14)]
DRIV=[('가격 주도',16),('물량 주도',18),('혼합',20)]
for _lab,_c in QUAD: head(w22,r,_c,_lab+' X'); head(w22,r,_c+1,_lab+' Y')
for _lab,_c in DRIV: head(w22,r,_c,_lab+' X'); head(w22,r,_c+1,_lab+' Y')
r+=1; A22=r
for i,(sc,nm,g,ds) in enumerate(SEC_LIVE):
    sr=S16+i
    put(w22,r,1,f"='분석16_섹터스코어보드'!B{sr}",sz=9)
    put(w22,r,2,f"=IFERROR('분석16_섹터스코어보드'!K{sr},\"\")",PCT,sz=9)
    put(w22,r,3,f"=IFERROR('분석16_섹터스코어보드'!M{sr}/100,\"\")",PCT,sz=9)
    put(w22,r,4,f"=IFERROR('분석16_섹터스코어보드'!O{sr},\"\")",PCT,sz=9)
    put(w22,r,5,f"=IFERROR('분석16_섹터스코어보드'!P{sr},\"\")",PCT,sz=9)
    put(w22,r,6,f"='분석16_섹터스코어보드'!D{sr}",NUM,sz=9)
    put(w22,r,7,f"=IFERROR('분석16_섹터스코어보드'!V{sr},\"\")",'0.0',sz=9)
    for _lab,_c in QUAD:
        put(w22,r,_c,  f'=IF(\'분석16_섹터스코어보드\'!N{sr}="{_lab}",B{r},"")',PCT,sz=9)
        put(w22,r,_c+1,f'=IF(\'분석16_섹터스코어보드\'!N{sr}="{_lab}",C{r},"")',PCT,sz=9)
    for _lab,_c in DRIV:
        put(w22,r,_c,  f'=IF(\'분석16_섹터스코어보드\'!Q{sr}="{_lab}",D{r},"")',PCT,sz=9)
        put(w22,r,_c+1,f'=IF(\'분석16_섹터스코어보드\'!Q{sr}="{_lab}",E{r},"")',PCT,sz=9)
    r+=1
E22A=r-1; r+=1
sec(w22,r,'보조표 B. 종합점수 상위 15 섹터',15); r+=1
B22H=r; head(w22,r,1,'섹터'); head(w22,r,2,'종합점수')
r+=1; B22=r
_sc_order=sorted(range(len(SEC_LIVE)),key=lambda i:-(0))  # 정렬은 수식 기준이 아니므로 규모순 사용
_by_size=sorted(range(len(SEC_LIVE)),key=lambda i:-sum(H4[c][m][1] for c in SEC_CODES[SEC_LIVE[i][0]] if c in H4 for m in M17_4['2026'] if m in H4[c]))[:15]
for i in _by_size:
    sr=S16+i
    put(w22,r,1,f"='분석16_섹터스코어보드'!B{sr}",sz=9)
    put(w22,r,2,f"=IFERROR('분석16_섹터스코어보드'!V{sr},0)",'0.0',sz=9)
    r+=1
E22B=r-1
CT22=E22B+3
def add22(ch,anchor,t,w=26,h=12,ylab=None,xlab=None,legend=True,off=0):
    flat(ch,off); ch.title=t; ch.width=w; ch.height=h; ch.style=2
    if ylab: ch.y_axis.title=ylab
    if xlab: ch.x_axis.title=xlab
    if not legend: ch.legend=None
    ch.y_axis.majorGridlines=ChartLines()
    w22.add_chart(ch,anchor)
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
def rowdata(ch,sh,rows,c0,c1,labcol=1):
    """행 단위 계열 추가 : 값은 c0~c1 열에서만, 계열명은 labcol 셀에서 가져온다.
    (라벨 열이 2개 이상인 표에서 titles_from_data를 쓰면 남은 텍스트 열이 데이터로 섞여
     계열이 월 축과 어긋나기 때문에 값 범위와 계열명을 분리한다.)"""
    for rr in rows:
        ch.add_data(Reference(sh,min_col=c0,max_col=c1,min_row=rr,max_row=rr),from_rows=True,titles_from_data=False)
        ch.series[-1].tx=SeriesLabel(strRef=StrRef("'%s'!$%s$%d"%(sh.title,L(labcol),rr)))
# 1. 밸류체인 단계별 월별 YoY
l1=LineChart()
rowdata(l1,w18,range(B18,B18+6),4,3+(len(M4)-12))
l1.set_categories(Reference(w18,min_col=4,max_col=3+(len(M4)-12),min_row=B18H))
add22(l1,f'A{CT22}','1. 반도체 밸류체인 단계별 전년동월비 — 소자·모듈·저장·개별소자·장비·검사',ylab='전년동월비')
# 2. 소자-장비 격차
l2=LineChart()
rowdata(l2,w18,range(GAP18,GAP18+2),4,3+(len(M4)-12))
l2.set_categories(Reference(w18,min_col=4,max_col=3+(len(M4)-12),min_row=B18H))
add22(l2,f'A{CT22+24}','2. 소자 − 장비 / 소자 − 검사 격차 (%p) — 좁혀지면 장비 사이클 진입',ylab='%p',off=2)
# 3. 산점도 : 3M YoY vs 가속도 (사분면)
sc3=ScatterChart(); sc3.x_axis.title='최근 3개월 전년동월비'; sc3.y_axis.title='가속도 (3M − 직전 3M)'
sc3.style=13
QCOL={'가속 성장':'2E75B6','바닥 통과':'70AD47','성장 둔화':'ED7D31','침체 심화':'C00000'}
for _lab,_c in QUAD:
    s=Series(Reference(w22,min_col=_c+1,min_row=A22,max_row=E22A),
             Reference(w22,min_col=_c,  min_row=A22,max_row=E22A),title=_lab)
    s.marker.symbol='circle'; s.marker.size=9
    s.marker.graphicalProperties.solidFill=QCOL[_lab]; s.marker.graphicalProperties.line.solidFill='1F3864'
    s.graphicalProperties.line.noFill=True
    sc3.series.append(s)
sc3.title='3. 섹터 국면 지도 — 오른쪽 위 = 가속 성장, 왼쪽 위 = 바닥 통과'
sc3.width=28; sc3.height=15; sc3.y_axis.majorGridlines=ChartLines()
w22.add_chart(sc3,f'A{CT22+48}')
# 4. 종합점수
b4=BarChart(); b4.type='bar'
b4.add_data(Reference(w22,min_col=2,min_row=B22H,max_row=E22B),titles_from_data=True)
b4.set_categories(Reference(w22,min_col=1,min_row=B22,max_row=E22B))
add22(b4,f'A{CT22+80}','4. 주요 섹터 종합점수 (규모 상위 15)',h=14,ylab='점수',legend=False,off=1)
# 5. 반도체 3형제 월별 수출
l5=LineChart()
rowdata(l5,w18,range(A18,A18+3),4,3+len(M4))
l5.set_categories(Reference(w18,min_col=4,max_col=3+len(M4),min_row=A18H))
add22(l5,f'A{CT22+108}','5. 반도체 3형제 월별 수출금액 (8542 · 8473 · 8523)',ylab='천달러')
# 6. 마진 프록시
l6=LineChart()
rowdata(l6,w20,range(A20,A20+3),3,2+len(M4))
l6.set_categories(Reference(w20,min_col=3,max_col=2+len(M4),min_row=A20H))
add22(l6,f'A{CT22+132}','6. 마진 프록시 — 정유 · 석유화학 · 철강 스프레드',ylab='배수',off=3)
# 7. 반도체 ASP 프록시
l7=LineChart()
rowdata(l7,w20,range(A20+3,A20+5),3,2+len(M4))
l7.set_categories(Reference(w20,min_col=3,max_col=2+len(M4),min_row=A20H))
add22(l7,f'A{CT22+156}','7. 반도체 단가 프록시 (8542 · 8473 톤당 단가)',ylab='천달러/톤',off=1)
# 8. 산점도 : 단가Δ vs 물량Δ (성장의 질)
sc8=ScatterChart(); sc8.x_axis.title='단가 증감률'; sc8.y_axis.title='물량 증감률'
sc8.style=13
DCOL={'가격 주도':'C00000','물량 주도':'548235','혼합':'BF9000'}
for _lab,_c in DRIV:
    s8=Series(Reference(w22,min_col=_c+1,min_row=A22,max_row=E22A),
              Reference(w22,min_col=_c,  min_row=A22,max_row=E22A),title=_lab)
    s8.marker.symbol='circle'; s8.marker.size=9
    s8.marker.graphicalProperties.solidFill=DCOL[_lab]; s8.marker.graphicalProperties.line.solidFill='1F3864'
    s8.graphicalProperties.line.noFill=True
    sc8.series.append(s8)
sc8.title='8. 성장의 질 — 오른쪽 아래 = 가격 주도(사이클), 왼쪽 위 = 물량 주도(실수요)'
sc8.width=28; sc8.height=15; sc8.y_axis.majorGridlines=ChartLines()
w22.add_chart(sc8,f'A{CT22+180}')
# 9. 자동차 완성차 vs 부품
i15=IDX_OF.get('S15'); i16=IDX_OF.get('S16')
l9=LineChart()
l9.add_data(Reference(w15,min_col=2,max_col=2+(len(M4)-12),min_row=B15+i15,max_row=B15+i15),from_rows=True,titles_from_data=True)
l9.add_data(Reference(w15,min_col=2,max_col=2+(len(M4)-12),min_row=B15+i16,max_row=B15+i16),from_rows=True,titles_from_data=True)
l9.set_categories(Reference(w15,min_col=3,max_col=2+(len(M4)-12),min_row=B15H))
add22(l9,f'A{CT22+212}','9. 자동차 완성차 vs 부품 — 월별 전년동월비',ylab='전년동월비',off=4)
# 10. 그룹별 구성
sec(w22,CT22+236,'보조표 C. 그룹별 2026 1~7월 수출 (천달러)',15)
gh=CT22+237; head(w22,gh,1,'그룹'); head(w22,gh,2,'수출')
grows=gh+1
GRPS=[]
for sc,nm,g,ds in SEC_LIVE:
    if g not in GRPS: GRPS.append(g)
for i,g in enumerate(GRPS):
    idxs=[j for j,(s2,n2,g2,d2) in enumerate(SEC_LIVE) if g2==g]
    put(w22,grows+i,1,g,sz=9)
    put(w22,grows+i,2,'='+'+'.join(f"'분석16_섹터스코어보드'!D{S16+j}" for j in idxs),NUM,sz=9)
b10=BarChart(); b10.type='bar'
b10.add_data(Reference(w22,min_col=2,min_row=gh,max_row=grows+len(GRPS)-1),titles_from_data=True)
b10.set_categories(Reference(w22,min_col=1,min_row=grows,max_row=grows+len(GRPS)-1))
add22(b10,f'A{CT22+260}','10. 산업 그룹별 2026년 1~7월 수출 규모',h=11,ylab='천달러',legend=False,off=5)
print('분석22 완료')
