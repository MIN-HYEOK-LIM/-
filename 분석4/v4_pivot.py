# -*- coding: utf-8 -*-
"""HS4 피벗 시트 : 원본 셀을 직접 참조하는 수식으로 코드×월 매트릭스 생성"""
import openpyxl, warnings, collections
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L
warnings.filterwarnings('ignore')
FONT='맑은 고딕'; HDR=PatternFill('solid',fgColor='C55A11')
NUM='#,##0'; NUM1='#,##0.0'
from sectors import SECTORS, SEC_OF
SEC_NAME={sc:nm for sc,nm,g,c,d in SECTORS}
SEC_GRP={sc:g for sc,nm,g,c,d in SECTORS}

wb=openpyxl.load_workbook('s2.xlsx')
SRC=[('HS4_85_87_90','hs4/0c8f703f-__________20260824_858790.xlsx'),
     ('HS4_84','hs4/9944d4ca-__________20260824_84.xlsx'),
     ('HS4_27_29_71','hs4/bad05d35-__________20260824_272971.xlsx'),
     ('HS4_39_72_89','hs4/cb89793f-__________20260824_723989.xlsx')]
# (코드, 월) -> (시트, 행)
loc={}; NAME={}; MONTHS4=set()
for title,f in SRC:
    ws=openpyxl.load_workbook(f).active
    for r in range(7,ws.max_row+1):
        p=(ws.cell(r,1).value or '').strip(); c=(ws.cell(r,2).value or '').strip()
        if not p or not c: continue
        loc[(c,p)]=(title,r); NAME[c]=(ws.cell(r,3).value or '').strip(); MONTHS4.add(p)
MONTHS4=sorted(MONTHS4); CODES=sorted(NAME)
print('4단위 코드',len(CODES),'월',len(MONTHS4))
def V(sheet,col,row): return f'=IFERROR(VALUE(SUBSTITUTE(TRIM(\'{sheet}\'!{col}{row}),",","")),0)'
NOTE=('원본 4단위 시트의 해당 셀을 직접 참조하는 수식입니다(코드×월 피벗). 원본이 바뀌면 이 표와 이후 모든 섹터 분석이 함께 갱신됩니다.')
def pivot(name, srccol, fmt):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    c=ws.cell(1,1,NOTE); c.font=Font(name=FONT,sz=9,i=True,color='595959')
    hdr=['HS4','품목명','2단위','섹터코드','섹터명','그룹']
    for i,h in enumerate(hdr):
        cc=ws.cell(2,i+1,h); cc.font=Font(name=FONT,sz=10,b=True,color='FFFFFF'); cc.fill=HDR
        cc.alignment=Alignment(horizontal='center')
    for k,m in enumerate(MONTHS4):
        cc=ws.cell(2,7+k,m); cc.font=Font(name=FONT,sz=9,b=True,color='FFFFFF'); cc.fill=HDR
        cc.alignment=Alignment(horizontal='center')
    for i,w_ in enumerate([8,30,8,9,20,12]): ws.column_dimensions[L(i+1)].width=w_
    for k in range(len(MONTHS4)): ws.column_dimensions[L(7+k)].width=12
    for i,code in enumerate(CODES):
        r=3+i
        ws.cell(r,1,code).number_format='@'; ws.cell(r,1).font=Font(name=FONT,sz=9)
        ws.cell(r,2,NAME[code]).font=Font(name=FONT,sz=8)
        ws.cell(r,3,code[:2]).number_format='@'; ws.cell(r,3).font=Font(name=FONT,sz=9)
        sc=SEC_OF.get(code,'-')
        ws.cell(r,4,sc).font=Font(name=FONT,sz=9)
        ws.cell(r,5,SEC_NAME.get(sc,'-')).font=Font(name=FONT,sz=9)
        ws.cell(r,6,SEC_GRP.get(sc,'-')).font=Font(name=FONT,sz=9)
        for k,m in enumerate(MONTHS4):
            cell=ws.cell(r,7+k)
            if (code,m) in loc:
                sheet,row=loc[(code,m)]; cell.value=V(sheet,srccol,row)
            else: cell.value=0
            cell.number_format=fmt; cell.font=Font(name=FONT,sz=8)
    ws.freeze_panes='G3'
    return 3, 3+len(CODES)-1
P_EX=pivot('데이터_HS4수출','E',NUM)
P_IM=pivot('데이터_HS4수입','G',NUM)
P_WT=pivot('데이터_HS4중량','D',NUM1)
P_WI=pivot('데이터_HS4수입중량','F',NUM1)
wb.save('s2b.xlsx')
print('피벗 저장', P_EX, len(CODES)*len(MONTHS4)*4, '수식')
