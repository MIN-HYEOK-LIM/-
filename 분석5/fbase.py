# -*- coding: utf-8 -*-
"""공통 기반 : 스타일 헬퍼 · 원본 참조 · 섹터/HS4 좌표"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
warnings.filterwarnings('ignore')
from data3 import MONTHS, TOT, HS, HSNAME, Y, M17
from data4 import H4, H4NAME
from hs_meta import SECTIONS, CH2SEC
from sectors import SECTORS, SEC_OF

FONT='맑은 고딕'; NAVY='1F3864'; BLUE='0000FF'; GREEN='008000'; RED='C00000'
HDR=PatternFill('solid',fgColor='1F3864'); SUB=PatternFill('solid',fgColor='D9E2F3')
HDR2=PatternFill('solid',fgColor='548235'); HDR3=PatternFill('solid',fgColor='C55A11')
IN=PatternFill('solid',fgColor='FFFF00'); WARN=PatternFill('solid',fgColor='FCE4D6')
GOOD=PatternFill('solid',fgColor='E2EFDA'); BAD=PatternFill('solid',fgColor='FBE5E5')
THIN=Side(style='thin',color='BFBFBF'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
NUM='#,##0'; NUM1='#,##0.0'; PCT='0.0%;[RED]-0.0%;-'; PCT2='0.00%'; UNIT='#,##0.000'
PP='+0.00"%p";[RED]-0.00"%p";-'; PP0='+0.0"%p";[RED]-0.0"%p";-'

wb=openpyxl.load_workbook('s2b.xlsx')

# ── 정제 시트 좌표 ──────────────────────────────────────────────
DT="'데이터_총괄'";      T0,T1=3,33
DH="'데이터_품목별'";    H0,H1=3,2984
NM=len(MONTHS)
def trow(k): return T0+k                       # 월 index k → 데이터_총괄 행
def TX(k):   return f"{DT}!H{trow(k)}"         # 수출 금액
def TI(k):   return f"{DT}!K{trow(k)}"         # 수입 금액
def TD(k):   return f"{DT}!E{trow(k)}"         # 조업일수
def TWX(k):  return f"{DT}!G{trow(k)}"         # 수출 중량

# ── HS4 피벗 좌표 ───────────────────────────────────────────────
_pv=openpyxl.load_workbook('s2b.xlsx',read_only=True)['데이터_HS4수출']
CODES=[]; CNAME={}
for r in range(3,10000):
    v=_pv.cell(r,1).value
    if not v: break
    CODES.append(str(v)); CNAME[str(v)]=_pv.cell(r,2).value
_pv.parent.close()
CROW={c:3+i for i,c in enumerate(CODES)}
PEX="'데이터_HS4수출'"; PIM="'데이터_HS4수입'"; PWT="'데이터_HS4중량'"; PWI="'데이터_HS4수입중량'"
def pcol(k): return L(7+k)
def pref(sheet,code,k): return f'{sheet}!{pcol(k)}{CROW[code]}'
def csum(sheet,codes,k):
    codes=[c for c in codes if c in CROW]
    return '+'.join(pref(sheet,c,k) for c in codes) if codes else '0'

SEC_CODES={sc:[c for c in cd if c in CROW] for sc,nm,g,cd,ds in SECTORS}
SEC_LIVE=[(sc,nm,g,ds) for sc,nm,g,cd,ds in SECTORS if SEC_CODES[sc]]
SEC_NAME={sc:nm for sc,nm,g,cd,ds in SECTORS}
IDX_OF={sc:i for i,(sc,nm,g,ds) in enumerate(SEC_LIVE)}
def ssum(sheet,sc,k): return csum(sheet,SEC_CODES[sc],k)
SEMI={'S01','S02','S03','S04'}                 # 반도체 본체 4개 섹터
SEMI_CODES=[c for sc in SEMI for c in SEC_CODES[sc]]

# ── 월 index 집합 ───────────────────────────────────────────────
K24=[MONTHS.index(m) for m in M17['2024']]
K25=[MONTHS.index(m) for m in M17['2025']]
K26=[MONTHS.index(m) for m in M17['2026']]
KY=[k for k in range(12,NM)]                   # YoY 산출 가능한 월 (2025.01~)
KL3=[NM-3,NM-2,NM-1]; KL3P=[k-12 for k in KL3]
KL3B=[NM-6,NM-5,NM-4]; KL3BP=[k-12 for k in KL3B]

# ── 스타일 헬퍼 ─────────────────────────────────────────────────
def title(ws,row,text,span):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=14,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); ws.row_dimensions[row].height=22
def note(ws,row,col,text,span=10,color='595959'):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,i=True,color=color)
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    c.alignment=Alignment(vertical='center',wrap_text=True); return c
def sec(ws,row,text,span=10):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=11,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); c.fill=SUB; return c
def head(ws,row,col,text,fill=None):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,b=True,color='FFFFFF'); c.fill=fill or HDR
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BOX; return c
def put(ws,row,col,v,fmt=None,bold=False,color='000000',align=None,sz=10,border=True,fill=None,wrap=False):
    c=ws.cell(row,col,v); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    if fmt: c.number_format=fmt
    if align or wrap: c.alignment=Alignment(horizontal=align or 'general',vertical='center',wrap_text=wrap)
    if border: c.border=BOX
    if fill: c.fill=fill
    return c
def widths(ws,ws_w):
    for i,w_ in enumerate(ws_w): ws.column_dimensions[L(i+1)].width=w_
def newsheet(name,w):
    s=wb.create_sheet(name); s.sheet_view.showGridLines=False; widths(s,w); return s

PALETTE=['4472C4','ED7D31','548235','FFC000','7030A0','C00000','2E75B6','00B0A0',
         '843C0C','767171','BF9000','375623','9E480E','264478','636363','997300']
def flat(ch,off=0):
    for i,s_ in enumerate(ch.series):
        col=PALETTE[(i+off)%len(PALETTE)]
        if isinstance(ch,BarChart): s_.graphicalProperties.solidFill=col; s_.graphicalProperties.line.solidFill=col
        else:
            s_.graphicalProperties.line.solidFill=col; s_.graphicalProperties.line.width=22000
            s_.smooth=False
def rowdata(ch,sh,rows,c0,c1,labcol=1):
    """행 단위 계열 : 값 범위와 계열명을 분리해 라벨 열이 데이터로 섞이지 않게 한다."""
    for rr in rows:
        ch.add_data(Reference(sh,min_col=c0,max_col=c1,min_row=rr,max_row=rr),from_rows=True,titles_from_data=False)
        ch.series[-1].tx=SeriesLabel(strRef=StrRef("'%s'!$%s$%d"%(sh.title,L(labcol),rr)))
