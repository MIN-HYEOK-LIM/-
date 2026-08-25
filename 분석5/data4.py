# -*- coding: utf-8 -*-
import openpyxl, warnings, collections
warnings.filterwarnings('ignore')
def num(v):
    s=(v or '').strip().replace(',','')
    if s in ('','-'): return 0.0
    try: return float(s)
    except: return 0.0
def txt(v): return (v or '').strip()
F4=['hs4/0c8f703f-__________20260824_858790.xlsx','hs4/9944d4ca-__________20260824_84.xlsx',
    'hs4/bad05d35-__________20260824_272971.xlsx','hs4/cb89793f-__________20260824_723989.xlsx']
H4=collections.defaultdict(dict); H4NAME={}
for f in F4:
    ws=openpyxl.load_workbook(f).active
    for r in range(7,ws.max_row+1):
        p=txt(ws.cell(r,1).value); c=txt(ws.cell(r,2).value)
        if not p or not c: continue
        H4NAME[c]=txt(ws.cell(r,3).value)
        H4[c][p]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value),num(ws.cell(r,6).value),num(ws.cell(r,7).value))
MONTHS4=sorted({m for v in H4.values() for m in v})
Y4=['2024','2025','2026']
M17_4={y:[f'{y}.{i:02d}' for i in range(1,8)] for y in Y4}
