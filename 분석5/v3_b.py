# -*- coding: utf-8 -*-
"""2단계 : 숫자 정제 시트(전부 수식) — 기간/연도/월/분기 파생 포함"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L
warnings.filterwarnings('ignore')
FONT='맑은 고딕'; HDR=PatternFill('solid',fgColor='548235')
NUM='#,##0'; NUM1='#,##0.0'
wb=openpyxl.load_workbook('s1.xlsx')
NOTE=('이 시트의 모든 셀은 원본 시트를 참조하는 수식입니다. 원본은 숫자가 텍스트(쉼표 포함)로 저장되어 있어 '
      'VALUE/SUBSTITUTE로 숫자화했고, 연도·월·분기 열을 파생시켜 MoM/QoQ/YoY 집계가 가능하도록 했습니다. 직접 입력하지 마세요.')
def V(ref): return f'=IFERROR(VALUE(SUBSTITUTE(TRIM({ref}),",","")),"")'
def make(name, src, first, last, cols):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    c=ws.cell(1,1,NOTE); c.font=Font(name=FONT,sz=9,i=True,color='595959')
    for i,(h,sc,kind) in enumerate(cols):
        cc=ws.cell(2,i+1,h); cc.font=Font(name=FONT,sz=10,b=True,color='FFFFFF'); cc.fill=HDR
        cc.alignment=Alignment(horizontal='center')
        ws.column_dimensions[L(i+1)].width = 34 if kind=='text' and i>3 else 12
    r0=3
    for k,sr in enumerate(range(first,last+1)):
        r=r0+k
        for i,(h,sc,kind) in enumerate(cols):
            col=i+1
            if kind=='text':
                cc=ws.cell(r,col,f"=TRIM('{src}'!{sc}{sr}))".replace('))',')')); cc.number_format='@'
            elif kind=='year':  cc=ws.cell(r,col,f'=IFERROR(VALUE(LEFT($A{r},4)),"")'); cc.number_format='0'
            elif kind=='month': cc=ws.cell(r,col,f'=IFERROR(VALUE(RIGHT($A{r},2)),"")'); cc.number_format='0'
            elif kind=='qtr':   cc=ws.cell(r,col,f'=IFERROR(ROUNDUP($C{r}/3,0),"")'); cc.number_format='0'
            elif kind=='num':   cc=ws.cell(r,col,V(f"'{src}'!{sc}{sr}")); cc.number_format=NUM
            elif kind=='num1':  cc=ws.cell(r,col,V(f"'{src}'!{sc}{sr}")); cc.number_format=NUM1
            cc.font=Font(name=FONT,sz=10)
    ws.freeze_panes='A3'
    print(name, r0, r0+(last-first))
    return r0, r0+(last-first)
KEY=[('기간','A','text'),('연도',None,'year'),('월',None,'month'),('분기',None,'qtr')]
R={}
R['총괄']=make('데이터_총괄','수출입 총괄',7,37, KEY+[
    ('조업일수','B','num1'),('수출 건수','C','num'),('수출 중량','D','num1'),('수출 금액','E','num'),
    ('수입 건수','F','num'),('수입 중량','G','num1'),('수입 금액','H','num'),('무역수지','I','num')])
R['품목별']=make('데이터_품목별','수출입 실적(품목별)',7,2988, KEY+[
    ('HS코드','B','text'),('품목명','C','text'),('수출 중량','D','num1'),('수출 금액','E','num'),
    ('수입 중량','F','num1'),('수입 금액','G','num'),('무역수지','H','num')])
R['성질수출']=make('데이터_성질별수출','수출입 실적(성질별_수출)',7,657, KEY+[
    ('성질명','C','text'),('수출 중량','D','num1'),('수출 금액','E','num')])
R['성질수입']=make('데이터_성질별수입','수출입 실적(성질별_수입)',7,631, KEY+[
    ('성질명','C','text'),('수입 중량','D','num1'),('수입 금액','E','num')])
R['신성질']=make('데이터_신성질별','수출입 실적(신성질별)',7,533, KEY+[
    ('신성질명','B','text'),('수출 중량','C','num1'),('수출 금액','D','num'),
    ('수입 중량','E','num1'),('수입 금액','F','num')])
wb.save('s2.xlsx'); print('saved s2', R)
