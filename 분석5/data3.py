# -*- coding: utf-8 -*-
import openpyxl, warnings, collections
warnings.filterwarnings('ignore')
def num(v):
    s=(v or '').strip().replace(',','')
    if s in ('','-'): return 0.0
    try: return float(s)
    except: return 0.0
def txt(v): return (v or '').strip()
F={'총괄':'9997c13b-_______20260824.xlsx','품목별':'f5647f1d-__________20260824.xlsx',
   '성질별_수출':'35650bd7-__________20260824___.xlsx','성질별_수입':'eeda0a5d-__________20260824___.xlsx',
   '신성질별':'61119590-___________20260824.xlsx'}
def sheet(f): return openpyxl.load_workbook(f).active
ws=sheet(F['총괄']); TOT={}; MONTHS=[]
for r in range(7,ws.max_row+1):
    p=txt(ws.cell(r,1).value)
    if not p: continue
    MONTHS.append(p); TOT[p]=[num(ws.cell(r,c).value) for c in range(2,10)]
# 0조업 1수출건수 2수출중량 3수출금액 4수입건수 5수입중량 6수입금액 7수지
ws=sheet(F['품목별']); HS=collections.defaultdict(dict); HSNAME={}
for r in range(7,ws.max_row+1):
    code=txt(ws.cell(r,2).value)
    if not code: continue
    HSNAME[code]=txt(ws.cell(r,3).value)
    HS[code][txt(ws.cell(r,1).value)]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value),
                                       num(ws.cell(r,6).value),num(ws.cell(r,7).value))
# 0수출중량 1수출금액 2수입중량 3수입금액
ws=sheet(F['성질별_수출']); OLDX=collections.defaultdict(dict); OLDX_ITEMS=[]
for r in range(7,ws.max_row+1):
    n=txt(ws.cell(r,3).value)
    if not n: continue
    if n not in OLDX_ITEMS: OLDX_ITEMS.append(n)
    OLDX[n][txt(ws.cell(r,1).value)]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value))
ws=sheet(F['성질별_수입']); OLDM=collections.defaultdict(dict); OLDM_ITEMS=[]
for r in range(7,ws.max_row+1):
    n=txt(ws.cell(r,3).value)
    if not n: continue
    if n not in OLDM_ITEMS: OLDM_ITEMS.append(n)
    OLDM[n][txt(ws.cell(r,1).value)]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value))
ws=sheet(F['신성질별']); NEW=collections.defaultdict(dict); NEW_ITEMS=[]
for r in range(7,ws.max_row+1):
    n=txt(ws.cell(r,2).value)
    if not n: continue
    if n not in NEW_ITEMS: NEW_ITEMS.append(n)
    NEW[n][txt(ws.cell(r,1).value)]=(num(ws.cell(r,3).value),num(ws.cell(r,4).value),
                                     num(ws.cell(r,5).value),num(ws.cell(r,6).value))
Y=['2024','2025','2026']
M24=[m for m in MONTHS if m.startswith('2024')]
M25=[m for m in MONTHS if m.startswith('2025')]
M26=[m for m in MONTHS if m.startswith('2026')]
M17={y:[f'{y}.{i:02d}' for i in range(1,8)] for y in Y}
