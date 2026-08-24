# -*- coding: utf-8 -*-
"""3단계 : 분석 시트"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
warnings.filterwarnings('ignore')
from data3 import MONTHS, TOT, HS, HSNAME, OLDX, OLDM, NEW, OLDX_ITEMS, OLDM_ITEMS, NEW_ITEMS, Y, M17
from hs_meta import SECTIONS, CH2SEC, DESC

FONT='맑은 고딕'; NAVY='1F3864'; BLUE='0000FF'; GREEN='008000'; RED='C00000'
HDR=PatternFill('solid',fgColor='1F3864'); SUB=PatternFill('solid',fgColor='D9E2F3')
IN=PatternFill('solid',fgColor='FFFF00'); WARN=PatternFill('solid',fgColor='FCE4D6')
G24=PatternFill('solid',fgColor='EDEDED'); G26=PatternFill('solid',fgColor='E2EFDA')
THIN=Side(style='thin',color='BFBFBF'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
NUM='#,##0'; NUM1='#,##0.0'; PCT='0.0%;[RED]-0.0%;-'; PCT2='0.00%'; UNIT='#,##0.000'
PP='+0.00"%p";[RED]-0.00"%p";-'
wb=openpyxl.load_workbook('s2.xlsx')

DT="'데이터_총괄'";      T0,T1=3,33
DH="'데이터_품목별'";    H0,H1=3,2984
DX="'데이터_성질별수출'"; X0,X1=3,653
DM="'데이터_성질별수입'"; M0,M1=3,627
DN="'데이터_신성질별'";   N0,N1=3,529
NM=len(MONTHS)
YRS=[2024,2025,2026]
def rg(sh,c,a,b): return f'{sh}!${c}${a}:${c}${b}'
tR=lambda c: rg(DT,c,T0,T1); hR=lambda c: rg(DH,c,H0,H1)
xR=lambda c: rg(DX,c,X0,X1); mR=lambda c: rg(DM,c,M0,M1); nR=lambda c: rg(DN,c,N0,N1)

def title(ws,row,text,span):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=14,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); ws.row_dimensions[row].height=22
def note(ws,row,col,text,span=10,color='595959'):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,i=True,color=color)
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    c.alignment=Alignment(vertical='center',wrap_text=True); return c
def sec(ws,row,text,span=10):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=11,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); c.fill=SUB; return c
def head(ws,row,col,text,fill=None):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,b=True,color='FFFFFF'); c.fill=fill or HDR
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BOX; return c
def put(ws,row,col,v,fmt=None,bold=False,color='000000',align=None,sz=10,border=True,fill=None,wrap=False):
    c=ws.cell(row,col,v); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    if fmt: c.number_format=fmt
    if align or wrap: c.alignment=Alignment(horizontal=align or 'general',vertical='center',wrap_text=wrap)
    if border: c.border=BOX
    if fill: c.fill=fill
    return c
def widths(ws,ws_w):
    for i,w_ in enumerate(ws_w): ws.column_dimensions[L(i+1)].width=w_
def SRCTOT(col): return f'IFERROR(VALUE(SUBSTITUTE(TRIM(\'수출입 총괄\'!{col}6),",","")),"")'

# ══════════════════════════════════════════════════════════════════
# 분석1_정합성검증
# ══════════════════════════════════════════════════════════════════
ws=wb.create_sheet('분석1_정합성검증'); ws.sheet_view.showGridLines=False
widths(ws,[24,18,18,15,20,18,18,15,20,12])
title(ws,1,'① 합계 정합성 검증 — 개별항목의 합 vs 총합 (5개 원본 시트 교차)',10)
note(ws,2,1,'원본은 숫자가 텍스트로 저장되어 있어 「데이터_○○」 정제 시트(수식)를 거쳐 검증합니다. 원본 시트는 수정하지 않았습니다. 기간 : 2024.01~2026.07(31개월).',10)
put(ws,3,1,'허용오차(중량, 톤)',bold=True); put(ws,3,2,1.0,NUM1,color=BLUE,fill=IN)
put(ws,3,3,'허용오차(금액, 천달러)',bold=True); put(ws,3,4,50,NUM,color=BLUE,fill=IN)
put(ws,3,5,'품목별 허용오차',bold=True); put(ws,3,6,500,NUM,color=BLUE,fill=IN)
TW,TA,TB='$B$3','$D$3','$F$3'
r=5
sec(ws,r,'[검증 1] 「수출입 총괄」 총계행 vs 31개월 합계'); r+=1
for i,h in enumerate(['항목','총계행(원본)','월별 합계','차이','판정']): head(ws,r,i+1,h)
r+=1
for nm,sc,dc,fmt,tol in [('조업일수','B','E',NUM1,TW),('수출 건수','C','F',NUM,TA),('수출 중량(톤)','D','G',NUM1,TW),
        ('수출 금액(천$)','E','H',NUM,TA),('수입 건수','F','I',NUM,TA),('수입 중량(톤)','G','J',NUM1,TW),
        ('수입 금액(천$)','H','K',NUM,TA),('무역수지(천$)','I','L',NUM,TA)]:
    put(ws,r,1,nm,bold=True); put(ws,r,2,'='+SRCTOT(sc),fmt)
    put(ws,r,3,f'=SUM({tR(dc)})',fmt); put(ws,r,4,f'=B{r}-C{r}',NUM1)
    put(ws,r,5,f'=IF(ABS(D{r})<={tol},"일치(반올림오차 이내)","불일치")',align='center'); r+=1
r+=1
sec(ws,r,'[검증 2] 무역수지 항등식 및 시트 간 교차검증 (31개월 총계)'); r+=1
for i,h in enumerate(['구분','수출 금액','수입 금액','무역수지','수출 중량','비고']): head(ws,r,i+1,h)
r+=1
X=r
put(ws,r,1,'수출입 총괄',bold=True)
put(ws,r,2,f'=SUM({tR("H")})',NUM); put(ws,r,3,f'=SUM({tR("K")})',NUM)
put(ws,r,4,f'=B{r}-C{r}',NUM); put(ws,r,5,f'=SUM({tR("G")})',NUM1); put(ws,r,6,'기준값'); r+=1
_xm=['1. 식료 및 직접소비재','2. 원료 및 연료','3. 경공업품','4. 중화학 공업품']
_mm=['1. 소비재','2. 원자재','3. 자본재']
_nm=['1.소비재','2.원자재','3.자본재']
put(ws,r,1,'성질별(수출)',bold=True)
put(ws,r,2,'='+'+'.join(f'SUMIFS({xR("G")},{xR("E")},"{k}")' for k in _xm),NUM)
put(ws,r,3,'-',align='center'); put(ws,r,4,'-',align='center')
put(ws,r,5,'='+'+'.join(f'SUMIFS({xR("F")},{xR("E")},"{k}")' for k in _xm),NUM1)
put(ws,r,6,'대분류 4개 합계'); r+=1
put(ws,r,1,'성질별(수입)',bold=True)
put(ws,r,2,'-',align='center')
put(ws,r,3,'='+'+'.join(f'SUMIFS({mR("G")},{mR("E")},"{k}")' for k in _mm),NUM)
put(ws,r,4,'-',align='center'); put(ws,r,5,'-',align='center')
put(ws,r,6,'대분류 3개 합계'); r+=1
put(ws,r,1,'신성질별',bold=True)
put(ws,r,2,'='+'+'.join(f'SUMIFS({nR("G")},{nR("E")},"{k}")' for k in _nm),NUM)
put(ws,r,3,'='+'+'.join(f'SUMIFS({nR("I")},{nR("E")},"{k}")' for k in _nm),NUM)
put(ws,r,4,f'=B{r}-C{r}',NUM)
put(ws,r,5,'='+'+'.join(f'SUMIFS({nR("F")},{nR("E")},"{k}")' for k in _nm),NUM1)
put(ws,r,6,'대분류 3개 합계'); r+=1
put(ws,r,1,'품목별(HS)',bold=True)
put(ws,r,2,f'=SUM({hR("H")})',NUM,color=RED); put(ws,r,3,f'=SUM({hR("J")})',NUM,color=RED)
put(ws,r,4,f'=B{r}-C{r}',NUM); put(ws,r,5,f'=SUM({hR("G")})',NUM1,color=RED)
put(ws,r,6,'HS 01~97·99류만 → 과소',fill=WARN); r+=1
put(ws,r,1,'총괄 대비 최대 차이',bold=True)
put(ws,r,2,f'=MAX(ABS(B{X+1}-B{X}),ABS(B{X+3}-B{X}))',NUM,bold=True)
put(ws,r,3,f'=MAX(ABS(C{X+2}-C{X}),ABS(C{X+3}-C{X}))',NUM,bold=True)
put(ws,r,4,f'=ABS(D{X+3}-D{X})',NUM,bold=True)
put(ws,r,5,f'=MAX(ABS(E{X+1}-E{X}),ABS(E{X+3}-E{X}))',NUM1,bold=True)
put(ws,r,6,'품목별 제외 기준'); r+=1
put(ws,r,1,'판정',bold=True)
put(ws,r,2,f'=IF(B{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,3,f'=IF(C{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,4,f'=IF(D{r-1}<={TA},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,5,f'=IF(E{r-1}<={TW},"일치","불일치")',align='center',bold=True,color=GREEN)
r+=2
sec(ws,r,'[검증 3] 「품목별(HS)」 합 vs 「총괄」 — 연도별 차이'); r+=1
for i,h in enumerate(['연도','품목별 수출 합','총괄 수출금액','차이','차이 비율','품목별 수입 합','총괄 수입금액','차이','차이 비율']): head(ws,r,i+1,h)
r+=1
for y in YRS:
    put(ws,r,1,y,align='center',bold=True)
    put(ws,r,2,f'=SUMIFS({hR("H")},{hR("B")},$A{r})',NUM)
    put(ws,r,3,f'=SUMIFS({tR("H")},{tR("B")},$A{r})',NUM)
    put(ws,r,4,f'=B{r}-C{r}',NUM,color=RED); put(ws,r,5,f'=D{r}/C{r}',PCT2)
    put(ws,r,6,f'=SUMIFS({hR("J")},{hR("B")},$A{r})',NUM)
    put(ws,r,7,f'=SUMIFS({tR("K")},{tR("B")},$A{r})',NUM)
    put(ws,r,8,f'=F{r}-G{r}',NUM,color=RED); put(ws,r,9,f'=H{r}/G{r}',PCT2)
    r+=1
note(ws,r,1,'★ 품목별 시트는 총괄보다 매년 0.03~0.05% 작습니다. 이 시트에 실제로 존재하는 HS류는 01~97류 중 96개(77류는 HS 결번)와 값이 대부분 0인 99류뿐이고, '
            '98류(특수분류 — 여행자 휴대품·별송품·소액물품 등)가 통째로 빠져 있습니다. 품목 간 비교·순위 분석에는 영향이 없으나 총액 인용 시에는 총괄 시트를 쓰십시오.',9,color=RED)
r+=2
sec(ws,r,'[검증 4] 월별 항등식 (수출−수입=수지) 및 성질별 대분류 합 검증'); r+=1
for i,h in enumerate(['기간','수출−수입','원본 수지','차이','성질별(수출) 대분류합','총괄 수출','차이','신성질별 수입 합','총괄 수입','차이']): head(ws,r,i+1,h)
r+=1
V4=r
for k in range(NM):
    dr=T0+k; m=MONTHS[k]
    put(ws,r,1,f'={DT}!A{dr}',align='center')
    put(ws,r,2,f'={DT}!H{dr}-{DT}!K{dr}',NUM); put(ws,r,3,f'={DT}!L{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM)
    put(ws,r,5,'='+'+'.join(f'SUMIFS({xR("G")},{xR("E")},"{k2}",{xR("A")},$A{r})' for k2 in _xm),NUM)
    put(ws,r,6,f'={DT}!H{dr}',NUM); put(ws,r,7,f'=E{r}-F{r}',NUM1)
    put(ws,r,8,'='+'+'.join(f'SUMIFS({nR("I")},{nR("E")},"{k2}",{nR("A")},$A{r})' for k2 in _nm),NUM)
    put(ws,r,9,f'={DT}!K{dr}',NUM); put(ws,r,10,f'=H{r}-I{r}',NUM1)
    r+=1
put(ws,r,1,'최대 절대차이',bold=True)
for col,src in [(4,'D'),(7,'G'),(10,'J')]:
    put(ws,r,col,f'=MAX(MAX({src}{V4}:{src}{r-1}),-MIN({src}{V4}:{src}{r-1}))',NUM1,bold=True)
r+=1
put(ws,r,1,'판정',bold=True)
put(ws,r,4,f'=IF(D{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
put(ws,r,7,f'=IF(G{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
put(ws,r,10,f'=IF(J{r-1}<={TA},"일치","불일치")',align='center',color=GREEN,bold=True)
ws.freeze_panes='A5'
print('분석1 완료', r)

# ══════════════════════════════════════════════════════════════════
# 분석2_월별추이 (31개월)
# ══════════════════════════════════════════════════════════════════
w2=wb.create_sheet('분석2_월별추이'); w2.sheet_view.showGridLines=False
widths(w2,[12,9,12,14,15,15,15,10,10,10,10,13,14,10,12,12,12,15,15,14])
title(w2,1,'② 월별 추이 (2024.01 ~ 2026.07, 31개월)',20)
note(w2,2,1,'「데이터_총괄」 참조. 일평균 = 금액÷조업일수, 단가 = 금액÷중량, 순상품교역조건 = 수출단가÷수입단가, 12개월 이동합 = 최근 12개월 누계(계절성 제거 추세).',20)
H2=['기간','조업일수','수출 건수','수출 중량(톤)','수출 금액(천$)','수입 금액(천$)','무역수지(천$)',
    '수출 전월비','수입 전월비','수출 전년동월비','수입 전년동월비','수지 전년동월증감','일평균 수출(천$)',
    '일평균 전년동월비','수출단가(천$/톤)','수입단가(천$/톤)','순상품교역조건','수출 12개월 이동합','수지 12개월 이동합','연내 누계 수출']
HR2=4
for i,h in enumerate(H2): head(w2,HR2,i+1,h)
w2.row_dimensions[HR2].height=36
D2=HR2+1
for k in range(NM):
    r=D2+k; dr=T0+k
    put(w2,r,1,f'={DT}!A{dr}',align='center',bold=True)
    put(w2,r,2,f'={DT}!E{dr}',NUM1); put(w2,r,3,f'={DT}!F{dr}',NUM)
    put(w2,r,4,f'={DT}!G{dr}',NUM1); put(w2,r,5,f'={DT}!H{dr}',NUM)
    put(w2,r,6,f'={DT}!K{dr}',NUM); put(w2,r,7,f'={DT}!L{dr}',NUM)
    put(w2,r,8,'' if k==0 else f'=E{r}/E{r-1}-1',PCT)
    put(w2,r,9,'' if k==0 else f'=F{r}/F{r-1}-1',PCT)
    put(w2,r,10,'' if k<12 else f'=E{r}/E{r-12}-1',PCT)
    put(w2,r,11,'' if k<12 else f'=F{r}/F{r-12}-1',PCT)
    put(w2,r,12,'' if k<12 else f'=G{r}-G{r-12}',NUM)
    put(w2,r,13,f'=E{r}/B{r}',NUM)
    put(w2,r,14,'' if k<12 else f'=M{r}/M{r-12}-1',PCT)
    put(w2,r,15,f'=E{r}/D{r}',UNIT); put(w2,r,16,f'=F{r}/{DT}!J{dr}',UNIT)
    put(w2,r,17,f'=O{r}/P{r}','0.000')
    put(w2,r,18,'' if k<11 else f'=SUM(E{r-11}:E{r})',NUM)
    put(w2,r,19,'' if k<11 else f'=SUM(G{r-11}:G{r})',NUM)
    put(w2,r,20,f'=SUMIFS({tR("H")},{tR("B")},VALUE(LEFT($A{r},4)),{tR("C")},"<="&VALUE(RIGHT($A{r},2)))',NUM)
E2=D2+NM-1
r=E2+2
SR={}
for lbl,cond in [('2024년 연간',2024),('2025년 연간',2025),('2026년 1~7월',2026),
                 ('2024년 1~7월',(2024,7)),('2025년 1~7월',(2025,7))]:
    put(w2,r,1,lbl,bold=True,align='center'); SR[lbl]=r
    if isinstance(cond,tuple):
        y,mm=cond
        f=lambda c: f'=SUMIFS({tR(c)},{tR("B")},{y},{tR("C")},"<={mm}")'
    else:
        f=lambda c: f'=SUMIFS({tR(c)},{tR("B")},{cond})'
    for col,src,fmt in [(2,'E',NUM1),(3,'F',NUM),(4,'G',NUM1),(5,'H',NUM),(6,'K',NUM),(7,'L',NUM)]:
        put(w2,r,col,f(src),fmt,bold=True)
    put(w2,r,13,f'=E{r}/B{r}',NUM,bold=True)
    put(w2,r,15,f'=E{r}/D{r}',UNIT,bold=True)
    put(w2,r,17,f'=O{r}/P{r}','0.000',bold=True)
    r+=1
put(w2,r,1,'25→26 동기비(1~7월)',bold=True,align='center')
for col in [2,3,4,5,6,13,15]:
    put(w2,r,col,f'={L(col)}{SR["2026년 1~7월"]}/{L(col)}{SR["2025년 1~7월"]}-1',PCT,bold=True)
put(w2,r,7,f'=G{SR["2026년 1~7월"]}-G{SR["2025년 1~7월"]}',NUM,bold=True)
r+=1
put(w2,r,1,'24→25 동기비(1~7월)',bold=True,align='center')
for col in [2,3,4,5,6,13,15]:
    put(w2,r,col,f'={L(col)}{SR["2025년 1~7월"]}/{L(col)}{SR["2024년 1~7월"]}-1',PCT,bold=True)
put(w2,r,7,f'=G{SR["2025년 1~7월"]}-G{SR["2024년 1~7월"]}',NUM,bold=True)
# 수입단가 요약행 보정
for lbl in SR:
    rr=SR[lbl]
    if lbl.endswith('연간'):
        y=int(lbl[:4]); put(w2,rr,16,f'=F{rr}/SUMIFS({tR("J")},{tR("B")},{y})',UNIT,bold=True)
    else:
        y=int(lbl[:4]); put(w2,rr,16,f'=F{rr}/SUMIFS({tR("J")},{tR("B")},{y},{tR("C")},"<=7")',UNIT,bold=True)
w2.freeze_panes=f'B{D2}'
print('분석2 완료')

# ══════════════════════════════════════════════════════════════════
# 분석3_MoM·QoQ·YoY
# ══════════════════════════════════════════════════════════════════
w3=wb.create_sheet('분석3_MoM·QoQ·YoY'); w3.sheet_view.showGridLines=False
widths(w3,[13,15,11,11,11,11,15,11,11,15,11,11,13,13])
title(w3,1,'③ MoM · QoQ · YoY 분석',14)
note(w3,2,1,'MoM(전월비)은 조업일수 차이에 취약하므로 일평균 기준 MoM을 함께 제시했습니다. QoQ는 분기 대비, YoY는 전년 동기 대비입니다. 2026년은 7월까지만 있어 3분기는 7월 1개월만 반영된 부분분기입니다.',14)
r=4
sec(w3,r,'[A] 월별 MoM · YoY (2024.01~2026.07)',14); r+=1
for i,h in enumerate(['기간','수출 금액','MoM','일평균 MoM','YoY','YoY 3개월이평','수입 금액','MoM','YoY','무역수지','수지 MoM증감','수지 YoY증감']): head(w3,r,i+1,h)
w3.row_dimensions[r].height=30
r+=1
A3=r
for k in range(NM):
    rr=A3+k; d2=D2+k
    put(w3,rr,1,f"='분석2_월별추이'!A{d2}",align='center',bold=True)
    put(w3,rr,2,f"='분석2_월별추이'!E{d2}",NUM)
    put(w3,rr,3,'' if k==0 else f'=B{rr}/B{rr-1}-1',PCT)
    put(w3,rr,4,'' if k==0 else f"='분석2_월별추이'!M{d2}/'분석2_월별추이'!M{d2-1}-1",PCT)
    put(w3,rr,5,'' if k<12 else f'=B{rr}/B{rr-12}-1',PCT)
    put(w3,rr,6,'' if k<14 else f'=SUM(B{rr-2}:B{rr})/SUM(B{rr-14}:B{rr-12})-1',PCT)
    put(w3,rr,7,f"='분석2_월별추이'!F{d2}",NUM)
    put(w3,rr,8,'' if k==0 else f'=G{rr}/G{rr-1}-1',PCT)
    put(w3,rr,9,'' if k<12 else f'=G{rr}/G{rr-12}-1',PCT)
    put(w3,rr,10,f"='분석2_월별추이'!G{d2}",NUM)
    put(w3,rr,11,'' if k==0 else f'=J{rr}-J{rr-1}',NUM)
    put(w3,rr,12,'' if k<12 else f'=J{rr}-J{rr-12}',NUM)
E3=A3+NM-1
r=E3+2
sec(w3,r,'[B] 분기별 QoQ · YoY',14); r+=1
for i,h in enumerate(['분기','수출 금액','QoQ','YoY','수입 금액','QoQ','YoY','무역수지','QoQ 증감','YoY 증감','조업일수','일평균 수출','비고']): head(w3,r,i+1,h)
r+=1
Q3=r
QL=[(y,q) for y in YRS for q in range(1,5)]
QL=[(y,q) for (y,q) in QL if not (y==2026 and q==4)]
for i,(y,q) in enumerate(QL):
    rr=Q3+i
    part = (y==2026 and q==3)
    put(w3,rr,1,f'{y} Q{q}',align='center',bold=True,fill=G26 if y==2026 else (G24 if y==2024 else None))
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y},{tR("D")},{q})',NUM)
    put(w3,rr,5,f'=SUMIFS({tR("K")},{tR("B")},{y},{tR("D")},{q})',NUM)
    put(w3,rr,8,f'=B{rr}-E{rr}',NUM)
    put(w3,rr,11,f'=SUMIFS({tR("E")},{tR("B")},{y},{tR("D")},{q})',NUM1)
    put(w3,rr,12,f'=IFERROR(B{rr}/K{rr},"-")',NUM)
    if i>0 and not part and not (y==2026 and q==3):
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,6,f'=E{rr}/E{rr-1}-1',PCT)
        put(w3,rr,9,f'=H{rr}-H{rr-1}',NUM)
    if i>=4 and not part:
        put(w3,rr,4,f'=B{rr}/B{rr-4}-1',PCT); put(w3,rr,7,f'=E{rr}/E{rr-4}-1',PCT)
        put(w3,rr,10,f'=H{rr}-H{rr-4}',NUM)
    put(w3,rr,13,'7월만 포함된 부분분기 — 증감률 계산 제외' if part else '',sz=9)
E3Q=Q3+len(QL)-1
r=E3Q+2
sec(w3,r,'[C] 연도별 비교 (연간 및 1~7월 동기)',14); r+=1
for i,h in enumerate(['구분','수출 금액','YoY','수입 금액','YoY','무역수지','YoY 증감','수출 중량','YoY','수출단가','YoY']): head(w3,r,i+1,h)
r+=1
C3=r
for i,y in enumerate(YRS):
    rr=C3+i
    put(w3,rr,1,f'{y}년 연간'+(' (1~7월만)' if y==2026 else ''),bold=True,fill=G26 if y==2026 else None)
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y})',NUM)
    put(w3,rr,4,f'=SUMIFS({tR("K")},{tR("B")},{y})',NUM)
    put(w3,rr,6,f'=B{rr}-D{rr}',NUM)
    put(w3,rr,8,f'=SUMIFS({tR("G")},{tR("B")},{y})',NUM1)
    put(w3,rr,10,f'=B{rr}/H{rr}',UNIT)
    if i>0 and y!=2026:
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,5,f'=D{rr}/D{rr-1}-1',PCT)
        put(w3,rr,7,f'=F{rr}-F{rr-1}',NUM); put(w3,rr,9,f'=H{rr}/H{rr-1}-1',PCT)
        put(w3,rr,11,f'=J{rr}/J{rr-1}-1',PCT)
r=C3+3
for i,y in enumerate(YRS):
    rr=r+i
    put(w3,rr,1,f'{y}년 1~7월',bold=True,fill=G26 if y==2026 else None)
    put(w3,rr,2,f'=SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},"<=7")',NUM)
    put(w3,rr,4,f'=SUMIFS({tR("K")},{tR("B")},{y},{tR("C")},"<=7")',NUM)
    put(w3,rr,6,f'=B{rr}-D{rr}',NUM)
    put(w3,rr,8,f'=SUMIFS({tR("G")},{tR("B")},{y},{tR("C")},"<=7")',NUM1)
    put(w3,rr,10,f'=B{rr}/H{rr}',UNIT)
    if i>0:
        put(w3,rr,3,f'=B{rr}/B{rr-1}-1',PCT); put(w3,rr,5,f'=D{rr}/D{rr-1}-1',PCT)
        put(w3,rr,7,f'=F{rr}-F{rr-1}',NUM); put(w3,rr,9,f'=H{rr}/H{rr-1}-1',PCT)
        put(w3,rr,11,f'=J{rr}/J{rr-1}-1',PCT)
Y17=r
w3.freeze_panes='B6'
print('분석3 완료')

# ══════════════════════════════════════════════════════════════════
# 분석4_연도별계절비교 (1~12월 × 2024/2025/2026)
# ══════════════════════════════════════════════════════════════════
w4=wb.create_sheet('분석4_연도별계절비교'); w4.sheet_view.showGridLines=False
widths(w4,[10,15,15,15,11,11,13,11,11])
title(w4,1,'④ 연도별 계절 비교 — 1~12월을 한 축에 놓고 2024·2025·2026을 겹쳐 본다',9)
note(w4,2,1,'각 블록은 행 = 1~12월, 열 = 2024·2025·2026 구조입니다. 이 표가 분석10의 오버레이 그래프 원본이며, 같은 달끼리 3개 연도를 비교하므로 계절성이 자동으로 통제됩니다.',9)
note(w4,3,1,'2026년은 7월까지만 자료가 있어 8~12월은 수식 없이 비워 두었습니다(그래프에서 선이 끊깁니다). 새 달 자료가 추가되면 바로 윗 셀의 수식을 복사해 내려 채우십시오. YoY 열은 해당 연도와 직전 연도 같은 달의 증감률입니다.',9)
SEASON={}
def season_block(ws,r,name,unit,fx,fmt=NUM,ratio=False):
    sec(ws,r,f'{name} ({unit})',9); r+=1
    hr=r
    head(ws,r,1,'월')
    for i,y in enumerate(YRS): head(ws,r,2+i,f'{y}년',fill=PatternFill('solid',fgColor='7F7F7F') if y==2024 else (PatternFill('solid',fgColor='2E75B6') if y==2025 else PatternFill('solid',fgColor='548235')))
    head(ws,r,5,'2025 YoY'); head(ws,r,6,'2026 YoY'); head(ws,r,7,'26 vs 24'); head(ws,r,8,'25 비중'); head(ws,r,9,'26 비중')
    r+=1; r0=r
    for m in range(1,13):
        put(ws,r,1,f'{m}월',align='center',bold=True)
        for i,y in enumerate(YRS):
            if f'{y}.{m:02d}' in MONTHS: put(ws,r,2+i,fx(y,m),fmt)
            else: put(ws,r,2+i,None,fmt)
        put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT)
        put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"")',PCT)
        put(ws,r,7,f'=IFERROR(D{r}/B{r}-1,"")',PCT)
        put(ws,r,8,f'=IFERROR(C{r}/SUM(C${r0}:C${r0+11}),"")',PCT2)
        put(ws,r,9,f'=IFERROR(D{r}/SUM(D${r0}:D${r0+11}),"")',PCT2)
        r+=1
    put(ws,r,1,'연간 합',bold=True)
    for i in range(3): put(ws,r,2+i,f'=SUM({L(2+i)}{r0}:{L(2+i)}{r0+11})',fmt,bold=True)
    put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT,bold=True); put(ws,r,6,'',PCT)
    r+=1
    put(ws,r,1,'1~7월 합',bold=True)
    for i in range(3): put(ws,r,2+i,f'=SUM({L(2+i)}{r0}:{L(2+i)}{r0+6})',fmt,bold=True)
    put(ws,r,5,f'=IFERROR(C{r}/B{r}-1,"")',PCT,bold=True)
    put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"")',PCT,bold=True)
    put(ws,r,7,f'=IFERROR(D{r}/B{r}-1,"")',PCT,bold=True)
    SEASON[name]=(hr,r0,r0+11,r-1,r)
    return r+2

def F_TOT(col): return lambda y,m: f'=SUMIFS({tR(col)},{tR("B")},{y},{tR("C")},{m})'
def F_NEW(item,col='G'): return lambda y,m: f'=SUMIFS({nR(col)},{nR("E")},"{item}",{nR("B")},{y},{nR("C")},{m})'
def F_HS(code,col='H'): return lambda y,m: f'=SUMIFS({hR(col)},{hR("E")},"{code}",{hR("B")},{y},{hR("C")},{m})'
def F_X(item,col='G'): return lambda y,m: f'=SUMIFS({xR(col)},{xR("E")},"{item}",{xR("B")},{y},{xR("C")},{m})'
def F_M(item,col='G'): return lambda y,m: f'=SUMIFS({mR(col)},{mR("E")},"{item}",{mR("B")},{y},{mR("C")},{m})'
def F_NONIT(y,m):
    tot=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'
    it='+'.join(f'SUMIFS({nR("G")},{nR("E")},"{i}",{nR("B")},{y},{nR("C")},{m})' for i in ['라.IT부품','다.IT제품'])
    return f'=IF({tot}=0,"",{tot}-({it}))'
def F_UNIT(y,m):
    a=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'; b=f'SUMIFS({tR("G")},{tR("B")},{y},{tR("C")},{m})'
    return f'=IFERROR({a}/{b},"")'
def F_DPA(y,m):
    a=f'SUMIFS({tR("H")},{tR("B")},{y},{tR("C")},{m})'; b=f'SUMIFS({tR("E")},{tR("B")},{y},{tR("C")},{m})'
    return f'=IFERROR({a}/{b},"")'
r=5
r=season_block(w4,r,'수출 금액','천달러',F_TOT('H'))
r=season_block(w4,r,'수입 금액','천달러',F_TOT('K'))
r=season_block(w4,r,'무역수지','천달러',F_TOT('L'))
r=season_block(w4,r,'수출 중량','톤',F_TOT('G'),NUM1)
r=season_block(w4,r,'수출 단가','천달러/톤',F_UNIT,UNIT)
r=season_block(w4,r,'일평균 수출','천달러/일',F_DPA)
r=season_block(w4,r,'IT부품 수출','천달러',F_NEW('라.IT부품'))
r=season_block(w4,r,'IT제품 수출','천달러',F_NEW('다.IT제품'))
r=season_block(w4,r,'IT 제외 수출','천달러',F_NONIT)
r=season_block(w4,r,'HS85 전기기기 수출','천달러',F_HS('85'))
r=season_block(w4,r,'HS84 기계류 수출','천달러',F_HS('84'))
r=season_block(w4,r,'HS87 차량 수출','천달러',F_HS('87'))
r=season_block(w4,r,'HS27 광물성연료 수출','천달러',F_HS('27'))
r=season_block(w4,r,'HS89 선박 수출','천달러',F_HS('89'))
r=season_block(w4,r,'중화학공업품 수출','천달러',F_X('4. 중화학 공업품'))
r=season_block(w4,r,'수입 원자재','천달러',F_M('2. 원자재'))
r=season_block(w4,r,'수입 자본재','천달러',F_M('3. 자본재'))
w4.freeze_panes='B6'
print('분석4 완료 블록수',len(SEASON))

# ══════════════════════════════════════════════════════════════════
# 분석5_반도체vs비반도체
# ══════════════════════════════════════════════════════════════════
w5=wb.create_sheet('분석5_반도체vs비반도체'); w5.sheet_view.showGridLines=False
widths(w5,[22,15,15,15,11,11,11,11,13,13,13,11])
title(w5,1,'⑤ 반도체 vs 비반도체 — 어느 쪽이 좋았나',12)
note(w5,2,1,'원본 자료가 HS 2단위까지만 있어 반도체(HS 8541·8542)를 단독으로 뽑을 수 없습니다. 그래서 서로 다른 세 가지 대용 정의를 나란히 계산해 어느 정의를 쓰더라도 결론이 흔들리지 않는지 확인할 수 있게 했습니다.',12)
r=4
sec(w5,r,'[정의] 세 가지 대용 지표',12); r+=1
for i,h in enumerate(['정의','구성','포함되는 것','제외되는 것','성격']): head(w5,r,i+1,h)
r+=1
for a,b,cc,d,e in [
 ('① 협의 : IT부품','신성질별 「라.IT부품」','메모리·시스템반도체·집적회로 등 부품','완제품(스마트폰·PC), 디스플레이 패널 일부','반도체에 가장 가까운 대용치'),
 ('② 광의 : IT부품+IT제품','신성질별 「라.IT부품」+「다.IT제품」','반도체 부품 + IT 완제품(휴대폰·컴퓨터 등)','비IT 전자부품','IT 산업 전체'),
 ('③ HS85 기준','품목별 HS 85류','반도체·집적회로 + 디스플레이·통신기기·가전·전선·이차전지','반도체 제조장비(84류), 검사장비(90류)','전기·전자 전반 (가장 넓음)')]:
    put(w5,r,1,a,bold=True); put(w5,r,2,b); put(w5,r,3,cc,wrap=True); put(w5,r,4,d,wrap=True); put(w5,r,5,e)
    w5.merge_cells(start_row=r,start_column=5,end_row=r,end_column=12); r+=1
r+=1
DEFS=[('① IT부품(협의)',['라.IT부품']),('② IT부품+IT제품(광의)',['라.IT부품','다.IT제품'])]
def it_sum(keys,y,mmax=None,col='G'):
    cond=f',{nR("C")},"<={mmax}"' if mmax else ''
    return '+'.join(f'SUMIFS({nR(col)},{nR("E")},"{k}",{nR("B")},{y}{cond})' for k in keys)
def hs_sum(code,y,mmax=None,col='H'):
    cond=f',{hR("C")},"<={mmax}"' if mmax else ''
    return f'SUMIFS({hR(col)},{hR("E")},"{code}",{hR("B")},{y}{cond})'
def tot_sum(y,mmax=None,col='H'):
    cond=f',{tR("C")},"<={mmax}"' if mmax else ''
    return f'SUMIFS({tR(col)},{tR("B")},{y}{cond})'
sec(w5,r,'[A] 정의별 규모·비중·증감 (1~7월 동기 기준, 천달러)',12); r+=1
for i,h in enumerate(['구분','2024','2025','2026','25 YoY','26 YoY','24→26','26 비중','25 비중','24 비중','증가 기여도']): head(w5,r,i+1,h)
r+=1
A5=r
for nm,keys in DEFS:
    put(w5,r,1,nm+' 수출',bold=True)
    for i,y in enumerate(YRS): put(w5,r,2+i,'='+it_sum(keys,y,7),NUM)
    put(w5,r,5,f'=C{r}/B{r}-1',PCT); put(w5,r,6,f'=D{r}/C{r}-1',PCT); put(w5,r,7,f'=D{r}/B{r}-1',PCT)
    put(w5,r,8,f'=D{r}/'+tot_sum(2026,7),PCT2); put(w5,r,9,f'=C{r}/'+tot_sum(2025,7),PCT2)
    put(w5,r,10,f'=B{r}/'+tot_sum(2024,7),PCT2)
    put(w5,r,11,f'=(D{r}-C{r})/({tot_sum(2026,7)}-{tot_sum(2025,7)})',PCT2)
    r+=1
    put(w5,r,1,nm+' 제외 수출')
    for i,y in enumerate(YRS): put(w5,r,2+i,f'={tot_sum(y,7)}-{L(2+i)}{r-1}',NUM)
    put(w5,r,5,f'=C{r}/B{r}-1',PCT); put(w5,r,6,f'=D{r}/C{r}-1',PCT); put(w5,r,7,f'=D{r}/B{r}-1',PCT)
    put(w5,r,8,f'=D{r}/'+tot_sum(2026,7),PCT2); put(w5,r,9,f'=C{r}/'+tot_sum(2025,7),PCT2)
    put(w5,r,10,f'=B{r}/'+tot_sum(2024,7),PCT2)
    put(w5,r,11,f'=(D{r}-C{r})/({tot_sum(2026,7)}-{tot_sum(2025,7)})',PCT2)
    r+=1
put(w5,r,1,'③ HS85 수출',bold=True)
for i,y in enumerate(YRS): put(w5,r,2+i,'='+hs_sum('85',y,7),NUM)
put(w5,r,5,f'=C{r}/B{r}-1',PCT); put(w5,r,6,f'=D{r}/C{r}-1',PCT); put(w5,r,7,f'=D{r}/B{r}-1',PCT)
put(w5,r,8,f'=D{r}/'+tot_sum(2026,7),PCT2); put(w5,r,9,f'=C{r}/'+tot_sum(2025,7),PCT2)
put(w5,r,10,f'=B{r}/'+tot_sum(2024,7),PCT2)
put(w5,r,11,f'=(D{r}-C{r})/({tot_sum(2026,7)}-{tot_sum(2025,7)})',PCT2)
r+=1
put(w5,r,1,'③ HS85 제외 수출')
for i,y in enumerate(YRS): put(w5,r,2+i,f'={tot_sum(y,7)}-{L(2+i)}{r-1}',NUM)
put(w5,r,5,f'=C{r}/B{r}-1',PCT); put(w5,r,6,f'=D{r}/C{r}-1',PCT); put(w5,r,7,f'=D{r}/B{r}-1',PCT)
put(w5,r,8,f'=D{r}/'+tot_sum(2026,7),PCT2); put(w5,r,9,f'=C{r}/'+tot_sum(2025,7),PCT2)
put(w5,r,10,f'=B{r}/'+tot_sum(2024,7),PCT2)
put(w5,r,11,f'=(D{r}-C{r})/({tot_sum(2026,7)}-{tot_sum(2025,7)})',PCT2)
r+=1
put(w5,r,1,'전체 수출',bold=True)
for i,y in enumerate(YRS): put(w5,r,2+i,'='+tot_sum(y,7),NUM,bold=True)
put(w5,r,5,f'=C{r}/B{r}-1',PCT,bold=True); put(w5,r,6,f'=D{r}/C{r}-1',PCT,bold=True); put(w5,r,7,f'=D{r}/B{r}-1',PCT,bold=True)
TOT5=r; r+=2
sec(w5,r,'[B] 월별 추이 (31개월, 천달러 / 비중)',12); r+=1
MB=r
head(w5,r,1,'구분')
for k in range(NM): head(w5,r,2+k,MONTHS[k])
r+=1; MB0=r
rows5=[('IT부품(협의) 수출','G',['라.IT부품']),('IT부품+IT제품(광의) 수출','G',['라.IT부품','다.IT제품'])]
for nm,col,keys in rows5:
    put(w5,r,1,nm,bold=True)
    for k in range(NM):
        put(w5,r,2+k,'='+'+'.join(f'SUMIFS({nR(col)},{nR("E")},"{x}",{nR("A")},"{MONTHS[k]}")' for x in keys),NUM,sz=9)
    r+=1
put(w5,r,1,'HS85 수출',bold=True)
for k in range(NM): put(w5,r,2+k,f'=SUMIFS({hR("H")},{hR("E")},"85",{hR("A")},"{MONTHS[k]}")',NUM,sz=9)
r+=1
put(w5,r,1,'전체 수출',bold=True)
for k in range(NM): put(w5,r,2+k,f"='분석2_월별추이'!E{D2+k}",NUM,sz=9)
TOTROW=r; r+=1
put(w5,r,1,'비반도체(광의 기준) 수출',bold=True)
for k in range(NM): put(w5,r,2+k,f'={L(2+k)}{TOTROW}-{L(2+k)}{MB0+1}',NUM,sz=9)
NONIT=r; r+=1
put(w5,r,1,'반도체(광의) 비중',bold=True)
for k in range(NM): put(w5,r,2+k,f'={L(2+k)}{MB0+1}/{L(2+k)}{TOTROW}',PCT2,sz=9)
SHARE=r; r+=1
put(w5,r,1,'반도체(광의) 전년동월비',bold=True)
for k in range(NM):
    put(w5,r,2+k,'' if k<12 else f'={L(2+k)}{MB0+1}/{L(2+k-12)}{MB0+1}-1',PCT,sz=9)
ITYOY=r; r+=1
put(w5,r,1,'비반도체 전년동월비',bold=True)
for k in range(NM):
    put(w5,r,2+k,'' if k<12 else f'={L(2+k)}{NONIT}/{L(2+k-12)}{NONIT}-1',PCT,sz=9)
NONYOY=r; r+=2
sec(w5,r,'[C] 어느 쪽이 좋았나 — 판정',12); r+=1
for i,h in enumerate(['기간','반도체(광의) YoY','비반도체 YoY','격차(%p)','판정']): head(w5,r,i+1,h)
r+=1
J5=r
for lbl,y in [('2025년 1~7월',2025),('2026년 1~7월',2026)]:
    put(w5,r,1,lbl,bold=True)
    itc=it_sum(['라.IT부품','다.IT제품'],y,7); itp=it_sum(['라.IT부품','다.IT제품'],y-1,7)
    tc=tot_sum(y,7); tp=tot_sum(y-1,7)
    put(w5,r,2,f'=({itc})/({itp})-1',PCT)
    put(w5,r,3,f'=({tc}-({itc}))/({tp}-({itp}))-1',PCT)
    put(w5,r,4,f'=(B{r}-C{r})*100',PP)
    put(w5,r,5,f'=IF(B{r}>C{r},"반도체 우위","비반도체 우위")',align='center',bold=True)
    r+=1
note(w5,r,1,'2025년과 2026년의 판정이 다르면, 같은 호황이라도 국면이 바뀌었다는 뜻입니다. 아래 [D]에서 그 전환 시점을 월 단위로 확인할 수 있습니다.',12)
r+=2
sec(w5,r,'[D] 국면 전환 추적 — 매월 어느 쪽이 더 좋았나 (전년동월비 기준)',12); r+=1
head(w5,r,1,'기간')
for k in range(12,NM): head(w5,r,2+k-12,MONTHS[k])
r+=1
for lbl,src in [('반도체(광의) YoY',ITYOY),('비반도체 YoY',NONYOY)]:
    put(w5,r,1,lbl,bold=True)
    for k in range(12,NM): put(w5,r,2+k-12,f'={L(2+k)}{src}',PCT,sz=9)
    r+=1
put(w5,r,1,'우위',bold=True)
for k in range(12,NM):
    cl=L(2+k-12)
    put(w5,r,2+k-12,f'=IF({cl}{r-2}>{cl}{r-1},"반도체","비반도체")',align='center',sz=9)
r+=1
put(w5,r,1,'격차(%p)',bold=True)
for k in range(12,NM):
    cl=L(2+k-12)
    put(w5,r,2+k-12,f'=({cl}{r-3}-{cl}{r-2})*100',PP,sz=9)
w5.freeze_panes='B5'
print('분석5 완료')

# ══════════════════════════════════════════════════════════════════
# 분석6_HS부문별
# ══════════════════════════════════════════════════════════════════
def hs_y(code,y,col='H',mmax=7):
    return f'SUMIFS({hR(col)},{hR("E")},"{code}",{hR("B")},{y},{hR("C")},"<={mmax}")'
def sec_y(chs,y,col='H'):
    return '+'.join(hs_y(c,y,col) for c in chs)
_e26=tot_sum(2026,7); _e25=tot_sum(2025,7)
w6=wb.create_sheet('분석6_HS부문별'); w6.sheet_view.showGridLines=False
widths(w6,[7,24,30,14,14,14,10,10,10,10,14,14,10,26])
title(w6,1,'⑥ HS 부문(Section)별 분석 — 21개 부문 전체 (1~7월 동기 기준)',14)
note(w6,2,1,'HS 21개 부(Section)로 97개 류를 모두 묶었습니다. 세 번째 열에 각 부문에 어떤 류가 들어가는지 명시했으며, 류 단위 상세는 분석7에 있습니다. 정렬은 2026년 수출 증감액 순(작성 시점 고정).',14)
r=4
for i,h in enumerate(['부','부문 명칭','포함 HS류','수출 2024','수출 2025','수출 2026','25 YoY','26 YoY','26 비중','기여도','수입 2026','무역수지 2026','수지 방향','성격 판정']): head(w6,r,i+1,h)
w6.row_dimensions[r].height=32
r+=1
S6=r
from data3 import M17 as _M17
def _sv(chs,y):
    return sum(HS[c][m][1] for c in chs if c in HS for m in _M17[str(y)] if m in HS[c])
order6=sorted(SECTIONS,key=lambda s:-( _sv(s[2],2026)-_sv(s[2],2025)))
for s,n,chs in order6:
    chs_in=[c for c in chs if c in HS]
    if not chs_in: continue
    put(w6,r,1,s,align='center',bold=True); put(w6,r,2,n,bold=True)
    put(w6,r,3,', '.join(chs_in),sz=9,wrap=True)
    for i,y in enumerate(YRS): put(w6,r,4+i,'='+sec_y(chs_in,y),NUM)
    put(w6,r,7,f'=IFERROR(E{r}/D{r}-1,"-")',PCT)
    put(w6,r,8,f'=IFERROR(F{r}/E{r}-1,"-")',PCT)
    put(w6,r,9,f'=F{r}/{_e26}',PCT2)
    put(w6,r,10,f'=(F{r}-E{r})/({_e26}-{_e25})',PCT2)
    put(w6,r,11,'='+sec_y(chs_in,2026,'J'),NUM)
    put(w6,r,12,f'=F{r}-K{r}',NUM)
    put(w6,r,13,f'=IF(L{r}>0,"흑자","적자")',align='center')
    put(w6,r,14,f'=IF(J{r}>0.3,"전체 증가를 주도",IF(J{r}>0.03,"증가에 뚜렷이 기여",'
                f'IF(H{r}<0,"감소 — 부진 부문",IF(J{r}>0,"소폭 기여","영향 미미"))))',sz=9)
    r+=1
E6=r-1
put(w6,r,2,'합계(HS 전체)',bold=True)
for col in [4,5,6,11,12]: put(w6,r,col,f'=SUM({L(col)}{S6}:{L(col)}{E6})',NUM,bold=True)
put(w6,r,7,f'=E{r}/D{r}-1',PCT,bold=True); put(w6,r,8,f'=F{r}/E{r}-1',PCT,bold=True)
put(w6,r,9,f'=F{r}/{_e26}',PCT2,bold=True); put(w6,r,10,f'=(F{r}-E{r})/({_e26}-{_e25})',PCT2,bold=True)
SUM6=r; r+=2
sec(w6,r,'[참고] 부문별 월별 수출 (상위 6개 부문, 그래프용, 천달러)',14); r+=1
S6M=r; head(w6,r,1,'부문')
for k in range(NM): head(w6,r,2+k,MONTHS[k])
r+=1; S6M0=r
top6=[(s,n,[c for c in chs if c in HS]) for s,n,chs in order6[:6]]
for s,n,chs in top6:
    put(w6,r,1,f'{s}. {n}',bold=True,sz=9)
    for k in range(NM):
        put(w6,r,2+k,'='+'+'.join(f'SUMIFS({hR("H")},{hR("E")},"{c}",{hR("A")},"{MONTHS[k]}")' for c in chs),NUM,sz=8)
    r+=1
S6M1=r-1
w6.freeze_panes='D5'
print('분석6 완료')

# ══════════════════════════════════════════════════════════════════
# 분석7_HS류별상세 (97류 전수)
# ══════════════════════════════════════════════════════════════════
w7=wb.create_sheet('분석7_HS류별상세'); w7.sheet_view.showGridLines=False
widths(w7,[6,7,26,34,13,13,13,10,10,10,10,12,12,10,10,12,13,13,12,44])
title(w7,1,'⑦ HS 류(2단위)별 전수 분석 — 97개 류 (1~7월 동기 기준, 천달러)',20)
note(w7,2,1,'원본에 있는 HS류를 하나도 빼지 않고 모두 담았습니다. 「주요 구성 품목」은 해당 류에 실제로 들어가는 대표 품목이며, 「해석」 열은 증감 크기·유형·전년 대비 방향 전환을 수식으로 자동 판정한 문장입니다.',20)
note(w7,3,1,'성장 유형 : 물량효과와 가격효과 중 한쪽이 2배 이상이면 해당 유형으로, 아니면 혼합으로 분류. 중량이 극소인 류(단가 왜곡)는 별도 표기. 정렬은 2026년 수출 증감액 순.',20)
r=5
H7=['부','HS','품목명(원본)','주요 구성 품목','수출 2024','수출 2025','수출 2026','25 YoY','26 YoY','기여도','26 비중',
    '중량 2025','중량 2026','물량 증감','단가 증감','성장 유형','수입 2026','무역수지 2026','수지 증감','해석']
for i,h in enumerate(H7): head(w7,r,i+1,h)
w7.row_dimensions[r].height=34
r+=1
R7=r
codes=sorted(HS)
def _cv(c,y): return sum(HS[c][m][1] for m in _M17[str(y)] if m in HS[c])
order7=sorted(codes,key=lambda c:-(_cv(c,2026)-_cv(c,2025)))
for code in order7:
    s,n=CH2SEC.get(code,('-','-'))
    put(w7,r,1,s,align='center',sz=9); put(w7,r,2,code,align='center',bold=True).number_format='@'
    put(w7,r,3,f'=IFERROR(INDEX({hR("F")},MATCH($B{r},{hR("E")},0)),"")',sz=9,wrap=True)
    put(w7,r,4,DESC.get(code,''),sz=9,wrap=True)
    for i,y in enumerate(YRS): put(w7,r,5+i,'='+hs_y(code,y),NUM)
    put(w7,r,8,f'=IFERROR(F{r}/E{r}-1,"-")',PCT)
    put(w7,r,9,f'=IFERROR(G{r}/F{r}-1,"-")',PCT)
    put(w7,r,10,f'=(G{r}-F{r})/({_e26}-{_e25})',PCT2)
    put(w7,r,11,f'=G{r}/{_e26}',PCT2)
    put(w7,r,12,'='+hs_y(code,2025,'G'),NUM1); put(w7,r,13,'='+hs_y(code,2026,'G'),NUM1)
    put(w7,r,14,f'=IFERROR(M{r}/L{r}-1,"-")',PCT)
    put(w7,r,15,f'=IFERROR((G{r}/M{r})/(F{r}/L{r})-1,"-")',PCT)
    put(w7,r,16,f'=IF(OR(L{r}<100,M{r}<100),"단가 불안정",'
                f'IF(G{r}-F{r}<0,"감소",'
                f'IF(IFERROR(((G{r}/M{r})-(F{r}/L{r}))*L{r},0)>ABS(IFERROR((M{r}-L{r})*(F{r}/L{r}),0))*2,"가격 주도",'
                f'IF(IFERROR((M{r}-L{r})*(F{r}/L{r}),0)>ABS(IFERROR(((G{r}/M{r})-(F{r}/L{r}))*L{r},0))*2,"물량 주도","혼합"))))',align='center',sz=9)
    put(w7,r,17,'='+hs_y(code,2026,'J'),NUM)
    put(w7,r,18,f'=G{r}-Q{r}',NUM)
    put(w7,r,19,f'=(G{r}-Q{r})-(F{r}-'+hs_y(code,2025,'J')+')',NUM)
    put(w7,r,20,
        f'=IF(G{r}=0,"자료 없음",'
        f'IF(I{r}>1,"2배 이상 급증 : ",IF(I{r}>0.5,"급증(+50% 초과) : ",IF(I{r}>0.15,"뚜렷한 증가 : ",'
        f'IF(I{r}>0.02,"소폭 증가 : ",IF(I{r}>-0.02,"보합 : ",IF(I{r}>-0.15,"소폭 감소 : ","뚜렷한 감소 : "))))))'
        f'&IF(P{r}="가격 주도",IF(N{r}>0.02,"물량도 늘었지만 단가 상승이 증가를 주도했다. ",'
        f'IF(N{r}<-0.02,"물량은 오히려 줄었고 단가 상승이 금액을 끌어올렸다. ","물량은 거의 그대로인데 단가가 올라 금액이 늘었다. ")),'
        f'IF(P{r}="물량 주도","단가보다 물량이 늘어 실수요 확대 성격이 강하다. ",'
        f'IF(P{r}="혼합","물량과 단가가 함께 움직였다. ",'
        f'IF(P{r}="감소",IF(AND(N{r}<0,O{r}<0),"물량과 단가가 함께 줄었다. ",IF(N{r}<0,"물량 감소가 금액 감소를 이끌었다. ","단가 하락이 금액 감소를 이끌었다. ")),'
        f'"중량이 적어 단가 해석은 유보. "))))'
        f'&IF(AND(H{r}>0,I{r}>0),"2년 연속 증가. ",IF(AND(H{r}<0,I{r}>0),"25년 감소에서 26년 증가로 전환. ",'
        f'IF(AND(H{r}>0,I{r}<0),"25년 증가에서 26년 감소로 전환. ","2년 연속 감소. ")))'
        f'&IF(J{r}>0.05,"전체 수출 증가의 5% 이상을 설명하는 대형 품목.",'
        f'IF(J{r}>0.005,"전체 증가에 의미 있게 기여.",IF(K{r}<0.001,"규모가 작아 총량 영향은 미미.","총량 영향은 제한적.")))',sz=9,wrap=True)
    w7.row_dimensions[r].height=30
    r+=1
E7=r-1
put(w7,r,3,'합계(HS 전체)',bold=True)
for col in [5,6,7,12,13,17,18,19]: put(w7,r,col,f'=SUM({L(col)}{R7}:{L(col)}{E7})',NUM if col not in (12,13) else NUM1,bold=True)
put(w7,r,8,f'=F{r}/E{r}-1',PCT,bold=True); put(w7,r,9,f'=G{r}/F{r}-1',PCT,bold=True)
put(w7,r,10,f'=(G{r}-F{r})/({_e26}-{_e25})',PCT2,bold=True); put(w7,r,11,f'=G{r}/{_e26}',PCT2,bold=True)
SUM7=r; r+=2
sec(w7,r,'[요약] 성장 유형별 집계',20); r+=1
for i,h in enumerate(['유형','류 수','2026 수출 합','Δ금액 합','전체 증가분 대비']): head(w7,r,i+1,h)
r+=1
T7=r
for t in ['가격 주도','물량 주도','혼합','감소','단가 불안정']:
    put(w7,r,1,t,bold=True)
    put(w7,r,2,f'=COUNTIF($P${R7}:$P${E7},A{r})',NUM)
    put(w7,r,3,f'=SUMIF($P${R7}:$P${E7},A{r},$G${R7}:$G${E7})',NUM)
    put(w7,r,4,f'=SUMIF($P${R7}:$P${E7},A{r},$G${R7}:$G${E7})-SUMIF($P${R7}:$P${E7},A{r},$F${R7}:$F${E7})',NUM)
    put(w7,r,5,f'=D{r}/({_e26}-{_e25})',PCT2)
    r+=1
w7.freeze_panes='E6'
print('분석7 완료', R7, E7)
exec(open('v3_d.py').read())
wb.save('final3.xlsx'); print('saved final3')
