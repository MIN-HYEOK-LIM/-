# -*- coding: utf-8 -*-
# 분석8~12 (v3_c 네임스페이스에서 exec)

# ══════════════════════════════════════════════════════════════════
# 분석8_성질별심층 (수출·수입 양방향)
# ══════════════════════════════════════════════════════════════════
w8=wb.create_sheet('분석8_성질별심층'); w8.sheet_view.showGridLines=False
widths(w8,[8,26,14,14,14,10,10,10,10,10,14,14,14,10])
title(w8,1,'⑧ 성질별·신성질별 심층 (1~7월 동기 기준, 천달러)',14)
note(w8,2,1,'이번 자료는 성질별이 수출·수입 두 파일로 나뉘어 있어 양방향 분석이 가능합니다. 수출과 수입의 분류 체계가 서로 다르므로(수출 : 식료/원료/경공업/중화학, 수입 : 소비재/원자재/자본재) 직접 대응시키지 마십시오.',14)
def blk_items(ws,r,title_,items,majors,rng,valcol,label='수출'):
    sec(ws,r,title_,14); r+=1
    for i,h in enumerate(['구분','항목',f'{label} 2024',f'{label} 2025',f'{label} 2026','25 YoY','26 YoY','24→26','26 비중','기여도','중량 2026','단가 2025','단가 2026','단가 증감']): head(ws,r,i+1,h)
    ws.row_dimensions[r].height=28
    r+=1; r0=r
    for nm in items:
        isM = nm in majors
        put(ws,r,1,'대분류' if isM else '중분류',align='center',bold=isM,fill=SUB if isM else None)
        put(ws,r,2,nm,bold=isM,fill=SUB if isM else None)
        for i,y in enumerate(YRS):
            put(ws,r,3+i,f'=SUMIFS({rng(valcol)},{rng("E")},$B{r},{rng("B")},{y},{rng("C")},"<=7")',NUM,bold=isM)
        put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
        put(ws,r,7,f'=IFERROR(E{r}/D{r}-1,"-")',PCT,bold=isM)
        put(ws,r,8,f'=IFERROR(E{r}/C{r}-1,"-")',PCT)
        put(ws,r,9,f'=IFERROR(E{r}/{_e26},"-")',PCT2)
        put(ws,r,10,f'=IFERROR((E{r}-D{r})/({_e26}-{_e25}),"-")',PCT2)
        wcol='F'
        put(ws,r,11,f'=SUMIFS({rng(wcol)},{rng("E")},$B{r},{rng("B")},2026,{rng("C")},"<=7")',NUM1)
        put(ws,r,12,f'=IFERROR(D{r}/SUMIFS({rng(wcol)},{rng("E")},$B{r},{rng("B")},2025,{rng("C")},"<=7"),"-")',UNIT)
        put(ws,r,13,f'=IFERROR(E{r}/K{r},"-")',UNIT)
        put(ws,r,14,f'=IFERROR(M{r}/L{r}-1,"-")',PCT)
        r+=1
    return r+1, r0
r=4
r,X8=blk_items(w8,r,'[A] 성질별 — 수출',OLDX_ITEMS,['1. 식료 및 직접소비재','2. 원료 및 연료','3. 경공업품','4. 중화학 공업품'],xR,'G','수출')
r,M8=blk_items(w8,r,'[B] 성질별 — 수입',OLDM_ITEMS,['1. 소비재','2. 원자재','3. 자본재'],mR,'G','수입')
sec(w8,r,'[C] 신성질별 — 수출·수입·무역수지',14); r+=1
for i,h in enumerate(['구분','항목','수출 2024','수출 2025','수출 2026','25 YoY','26 YoY','수입 2024','수입 2025','수입 2026','수입 26 YoY','수지 2025','수지 2026','수지 증감']): head(w8,r,i+1,h)
w8.row_dimensions[r].height=28
r+=1; N8=r
for nm in NEW_ITEMS:
    isM = nm in ['1.소비재','2.원자재','3.자본재']
    put(w8,r,1,'대분류' if isM else '중분류',align='center',bold=isM,fill=SUB if isM else None)
    put(w8,r,2,nm,bold=isM,fill=SUB if isM else None)
    for i,y in enumerate(YRS):
        put(w8,r,3+i,f'=SUMIFS({nR("G")},{nR("E")},$B{r},{nR("B")},{y},{nR("C")},"<=7")',NUM,bold=isM)
        put(w8,r,8+i,f'=SUMIFS({nR("I")},{nR("E")},$B{r},{nR("B")},{y},{nR("C")},"<=7")',NUM)
    put(w8,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(w8,r,7,f'=IFERROR(E{r}/D{r}-1,"-")',PCT,bold=isM)
    put(w8,r,11,f'=IFERROR(J{r}/I{r}-1,"-")',PCT)
    put(w8,r,12,f'=D{r}-I{r}',NUM); put(w8,r,13,f'=E{r}-J{r}',NUM,bold=isM)
    put(w8,r,14,f'=M{r}-L{r}',NUM)
    r+=1
E8N=r-1
w8.freeze_panes='C5'
print('분석8 완료')

# ══════════════════════════════════════════════════════════════════
# 분석9_물량가격분해
# ══════════════════════════════════════════════════════════════════
w9=wb.create_sheet('분석9_물량·가격분해'); w9.sheet_view.showGridLines=False
widths(w9,[26,16,16,16,14,14,16,16,16,14])
title(w9,1,'⑨ 물량·가격 분해 — 늘어난 것은 양인가 값인가',10)
note(w9,2,1,'금액 증감 = 물량효과 (Q₁−Q₀)×P₀ + 가격효과 (P₁−P₀)×Q₀ + 교차효과. 세 값의 합은 실제 증감액과 일치합니다(항등식 검증). 2024→2025 구간과 2025→2026 구간을 나란히 비교했습니다.',10)
r=4
sec(w9,r,'[A] 총괄 수출 — 구간별 3효과 분해 (1~7월 동기)',10); r+=1
for i,h in enumerate(['구간','금액 증감','물량효과','가격효과','교차효과','물량 기여','가격 기여','물량 증감률','단가 증감률','검증']): head(w9,r,i+1,h)
r+=1
A9=r
for lbl,y0,y1 in [('2024 → 2025',2024,2025),('2025 → 2026',2025,2026)]:
    v0=tot_sum(y0,7); v1=tot_sum(y1,7)
    q0=tot_sum(y0,7,'G'); q1=tot_sum(y1,7,'G')
    put(w9,r,1,lbl,bold=True)
    put(w9,r,2,f'={v1}-{v0}',NUM)
    put(w9,r,3,f'=({q1}-{q0})*({v0}/{q0})',NUM)
    put(w9,r,4,f'=(({v1}/{q1})-({v0}/{q0}))*{q0}',NUM)
    put(w9,r,5,f'=({q1}-{q0})*(({v1}/{q1})-({v0}/{q0}))',NUM)
    put(w9,r,6,f'=C{r}/B{r}',PCT); put(w9,r,7,f'=D{r}/B{r}',PCT)
    put(w9,r,8,f'={q1}/{q0}-1',PCT); put(w9,r,9,f'=({v1}/{q1})/({v0}/{q0})-1',PCT)
    put(w9,r,10,f'=IF(ABS(C{r}+D{r}+E{r}-B{r})<1,"OK","불일치")',align='center')
    r+=1
r+=1
sec(w9,r,'[B] 신성질별 중분류별 분해 (2025 → 2026, 1~7월)',10); r+=1
for i,h in enumerate(['품목','중량 2025','금액 2025','중량 2026','금액 2026','물량 증감률','단가 증감률','물량효과','가격효과','Δ금액']): head(w9,r,i+1,h)
r+=1
B9=r
SUBS=[n for n in NEW_ITEMS if n not in ['1.소비재','2.원자재','3.자본재']]
for nm in SUBS:
    put(w9,r,1,nm,bold=True)
    put(w9,r,2,f'=SUMIFS({nR("F")},{nR("E")},"{nm}",{nR("B")},2025,{nR("C")},"<=7")',NUM1)
    put(w9,r,3,f'=SUMIFS({nR("G")},{nR("E")},"{nm}",{nR("B")},2025,{nR("C")},"<=7")',NUM)
    put(w9,r,4,f'=SUMIFS({nR("F")},{nR("E")},"{nm}",{nR("B")},2026,{nR("C")},"<=7")',NUM1)
    put(w9,r,5,f'=SUMIFS({nR("G")},{nR("E")},"{nm}",{nR("B")},2026,{nR("C")},"<=7")',NUM)
    put(w9,r,6,f'=IFERROR(D{r}/B{r}-1,"-")',PCT)
    put(w9,r,7,f'=IFERROR((E{r}/D{r})/(C{r}/B{r})-1,"-")',PCT)
    put(w9,r,8,f'=IFERROR((D{r}-B{r})*(C{r}/B{r}),"-")',NUM)
    put(w9,r,9,f'=IFERROR(((E{r}/D{r})-(C{r}/B{r}))*B{r},"-")',NUM)
    put(w9,r,10,f'=E{r}-C{r}',NUM)
    r+=1
E9=r-1
put(w9,r,1,'합계',bold=True)
for col in [2,3,4,5,8,9,10]: put(w9,r,col,f'=SUM({L(col)}{B9}:{L(col)}{E9})',NUM1 if col in (2,4) else NUM,bold=True)
put(w9,r,6,f'=D{r}/B{r}-1',PCT,bold=True); put(w9,r,7,f'=(E{r}/D{r})/(C{r}/B{r})-1',PCT,bold=True)
S9=r; r+=2
sec(w9,r,'[C] 지수 분해 (라스파이레스·파셰·피셔, 2025→2026)',10); r+=1
for i,h in enumerate(['지수','값','증감률','의미']): head(w9,r,i+1,h)
r+=1
I9=r
q0p0=f'SUMPRODUCT($B${B9}:$B${E9},$C${B9}:$C${E9}/$B${B9}:$B${E9})'
q1p0=f'SUMPRODUCT($D${B9}:$D${E9},$C${B9}:$C${E9}/$B${B9}:$B${E9})'
q1p1=f'SUM($E${B9}:$E${E9})'
q0p1=f'SUMPRODUCT($B${B9}:$B${E9},$E${B9}:$E${E9}/$D${B9}:$D${E9})'
for nm,f_,mean in [('라스파이레스 물량지수',f'={q1p0}/{q0p0}','기준연도 가격 고정, 물량만 변화'),
                   ('파셰 물량지수',f'={q1p1}/{q0p1}','비교연도 가격 고정, 물량만 변화'),
                   ('피셔 물량지수',None,'두 지수의 기하평균 (권장)'),
                   ('라스파이레스 가격지수',f'={q0p1}/{q0p0}','기준연도 물량 고정, 가격만 변화'),
                   ('파셰 가격지수',f'={q1p1}/{q1p0}','비교연도 물량 고정, 가격만 변화'),
                   ('피셔 가격지수',None,'두 지수의 기하평균 (권장)'),
                   ('금액지수',f'={q1p1}/{q0p0}','피셔 물량 × 피셔 가격과 일치해야 함')]:
    put(w9,r,1,nm,bold=True)
    if nm=='피셔 물량지수': put(w9,r,2,f'=SQRT(B{I9}*B{I9+1})','0.0000',bold=True)
    elif nm=='피셔 가격지수': put(w9,r,2,f'=SQRT(B{I9+3}*B{I9+4})','0.0000',bold=True)
    else: put(w9,r,2,f_,'0.0000')
    put(w9,r,3,f'=B{r}-1',PCT); put(w9,r,4,mean)
    r+=1
put(w9,r,1,'구성(믹스)효과',bold=True)
put(w9,r,2,f'=((E{S9}/D{S9})/(C{S9}/B{S9}))/B{I9+5}','0.0000')
put(w9,r,3,f'=B{r}-1',PCT,color=RED); put(w9,r,4,'단순 단가 상승분 중 품목 구성 변화가 만든 몫')
print('분석9 완료')

# ══════════════════════════════════════════════════════════════════
# 분석10_그래프_연도오버레이
# ══════════════════════════════════════════════════════════════════
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import ChartLines
PALETTE=['A6A6A6','2E75B6','548235','ED7D31','C00000','7030A0','1F3864','BF8F00',
         '4472C4','70AD47','FFC000','9E480E','636363','255E91','43682B','7CAFDD']
def flat(ch,off=0):
    isbar=isinstance(ch,BarChart)
    for i,s_ in enumerate(ch.series):
        col=PALETTE[(i+off)%len(PALETTE)]
        if isbar:
            s_.graphicalProperties.solidFill=col; s_.graphicalProperties.line.noFill=True
        else:
            s_.smooth=False
            s_.graphicalProperties.line.solidFill=col
            s_.graphicalProperties.line.width=24000
def subax(ch,off=0):
    ch.x_axis.delete=True; flat(ch,off)

w10=wb.create_sheet('분석10_그래프_연도오버레이'); w10.sheet_view.showGridLines=False
widths(w10,[16]+[12]*9)
title(w10,1,'⑩ 연도 오버레이 그래프 — 1~12월을 한 축에 놓고 2024·2025·2026을 겹쳐 비교',10)
note(w10,2,1,'모든 그래프의 가로축은 1~12월이고, 회색 = 2024년, 파랑 = 2025년, 초록 = 2026년입니다. 2026년은 7월에서 선이 끊깁니다. 같은 달끼리 비교하므로 계절성이 자동으로 통제되며, 세 선의 간격이 곧 전년 대비 증감입니다.',10)
note(w10,3,1,'원본 데이터는 「분석4_연도별계절비교」의 각 블록이며, 값이 바뀌면 그래프도 함께 갱신됩니다.',10)
CT=5
def overlay(name,anchor,ylab,fmt_pct=False):
    hr,r0,r1,ytot,y7=SEASON[name]
    ch=LineChart()
    ch.add_data(Reference(w4,min_col=2,max_col=4,min_row=hr,max_row=r1),titles_from_data=True)
    ch.set_categories(Reference(w4,min_col=1,min_row=r0,max_row=r1))
    flat(ch); ch.dispBlanksAs='gap'; ch.title=f'{name} — 연도 비교 (2024·2025·2026)'
    ch.width=24; ch.height=11; ch.style=2
    ch.y_axis.title=ylab; ch.y_axis.majorGridlines=ChartLines()
    w10.add_chart(ch,anchor)
ORDER10=[('수출 금액','천달러'),('수입 금액','천달러'),('무역수지','천달러'),('수출 단가','천달러/톤'),
         ('일평균 수출','천달러/일'),('수출 중량','톤'),('IT부품 수출','천달러'),('IT 제외 수출','천달러'),
         ('HS85 전기기기 수출','천달러'),('HS84 기계류 수출','천달러'),('HS87 차량 수출','천달러'),
         ('HS27 광물성연료 수출','천달러'),('HS89 선박 수출','천달러'),('중화학공업품 수출','천달러'),
         ('수입 원자재','천달러'),('수입 자본재','천달러'),('IT제품 수출','천달러')]
for i,(nm,ylab) in enumerate(ORDER10):
    overlay(nm,f'A{CT+i*22}',ylab)
print('분석10 완료 : 오버레이 그래프',len(ORDER10),'개')

# ══════════════════════════════════════════════════════════════════
# 분석11_그래프_구조분석
# ══════════════════════════════════════════════════════════════════
w11=wb.create_sheet('분석11_그래프_구조분석'); w11.sheet_view.showGridLines=False
widths(w11,[26]+[12]*12)
title(w11,1,'⑪ 구조 분석 그래프 — 반도체 vs 비반도체 · 부문 · 분기 · 분해',13)
note(w11,2,1,'보조표는 분석3·5·6·7·9를 참조하는 수식입니다.',13)
r=4
sec(w11,r,'보조표 A. HS 부문별 1~7월 수출 증감액 (2026 vs 2025, 천달러)',13); r+=1
A11H=r; head(w11,r,1,'부문'); head(w11,r,2,'증감액'); head(w11,r,3,'26 YoY')
r+=1; A11=r
for i,(s,n,chs) in enumerate(order6):
    if not [c for c in chs if c in HS]: continue
    put(w11,r,1,f'{s}. {n}',sz=9)
    put(w11,r,2,f"='분석6_HS부문별'!F{S6+i}-'분석6_HS부문별'!E{S6+i}",NUM)
    put(w11,r,3,f"='분석6_HS부문별'!H{S6+i}",PCT)
    r+=1
A11E=r-1; r+=1
sec(w11,r,'보조표 B. 성장 유형별 Δ금액 (HS 97류, 천달러)',13); r+=1
B11H=r; head(w11,r,1,'유형'); head(w11,r,2,'Δ금액'); head(w11,r,3,'류 수')
r+=1; B11=r
for i,t in enumerate(['가격 주도','물량 주도','혼합','감소','단가 불안정']):
    put(w11,r,1,t); put(w11,r,2,f"='분석7_HS류별상세'!D{T7+i}",NUM); put(w11,r,3,f"='분석7_HS류별상세'!B{T7+i}",NUM)
    r+=1
B11E=r-1; r+=1
sec(w11,r,'보조표 C. HS 류별 1~7월 수출 증감액 상위 15 (천달러)',13); r+=1
C11H=r; head(w11,r,1,'HS류'); head(w11,r,2,'증감액'); head(w11,r,3,'26 YoY')
r+=1; C11=r
for i,code in enumerate(order7[:15]):
    put(w11,r,1,f'{code}. {HSNAME[code][:20]}',sz=9)
    put(w11,r,2,f"='분석7_HS류별상세'!G{R7+i}-'분석7_HS류별상세'!F{R7+i}",NUM)
    put(w11,r,3,f"='분석7_HS류별상세'!I{R7+i}",PCT)
    r+=1
C11E=r-1; r+=1
sec(w11,r,'보조표 D. HS 류별 1~7월 수출 감소액 하위 10 (천달러)',13); r+=1
D11H=r; head(w11,r,1,'HS류'); head(w11,r,2,'증감액'); head(w11,r,3,'26 YoY')
r+=1; D11=r
for i,code in enumerate(order7[-10:]):
    idx=order7.index(code)
    put(w11,r,1,f'{code}. {HSNAME[code][:20]}',sz=9)
    put(w11,r,2,f"='분석7_HS류별상세'!G{R7+idx}-'분석7_HS류별상세'!F{R7+idx}",NUM)
    put(w11,r,3,f"='분석7_HS류별상세'!I{R7+idx}",PCT)
    r+=1
D11E=r-1
CT11=D11E+3
def add11(ch,anchor,t,w=26,h=12,ylab=None,xlab=None,legend=True,off=0):
    flat(ch,off); ch.title=t; ch.width=w; ch.height=h; ch.style=2
    if ylab: ch.y_axis.title=ylab
    if xlab: ch.x_axis.title=xlab
    if not legend: ch.legend=None
    ch.y_axis.majorGridlines=ChartLines()
    w11.add_chart(ch,anchor)
# 1. 반도체 vs 비반도체 월별 YoY
l1=LineChart()
l1.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=ITYOY,max_row=ITYOY),from_rows=True,titles_from_data=True)
l1.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=NONYOY,max_row=NONYOY),from_rows=True,titles_from_data=True)
l1.set_categories(Reference(w5,min_col=2,max_col=1+NM,min_row=MB))
add11(l1,f'A{CT11}','1. 반도체(광의) vs 비반도체 — 전년 동월 대비 증감률',ylab='전년동월비',off=3)
# 2. 반도체 비중 추이
l2=LineChart()
l2.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=SHARE,max_row=SHARE),from_rows=True,titles_from_data=True)
l2.set_categories(Reference(w5,min_col=2,max_col=1+NM,min_row=MB))
add11(l2,f'A{CT11+24}','2. 반도체(광의) 비중 추이 (2024.01~2026.07)',ylab='전체 수출 대비 비중',off=4)
# 3. 월별 수출 : 반도체/비반도체 금액
l3=LineChart()
l3.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=MB0+1,max_row=MB0+1),from_rows=True,titles_from_data=True)
l3.add_data(Reference(w5,min_col=1,max_col=1+NM,min_row=NONIT,max_row=NONIT),from_rows=True,titles_from_data=True)
l3.set_categories(Reference(w5,min_col=2,max_col=1+NM,min_row=MB))
add11(l3,f'A{CT11+48}','3. 반도체(광의)와 비반도체 수출 금액 추이',ylab='천달러',off=1)
# 4. HS 부문별 증감액
b4=BarChart(); b4.type='bar'
b4.add_data(Reference(w11,min_col=2,min_row=A11H,max_row=A11E),titles_from_data=True)
b4.set_categories(Reference(w11,min_col=1,min_row=A11,max_row=A11E))
add11(b4,f'A{CT11+72}','4. HS 부문별 1~7월 수출 증감액 (2026 vs 2025)',h=15,ylab='증감액(천달러)',legend=False,off=1)
# 5. HS 상위 15
b5=BarChart(); b5.type='bar'
b5.add_data(Reference(w11,min_col=2,min_row=C11H,max_row=C11E),titles_from_data=True)
b5.set_categories(Reference(w11,min_col=1,min_row=C11,max_row=C11E))
add11(b5,f'A{CT11+102}','5. HS 류별 수출 증감액 상위 15',h=14,ylab='증감액(천달러)',legend=False,off=1)
# 6. HS 하위 10
b6=BarChart(); b6.type='bar'
b6.add_data(Reference(w11,min_col=2,min_row=D11H,max_row=D11E),titles_from_data=True)
b6.set_categories(Reference(w11,min_col=1,min_row=D11,max_row=D11E))
add11(b6,f'A{CT11+130}','6. HS 류별 수출 감소액 하위 10',h=12,ylab='증감액(천달러)',legend=False,off=4)
# 7. 성장 유형별
b7=BarChart(); b7.type='col'
b7.add_data(Reference(w11,min_col=2,min_row=B11H,max_row=B11E),titles_from_data=True)
b7.set_categories(Reference(w11,min_col=1,min_row=B11,max_row=B11E))
add11(b7,f'A{CT11+154}','7. 성장 유형별 수출 증감액 (HS 97류 분류)',h=11,ylab='증감액(천달러)',legend=False,off=1)
# 8. 분기별 수출/수입 + QoQ
b8=BarChart(); b8.type='col'
b8.add_data(Reference(w3,min_col=2,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
b8.add_data(Reference(w3,min_col=5,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
b8.set_categories(Reference(w3,min_col=1,min_row=Q3,max_row=E3Q))
l8=LineChart()
l8.add_data(Reference(w3,min_col=8,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l8.y_axis.axId=200; l8.y_axis.title='무역수지(천$)'; l8.y_axis.crosses='max'; subax(l8,4); b8+=l8
add11(b8,f'A{CT11+178}','8. 분기별 수출·수입·무역수지 (2024 Q1~2026 Q3)',ylab='금액(천달러)',off=1)
# 9. 분기 YoY
l9=LineChart()
l9.add_data(Reference(w3,min_col=4,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l9.add_data(Reference(w3,min_col=3,min_row=Q3-1,max_row=E3Q),titles_from_data=True)
l9.set_categories(Reference(w3,min_col=1,min_row=Q3,max_row=E3Q))
add11(l9,f'A{CT11+202}','9. 분기별 수출 YoY와 QoQ',ylab='증감률',off=1)
# 10. 물량·가격 분해
b10=BarChart(); b10.type='col'; b10.grouping='clustered'
b10.add_data(Reference(w9,min_col=3,max_col=5,min_row=A9-1,max_row=A9+1),titles_from_data=True)
b10.set_categories(Reference(w9,min_col=1,min_row=A9,max_row=A9+1))
add11(b10,f'A{CT11+226}','10. 물량효과·가격효과 — 구간 비교(24→25 vs 25→26)',h=11,ylab='금액(천달러)',off=1)
print('분석11 완료')

# ══════════════════════════════════════════════════════════════════
# 분석12_해설
# ══════════════════════════════════════════════════════════════════
w12=wb.create_sheet('분석12_해설'); w12.sheet_view.showGridLines=False
widths(w12,[30,16,16,16,14,14,14,14,14])
title(w12,1,'⑫ 해설 — 3년치(2024~2026) 자료가 새로 말해주는 것',9)
note(w12,2,1,'[핵심 수치] 표는 모두 수식이며 원본이 바뀌면 자동 갱신됩니다. 본문 서술의 숫자는 2024.01~2026.07 자료 기준이며, 별도 표기가 없으면 1~7월 동기 비교입니다.',9)
r=4
sec(w12,r,'[핵심 수치] 1~7월 동기 비교 (천달러)',9); r+=1
for i,h in enumerate(['지표','2024','2025','2026','25 YoY','26 YoY','24→26']): head(w12,r,i+1,h)
r+=1
K12=r
for nm,col,fmt in [('수출 금액','H',NUM),('수입 금액','K',NUM),('무역수지','L',NUM),
                   ('수출 중량(톤)','G',NUM1),('수출 건수','F',NUM),('조업일수','E',NUM1)]:
    put(w12,r,1,nm,bold=True)
    for i,y in enumerate(YRS): put(w12,r,2+i,f'=SUMIFS({tR(col)},{tR("B")},{y},{tR("C")},"<=7")',fmt)
    put(w12,r,5,f'=IFERROR(C{r}/B{r}-1,"-")',PCT); put(w12,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(w12,r,7,f'=IFERROR(D{r}/B{r}-1,"-")',PCT)
    r+=1
put(w12,r,1,'수출 단가(천$/톤)',bold=True)
for i in range(3): put(w12,r,2+i,f'={L(2+i)}{K12}/{L(2+i)}{K12+3}',UNIT)
put(w12,r,5,f'=C{r}/B{r}-1',PCT); put(w12,r,6,f'=D{r}/C{r}-1',PCT); put(w12,r,7,f'=D{r}/B{r}-1',PCT); r+=1
put(w12,r,1,'일평균 수출',bold=True)
for i in range(3): put(w12,r,2+i,f'={L(2+i)}{K12}/{L(2+i)}{K12+5}',NUM)
put(w12,r,5,f'=C{r}/B{r}-1',PCT); put(w12,r,6,f'=D{r}/C{r}-1',PCT); put(w12,r,7,f'=D{r}/B{r}-1',PCT); r+=1
put(w12,r,1,'반도체(광의) 수출',bold=True)
for i,y in enumerate(YRS): put(w12,r,2+i,'='+it_sum(['라.IT부품','다.IT제품'],y,7),NUM)
put(w12,r,5,f'=C{r}/B{r}-1',PCT); put(w12,r,6,f'=D{r}/C{r}-1',PCT); put(w12,r,7,f'=D{r}/B{r}-1',PCT); IT12=r; r+=1
put(w12,r,1,'비반도체 수출',bold=True)
for i in range(3): put(w12,r,2+i,f'={L(2+i)}{K12}-{L(2+i)}{IT12}',NUM)
put(w12,r,5,f'=C{r}/B{r}-1',PCT); put(w12,r,6,f'=D{r}/C{r}-1',PCT); put(w12,r,7,f'=D{r}/B{r}-1',PCT); r+=2

def para(ws,row,text,span=9,sz=10,color='000000',bold=False):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    c.alignment=Alignment(vertical='top',wrap_text=True)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    ws.row_dimensions[row].height=max(16,15*(len(text)//64+1))
SECT=[
 ('1. 이번 자료로 새로 알 수 있게 된 것',[
  '· 기간이 2024.01~2026.07(31개월)로 늘어 같은 달을 세 해에 걸쳐 비교할 수 있게 되었습니다. 이전 자료(19개월)로는 2026년의 급증만 보였지만, 이제 그 앞에 2024년과 2025년이 놓이면서 "이 급증이 어디에서 출발했는가"를 볼 수 있습니다.',
  '· 성질별이 수출·수입 두 파일로 나뉘어 들어와 수입 쪽 성질별 분해가 처음으로 가능해졌습니다. 다만 수출과 수입의 분류 체계가 서로 달라(수출 : 식료/원료/경공업/중화학, 수입 : 소비재/원자재/자본재) 항목을 1:1로 대응시킬 수는 없습니다.',
  '· 국가별 자료는 이번 묶음에 없습니다. 따라서 이전 분석의 국가 동조성·집중도 항목은 갱신되지 않았습니다.',
 ]),
 ('2. 3년 그림 — 2025년은 사실상 정체, 2026년에 점프',[
  '· 1~7월 수출은 2024년 3,922.7억 달러 → 2025년 3,953.9억 달러(+0.8%) → 2026년 5,950.6억 달러(+50.5%)입니다. 2025년은 사실상 제자리였고, 도약은 2026년에 몰려 있습니다.',
  '· 연간으로 보면 2024년 6,836.1억 달러, 2025년 7,093.3억 달러(+3.8%)입니다. 2025년 증가분은 대부분 하반기에 나왔습니다. 월별 전년동월비를 보면 2025년 1월 −10.1%로 시작해 9월 +12.6%, 12월 +13.3%로 올라갑니다. 즉 2025년 상반기 부진 → 하반기 회복 → 2026년 가속의 3단계입니다.',
  '· 분기로 보면 전환점이 뚜렷합니다. 2025 Q1은 전년 대비 −2.3%로 이 기간의 저점이고, 이후 Q2 +2.1%, Q3 +6.5%, Q4 +8.4%로 계단을 밟아 2026 Q1 +38.4%, Q2 +57.2%가 됩니다. QoQ로도 2026 Q1 +16.4%, Q2 +24.7%로 두 분기 연속 두 자릿수입니다.',
  '· 무역수지는 2024년 1~7월 265.8억 → 2025년 339.3억 → 2026년 1,677.6억 달러입니다. 2년 만에 6.3배가 되었는데, 그 대부분이 마지막 7개월에 만들어졌습니다.',
 ]),
 ('3. 반도체 vs 비반도체 — 19개월 연속 반도체 우위',[
  '· 정의를 세 가지로 나눠 계산했습니다(분석5). ① 협의 : 신성질별 IT부품 ② 광의 : IT부품+IT제품 ③ HS85 기준. 어느 정의를 쓰더라도 결론은 같습니다.',
  '· 광의 기준으로 1~7월 반도체 수출은 2024년 1,112.8억 → 2025년 1,215.8억(+9.3%) → 2026년 2,908.1억 달러(+139.2%)입니다. 같은 기간 비반도체는 2,809.9억 → 2,738.1억(−2.6%) → 3,042.5억 달러(+11.1%)입니다.',
  '· ★ 2025년 1월부터 2026년 7월까지 19개월 내내 반도체의 전년동월비가 비반도체를 앞섰습니다. 단 한 달의 예외도 없습니다(분석5-D). 특히 2025년은 반도체 +9.3%, 비반도체 −2.6%로, 그해 수출이 그나마 플러스를 지킨 것은 전적으로 반도체 덕분이었습니다.',
  '· 다만 2026년 들어 비반도체도 살아났습니다. 2025년 −2.6%에서 2026년 +11.1%로 돌아섰고, 6·7월 전년동월비는 +17.7%, +17.8%로 가속 중입니다. "반도체만 좋다"는 서술은 2025년에는 정확했지만 2026년에는 절반만 맞습니다.',
  '· 반도체 비중은 2024년 1~7월 28.4% → 2025년 30.7% → 2026년 48.9%입니다. HS85 기준으로는 29.8% → 31.1% → 43.2%입니다. 2년 사이 수출 구조에서 반도체가 차지하는 자리가 3분의 1에서 절반으로 바뀌었습니다.',
 ]),
 ('4. HS 부문(Section)별 — 증가의 85%가 한 부문에서',[
  '· HS 97개 류를 21개 부문으로 묶었습니다(분석6). 1~7월 기준 XVI부(기계·전기전자기기, HS 84·85류)가 1,710.7억 → 3,410.1억 달러로 +99.3% 늘며 전체 증가분의 85.1%를 만들었습니다. 이 부문 하나가 수출의 57.3%입니다.',
  '· 그다음은 V부(광물성 생산품, 25~27류) 기여도 4.8%, XIV부(귀금속·보석, 71류) 3.0%, VI부(화학공업, 28~38류) 2.8%, XVII부(수송기기, 86~89류) 1.6% 순입니다. 상위 5개 부문이 증가분의 97%를 설명합니다.',
  '· 반대로 XI부(방직용 섬유, 50~63류)는 −1.7%, XIII부(석·시멘트·도자·유리)는 −1.2%, II부(식물성 생산품)는 −5.9%로 줄었습니다. 규모는 작지만 2년 연속 감소한 부문들이라 구조적 후퇴로 볼 여지가 있습니다.',
  '· 주목할 점은 V부(광물성 생산품)의 반전입니다. 2025년 −16.3%로 크게 빠졌다가 2026년 +34.5%로 돌아섰습니다. 유가·석유제품 가격 국면이 바뀐 것으로 보이며, 같은 반전이 28류(무기화학품, −12.9%→+22.7%), 75류(니켈, −18.6%→+40.6%), 79류(아연, −14.0%→+31.1%)에서도 나타납니다. 원자재·소재 가격이 2025년 바닥을 찍고 2026년 반등한 흐름이 여러 류에서 동시에 확인됩니다.',
 ]),
 ('5. HS 류별 — 무엇이 얼마나, 그리고 그 의미',[
  '· 분석7에 97개 류를 하나도 빼지 않고 담았습니다. 각 류마다 어떤 품목이 들어가는지(주요 구성 품목), 물량·단가 중 무엇이 움직였는지(성장 유형), 2025년과 2026년의 방향이 같은지 다른지를 자동 판정한 해석 문장을 붙였습니다.',
  '· 증가액 1위는 HS85(반도체·집적회로·디스플레이·통신기기·가전)로 1,231.3억 → 2,568.7억 달러(+108.6%)입니다. 2025년에는 +5.4%에 그쳤던 품목이 이듬해 두 배가 되었습니다. 2위 HS84(컴퓨터·SSD·반도체 제조장비·일반기계)는 +75.5%인데, 2025년에는 +2.2%였습니다. 두 류 모두 2025년까지는 평범했다가 2026년에 폭발한 형태입니다.',
  '· 3위 HS27(원유·석유제품·LNG·석탄)은 2025년 −16.1% → 2026년 +34.3%로 방향을 바꿨고, 4위 HS71(금·은·백금·다이아몬드)은 2025년 +23.1%에 이어 2026년 +154.9%로 2년 연속 급증했습니다. 5위 HS89(선박·해양구조물)는 +27.2% → +23.7%로 2년 연속 안정적으로 늘었습니다. 선박은 수주 시점과 인도 시점이 몇 년 차이 나므로, 이 증가는 과거 고선가 수주분이 지금 인도되고 있다는 뜻으로 읽는 편이 정확합니다.',
  '· 감소 쪽에서 규모가 가장 큰 것은 HS87(승용차·부품)입니다. 547.1억 → 544.6억 → 539.0억 달러로 2년 연속 완만하게 줄었습니다. 수출 전체가 50% 늘어난 해에 최대 소비재 품목이 역성장했다는 사실은 이번 호황의 성격을 압축해 보여줍니다. 그밖에 HS73(강관·구조물) −4.2%, HS70(판유리·디스플레이용 유리) −11.6%, HS60(니트 원단) −7.9%도 2년 연속 감소입니다.',
  '· 흥미로운 반전 품목도 있습니다. HS86(철도차량)은 2025년 −52.0%에서 2026년 +105.4%로, HS36(화약류)은 −16.1%에서 +83.4%로 뒤집혔습니다. 규모는 작지만 수주·인도 산업의 특성상 이런 진폭이 정상이며, 추세로 읽으면 안 되는 대표적인 사례입니다.',
 ]),
 ('6. 계절성 — 같은 달끼리 겹쳐 보면 보이는 것',[
  '· 분석4와 분석10은 1~12월을 가로축에 놓고 2024·2025·2026 세 해를 겹쳐 그립니다. 이렇게 보면 계절성이 자동으로 통제되어 세 선의 간격이 곧 전년 대비 증감이 됩니다.',
  '· 2024년과 2025년 선은 대체로 붙어 있습니다(월별 격차 −10%~+13%). 반면 2026년 선은 1월부터 위로 떨어져 나가기 시작해 6월에 가장 벌어집니다. 이 "벌어짐의 시작"이 2026년 1월이라는 점이 중요합니다. 반도체 단가가 오르기 시작한 시점과 맞물립니다.',
  '· 계절 패턴 자체는 세 해가 비슷합니다. 2월이 낮고(조업일수), 12월이 높습니다. 2024년 12월 613.6억, 2025년 12월 695.1억 달러로 두 해 모두 연중 최고였습니다. 2026년 8~12월을 볼 때 이 계절성을 감안하지 않으면 과대평가하기 쉽습니다.',
  '· 수입 쪽 오버레이는 완전히 다른 그림입니다. 세 해 선이 거의 포개져 있습니다. 수출만 위로 떠오르고 수입은 제자리라는 것이 무역수지 급증의 직접적 원인입니다.',
 ]),
 ('7. 물량인가 가격인가 — 두 구간의 대조',[
  '· 2024→2025 구간과 2025→2026 구간을 같은 방식으로 분해했습니다(분석9). 두 구간의 성격이 완전히 다릅니다.',
  '· 2025→2026 : 금액 +1,996.7억 달러 가운데 가격효과가 +2,260.5억(113.2%), 물량효과가 −167.8억(−8.4%), 교차효과가 −95.9억 달러(−4.8%)입니다. 물량은 오히려 줄었고 증가는 전부 단가에서 나왔습니다.',
  '· 신성질별 중분류로 계산한 피셔 가격지수는 1.514(+51.4%), 피셔 물량지수는 0.994(−0.6%)입니다. 구성(믹스)효과는 +3.8%p에 불과해, "비싼 품목으로 갈아탄 결과"라는 설명은 성립하지 않습니다.',
  '· 실무적 함의는 분명합니다. 물량이 늘어난 성장은 설비·고용·물류를 남기지만 가격이 오른 성장은 그 대부분을 건너뜁니다. 실제로 이 기간 수출 건수는 +1.5%, 수출 중량은 −4.2%, 수입 중량은 +1.2%로 실물 활동은 거의 늘지 않았습니다. 동시에 가격 성장은 되돌림도 빠릅니다.',
 ]),
 ('8. 이 파일을 쓰는 방법과 한계',[
  '· 매월 갱신 : 원본 5개 시트에 새 달 데이터를 같은 형식으로 붙이고 「데이터_○○」 정제 시트의 참조 범위만 늘리면 12개 분석 시트와 27개 그래프가 모두 자동으로 다시 계산됩니다.',
  '· 반도체 정의 : HS 2단위 자료라 반도체(HS 8541·8542)를 단독으로 뽑을 수 없습니다. 세 가지 대용 정의를 병기했으니 목적에 맞는 것을 고르십시오. 정밀한 반도체 분석이 필요하면 HS 4단위 또는 6단위 자료를 별도로 받아야 합니다.',
  '· 품목별 시트의 과소집계 : HS 01~97류(77류는 결번)와 값이 대부분 0인 99류만 있고 98류(특수분류)가 없어 총괄보다 매년 0.03~0.05% 작습니다. 품목 간 비교에는 영향이 없지만 총액 인용 시에는 총괄 시트를 쓰십시오.',
  '· 2026년은 7개월치입니다. 연간 비교는 성립하지 않으며, 분기 분석에서 2026 Q3는 7월만 포함된 부분분기이므로 증감률 계산에서 제외했습니다.',
  '· 단가는 톤당 금액입니다. 같은 류의 시점 비교에는 유효하지만 류 간 절대 비교(예: HS85 vs HS27)는 의미가 없습니다. 중량이 극소인 류는 분석7에서 「단가 불안정」으로 따로 표시했습니다.',
  '· 국가별 자료가 없어 지역·상대국 분석은 이번 파일에 포함되지 않았습니다.',
 ]),
]
for t,lines in SECT:
    sec(w12,r,t,9); r+=1
    for ln in lines:
        para(w12,r,ln); r+=1
    r+=1
note(w12,r,1,'작성 : 업로드된 5개 원본 파일만을 근거로 산출했으며 외부 자료를 참조하지 않았습니다. 원인에 대한 해석은 데이터 밖 정보가 섞인 가설이므로 별도 검증이 필요합니다.',9)
print('분석12 완료')

# ══════════════════════════════════════════════════════════════════
# 0_안내
# ══════════════════════════════════════════════════════════════════
w0=wb.create_sheet('0_안내',0); w0.sheet_view.showGridLines=False
widths(w0,[26,70,22])
title(w0,1,'수출입 실적 통합 분석 (2024.01 ~ 2026.07, 31개월)',3)
note(w0,2,1,'업로드하신 5개 원본 파일(총괄·품목별·성질별 수출·성질별 수입·신성질별)을 한 파일로 모으고, 원본은 수정하지 않은 채 분석 시트를 추가했습니다.',3)
r=4
sec(w0,r,'시트 구성',3); r+=1
for i,h in enumerate(['시트','내용','비고']): head(w0,r,i+1,h)
r+=1
ROWS0=[('수출입 총괄','2024.01~2026.07 월별 총괄 (원본 그대로)','원본 · 수정 없음'),
 ('수출입 실적(품목별)','HS 2단위 품목별 수출입 (원본 그대로)','원본 · 수정 없음'),
 ('수출입 실적(성질별_수출)','성질별 수출 (원본 그대로)','원본 · 수정 없음'),
 ('수출입 실적(성질별_수입)','성질별 수입 (원본 그대로)','원본 · 수정 없음'),
 ('수출입 실적(신성질별)','신성질별 수출입 (원본 그대로)','원본 · 수정 없음'),
 ('데이터_○○ 5개','원본의 텍스트 숫자를 VALUE 수식으로 숫자화 + 연도·월·분기 파생','전부 수식 · 직접 입력 금지'),
 ('분석1_정합성검증','총계 vs 합계, 시트 간 교차, 월별 항등식 검증','품목별 과소집계 확인'),
 ('분석2_월별추이','31개월 추이 · MoM · YoY · 일평균 · 단가 · 12개월 이동합',''),
 ('분석3_MoM·QoQ·YoY','월별 MoM/YoY, 분기별 QoQ/YoY, 연도별 비교','요청 항목'),
 ('분석4_연도별계절비교','1~12월 × 2024/2025/2026 매트릭스 17종','오버레이 그래프 원본'),
 ('분석5_반도체vs비반도체','정의 3종 · 월별 국면 전환 추적 · 우위 판정','요청 항목'),
 ('분석6_HS부문별','HS 21개 부문 전체 + 포함 류 명시','요청 항목'),
 ('분석7_HS류별상세','HS 97개 류 전수 · 구성 품목 · 성장 유형 · 자동 해석','요청 항목'),
 ('분석8_성질별심층','성질별 수출·수입 양방향, 신성질별 수지',''),
 ('분석9_물량·가격분해','3효과 분해 + 피셔 지수, 두 구간 대조',''),
 ('분석10_그래프_연도오버레이','1~12월 축에 3개 연도를 겹친 그래프 17종','요청 항목'),
 ('분석11_그래프_구조분석','반도체/비반도체, 부문별, 분기, 분해 그래프 10종',''),
 ('분석12_해설','핵심 수치와 해석 · 사용법 · 한계','먼저 읽으시면 좋습니다')]
for a,b,c in ROWS0:
    put(w0,r,1,a,bold=True); put(w0,r,2,b); put(w0,r,3,c,align='center'); r+=1
r+=1
sec(w0,r,'이번 분석에서 새로 들어간 것',3); r+=1
for t in ['① MoM(전월비) · QoQ(전분기비) · YoY(전년동기비) 3종 세트 — 분석3',
          '② 1~12월을 한 축에 놓고 2024·2025·2026을 겹쳐 보는 오버레이 그래프 17종 — 분석4·분석10',
          '③ 반도체 vs 비반도체 구분 분석 (정의 3종, 월별 우위 판정) — 분석5',
          '④ HS 부문(Section) 21개 전체 + HS 류 97개 전수 분석 (구성 품목·증감 의미 포함) — 분석6·분석7',
          '⑤ 성질별 수입 자료가 새로 들어와 수출·수입 양방향 성질별 분해 — 분석8']:
    put(w0,r,1,t,border=False); w0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
r+=1
note(w0,r,1,'※ 모든 분석 셀은 원본을 참조하는 수식입니다. 원본 시트에 다음 달 데이터를 같은 형식으로 추가하고 정제 시트의 참조 범위를 늘리면 분석 전체가 그대로 갱신됩니다.',3)
print('0_안내 완료')
