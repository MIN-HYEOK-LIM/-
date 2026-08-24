# -*- coding: utf-8 -*-
"""1단계 : 원본 5개 시트를 무수정 통합"""
import openpyxl, warnings
from openpyxl.styles import Font
warnings.filterwarnings('ignore')
FONT='맑은 고딕'
SRC=[('수출입 총괄','9997c13b-_______20260824.xlsx'),
     ('수출입 실적(품목별)','f5647f1d-__________20260824.xlsx'),
     ('수출입 실적(성질별_수출)','35650bd7-__________20260824___.xlsx'),
     ('수출입 실적(성질별_수입)','eeda0a5d-__________20260824___.xlsx'),
     ('수출입 실적(신성질별)','61119590-___________20260824.xlsx')]
out=openpyxl.Workbook(); out.remove(out.active)
for title,f in SRC:
    s=openpyxl.load_workbook(f).active
    d=out.create_sheet(title)
    for r in range(1,s.max_row+1):
        for c in range(1,s.max_column+1):
            v=s.cell(r,c).value
            if v is not None:
                cc=d.cell(r,c,v); cc.font=Font(name=FONT,sz=10,b=(r==5)); cc.number_format='@'
    for k,dim in s.column_dimensions.items():
        if dim.width: d.column_dimensions[k].width=dim.width
    d.freeze_panes='A6'
    print(title, s.max_row, s.max_column)
out.save('s1.xlsx'); print('saved s1')
