import openpyxl, warnings, collections, json
warnings.filterwarnings('ignore')
def num(v):
    s=(v or '').strip().replace(',','')
    if s in ('','-'): return 0.0
    return float(s)
def txt(v): return (v or '').strip()
def sheet(f):
    return openpyxl.load_workbook(f).active

F={'총괄':'d6c978fa-_______20260823.xlsx','품목별':'9acf3c62-__________20260823.xlsx',
   '성질별':'58654b7f-__________20260823.xlsx','신성질별':'860505c9-___________20260823.xlsx',
   '국가별':'6f515e2a-__________20260823.xlsx'}

# 총괄
ws=sheet(F['총괄'])
TOT={}; MONTHS=[]
for r in range(7,ws.max_row+1):
    p=txt(ws.cell(r,1).value); MONTHS.append(p)
    TOT[p]=[num(ws.cell(r,c).value) for c in range(2,10)]
TOTAL_ROW=[num(ws.cell(6,c).value) for c in range(2,10)]

# 성질별
ws=sheet(F['성질별']); OLD=collections.defaultdict(dict)
for r in range(7,ws.max_row+1):
    OLD[txt(ws.cell(r,3).value)][txt(ws.cell(r,1).value)]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value))
OLD_TOTAL=(num(ws.cell(6,4).value),num(ws.cell(6,5).value))
# 신성질별
ws=sheet(F['신성질별']); NEW=collections.defaultdict(dict)
for r in range(7,ws.max_row+1):
    NEW[txt(ws.cell(r,2).value)][txt(ws.cell(r,1).value)]=(num(ws.cell(r,3).value),num(ws.cell(r,4).value),num(ws.cell(r,5).value),num(ws.cell(r,6).value))
NEW_TOTAL=[num(ws.cell(6,c).value) for c in range(3,7)]
# 품목별
ws=sheet(F['품목별']); HS=collections.defaultdict(dict); HSNAME={}
for r in range(7,ws.max_row+1):
    code=txt(ws.cell(r,2).value); HSNAME[code]=txt(ws.cell(r,3).value)
    HS[code][txt(ws.cell(r,1).value)]=(num(ws.cell(r,4).value),num(ws.cell(r,5).value),num(ws.cell(r,6).value),num(ws.cell(r,7).value))
HS_TOTAL=[num(ws.cell(6,c).value) for c in range(4,9)]
# 국가별
ws=sheet(F['국가별']); CTY=collections.defaultdict(dict)
for r in range(7,ws.max_row+1):
    CTY[txt(ws.cell(r,2).value)][txt(ws.cell(r,1).value)]=(num(ws.cell(r,3).value),num(ws.cell(r,4).value),num(ws.cell(r,5).value),num(ws.cell(r,6).value),num(ws.cell(r,7).value))
CTY_TOTAL=[num(ws.cell(6,c).value) for c in range(3,8)]

