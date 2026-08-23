# -*- coding: utf-8 -*-
"""2단계 : 숫자 정제 시트(전부 수식) 추가"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment
warnings.filterwarnings('ignore')
FONT='맑은 고딕'
HDR=PatternFill('solid', fgColor='548235')
NUM='#,##0'; NUM1='#,##0.0'

wb=openpyxl.load_workbook('stage1.xlsx')

def V(ref):   # 텍스트 숫자 -> 숫자
    return f'=IFERROR(VALUE(SUBSTITUTE(TRIM({ref}),",","")),"")'
def T(ref):
    return f'=TRIM({ref})'

def make(name, src, first, last, cols, note):
    """cols: [(헤더, 원본열 또는 None, kind)] kind: 'text'|'num'|'formula'"""
    ws=wb.create_sheet(name)
    ws.sheet_view.showGridLines=False
    c=ws.cell(1,1,note); c.font=Font(name=FONT, sz=9, i=True, color='595959')
    for i,(h,sc,kind) in enumerate(cols):
        cc=ws.cell(2,i+1,h); cc.font=Font(name=FONT,sz=10,b=True,color='FFFFFF'); cc.fill=HDR
        cc.alignment=Alignment(horizontal='center')
    r0=3
    for k,sr in enumerate(range(first,last+1)):
        r=r0+k
        for i,(h,sc,kind) in enumerate(cols):
            col=i+1
            if kind=='text':
                cc=ws.cell(r,col,T(f"'{src}'!{sc}{sr}")); cc.number_format='@'
            elif kind=='year':
                cc=ws.cell(r,col,f'=IFERROR(VALUE(LEFT(A{r},4)),"")'); cc.number_format='0'
            elif kind=='month':
                cc=ws.cell(r,col,f'=IFERROR(VALUE(RIGHT(A{r},2)),"")'); cc.number_format='0'
            elif kind=='num':
                cc=ws.cell(r,col,V(f"'{src}'!{sc}{sr}")); cc.number_format=NUM
            elif kind=='num1':
                cc=ws.cell(r,col,V(f"'{src}'!{sc}{sr}")); cc.number_format=NUM1
            elif kind=='diff':
                cc=ws.cell(r,col,f'={sc}{r}'); cc.number_format=NUM
            cc.font=Font(name=FONT,sz=10)
    ws.freeze_panes='A3'
    for i,(h,sc,kind) in enumerate(cols):
        ws.column_dimensions[get_col(i+1)].width = 30 if kind=='text' and i>2 else 13
    return ws, r0, r0+(last-first)

from openpyxl.utils import get_column_letter as get_col

NOTE='이 시트의 모든 셀은 원본 시트를 참조하는 수식입니다. 원본은 숫자가 텍스트(쉼표 포함)로 저장되어 있어 VALUE/SUBSTITUTE로 숫자화했습니다. 직접 입력하지 마세요.'

ws,a,b=make('데이터_총괄','수출입 총괄',7,25,
    [('기간','A','text'),('연도',None,'year'),('월',None,'month'),
     ('조업일수','B','num1'),('수출 건수','C','num'),('수출 중량','D','num1'),('수출 금액','E','num'),
     ('수입 건수','F','num'),('수입 중량','G','num1'),('수입 금액','H','num'),('무역수지','I','num')], NOTE)
print('총괄',a,b)
ws,a,b=make('데이터_성질별','수출입 실적(성질별)',7,405,
    [('기간','A','text'),('연도',None,'year'),('월',None,'month'),
     ('성질명','C','text'),('수출 중량','D','num1'),('수출 금액','E','num')], NOTE)
print('성질별',a,b)
ws,a,b=make('데이터_신성질별','수출입 실적(신성질별)',7,329,
    [('기간','A','text'),('연도',None,'year'),('월',None,'month'),
     ('신성질명','B','text'),('수출 중량','C','num1'),('수출 금액','D','num'),
     ('수입 중량','E','num1'),('수입 금액','F','num')], NOTE)
print('신성질별',a,b)
ws,a,b=make('데이터_품목별','수출입 실적(품목별)',7,1832,
    [('기간','A','text'),('연도',None,'year'),('월',None,'month'),
     ('HS코드','B','text'),('품목명','C','text'),('수출 금액','E','num'),
     ('수입 금액','G','num'),('무역수지','H','num'),('수출 중량','D','num1')], NOTE)
print('품목별',a,b)
ws,a,b=make('데이터_국가별','수출입 실적(국가별)',7,4481,
    [('기간','A','text'),('연도',None,'year'),('월',None,'month'),
     ('국가','B','text'),('수출 금액','D','num'),('수입 금액','F','num'),('무역수지','G','num'),
     ('수출 건수','C','num'),('수입 건수','E','num')], NOTE)
print('국가별',a,b)
wb.save('stage2.xlsx')
print('saved')
