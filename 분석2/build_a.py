# -*- coding: utf-8 -*-
"""1단계 : 원본 5개 시트 통합 + 숫자 정제 시트(수식) 생성"""
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

FONT='맑은 고딕'
SRC={'수출입 총괄':'d6c978fa-_______20260823.xlsx',
     '수출입 실적(품목별)':'9acf3c62-__________20260823.xlsx',
     '수출입 실적(성질별)':'58654b7f-__________20260823.xlsx',
     '수출입 실적(신성질별)':'860505c9-___________20260823.xlsx',
     '수출입 실적(국가별)':'6f515e2a-__________20260823.xlsx'}

out=openpyxl.Workbook(); out.remove(out.active)
meta={}
for title,f in SRC.items():
    s=openpyxl.load_workbook(f).active
    d=out.create_sheet(title)
    for r in range(1,s.max_row+1):
        for c in range(1,s.max_column+1):
            v=s.cell(r,c).value
            if v is not None:
                cc=d.cell(r,c,v); cc.font=Font(name=FONT, sz=10, b=(r==5))
                cc.number_format='@'
    for k,dim in s.column_dimensions.items():
        if dim.width: d.column_dimensions[k].width=dim.width
    d.freeze_panes='A6'
    meta[title]=(s.max_row, s.max_column)
    print(title, s.max_row, s.max_column)
out.save('stage1.xlsx')
print(meta)
