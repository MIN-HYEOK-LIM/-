# -*- coding: utf-8 -*-
# 수출 심층 분석 시트 (build_c.py 네임스페이스에서 exec)
NEW_SUB=[n for i,n in enumerate(NEW_ITEMS) if i not in NEW_MAJ]
IT=['다.IT제품','라.IT부품']
GC=['중국','홍콩','대만']
nrng=lambda c: f'{DN}!${c}${DN_R[0]}:${c}${DN_R[1]}'
hrng=lambda c: f'{DH}!${c}${DH_R[0]}:${c}${DH_R[1]}'
crng=lambda c: f'{DC}!${c}${DC_R[0]}:${c}${DC_R[1]}'
def S_N(col,name,yr):  # 신성질별 기간합
    return f'SUMIFS({nrng(col)},{nrng("D")},"{name}",{nrng("B")},{yr},{nrng("C")},"<=7")'
def S_Nm(col,name,m):
    return f'SUMIFS({nrng(col)},{nrng("D")},"{name}",{nrng("A")},"{m}")'
def S_H(col,code,yr):
    return f'SUMIFS({hrng(col)},{hrng("D")},$A{{r}},{hrng("B")},{yr},{hrng("C")},"<=7")'
EXM=lambda k: f"'{A2}'!E{D2+k}"      # 분석2 월별 수출
E25=f"'{A2}'!E{SUMROWS['2025.01~07']}"; E26=f"'{A2}'!E{SUMROWS['2026.01~07']}"
W25=f"'{A2}'!D{SUMROWS['2025.01~07']}"; W26=f"'{A2}'!D{SUMROWS['2026.01~07']}"

# ══════════════════════════════════════════════════════════════════
# 분석8_물량가격분해
# ══════════════════════════════════════════════════════════════════
w8=wb.create_sheet('분석8_물량·가격분해'); w8.sheet_view.showGridLines=False
widths(w8,[24,15,15,15,12,12,12,15,15,15,15,12])
title(w8,1,'⑧ 수출 증가의 분해 — 물량이 늘어서인가, 값이 올라서인가',12)
note(w8,2,1,'금액 = 물량 × 단가이므로 금액 변화는 물량효과 (Q₁−Q₀)×P₀ + 가격효과 (P₁−P₀)×Q₀ + 교차효과 (Q₁−Q₀)×(P₁−P₀) 로 정확히 분해됩니다. 세 값의 합은 금액 증감과 일치합니다.',12)
note(w8,3,1,'단가 = 금액(천달러) ÷ 중량(톤). 톤당 단가이므로 제품 가격 그 자체는 아니지만, 같은 품목 안에서 시점 간 비교에는 유효한 대리지표입니다.',12)
r=5
sec(w8,r,'[A] 총괄 수출 (2026.01~07 vs 2025.01~07)',12); r+=1
for i,h in enumerate(['구분','2025.01~07','2026.01~07','증감','증감률']): head(w8,r,i+1,h)
r+=1
A8=r
for nm,a,b,fmt in [('수출 금액(천$)',E25,E26,NUM),('수출 중량(톤)',W25,W26,NUM1)]:
    put(w8,r,1,nm,bold=True); put(w8,r,2,'='+a,fmt); put(w8,r,3,'='+b,fmt)
    put(w8,r,4,f'=C{r}-B{r}',fmt); put(w8,r,5,f'=C{r}/B{r}-1',PCT); r+=1
put(w8,r,1,'단가(천$/톤)',bold=True); put(w8,r,2,f'=B{A8}/B{A8+1}',UNIT); put(w8,r,3,f'=C{A8}/C{A8+1}',UNIT)
put(w8,r,4,f'=C{r}-B{r}',UNIT); put(w8,r,5,f'=C{r}/B{r}-1',PCT)
P_ROW=r; r+=2
for i,h in enumerate(['효과','금액(천$)','금액 증감 대비','해석']): head(w8,r,i+1,h)
r+=1
DEC=r
put(w8,r,1,'물량효과 (Q₁−Q₀)×P₀',bold=True); put(w8,r,2,f'=(C{A8+1}-B{A8+1})*B{P_ROW}',NUM)
put(w8,r,3,f'=B{r}/(C{A8}-B{A8})',PCT); put(w8,r,4,'물량이 그대로였다면 얼마였을까'); r+=1
put(w8,r,1,'가격효과 (P₁−P₀)×Q₀',bold=True); put(w8,r,2,f'=(C{P_ROW}-B{P_ROW})*B{A8+1}',NUM)
put(w8,r,3,f'=B{r}/(C{A8}-B{A8})',PCT); put(w8,r,4,'단가만 올랐다면 얼마였을까'); r+=1
put(w8,r,1,'교차효과',bold=True); put(w8,r,2,f'=(C{A8+1}-B{A8+1})*(C{P_ROW}-B{P_ROW})',NUM)
put(w8,r,3,f'=B{r}/(C{A8}-B{A8})',PCT); put(w8,r,4,'물량·단가 동시 변화분'); r+=1
put(w8,r,1,'합계 = 금액 증감',bold=True); put(w8,r,2,f'=SUM(B{DEC}:B{r-1})',NUM,bold=True)
put(w8,r,3,f'=B{r}/(C{A8}-B{A8})',PCT,bold=True); put(w8,r,4,f'=IF(ABS(B{r}-(C{A8}-B{A8}))<1,"분해 검증 OK","불일치")',align='center')
r+=2
sec(w8,r,'[B] 신성질별 중분류별 분해 (수출, 중량 보유 품목)',12); r+=1
for i,h in enumerate(['품목','중량 25(톤)','금액 25','중량 26(톤)','금액 26','물량 증감률','단가 증감률','금액 증감률',
                      '물량효과','가격효과','교차효과','Δ금액']): head(w8,r,i+1,h)
w8.row_dimensions[r].height=30
r+=1
B8=r
for nm in NEW_SUB:
    put(w8,r,1,nm,bold=True)
    put(w8,r,2,'='+S_N('E',nm,2025),NUM1); put(w8,r,3,'='+S_N('F',nm,2025),NUM)
    put(w8,r,4,'='+S_N('E',nm,2026),NUM1); put(w8,r,5,'='+S_N('F',nm,2026),NUM)
    put(w8,r,6,f'=IFERROR(D{r}/B{r}-1,"-")',PCT)
    put(w8,r,7,f'=IFERROR((E{r}/D{r})/(C{r}/B{r})-1,"-")',PCT)
    put(w8,r,8,f'=IFERROR(E{r}/C{r}-1,"-")',PCT)
    put(w8,r,9,f'=IFERROR((D{r}-B{r})*(C{r}/B{r}),"-")',NUM)
    put(w8,r,10,f'=IFERROR(((E{r}/D{r})-(C{r}/B{r}))*B{r},"-")',NUM)
    put(w8,r,11,f'=IFERROR((D{r}-B{r})*((E{r}/D{r})-(C{r}/B{r})),"-")',NUM)
    put(w8,r,12,f'=E{r}-C{r}',NUM)
    r+=1
E8=r-1
put(w8,r,1,'합계',bold=True)
for col in [2,3,4,5,9,10,11,12]:
    put(w8,r,col,f'=SUM({L(col)}{B8}:{L(col)}{E8})', NUM1 if col in (2,4) else NUM, bold=True)
put(w8,r,6,f'=D{r}/B{r}-1',PCT,bold=True); put(w8,r,7,f'=(E{r}/D{r})/(C{r}/B{r})-1',PCT,bold=True)
put(w8,r,8,f'=E{r}/C{r}-1',PCT,bold=True)
SUM8=r; r+=1
note(w8,r,1,'합계행의 단가 증감률(G열)은 중량 단순합 기준이므로, 저가·중량물(광산물 등)의 비중 변화에 크게 좌우됩니다. 구성 변화를 제거한 값은 아래 [C]의 물량·가격지수를 보십시오.',12)
r+=2
sec(w8,r,'[C] 지수 분해 — 구성 변화를 제거한 순수 물량·가격 (라스파이레스·파셰·피셔)',12); r+=1
for i,h in enumerate(['지수','값','증감률','산식','의미']): head(w8,r,i+1,h)
r+=1
IDX8=r
q0p0=f'SUMPRODUCT($B${B8}:$B${E8},$C${B8}:$C${E8}/$B${B8}:$B${E8})'
q1p0=f'SUMPRODUCT($D${B8}:$D${E8},$C${B8}:$C${E8}/$B${B8}:$B${E8})'
q1p1=f'SUM($E${B8}:$E${E8})'
q0p1=f'SUMPRODUCT($B${B8}:$B${E8},$E${B8}:$E${E8}/$D${B8}:$D${E8})'
rows8=[('라스파이레스 물량지수',f'={q1p0}/{q0p0}','기준연도 가격 고정, 물량만 변화'),
       ('파셰 물량지수',      f'={q1p1}/{q0p1}','비교연도 가격 고정, 물량만 변화'),
       ('피셔 물량지수',      None,'두 지수의 기하평균(권장)'),
       ('라스파이레스 가격지수',f'={q0p1}/{q0p0}','기준연도 물량 고정, 가격만 변화'),
       ('파셰 가격지수',      f'={q1p1}/{q1p0}','비교연도 물량 고정, 가격만 변화'),
       ('피셔 가격지수',      None,'두 지수의 기하평균(권장)'),
       ('금액지수',           f'={q1p1}/{q0p0}','피셔 물량지수 × 피셔 가격지수와 같아야 함'),
       ]
for i,(nm,f_,mean) in enumerate(rows8):
    put(w8,r,1,nm,bold=True)
    if nm=='피셔 물량지수': put(w8,r,2,f'=SQRT(B{IDX8}*B{IDX8+1})','0.0000',bold=True)
    elif nm=='피셔 가격지수': put(w8,r,2,f'=SQRT(B{IDX8+3}*B{IDX8+4})','0.0000',bold=True)
    else: put(w8,r,2,f_,'0.0000')
    put(w8,r,3,f'=B{r}-1',PCT)
    put(w8,r,4,'Σ(q₁p₀)/Σ(q₀p₀) 형태의 표준 지수식')
    put(w8,r,5,mean)
    r+=1
FQ=IDX8+2; FP=IDX8+5
put(w8,r,1,'검증 : 피셔물량×피셔가격',bold=True); put(w8,r,2,f'=B{FQ}*B{FP}','0.0000')
put(w8,r,3,f'=IF(ABS(B{r}-B{IDX8+6})<0.001,"금액지수와 일치","불일치")',align='center'); r+=1
put(w8,r,1,'단순 단가 변화(중량합 기준)',bold=True); put(w8,r,2,f'=(E{SUM8}/D{SUM8})/(C{SUM8}/B{SUM8})','0.0000')
put(w8,r,3,f'=B{r}-1',PCT); put(w8,r,5,'구성 변화가 섞여 있는 값'); r+=1
put(w8,r,1,'구성(믹스)효과',bold=True); put(w8,r,2,f'=B{r-1}/B{FP}','0.0000')
put(w8,r,3,f'=B{r}-1',PCT,color=RED); put(w8,r,5,'단순 단가 상승분 중 품목 구성 변화가 만든 몫'); r+=2
sec(w8,r,'[D] 월별 단가 추이 (천달러/톤) — 전체·IT·비IT',12); r+=1
U8H=r; head(w8,r,1,'구분')
for k in range(NM): head(w8,r,2+k,M[k])
r+=1
U8=r
put(w8,r,1,'전체 수출 단가',bold=True)
for k in range(NM): put(w8,r,2+k,f"='{A2}'!O{D2+k}",UNIT)
r+=1
for nm in IT:
    put(w8,r,1,f'{nm} 단가',bold=True)
    for k in range(NM): put(w8,r,2+k,f'=IFERROR({S_Nm("F",nm,M[k])}/{S_Nm("E",nm,M[k])},"")',UNIT)
    r+=1
put(w8,r,1,'IT 제외 단가',bold=True)
for k in range(NM):
    ex='+'.join(S_Nm('F',n,M[k]) for n in IT); wt='+'.join(S_Nm('E',n,M[k]) for n in IT)
    put(w8,r,2+k,f"=IFERROR(({EXM(k)}-({ex}))/(('{A2}'!D{D2+k})-({wt})),\"\")",UNIT)
r+=1
put(w8,r,1,'전체 단가 지수(2025.01=100)',bold=True)
for k in range(NM): put(w8,r,2+k,f'=B{U8}*0+{L(2+k)}{U8}/$B${U8}*100','0.0')
U8E=r
w8.freeze_panes='B6'

# ══════════════════════════════════════════════════════════════════
# 분석9_품목단가분해(HS)
# ══════════════════════════════════════════════════════════════════
w9=wb.create_sheet('분석9_품목단가분해(HS)'); w9.sheet_view.showGridLines=False
widths(w9,[7,34,14,14,14,14,11,11,11,11,15,15,15,15,10])
title(w9,1,'⑨ 품목별(HS) 물량·단가 분해 — 어떤 품목이 값으로 올랐고 어떤 품목이 물량으로 늘었나',15)
note(w9,2,1,'HS 2단위 기준 수출입니다. 단가 = 수출금액 ÷ 수출중량. 정렬은 금액 증감액 순(작성 시점 고정). 중량이 0이거나 극소인 품목은 단가가 불안정하므로 판정열에 표시했습니다.',15)
r=4
for i,h in enumerate(['HS','품목명','중량 25(톤)','금액 25','중량 26(톤)','금액 26','물량 증감률','단가 증감률',
                      '금액 증감률','기여도','물량효과','가격효과','교차효과','Δ금액','성장 유형']): head(w9,r,i+1,h)
w9.row_dimensions[r].height=32
r+=1
H9=r
hs_order9=[k for _,k in sorted(((sum(HS[k][m][1] for m in M26 if m in HS[k])-sum(HS[k][m][1] for m in M25_7 if m in HS[k]),k) for k in HS), reverse=True)]
for code in hs_order9:
    put(w9,r,1,code,align='center').number_format='@'
    put(w9,r,2,f'=IFERROR(INDEX({hrng("E")},MATCH($A{r},{hrng("D")},0)),"")')
    put(w9,r,3,f'=SUMIFS({hrng("I")},{hrng("D")},$A{r},{hrng("B")},2025,{hrng("C")},"<=7")',NUM1)
    put(w9,r,4,f'=SUMIFS({hrng("F")},{hrng("D")},$A{r},{hrng("B")},2025,{hrng("C")},"<=7")',NUM)
    put(w9,r,5,f'=SUMIFS({hrng("I")},{hrng("D")},$A{r},{hrng("B")},2026,{hrng("C")},"<=7")',NUM1)
    put(w9,r,6,f'=SUMIFS({hrng("F")},{hrng("D")},$A{r},{hrng("B")},2026,{hrng("C")},"<=7")',NUM)
    put(w9,r,7,f'=IFERROR(E{r}/C{r}-1,"-")',PCT)
    put(w9,r,8,f'=IFERROR((F{r}/E{r})/(D{r}/C{r})-1,"-")',PCT)
    put(w9,r,9,f'=IFERROR(F{r}/D{r}-1,"-")',PCT)
    put(w9,r,10,f'=IFERROR((F{r}-D{r})/({E26}-{E25}),"-")',PCT2)
    put(w9,r,11,f'=IFERROR((E{r}-C{r})*(D{r}/C{r}),"-")',NUM)
    put(w9,r,12,f'=IFERROR(((F{r}/E{r})-(D{r}/C{r}))*C{r},"-")',NUM)
    put(w9,r,13,f'=IFERROR((E{r}-C{r})*((F{r}/E{r})-(D{r}/C{r})),"-")',NUM)
    put(w9,r,14,f'=F{r}-D{r}',NUM)
    put(w9,r,15,f'=IF(OR(C{r}<100,E{r}<100),"단가 불안정(중량 미미)",'
                f'IF(N{r}<0,"감소",IF(L{r}>ABS(K{r})*2,"가격 주도",IF(K{r}>ABS(L{r})*2,"물량 주도","혼합"))))',align='center')
    r+=1
E9=r-1
put(w9,r,2,'합계',bold=True)
for col in [3,5]: put(w9,r,col,f'=SUM({L(col)}{H9}:{L(col)}{E9})',NUM1,bold=True)
for col in [4,6,11,12,13,14]: put(w9,r,col,f'=SUM({L(col)}{H9}:{L(col)}{E9})',NUM,bold=True)
put(w9,r,7,f'=E{r}/C{r}-1',PCT,bold=True); put(w9,r,8,f'=(F{r}/E{r})/(D{r}/C{r})-1',PCT,bold=True)
put(w9,r,9,f'=F{r}/D{r}-1',PCT,bold=True); put(w9,r,10,f'=(F{r}-D{r})/({E26}-{E25})',PCT2,bold=True)
SUM9=r; r+=2
sec(w9,r,'[요약] 성장 유형별 집계',15); r+=1
for i,h in enumerate(['유형','품목 수','Δ금액 합','전체 증가분 대비']): head(w9,r,i+1,h)
r+=1
T9=r
for t in ['가격 주도','물량 주도','혼합','감소','단가 불안정(중량 미미)']:
    put(w9,r,1,t,bold=True)
    put(w9,r,2,f'=COUNTIF($O${H9}:$O${E9},A{r})',NUM)
    put(w9,r,3,f'=SUMIF($O${H9}:$O${E9},A{r},$N${H9}:$N${E9})',NUM)
    put(w9,r,4,f'=IFERROR(C{r}/({E26}-{E25}),"-")',PCT2)
    r+=1
w9.freeze_panes='C5'

# ══════════════════════════════════════════════════════════════════
# 분석10_집중도·확산도
# ══════════════════════════════════════════════════════════════════
w10=wb.create_sheet('분석10_집중도·확산도'); w10.sheet_view.showGridLines=False
widths(w10,[12,12,12,12,12,12,12,12,11,11,11,12,12,12])
title(w10,1,'⑩ 집중도와 확산도 — 성장이 좁아졌는가, 넓어졌는가',14)
note(w10,2,1,'HHI(허핀달지수) = Σ(점유율²)×10,000. 값이 클수록 집중. 유효 품목수 = 10,000÷HHI로, "실질적으로 몇 개 품목이 수출을 지탱하는가"를 뜻합니다(동일 규모 품목 환산 개수).',14)
note(w10,3,1,'상위 N 비중은 2026.01~07 기준 고정 바스켓입니다(월별 상위 재계산이 아님). 확산지수는 전년동월 대비 증가한 품목/국가의 비율로, 성장이 몇 개 품목에 몰렸는지를 봅니다.',14)
r=5
sec(w10,r,'[A] 월별 집중도',14); r+=1
for i,h in enumerate(['기간','품목 HHI','유효 품목수','국가 HHI','유효 국가수','상위3 품목 비중','상위5 국가 비중',
                      '상위10 국가 비중','IT 비중','중화권 비중','미국 비중','수출 총액']): head(w10,r,i+1,h)
w10.row_dimensions[r].height=30
r+=1
C10=r
top3hs=HS_ORDER[:3]; top5c=CTY_ORDER[:5]; top10c=CTY_ORDER[:10]
for k in range(NM):
    m=M[k]; tot=EXM(k)
    put(w10,r,1,f"='{A2}'!A{D2+k}",align='center',bold=True)
    put(w10,r,2,f'=SUMPRODUCT(({hrng("A")}="{m}")*({hrng("F")})^2)/({tot})^2*10000',NUM)
    put(w10,r,3,f'=10000/B{r}','0.0')
    put(w10,r,4,f'=SUMPRODUCT(({crng("A")}="{m}")*({crng("E")})^2)/({tot})^2*10000',NUM)
    put(w10,r,5,f'=10000/D{r}','0.0')
    s3='+'.join(f'SUMIFS({hrng("F")},{hrng("D")},"{c}",{hrng("A")},"{m}")' for c in top3hs)
    put(w10,r,6,f'=({s3})/{tot}',PCT2)
    s5='+'.join(f'SUMIFS({crng("E")},{crng("D")},"{c}",{crng("A")},"{m}")' for c in top5c)
    put(w10,r,7,f'=({s5})/{tot}',PCT2)
    s10='+'.join(f'SUMIFS({crng("E")},{crng("D")},"{c}",{crng("A")},"{m}")' for c in top10c)
    put(w10,r,8,f'=({s10})/{tot}',PCT2)
    sit='+'.join(S_Nm('F',n,m) for n in IT)
    put(w10,r,9,f'=({sit})/{tot}',PCT2)
    sgc='+'.join(f'SUMIFS({crng("E")},{crng("D")},"{c}",{crng("A")},"{m}")' for c in GC)
    put(w10,r,10,f'=({sgc})/{tot}',PCT2)
    put(w10,r,11,f'=SUMIFS({crng("E")},{crng("D")},"미국",{crng("A")},"{m}")/{tot}',PCT2)
    put(w10,r,12,f'={tot}',NUM)
    r+=1
E10=r-1
put(w10,r,1,'2025.01→2026.07 변화',bold=True)
for col in [2,3,4,5]: put(w10,r,col,f'={L(col)}{E10}-{L(col)}{C10}','+#,##0.0;-#,##0.0')
for col in [6,7,8,9,10,11]: put(w10,r,col,f'=({L(col)}{E10}-{L(col)}{C10})*100',PP)
r+=2
sec(w10,r,'[B] 확산지수 — 전년 동월 대비 증가한 품목·국가의 비율 (2026.01~07)',14); r+=1
for i,h in enumerate(['기간','증가 품목 수','전체 품목 수','품목 확산지수','금액가중 확산지수','증가 국가 수(상위50)','국가 확산지수','해석']): head(w10,r,i+1,h)
r+=1
DIF=r
HS_D=[c for c in HS_ORDER]
CT_D=CTY_ORDER[:50]
# 헬퍼 블록 위치는 아래에 생성 (품목 96 x 7, 국가 50 x 7)
HELP_HS=E10+40; HELP_CT=HELP_HS+len(HS_D)+6
for j,m in enumerate(M26):
    cl=L(2+j)
    put(w10,r,1,m,align='center',bold=True)
    put(w10,r,2,f'=SUM({cl}${HELP_HS+2}:{cl}${HELP_HS+1+len(HS_D)})',NUM)
    put(w10,r,3,f'=COUNT({cl}${HELP_HS+2}:{cl}${HELP_HS+1+len(HS_D)})',NUM)
    put(w10,r,4,f'=B{r}/C{r}',PCT2)
    wcl=L(2+7+j)
    put(w10,r,5,f'=SUMPRODUCT({cl}${HELP_HS+2}:{cl}${HELP_HS+1+len(HS_D)},{wcl}${HELP_HS+2}:{wcl}${HELP_HS+1+len(HS_D)})'
                f'/SUM({wcl}${HELP_HS+2}:{wcl}${HELP_HS+1+len(HS_D)})',PCT2)
    put(w10,r,6,f'=SUM({cl}${HELP_CT+2}:{cl}${HELP_CT+1+len(CT_D)})',NUM)
    put(w10,r,7,f'=F{r}/50',PCT2)
    put(w10,r,8,f'=IF(E{r}>0.9,"금액 기준으로는 사실상 전면 확산",IF(D{r}<0.5,"증가 품목이 절반 미만 - 편중","혼재"))')
    r+=1
note(w10,r,1,'품목 확산지수(개수 기준)와 금액가중 확산지수(금액 기준)를 함께 보십시오. 개수는 낮은데 금액가중이 높으면 "큰 품목은 다 늘고 작은 품목만 줄었다"는 뜻입니다.',14)
r+=2
# 헬퍼 블록
sec(w10,HELP_HS,'[보조] 품목별 전년동월 대비 증가 여부(1/0)와 당월 금액 — 확산지수 계산용',16)
hr=HELP_HS+1
head(w10,hr,1,'HS')
for j,m in enumerate(M26): head(w10,hr,2+j,m+' 증가')
for j,m in enumerate(M26): head(w10,hr,9+j,m+' 금액')
for i,code in enumerate(HS_D):
    rr=hr+1+i
    put(w10,rr,1,code,align='center',sz=9).number_format='@'
    for j,m in enumerate(M26):
        prev='2025.'+m[5:]
        cur=f'SUMIFS({hrng("F")},{hrng("D")},$A{rr},{hrng("A")},"{m}")'
        pre=f'SUMIFS({hrng("F")},{hrng("D")},$A{rr},{hrng("A")},"{prev}")'
        put(w10,rr,2+j,f'=IF({pre}<=0,"",IF({cur}>{pre},1,0))','0',sz=9)
        put(w10,rr,9+j,f'={cur}',NUM,sz=9)
sec(w10,HELP_CT,'[보조] 국가별(상위 50) 전년동월 대비 증가 여부(1/0) — 확산지수 계산용',16)
cr=HELP_CT+1
head(w10,cr,1,'국가')
for j,m in enumerate(M26): head(w10,cr,2+j,m+' 증가')
for i,c in enumerate(CT_D):
    rr=cr+1+i
    put(w10,rr,1,c,sz=9)
    for j,m in enumerate(M26):
        prev='2025.'+m[5:]
        cur=f'SUMIFS({crng("E")},{crng("D")},$A{rr},{crng("A")},"{m}")'
        pre=f'SUMIFS({crng("E")},{crng("D")},$A{rr},{crng("A")},"{prev}")'
        put(w10,rr,2+j,f'=IF({pre}<=0,0,IF({cur}>{pre},1,0))','0',sz=9)
w10.freeze_panes='B7'

# ══════════════════════════════════════════════════════════════════
# 분석11_모멘텀·계절성
# ══════════════════════════════════════════════════════════════════
w11=wb.create_sheet('분석11_모멘텀·계절성'); w11.sheet_view.showGridLines=False
widths(w11,[12,15,10,11,14,11,10,12,11,14,12,15,15,11,11,12,10])
title(w11,1,'⑪ 모멘텀과 계절성 — 지금 속도는 빨라지고 있는가',17)
note(w11,2,1,'3개월 이동평균 전년동월비는 단월 변동을 걷어낸 추세, 전월비 연율화는 이 속도가 1년 지속될 경우의 연간 증가율입니다(과잉 해석 주의).',17)
note(w11,3,1,'계절지수는 2025년 12개월의 일평균 수출을 100으로 환산한 값이며, 2026년에는 같은 달의 2025년 지수를 적용했습니다. 자료가 19개월뿐이라 정밀 계절조정(X-13 등)은 불가능하므로 근사치입니다.',17)
H11=['기간','수출 금액(천$)','전년동월비','3개월이평 전년동월비','일평균 수출(천$)','일평균 전년동월비',
     '전월비','전월비 연율화','계절지수','계절조정 일평균','계절조정 전월비','IT 수출','IT 제외 수출',
     'IT 전년동월비','IT 제외 전년동월비','IT 제외 3개월이평 YoY','IT 비중']
r=5
for i,h in enumerate(H11): head(w11,r,i+1,h)
w11.row_dimensions[r].height=40
r+=1
D11=r
dt=lambda k,c: f'{DT}!{c}{DT_R[0]+k}'
for k in range(NM):
    m=M[k]; rr=D11+k
    put(w11,rr,1,f"='{A2}'!A{D2+k}",align='center',bold=True)
    put(w11,rr,2,f'={EXM(k)}',NUM)
    put(w11,rr,3,'' if k<12 else f'=B{rr}/B{rr-12}-1',PCT)
    put(w11,rr,4,'' if k<14 else f'=SUM(B{rr-2}:B{rr})/SUM(B{rr-14}:B{rr-12})-1',PCT)
    put(w11,rr,5,f'=B{rr}/{dt(k,"D")}',NUM)
    put(w11,rr,6,'' if k<12 else f'=E{rr}/E{rr-12}-1',PCT)
    put(w11,rr,7,'' if k==0 else f'=B{rr}/B{rr-1}-1',PCT)
    put(w11,rr,8,'' if k==0 else f'=(E{rr}/E{rr-1})^12-1',PCT)
    mn=int(m[5:])
    y25=f'SUMIFS({DT}!G${DT_R[0]}:G${DT_R[1]},{DT}!B${DT_R[0]}:B${DT_R[1]},2025,{DT}!C${DT_R[0]}:C${DT_R[1]},{mn})'
    d25=f'SUMIFS({DT}!D${DT_R[0]}:D${DT_R[1]},{DT}!B${DT_R[0]}:B${DT_R[1]},2025,{DT}!C${DT_R[0]}:C${DT_R[1]},{mn})'
    yy=f'SUMIFS({DT}!G${DT_R[0]}:G${DT_R[1]},{DT}!B${DT_R[0]}:B${DT_R[1]},2025)'
    dd=f'SUMIFS({DT}!D${DT_R[0]}:D${DT_R[1]},{DT}!B${DT_R[0]}:B${DT_R[1]},2025)'
    put(w11,rr,9,f'=(({y25})/({d25}))/(({yy})/({dd}))*100','0.0')
    put(w11,rr,10,f'=E{rr}/I{rr}*100',NUM)
    put(w11,rr,11,'' if k==0 else f'=J{rr}/J{rr-1}-1',PCT)
    sit='+'.join(S_Nm('F',n,m) for n in IT)
    put(w11,rr,12,f'={sit}',NUM)
    put(w11,rr,13,f'=B{rr}-L{rr}',NUM)
    put(w11,rr,14,'' if k<12 else f'=L{rr}/L{rr-12}-1',PCT)
    put(w11,rr,15,'' if k<12 else f'=M{rr}/M{rr-12}-1',PCT)
    put(w11,rr,16,'' if k<14 else f'=SUM(M{rr-2}:M{rr})/SUM(M{rr-14}:M{rr-12})-1',PCT)
    put(w11,rr,17,f'=L{rr}/B{rr}',PCT2)
E11=D11+NM-1
r=E11+1
for lbl,a,b in [('2025.01~07',D11,D11+6),('2026.01~07',D11+12,E11),('2025년 연간',D11,D11+11)]:
    put(w11,r,1,lbl,bold=True,align='center')
    for col in [2,12,13]: put(w11,r,col,f'=SUM({L(col)}{a}:{L(col)}{b})',NUM,bold=True)
    put(w11,r,5,f'=B{r}/SUM({DT}!D{DT_R[0]+(a-D11)}:D{DT_R[0]+(b-D11)})',NUM,bold=True)
    put(w11,r,17,f'=L{r}/B{r}',PCT2,bold=True)
    r+=1
S25,S26,S25Y=E11+1,E11+2,E11+3
put(w11,r,1,'동기 증감률',bold=True,align='center')
for col in [2,5,12,13]: put(w11,r,col,f'={L(col)}{S26}/{L(col)}{S25}-1',PCT,bold=True)
put(w11,r,17,f'=(Q{S26}-Q{S25})*100',PP,bold=True)
r+=2
sec(w11,r,'[B] 잔여기간 시나리오 — 2026년 8~12월을 어떻게 가정하는가',17); r+=1
note(w11,r,1,'노란색 셀만 바꾸면 아래 표가 전부 다시 계산됩니다. 기준선은 2026.05~07 3개월 평균입니다.',17); r+=1
put(w11,r,1,'기준선(최근 3개월 평균 수출)',bold=True)
put(w11,r,2,f'=AVERAGE(B{E11-2}:B{E11})',NUM)
BASE11=r; r+=1
put(w11,r,1,'2025년 연간 수출',bold=True); put(w11,r,2,f'=B{S25Y}',NUM); Y25=r; r+=1
put(w11,r,1,'2026년 1~7월 누계',bold=True); put(w11,r,2,f'=B{S26}',NUM); Y26=r; r+=2
for i,h in enumerate(['시나리오','8~12월 월평균 가정(기준선 대비)','8~12월 월평균(천$)','2026년 연간 수출','2025년 대비','연간 증감액']): head(w11,r,i+1,h)
r+=1
SC=r
for nm,v in [('낙관 : 기준선 +10%',0.10),('기준 : 기준선 유지',0.0),('완만한 조정 : −10%',-0.10),
             ('뚜렷한 조정 : −25%',-0.25),('급락 : −40%',-0.40)]:
    put(w11,r,1,nm,bold=True)
    c=put(w11,r,2,v,PCT,color=BLUE); c.fill=IN
    put(w11,r,3,f'=$B${BASE11}*(1+B{r})',NUM)
    put(w11,r,4,f'=$B${Y26}+C{r}*5',NUM)
    put(w11,r,5,f'=D{r}/$B${Y25}-1',PCT)
    put(w11,r,6,f'=D{r}-$B${Y25}',NUM)
    r+=1
note(w11,r,1,'참고 : 2026년 1~7월 누계만으로도 2025년 연간 수출의 83.9% 수준입니다(=1~7월÷2025년 연간). 8~12월이 40% 급락해도 연간으로는 큰 폭 증가가 유지된다는 뜻이며, 이것이 이른바 캐리오버(이월) 효과입니다.',17)
CARRY=r; r+=1
put(w11,r,1,'1~7월 누계 ÷ 2025년 연간',bold=True); put(w11,r,2,f'=B{Y26}/B{Y25}',PCT)
w11.freeze_panes='B6'

# ══════════════════════════════════════════════════════════════════
# 분석12_국가심층
# ══════════════════════════════════════════════════════════════════
w12=wb.create_sheet('분석12_국가심층'); w12.sheet_view.showGridLines=False
widths(w12,[6,18,15,11,11,11,12,12,11,11,11,11,9,9,9,24])
title(w12,1,'⑫ 국가 심층 — 지역 분산이 품목 분산을 대신할 수 있는가',16)
note(w12,2,1,'「HS85 동조성」은 각국 월별 수출과 HS85류(전기기기·반도체) 월별 수출의 19개월 상관계수입니다. 1에 가까울수록 그 나라 수출이 반도체 사이클과 함께 움직였다는 뜻이며, 값이 높은 나라들로만 분산해서는 품목 리스크가 줄지 않습니다.',16)
note(w12,3,1,'건당 수출금액 = 수출금액 ÷ 수출건수. 건수가 줄면서 건당 금액이 오르면 고가품 위주로 재편되었다는 신호입니다. 순위는 분석5의 상위 40개국 바스켓 내 순위입니다.',16)
r=5
for i,h in enumerate(['순위','국가','수출 26.1~7','증감률','점유율 26','점유율 변화','건당금액 25','건당금액 26',
                      '건당금액 변화','건수 증감률','HS85 동조성','월증감 변동성','25순위','26순위','순위변동','성격']): head(w12,r,i+1,h)
w12.row_dimensions[r].height=32
r+=1
C12=r
TOP12=CTY_ORDER[:20]
MAT12=C12+len(TOP12)+4
R5C,R5E=C5,E5     # 분석5 상위 40 범위
for i,c in enumerate(TOP12):
    rr=C12+i; mrow=MAT12+2+i
    q25=f'SUMIFS({crng("H")},{crng("D")},$B{rr},{crng("B")},2025,{crng("C")},"<=7")'
    q26=f'SUMIFS({crng("H")},{crng("D")},$B{rr},{crng("B")},2026,{crng("C")},"<=7")'
    put(w12,rr,1,i+1,align='center'); put(w12,rr,2,c,bold=True)
    put(w12,rr,3,f"='{A5}'!D{cty_row[c]}",NUM)
    put(w12,rr,4,f"='{A5}'!F{cty_row[c]}",PCT)
    put(w12,rr,5,f"='{A5}'!I{cty_row[c]}",PCT2)
    put(w12,rr,6,f"='{A5}'!J{cty_row[c]}",PP)
    put(w12,rr,7,f"=IFERROR('{A5}'!C{cty_row[c]}/({q25}),\"-\")",'0.0')
    put(w12,rr,8,f"=IFERROR('{A5}'!D{cty_row[c]}/({q26}),\"-\")",'0.0')
    put(w12,rr,9,f'=IFERROR(H{rr}/G{rr}-1,"-")',PCT)
    put(w12,rr,10,f'=IFERROR(({q26})/({q25})-1,"-")',PCT)
    put(w12,rr,11,f'=IFERROR(CORREL(${L(2)}${mrow}:${L(1+NM)}${mrow},${L(2)}${MAT12+1}:${L(1+NM)}${MAT12+1}),"-")','0.000')
    put(w12,rr,12,f'=STDEV({L(2)}{mrow}:{L(1+NM)}{mrow})/AVERAGE({L(2)}{mrow}:{L(1+NM)}{mrow})',PCT)
    put(w12,rr,13,f"=RANK('{A5}'!C{cty_row[c]},'{A5}'!$C${R5C}:$C${R5E})",'0')
    put(w12,rr,14,f"=RANK('{A5}'!D{cty_row[c]},'{A5}'!$D${R5C}:$D${R5E})",'0')
    put(w12,rr,15,f'=M{rr}-N{rr}','+0;-0;0')
    put(w12,rr,16,f'=IF(K{rr}>0.9,"반도체 동조 매우 강함",IF(K{rr}>0.8,"반도체 동조 강함",'
                  f'IF(K{rr}>0.6,"부분 동조","독립적 흐름")))',align='center')
E12=C12+len(TOP12)-1
r=E12+2
sec(w12,r,'[보조] 상위 20개국 및 HS85류 월별 수출 (상관계수·변동성 계산용, 천달러)',1+NM); r+=1
head(w12,MAT12,1,'구분')
for k in range(NM): head(w12,MAT12,2+k,M[k])
put(w12,MAT12+1,1,'HS85 전기기기',bold=True)
for k in range(NM):
    put(w12,MAT12+1,2+k,f'=SUMIFS({hrng("F")},{hrng("D")},"85",{hrng("A")},"{M[k]}")',NUM,sz=9)
for i,c in enumerate(TOP12):
    rr=MAT12+2+i
    put(w12,rr,1,c,sz=9)
    for k in range(NM):
        put(w12,rr,2+k,f'=SUMIFS({crng("E")},{crng("D")},$A{rr},{crng("A")},"{M[k]}")',NUM,sz=9)
MAT12_END=MAT12+1+len(TOP12)
r=MAT12_END+2
sec(w12,r,'[요약] 동조성 구간별 수출 집계',16); r+=1
for i,h in enumerate(['구간','국가 수','수출 26.1~7 합','상위20 대비 비중']): head(w12,r,i+1,h)
r+=1
for lo,hi,nm in [(0.9,9,'0.90 이상 (매우 강함)'),(0.8,0.9,'0.80~0.90'),(0.6,0.8,'0.60~0.80'),(-9,0.6,'0.60 미만')]:
    put(w12,r,1,nm,bold=True)
    put(w12,r,2,f'=COUNTIFS($K${C12}:$K${E12},">={lo}",$K${C12}:$K${E12},"<{hi}")',NUM)
    put(w12,r,3,f'=SUMIFS($C${C12}:$C${E12},$K${C12}:$K${E12},">={lo}",$K${C12}:$K${E12},"<{hi}")',NUM)
    put(w12,r,4,f'=C{r}/SUM($C${C12}:$C${E12})',PCT2)
    r+=1
w12.freeze_panes='C6'

# ══════════════════════════════════════════════════════════════════
# 분석13_민감도
# ══════════════════════════════════════════════════════════════════
w13=wb.create_sheet('분석13_민감도'); w13.sheet_view.showGridLines=False
widths(w13,[30,16,16,16,16,16,14,14])
title(w13,1,'⑬ 민감도 — 무엇이 흔들리면 얼마나 흔들리는가 (2026.01~07 실적 기준)',8)
note(w13,2,1,'노란색 셀만 바꾸면 전체가 다시 계산됩니다. 단가 충격은 물량이 그대로일 때 금액이 그만큼 줄어드는 것으로 가정한 단순 시뮬레이션이며, 수요·환율의 2차 효과는 반영하지 않았습니다.',8)
r=4
sec(w13,r,'[A] 기준 실적',8); r+=1
base=[('수출 (26.1~7)',f'={E26}',NUM),('수입 (26.1~7)',f"='{A2}'!F{SUMROWS['2026.01~07']}",NUM),
      ('무역수지',None,NUM),('IT(IT부품+IT제품) 수출',None,NUM),('중화권(중국·홍콩·대만) 수출',None,NUM),
      ('IT 제외 수출',None,NUM)]
B13=r
for nm,f_,fmt in base:
    put(w13,r,1,nm,bold=True); r+=1
put(w13,B13,2,f'={E26}',NUM); put(w13,B13+1,2,f"='{A2}'!F{SUMROWS['2026.01~07']}",NUM)
put(w13,B13+2,2,f'=B{B13}-B{B13+1}',NUM)
put(w13,B13+3,2,f"='{A2}'!E{SUMROWS['2026.01~07']}*0+"+'+'.join(S_N('F',n,2026) for n in IT),NUM)
put(w13,B13+4,2,'='+'+'.join(f'SUMIFS({crng("E")},{crng("D")},"{c}",{crng("B")},2026,{crng("C")},"<=7")' for c in GC),NUM)
put(w13,B13+5,2,f'=B{B13}-B{B13+3}',NUM)
for k in range(6): put(w13,B13+k,3,f'=B{B13+k}/B{B13}',PCT2)
r=B13+7
sec(w13,r,'[B] 사용자 시나리오 (노란색 입력)',8); r+=1
IN13=r
inputs=[('IT 수출단가 변화',-0.20),('IT 제외 수출 변화',0.05),('중화권 수요 변화(중복 반영 방지용, 0 권장)',0.0),('수입 변화',0.0)]
for nm,v in inputs:
    put(w13,r,1,nm,bold=True); c=put(w13,r,2,v,PCT,color=BLUE); c.fill=IN; r+=1
put(w13,r,1,'→ 조정 후 수출',bold=True)
put(w13,r,2,f'=B{B13+3}*(1+B{IN13})+B{B13+5}*(1+B{IN13+1})+B{B13+4}*B{IN13+2}',NUM,bold=True)
put(w13,r,3,f'=B{r}/B{B13}-1',PCT,bold=True); OUT13=r; r+=1
put(w13,r,1,'→ 조정 후 수입',bold=True); put(w13,r,2,f'=B{B13+1}*(1+B{IN13+3})',NUM)
put(w13,r,3,f'=B{r}/B{B13+1}-1',PCT); r+=1
put(w13,r,1,'→ 조정 후 무역수지',bold=True); put(w13,r,2,f'=B{OUT13}-B{r-1}',NUM,bold=True)
put(w13,r,3,f'=B{r}/B{B13+2}-1',PCT,bold=True,color=RED); r+=1
put(w13,r,1,'→ 전년 동기(25.1~7) 대비 수출',bold=True)
put(w13,r,2,f'=B{OUT13}/{E25}-1',PCT,bold=True); r+=2
sec(w13,r,'[C] 표준 시나리오 표 — IT 수출단가 충격만 반영',8); r+=1
for i,h in enumerate(['IT 단가 충격','수출','수출 증감률','전년동기비','무역수지','수지 증감률','수지 감소액']): head(w13,r,i+1,h)
r+=1
for sh in [0,-0.10,-0.20,-0.30,-0.40,-0.50]:
    put(w13,r,1,sh,PCT,bold=True)
    put(w13,r,2,f'=$B${B13}+$B${B13+3}*A{r}',NUM)
    put(w13,r,3,f'=B{r}/$B${B13}-1',PCT)
    put(w13,r,4,f'=B{r}/{E25}-1',PCT)
    put(w13,r,5,f'=B{r}-$B${B13+1}',NUM)
    put(w13,r,6,f'=E{r}/$B${B13+2}-1',PCT)
    put(w13,r,7,f'=E{r}-$B${B13+2}',NUM)
    r+=1
r+=1
sec(w13,r,'[D] 임계점 계산',8); r+=1
for nm,f_,fmt,desc in [
    ('무역수지가 0이 되는 IT 단가 하락폭',f'=-B{B13+2}/B{B13+3}',PCT,'IT 수출금액이 이만큼 줄면 흑자가 사라집니다'),
    ('수출이 전년 동기 수준으로 돌아가는 IT 단가 하락폭',f'=({E25}-B{B13})/B{B13+3}',PCT,'이만큼 줄어야 증가율이 0%가 됩니다'),
    ('무역수지가 2025년 동기(수준)로 돌아가는 하락폭',f"=(('{A2}'!G{SUMROWS['2025.01~07']}) - B{B13+2})/B{B13+3}",PCT,'전년 동기 흑자 수준으로의 회귀'),
    ('중화권 수요가 0이 될 때 수출 감소율',f'=-B{B13+4}/B{B13}',PCT,'극단 가정 - 상한 참고치')]:
    put(w13,r,1,nm,bold=True); put(w13,r,2,f_,fmt); put(w13,r,3,desc); r+=1

# ══════════════════════════════════════════════════════════════════
# 분석14_원인분석
# ══════════════════════════════════════════════════════════════════
w14=wb.create_sheet('분석14_원인분석'); w14.sheet_view.showGridLines=False
widths(w14,[26,20,34,34,30,10])
title(w14,1,'⑭ 원인 분석 — 왜 이런 숫자가 나왔는가',6)
note(w14,2,1,'왼쪽 두 열은 이 파일의 데이터로 확인된 사실입니다. 세 번째 열부터는 데이터 밖의 환경 요인에 대한 해석(가설)이며, 다섯 번째 열에 검증 방법을 함께 적었습니다. 사실과 해석을 구분해서 보십시오.',6)
note(w14,4,1,'「대표 수치」 열은 왼쪽 문장에 나오는 수치 중 하나를 해당 분석 시트에서 수식으로 가져온 것입니다(문장 속 모든 수치를 담지는 않습니다).',6)
note(w14,3,1,'작성자 주 : 외부 환경 서술은 일반적으로 알려진 시장 흐름에 근거한 것이며, 2026년 6~7월의 개별 사건까지 확인한 것은 아닙니다. 신뢰도 열은 데이터 정합성과 일반적 인과 관계를 함께 고려한 주관적 평가입니다.',6)
r=6
for i,h in enumerate(['관측된 사실','대표 수치(수식)','데이터 내부 근거','유력한 원인(해석)','검증 방법','신뢰도']): head(w14,r,i+1,h)
w14.row_dimensions[r].height=28
r+=1
def cw(row, fact, ref, fmt, inside, cause, verify, conf):
    put(w14,row,1,fact,bold=True).alignment=Alignment(vertical='top',wrap_text=True)
    put(w14,row,2,ref,fmt)
    for col,txt in [(3,inside),(4,cause),(5,verify)]:
        c=put(w14,row,col,txt); c.alignment=Alignment(vertical='top',wrap_text=True); c.font=Font(name=FONT,sz=9)
    put(w14,row,6,conf,align='center')
    w14.row_dimensions[row].height=max(30,14*(max(len(inside),len(cause),len(verify))//34+1))

HS85_ROW=[i for i,c in enumerate(hs_order9) if c=='85'][0]+H9
HS84_ROW=[i for i,c in enumerate(hs_order9) if c=='84'][0]+H9
HS71_ROW=[i for i,c in enumerate(hs_order9) if c=='71'][0]+H9
HS27_ROW=[i for i,c in enumerate(hs_order9) if c=='27'][0]+H9
HS87_ROW=[i for i,c in enumerate(hs_order9) if c=='87'][0]+H9
HS89_ROW=[i for i,c in enumerate(hs_order9) if c=='89'][0]+H9
W9="'분석9_품목단가분해(HS)'"; W8="'분석8_물량·가격분해'"; W10="'분석10_집중도·확산도'"
W11="'분석11_모멘텀·계절성'"; W12="'분석12_국가심층'"; W13="'분석13_민감도'"
CASES=[
 ('수출 +50.5% · 중량 −4.2% · 단가 +57.2%', f'={W8}!E9', PCT,
  '물량효과 −168억 달러, 가격효과 +2,260억 달러, 교차효과 −96억 달러. 금액 증가의 113%가 가격효과(분석8-A).',
  '수출 증가는 더 많이 판 결과가 아니라 같은 물량을 훨씬 비싸게 판 결과입니다. 단가가 사이클성 변수인 만큼 이 증가는 구조적 체력 향상과 구분해야 합니다.',
  '분석8-A의 3효과 분해와 [C]의 피셔지수(물량 −0.6%, 가격 +51.4%)를 함께 보십시오.','높음'),
 ('구성효과는 +3.8%p뿐, 순수 가격효과가 +51.4%', f'={W8}!C{IDX8+5}', PCT,
  '피셔 가격지수 1.514, 피셔 물량지수 0.994. 단순 단가 상승(+57.2%) 중 품목 구성 변화 몫은 3.8%p에 불과(분석8-C).',
  '"비싼 품목으로 갈아탄 것"보다 "팔던 품목의 값이 오른 것"이 지배적입니다. 즉 수출 구조 고도화보다 가격 사이클의 영향이 큽니다.',
  '지수 계산은 신성질별 중분류 14개 기준입니다. 품목 세분화 수준을 HS 4단위로 낮추면 구성효과가 다소 커질 수 있습니다.','높음'),
 ('HS85류 : 물량 −0.7%, 단가 +110.0%', f'={W9}!H{HS85_ROW}', PCT,
  '금액 +1,337억 달러 중 가격효과 +1,355억, 물량효과 −8억(분석9). 신성질별 IT부품 단가도 +141.3%.',
  '메모리 반도체 가격 급등이 사실상 유일한 설명 변수입니다. AI 데이터센터 투자 확대에 따른 HBM·서버용 DRAM 수요 급증과, 범용 DRAM 생산능력이 HBM으로 전환되며 발생한 공급 부족이 함께 작용한 것으로 봅니다.',
  'DRAM·NAND 고정거래가격 지수, 반도체 수출물량·단가 지수(한국은행·산업부), 주요 메모리 업체 분기 실적의 ASP 추이와 대조하십시오.','높음'),
 ('HS84류(기계)도 단가 +74.0%', f'={W9}!H{HS84_ROW}', PCT,
  '물량은 +0.9%에 그침. 반면 성질별 「기계류와 정밀기기」 단가는 −0.7%로 정반대(분석3·분석9).',
  'HS84에는 컴퓨터·SSD(8471)와 반도체 제조장비가 포함됩니다. 메모리 가격 상승이 저장장치·모듈 형태로 84류에도 반영된 것으로 보이며, 순수 일반기계는 정체 상태로 판단됩니다.',
  'HS 4단위(8471, 8473, 8486)로 재조회해 84류 내부를 분리해 보십시오. 이 파일은 2단위까지만 있어 확정할 수 없습니다.','중간'),
 ('HS71류 : 단가 +215.6%, 물량 −19.2%', f'={W9}!H{HS71_ROW}', PCT,
  '금액 +60억 달러(기여도 3.0%), 순위 16위 → 10위. 월별 변동성 26.6%로 매우 큼(분석9·분석12 방식).',
  '국제 금 가격 급등이 유력합니다. 물량이 오히려 20% 줄었는데 금액이 155% 늘어난 패턴은 가격 요인 외에 설명하기 어렵습니다.',
  '국제 금 현물가격, HS 7108(금) 세부 실적과 대조. 귀금속은 재수출·중계 성격도 있어 방향성 해석에 주의가 필요합니다.','중간'),
 ('HS27류 : 단가 +46.7%, 물량 −8.5%', f'={W9}!H{HS27_ROW}', PCT,
  '금액 +93억 달러. 원자재 수입단가(+16.8%)보다 수출단가 상승폭이 큼.',
  '석유제품 수출단가 상승(유가·정제마진)과 정기보수에 따른 물량 감소가 동시에 작용한 것으로 보입니다.',
  '두바이유 가격, 복합정제마진, 정유사 가동률·정기보수 일정과 대조하십시오.','중간'),
 ('HS87류(차량) : −1.0%, 물량 −2.2%', f'={W9}!I{HS87_ROW}', PCT,
  '신성질별 내구소비재도 −7.2%. 수출 상위 10개 품목 중 유일한 감소(분석4·분석9).',
  '주요 시장의 관세 부담과 현지생산 확대가 겹친 결과로 해석됩니다. 반도체 호황이 자동차로 전이되지 않았다는 점이 중요합니다.',
  '대미·대EU 자동차 수출 단가와 물량, 현지공장 생산 실적을 함께 보십시오.','중간'),
 ('홍콩 +161.9%, 건당금액 +126.6%', f"='{A5}'!F{cty_row['홍콩']}", PCT,
  'HS85 동조성 0.949, 월 변동성 49.7%로 상위 20개국 중 최고 수준(분석12).',
  '최종 소비지라기보다 반도체의 중계·환적 경로로 보는 편이 타당합니다. 급증분에는 실수요와 재고·통관 이동이 섞여 있을 수 있습니다.',
  '홍콩 경유 물량의 최종 목적지 통계, 중국 본토 직수출과의 합산 추이를 함께 보십시오.','중간'),
 ('미국 : 금액 +53.1%인데 건수 −11.0%', f'={W12}!J{C12+1}', PCT,
  '건당 수출금액이 46.2 → 79.5 천달러로 +72.0%(분석12).',
  '건수가 줄고 건당 금액이 커지는 패턴은 소액·다건 화물이 줄고 고가품 중심으로 재편되었음을 시사합니다. 관세 환경 변화와 소액 화물 취급 축소가 배경일 수 있습니다.',
  '대미 수출을 HS별·금액구간별로 분해하고, 관세 부과 대상 품목과 예외 품목을 나누어 비교하십시오.','중간'),
 ('상위 5개국 비중 54.8% → 62.1%', f'={W10}!G{E10}', PCT2,
  '국가 HHI 896 → 1,028, 유효 국가수 11.2 → 9.7개. 품목 HHI는 1,334 → 2,444, 유효 품목수 7.5 → 4.1개(분석10).',
  '품목 집중이 지역 집중을 끌고 간 구조입니다. 반도체 공급망이 중화권·동남아에 몰려 있어, 반도체가 커질수록 상대국도 자동으로 좁아집니다.',
  '분석12의 HS85 동조성을 보십시오. 상위 20개국 중 13개국이 상관계수 0.8 이상이며, 이들이 상위 20개국 수출의 90%를 차지합니다.','높음'),
 ('품목 확산지수는 낮은데 금액가중 확산은 99%', f'={W10}!E{DIF+5}', PCT2,
  '2026.06 기준 증가 품목 76/96개(79.2%)인데 금액가중 확산지수는 99.0%(분석10-B).',
  '"반도체만 좋다"는 통념과 달리, 금액이 큰 품목은 거의 모두 늘고 있습니다. 다만 개수 기준으로는 감소 품목이 33개(전체의 34%)로 적지 않습니다.',
  '분석9의 성장 유형 집계(가격 주도 32 / 물량 주도 11 / 혼합 12 / 감소 33)와 함께 보십시오.','높음'),
 ('IT 제외 수출도 +11.1%, 최근 2개월 +17.8%', f'={W11}!M{E11+4}', PCT,
  'IT 제외 수출 2,738억 → 3,043억 달러. 2026.06·07 전년동월비 +17.7%, +17.8%로 가속(분석11).',
  '수출 호조가 반도체 단일 요인만은 아니라는 반증입니다. 다만 절대 기여도는 IT가 압도적이며(증가분의 85%), 비IT는 낮은 기저에서의 회복 성격이 강합니다.',
  '분석11의 IT 제외 3개월이평 전년동월비가 계속 상승하는지 매월 확인하십시오. 이 지표가 꺾이면 호황이 실제로 반도체 단일 축이 됩니다.','높음'),
 ('12월 계절지수 109.6, 2월 89.9 (일평균 기준)', f'={W11}!I{D11+11}', '0.0',
  '조업일수를 보정한 뒤에도 연말이 높고 연초가 낮은 계절성이 남아 있음(분석11).',
  '연말 인도·회계연도 마감 효과가 남아 있는 것으로 보입니다. 월별 실적을 전월과 단순 비교하면 이 계절성 때문에 오독하기 쉽습니다.',
  '자료가 19개월뿐이라 계절지수는 2025년 1년치로만 산출한 근사치입니다. 2~3년치가 쌓이면 재추정하십시오.','중간'),
 ('무역수지 흑자의 57.7%가 IT 단가에 달려 있음', f'={W13}!B32', PCT,
  'IT 수출금액이 57.7% 줄면 무역수지 흑자가 0이 됩니다. 20% 감소만으로도 흑자는 34.7% 축소(분석13).',
  '흑자 규모 자체가 반도체 가격의 함수라는 뜻입니다. 재정·환율 정책 판단에서 이 흑자를 항상적 체력으로 보면 위험합니다.',
  '분석13의 노란색 셀에 실제 가격 전망치를 넣어 보십시오. 임계점은 자동으로 다시 계산됩니다.','높음'),
]
for fact,ref,fmt,inside,cause,verify,conf in CASES:
    cw(r,fact,ref,fmt,inside,cause,verify,conf); r+=1
r+=1
sec(w14,r,'해석에 붙이는 단서',6); r+=1
for t in ['· 이 표의 「원인」은 데이터 밖의 정보를 끌어와 붙인 해석입니다. 같은 데이터가 다른 원인으로도 설명될 수 있으므로, 다섯 번째 열의 검증 방법을 실제로 수행한 뒤 확정하시기 바랍니다.',
          '· 특히 2026년 6~7월에 벌어진 개별 사건(가격 협상, 관세 조치, 대형 인도 건 등)은 확인하지 않았습니다. 최근 2개월 해석은 잠정으로 두십시오.',
          '· 단가는 톤당 금액입니다. 같은 품목 안에서 시점 비교에는 유효하지만, 품목 간 절대 비교(예: IT부품 1,165 vs 광산물 1.19)는 의미가 없습니다.',
          '· 이 파일의 수치가 관세청 원자료와 일치하는지는 별도로 확인하지 않았습니다. 대외 공표용으로 쓰기 전에 원자료 대조를 권합니다.']:
    c=put(w14,r,1,t,border=False); c.alignment=Alignment(vertical='top',wrap_text=True)
    w14.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    w14.row_dimensions[r].height=max(16,14*(len(t)//80+1)); r+=1
w14.freeze_panes='A6'

# ══════════════════════════════════════════════════════════════════
# 분석15_그래프2 (심층 분석 그래프)
# ══════════════════════════════════════════════════════════════════
from openpyxl.chart import ScatterChart, Series
w15=wb.create_sheet('분석15_그래프2'); w15.sheet_view.showGridLines=False
widths(w15,[26]+[13]*20)
title(w15,1,'⑮ 심층 분석 그래프 — 분해·집중도·모멘텀·민감도',21)
note(w15,2,1,'보조표는 분석8~13을 참조하는 수식입니다.',21)
r=4
sec(w15,r,'보조표 A. 총괄 물량·가격 분해 (천달러)',21); r+=1
GA_H=r; head(w15,r,1,'효과'); head(w15,r,2,'금액')
r+=1; GA=r
for i,nm in enumerate(['물량효과','가격효과','교차효과']):
    put(w15,r,1,nm,bold=True); put(w15,r,2,f'={W8}!B{DEC+i}',NUM); r+=1
GA_E=r-1; r+=1
sec(w15,r,'보조표 B. 주요 품목 물량효과·가격효과 (신성질별 중분류, 천달러)',21); r+=1
GB_H=r; head(w15,r,1,'품목'); head(w15,r,2,'물량효과'); head(w15,r,3,'가격효과')
r+=1; GB=r
order_sub=sorted(range(len(NEW_SUB)), key=lambda i: -(sum(NEW[NEW_SUB[i]][m][1] for m in M26)-sum(NEW[NEW_SUB[i]][m][1] for m in M25_7)))[:8]
for i in order_sub:
    put(w15,r,1,NEW_SUB[i]); put(w15,r,2,f'={W8}!I{B8+i}',NUM); put(w15,r,3,f'={W8}!J{B8+i}',NUM); r+=1
GB_E=r-1; r+=1
sec(w15,r,'보조표 C. 국가별 HS85 동조성 vs 수출 증가율 (산점도용)',21); r+=1
GC_H=r; head(w15,r,1,'국가'); head(w15,r,2,'HS85 동조성'); head(w15,r,3,'수출 증가율')
r+=1; GC_S=r
for i,c in enumerate(TOP12):
    put(w15,r,1,c); put(w15,r,2,f'={W12}!K{C12+i}','0.000'); put(w15,r,3,f'={W12}!D{C12+i}',PCT); r+=1
GC_E=r-1; r+=1
sec(w15,r,'보조표 D. 성장 유형별 Δ금액 (천달러)',21); r+=1
GD_H=r; head(w15,r,1,'유형'); head(w15,r,2,'Δ금액')
r+=1; GD=r
for i,t in enumerate(['가격 주도','물량 주도','혼합','감소','단가 불안정(중량 미미)']):
    put(w15,r,1,t); put(w15,r,2,f'={W9}!C{T9+i}',NUM); r+=1
GD_E=r-1
CT15=GD_E+3

def add15(ch,anchor,t,w=30,h=13,ylab=None,xlab=None,legend=True,off=0):
    flat(ch,off); ch.title=t; ch.width=w; ch.height=h; ch.style=2
    if ylab: ch.y_axis.title=ylab
    if xlab: ch.x_axis.title=xlab
    if not legend: ch.legend=None
    ch.y_axis.majorGridlines=ChartLines()
    w15.add_chart(ch,anchor)

# 1. 물량·가격 분해
b=BarChart(); b.type='col'
b.add_data(Reference(w15,min_col=2,min_row=GA_H,max_row=GA_E),titles_from_data=True)
b.set_categories(Reference(w15,min_col=1,min_row=GA,max_row=GA_E)); b.legend=None
add15(b,f'A{CT15}','1. 수출 증가의 분해 — 물량효과 vs 가격효과 (26.1~7 vs 25.1~7)',ylab='금액(천달러)',legend=False)
# 2. 품목별 효과
b2=BarChart(); b2.type='bar'; b2.grouping='clustered'
b2.add_data(Reference(w15,min_col=2,max_col=3,min_row=GB_H,max_row=GB_E),titles_from_data=True)
b2.set_categories(Reference(w15,min_col=1,min_row=GB,max_row=GB_E))
add15(b2,f'A{CT15+26}','2. 주요 품목의 물량효과·가격효과',h=14,ylab='금액(천달러)')
# 3. 단가 추이
l3=LineChart()
l3.add_data(Reference(w8,min_col=1,max_col=1+NM,min_row=U8,max_row=U8+3),from_rows=True,titles_from_data=True)
l3.set_categories(Reference(w8,min_col=2,max_col=1+NM,min_row=U8H))
add15(l3,f'A{CT15+54}','3. 월별 수출단가 추이 — 전체 · IT부품 · IT제품 · IT 제외',ylab='천달러/톤')
# 4. 집중도
l4=LineChart()
l4.add_data(Reference(w10,min_col=2,min_row=C10-1,max_row=E10),titles_from_data=True)
l4.set_categories(Reference(w10,min_col=1,min_row=C10,max_row=E10))
l4b=LineChart()
l4b.add_data(Reference(w10,min_col=3,min_row=C10-1,max_row=E10),titles_from_data=True)
l4b.y_axis.axId=200; l4b.y_axis.title='유효 품목수(개)'; l4b.y_axis.crosses='max'; sub(l4b,3); l4+=l4b
add15(l4,f'A{CT15+80}','4. 품목 집중도 — HHI와 유효 품목수',ylab='HHI')
# 5. 확산지수
l5=LineChart()
l5.add_data(Reference(w10,min_col=4,max_col=5,min_row=DIF-1,max_row=DIF+6),titles_from_data=True)
l5.add_data(Reference(w10,min_col=7,min_row=DIF-1,max_row=DIF+6),titles_from_data=True)
l5.set_categories(Reference(w10,min_col=1,min_row=DIF,max_row=DIF+6))
add15(l5,f'A{CT15+106}','5. 확산지수 — 개수 기준 vs 금액가중 기준 (2026.01~07)',ylab='비율')
# 6. IT vs 비IT
l6=LineChart()
l6.add_data(Reference(w11,min_col=14,max_col=15,min_row=5,max_row=E11),titles_from_data=True)
l6.set_categories(Reference(w11,min_col=1,min_row=D11,max_row=E11))
add15(l6,f'A{CT15+132}','6. IT vs IT 제외 — 전년 동월 대비 증감률',ylab='전년동월비')
# 7. 계절조정
l7=LineChart()
l7.add_data(Reference(w11,min_col=5,min_row=5,max_row=E11),titles_from_data=True)
l7.add_data(Reference(w11,min_col=10,min_row=5,max_row=E11),titles_from_data=True)
l7.set_categories(Reference(w11,min_col=1,min_row=D11,max_row=E11))
add15(l7,f'A{CT15+158}','7. 일평균 수출과 계절조정 일평균',ylab='천달러/일')
# 8. 민감도
b8=BarChart(); b8.type='col'
b8.add_data(Reference(w13,min_col=2,min_row=23,max_row=29),titles_from_data=True)
b8.add_data(Reference(w13,min_col=5,min_row=23,max_row=29),titles_from_data=True)
b8.set_categories(Reference(w13,min_col=1,min_row=24,max_row=29))
add15(b8,f'A{CT15+184}','8. IT 수출단가 충격에 따른 수출·무역수지 (26.1~7 기준)',ylab='금액(천달러)',xlab='IT 단가 충격')
# 9. 산점도
sc=ScatterChart(); sc.x_axis.title='HS85 동조성(상관계수)'; sc.y_axis.title='수출 증가율'
sc.style=13
xs=Reference(w15,min_col=2,min_row=GC_S,max_row=GC_E)
ys=Reference(w15,min_col=3,min_row=GC_S,max_row=GC_E)
s=Series(ys,xs,title='상위 20개국')
s.marker.symbol='circle'; s.marker.size=9
s.marker.graphicalProperties.solidFill='4472C4'
s.marker.graphicalProperties.line.solidFill='1F3864'
s.graphicalProperties.line.noFill=True
sc.series.append(s)
sc.title='9. 국가별 반도체 동조성과 수출 증가율 — 오른쪽 위에 몰릴수록 분산 효과 없음'
sc.width=30; sc.height=14; sc.y_axis.majorGridlines=ChartLines()
w15.add_chart(sc,f'A{CT15+210}')
# 10. 성장 유형
b10=BarChart(); b10.type='bar'
b10.add_data(Reference(w15,min_col=2,min_row=GD_H,max_row=GD_E),titles_from_data=True)
b10.set_categories(Reference(w15,min_col=1,min_row=GD,max_row=GD_E))
add15(b10,f'A{CT15+240}','10. 성장 유형별 수출 증감액 (HS 2단위 97개 분류)',h=11,ylab='금액(천달러)',legend=False)
