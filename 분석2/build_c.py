# -*- coding: utf-8 -*-
"""3단계 : 분석 시트 생성"""
import openpyxl, warnings, collections
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
warnings.filterwarnings('ignore')
from data import MONTHS, TOT, OLD, NEW, HS, HSNAME, CTY   # 빌드 시점 정렬용

FONT='맑은 고딕'; NAVY='1F3864'; BLUE='0000FF'; GREEN='008000'; RED='C00000'
HDR=PatternFill('solid', fgColor='1F3864'); SUB=PatternFill('solid', fgColor='D9E2F3')
IN=PatternFill('solid', fgColor='FFFF00'); WARN=PatternFill('solid', fgColor='FCE4D6')
THIN=Side(style='thin', color='BFBFBF'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
NUM='#,##0'; NUM1='#,##0.0'; PCT='0.0%;[RED]-0.0%;-'; PCT2='0.00%'; UNIT='#,##0.000'; PP='+0.00"%p";[RED]-0.00"%p";-'

wb=openpyxl.load_workbook('stage2.xlsx')

# ── 원본/정제 시트 좌표 ────────────────────────────────────────────────
SO='수출입 총괄'
DT="'데이터_총괄'";  DT_R=(3,21)
DO="'데이터_성질별'"; DO_R=(3,401)
DN="'데이터_신성질별'"; DN_R=(3,325)
DH="'데이터_품목별'"; DH_R=(3,1828)
DC="'데이터_국가별'"; DC_R=(3,4477)
M=MONTHS; NM=len(M)                       # 19개월
M25=[x for x in M if x.startswith('2025')]; M26=[x for x in M if x.startswith('2026')]
I26=M.index('2026.01')

OLD_ITEMS=['1. 식료 및 직접소비재','2. 원료 및 연료','3. 경공업품','가. 섬유원료','나. 섬유사','다. 직 물',
 '라. 기타 섬유제품','마. 의 류','바. 목제품','사. 가죽, 고무 및 신발류','아. 귀금속 및 보석류',
 '자. 기타 비금속 광물제품','차. 완구, 운동용구 및 악기','카. 기 타','4. 중화학 공업품','가. 화공품',
 '나. 철강제품','다. 기계류와 정밀기기','라. 전기, 전자제품','마. 수송장비','바. 기 타']
OLD_MAJ=[0,1,2,14]
NEW_ITEMS=['1.소비재','가.직접소비재','나.내구소비재','다.비내구소비재','라.간이세율적용분','2.원자재',
 '가.동식물성 연·원료','나.섬유류','다.광산물','라.철강 및 금속제품','마.화학공업제품','바.기타 원자재',
 '3.자본재','가.수송장비','나.기계류','다.IT제품','라.IT부품']
NEW_MAJ=[0,5,12]
OSTART=[3+21*i for i in range(NM)]
NSTART=[3+17*i for i in range(NM)]
# 정제 시트 구조 검증
w=wb['데이터_성질별']
for i,st in enumerate(OSTART):
    for k,nm in enumerate(OLD_ITEMS):
        assert w.cell(st+k,4).value==f"=TRIM('수출입 실적(성질별)'!C{7+21*i+k})", (i,k)
w=wb['데이터_신성질별']
for i,st in enumerate(NSTART):
    for k,nm in enumerate(NEW_ITEMS):
        assert w.cell(st+k,4).value==f"=TRIM('수출입 실적(신성질별)'!B{7+17*i+k})", (i,k)

def title(ws,row,text,span):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=14,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); ws.row_dimensions[row].height=22
def note(ws,row,col,text,span=10,color='595959'):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,i=True,color=color)
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    c.alignment=Alignment(vertical='center',wrap_text=True); return c
def sec(ws,row,text,span=10):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=11,b=True,color=NAVY)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); c.fill=SUB; return c
def head(ws,row,col,text):
    c=ws.cell(row,col,text); c.font=Font(name=FONT,sz=9,b=True,color='FFFFFF'); c.fill=HDR
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BOX; return c
def put(ws,row,col,v,fmt=None,bold=False,color='000000',align=None,sz=10,border=True,fill=None):
    c=ws.cell(row,col,v); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    if fmt: c.number_format=fmt
    if align: c.alignment=Alignment(horizontal=align,vertical='center')
    if border: c.border=BOX
    if fill: c.fill=fill
    return c
def widths(ws,ws_widths):
    for i,w_ in enumerate(ws_widths): ws.column_dimensions[L(i+1)].width=w_

# 원본 총계행(텍스트) 숫자화
def SRCTOT(col):
    return f'IFERROR(VALUE(SUBSTITUTE(TRIM(\'{SO}\'!{col}6),",","")),"")'
def raw(sheet,col,row):
    return f'IFERROR(VALUE(SUBSTITUTE(TRIM(\'{sheet}\'!{col}{row}),",","")),"")'

# ══════════════════════════════════════════════════════════════════
# 분석1_정합성검증
# ══════════════════════════════════════════════════════════════════
ws=wb.create_sheet('분석1_정합성검증'); ws.sheet_view.showGridLines=False
widths(ws,[22,18,18,14,20,18,18,14,20,12])
title(ws,1,'① 합계 정합성 검증 — 개별항목의 합 vs 총합 (5개 원본 시트 교차)',10)
note(ws,2,1,'원본 5개 시트는 숫자가 텍스트로 저장되어 있어 「데이터_○○」 정제 시트(수식)를 거쳐 검증합니다. 원본 시트는 수정하지 않았습니다.',10)
note(ws,3,1,'허용오차 : 원본이 반올림 제공되므로 아래 값 이내의 차이는 반올림 오차로 판정합니다. (노란색 = 입력 셀)',10)
put(ws,4,1,'허용오차(중량, 톤)',bold=True); put(ws,4,2,1.0,NUM1,color=BLUE,fill=IN)
put(ws,4,3,'허용오차(금액, 천달러)',bold=True); put(ws,4,4,50,NUM,color=BLUE,fill=IN)
put(ws,4,5,'국가별·품목별 허용오차',bold=True); put(ws,4,6,500,NUM,color=BLUE,fill=IN)
note(ws,4,7,'행 수가 많은 시트(국가별 4,475행·품목별 1,826행)는 반올림 누적이 커서 별도 허용오차를 둡니다.',4)
TW,TA,TB='$B$4','$D$4','$F$4'
r=6
sec(ws,r,'[검증 1] 「수출입 총괄」 총계행 vs 19개월(2025.01~2026.07) 합계'); r+=1
for i,h in enumerate(['항목','총계행(원본)','월별 합계','차이','판정']): head(ws,r,i+1,h)
r+=1
metrics=[('조업일수','B','D',NUM1,TW),('수출 건수','C','E',NUM,TA),('수출 중량(톤)','D','F',NUM1,TW),
         ('수출 금액(천$)','E','G',NUM,TA),('수입 건수','F','H',NUM,TA),('수입 중량(톤)','G','I',NUM1,TW),
         ('수입 금액(천$)','H','J',NUM,TA),('무역수지(천$)','I','K',NUM,TA)]
for nm,sc,dc,fmt,tol in metrics:
    put(ws,r,1,nm,bold=True)
    put(ws,r,2,'='+SRCTOT(sc),fmt)
    put(ws,r,3,f'=SUM({DT}!{dc}{DT_R[0]}:{dc}{DT_R[1]})',fmt)
    put(ws,r,4,f'=B{r}-C{r}',NUM1)
    put(ws,r,5,f'=IF(ABS(D{r})<={tol},"일치(반올림오차 이내)","불일치")',align='center')
    r+=1
r+=1
sec(ws,r,'[검증 2] 무역수지 항등식 : 수출금액 − 수입금액 = 무역수지'); r+=1
for i,h in enumerate(['기간','수출금액','수입금액','계산값','원본 무역수지','차이','판정']): head(ws,r,i+1,h)
r+=1
for k in range(NM):
    dr=DT_R[0]+k
    put(ws,r,1,f'={DT}!A{dr}',align='center')
    put(ws,r,2,f'={DT}!G{dr}',NUM); put(ws,r,3,f'={DT}!J{dr}',NUM)
    put(ws,r,4,f'=B{r}-C{r}',NUM); put(ws,r,5,f'={DT}!K{dr}',NUM)
    put(ws,r,6,f'=D{r}-E{r}',NUM)
    put(ws,r,7,f'=IF(ABS(F{r})<={TA},"일치","불일치")',align='center')
    r+=1
r+=1
sec(ws,r,'[검증 3] 「성질별」 대분류 4개 합 vs 「총괄」 수출 (월별)'); r+=1
for i,h in enumerate(['기간','성질별 금액 합','총괄 수출금액','차이','판정','성질별 중량 합','총괄 수출중량','차이','판정']): head(ws,r,i+1,h)
r+=1
V3=r
for k in range(NM):
    st=OSTART[k]; dr=DT_R[0]+k
    amt='+'.join(f'{DO}!F{st+o}' for o in OLD_MAJ); wgt='+'.join(f'{DO}!E{st+o}' for o in OLD_MAJ)
    put(ws,r,1,f'={DT}!A{dr}',align='center')
    put(ws,r,2,'='+amt,NUM); put(ws,r,3,f'={DT}!G{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM1)
    put(ws,r,5,f'=IF(ABS(D{r})<={TA},"일치","불일치")',align='center')
    put(ws,r,6,'='+wgt,NUM1); put(ws,r,7,f'={DT}!F{dr}',NUM1); put(ws,r,8,f'=F{r}-G{r}',NUM1)
    put(ws,r,9,f'=IF(ABS(H{r})<={TW},"일치","불일치")',align='center')
    r+=1
put(ws,r,1,'19개월 누계',bold=True)
put(ws,r,2,f'=SUM(B{V3}:B{r-1})',NUM,bold=True); put(ws,r,3,f'=SUM(C{V3}:C{r-1})',NUM,bold=True)
put(ws,r,4,f'=B{r}-C{r}',NUM1); put(ws,r,5,f'=IF(ABS(D{r})<={TA},"일치","불일치")',align='center')
put(ws,r,6,f'=SUM(F{V3}:F{r-1})',NUM1,bold=True); put(ws,r,7,f'=SUM(G{V3}:G{r-1})',NUM1,bold=True)
put(ws,r,8,f'=F{r}-G{r}',NUM1); put(ws,r,9,f'=IF(ABS(H{r})<={TW},"일치","불일치")',align='center')
r+=2
sec(ws,r,'[검증 4] 「성질별」 중분류 합 vs 대분류 (경공업품 가~카 / 중화학공업품 가~바)'); r+=1
for i,h in enumerate(['기간','3.경공업품','중분류 합','차이','판정','4.중화학공업품','중분류 합','차이','판정']): head(ws,r,i+1,h)
r+=1
for k in range(NM):
    st=OSTART[k]
    put(ws,r,1,f'={DO}!A{st}',align='center')
    put(ws,r,2,f'={DO}!F{st+2}',NUM); put(ws,r,3,f'=SUM({DO}!F{st+3}:F{st+13})',NUM)
    put(ws,r,4,f'=B{r}-C{r}',NUM1); put(ws,r,5,f'=IF(ABS(D{r})<={TA},"일치","불일치")',align='center')
    put(ws,r,6,f'={DO}!F{st+14}',NUM); put(ws,r,7,f'=SUM({DO}!F{st+15}:F{st+20})',NUM)
    put(ws,r,8,f'=F{r}-G{r}',NUM1); put(ws,r,9,f'=IF(ABS(H{r})<={TA},"일치","불일치")',align='center')
    r+=1
r+=1
sec(ws,r,'[검증 5] 「신성질별」 대분류 3개 합 vs 「총괄」 (수출·수입 양방향)'); r+=1
for i,h in enumerate(['기간','신성질별 수출 합','총괄 수출금액','차이','판정','신성질별 수입 합','총괄 수입금액','차이','판정']): head(ws,r,i+1,h)
r+=1
V5=r
for k in range(NM):
    st=NSTART[k]; dr=DT_R[0]+k
    ex='+'.join(f'{DN}!F{st+o}' for o in NEW_MAJ); im='+'.join(f'{DN}!H{st+o}' for o in NEW_MAJ)
    put(ws,r,1,f'={DN}!A{st}',align='center')
    put(ws,r,2,'='+ex,NUM); put(ws,r,3,f'={DT}!G{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM1)
    put(ws,r,5,f'=IF(ABS(D{r})<={TA},"일치","불일치")',align='center')
    put(ws,r,6,'='+im,NUM); put(ws,r,7,f'={DT}!J{dr}',NUM); put(ws,r,8,f'=F{r}-G{r}',NUM1)
    put(ws,r,9,f'=IF(ABS(H{r})<={TA},"일치","불일치")',align='center')
    r+=1
put(ws,r,1,'19개월 누계',bold=True)
for col,src in [(2,'B'),(3,'C'),(6,'F'),(7,'G')]:
    put(ws,r,col,f'=SUM({src}{V5}:{src}{r-1})',NUM,bold=True)
put(ws,r,4,f'=B{r}-C{r}',NUM1); put(ws,r,5,f'=IF(ABS(D{r})<={TA},"일치","불일치")',align='center')
put(ws,r,8,f'=F{r}-G{r}',NUM1); put(ws,r,9,f'=IF(ABS(H{r})<={TA},"일치","불일치")',align='center')
r+=2
sec(ws,r,'[검증 6] 「신성질별」 중분류 합 vs 대분류 (수출 기준 차이)'); r+=1
for i,h in enumerate(['기간','1.소비재 차이','2.원자재 차이','3.자본재 차이','판정']): head(ws,r,i+1,h)
r+=1
for k in range(NM):
    st=NSTART[k]
    put(ws,r,1,f'={DN}!A{st}',align='center')
    put(ws,r,2,f'={DN}!F{st}-SUM({DN}!F{st+1}:F{st+4})',NUM1)
    put(ws,r,3,f'={DN}!F{st+5}-SUM({DN}!F{st+6}:F{st+11})',NUM1)
    put(ws,r,4,f'={DN}!F{st+12}-SUM({DN}!F{st+13}:F{st+16})',NUM1)
    put(ws,r,5,f'=IF(MAX(ABS(B{r}),ABS(C{r}),ABS(D{r}))<={TA},"일치","불일치")',align='center')
    r+=1
r+=1
sec(ws,r,'[검증 7] 「품목별(HS)」 합 vs 「총괄」 — ★ 체계적 차이 발견 구간'); r+=1
for i,h in enumerate(['기간','품목별 수출 합','총괄 수출금액','차이','차이 비율','품목별 수입 합','총괄 수입금액','차이','차이 비율']): head(ws,r,i+1,h)
r+=1
V7=r
for k in range(NM):
    dr=DT_R[0]+k
    p=f'{DT}!A{dr}'
    put(ws,r,1,f'={p}',align='center')
    put(ws,r,2,f'=SUMIFS({DH}!F${DH_R[0]}:F${DH_R[1]},{DH}!A${DH_R[0]}:A${DH_R[1]},$A{r})',NUM)
    put(ws,r,3,f'={DT}!G{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM,color=RED)
    put(ws,r,5,f'=D{r}/C{r}',PCT2)
    put(ws,r,6,f'=SUMIFS({DH}!G${DH_R[0]}:G${DH_R[1]},{DH}!A${DH_R[0]}:A${DH_R[1]},$A{r})',NUM)
    put(ws,r,7,f'={DT}!J{dr}',NUM); put(ws,r,8,f'=F{r}-G{r}',NUM,color=RED)
    put(ws,r,9,f'=H{r}/G{r}',PCT2)
    r+=1
put(ws,r,1,'19개월 누계',bold=True)
for col,src in [(2,'B'),(3,'C'),(6,'F'),(7,'G')]: put(ws,r,col,f'=SUM({src}{V7}:{src}{r-1})',NUM,bold=True)
put(ws,r,4,f'=B{r}-C{r}',NUM,bold=True,color=RED); put(ws,r,5,f'=D{r}/C{r}',PCT2,bold=True)
put(ws,r,8,f'=F{r}-G{r}',NUM,bold=True,color=RED); put(ws,r,9,f'=H{r}/G{r}',PCT2,bold=True)
r+=1
note(ws,r,1,'★ 품목별 시트는 매월 총괄보다 수출 2~3천만 천달러(-0.04%), 수입 1~6천만 천달러 작습니다. '
            '반올림 오차(±100 수준)로 설명되지 않는 체계적 차이입니다. 이 시트에 실제로 들어 있는 HS류를 세어 보면 '
            '01~97류(77류는 HS 결번) 96개와 값이 모두 0인 99류뿐이고, 98류(특수분류 - 여행자 휴대품·별송품·소액물품 등)가 통째로 빠져 있습니다. '
            '그만큼이 차이로 나타나는 것으로 보입니다. 품목별 시트로 전체 합계를 대체하면 이만큼 과소집계됩니다.',10,color=RED)
r+=2
sec(ws,r,'[검증 8] 「국가별」 합 vs 「총괄」'); r+=1
for i,h in enumerate(['기간','국가별 수출 합','총괄 수출금액','차이','판정','국가별 수입 합','총괄 수입금액','차이','판정']): head(ws,r,i+1,h)
r+=1
V8=r
for k in range(NM):
    dr=DT_R[0]+k
    put(ws,r,1,f'={DT}!A{dr}',align='center')
    put(ws,r,2,f'=SUMIFS({DC}!E${DC_R[0]}:E${DC_R[1]},{DC}!A${DC_R[0]}:A${DC_R[1]},$A{r})',NUM)
    put(ws,r,3,f'={DT}!G{dr}',NUM); put(ws,r,4,f'=B{r}-C{r}',NUM)
    put(ws,r,5,f'=IF(ABS(D{r})<={TB},"일치(반올림오차)","불일치")',align='center')
    put(ws,r,6,f'=SUMIFS({DC}!F${DC_R[0]}:F${DC_R[1]},{DC}!A${DC_R[0]}:A${DC_R[1]},$A{r})',NUM)
    put(ws,r,7,f'={DT}!J{dr}',NUM); put(ws,r,8,f'=F{r}-G{r}',NUM)
    put(ws,r,9,f'=IF(ABS(H{r})<={TB},"일치(반올림오차)","불일치")',align='center')
    r+=1
put(ws,r,1,'19개월 누계',bold=True)
for col,src in [(2,'B'),(3,'C'),(6,'F'),(7,'G')]: put(ws,r,col,f'=SUM({src}{V8}:{src}{r-1})',NUM,bold=True)
put(ws,r,4,f'=B{r}-C{r}',NUM,bold=True); put(ws,r,5,f'=IF(ABS(D{r})<={TB},"일치(반올림오차)","불일치")',align='center')
put(ws,r,8,f'=F{r}-G{r}',NUM,bold=True); put(ws,r,9,f'=IF(ABS(H{r})<={TB},"일치(반올림오차)","불일치")',align='center')
r+=2
sec(ws,r,'[검증 9] 5개 시트 교차검증 — 19개월 총계'); r+=1
for i,h in enumerate(['구분','수출 금액(천$)','수입 금액(천$)','수출 중량(톤)','비고']): head(ws,r,i+1,h)
r+=1
X=r
put(ws,r,1,'수출입 총괄',bold=True); put(ws,r,2,f'=SUM({DT}!G{DT_R[0]}:G{DT_R[1]})',NUM)
put(ws,r,3,f'=SUM({DT}!J{DT_R[0]}:J{DT_R[1]})',NUM); put(ws,r,4,f'=SUM({DT}!F{DT_R[0]}:F{DT_R[1]})',NUM1)
put(ws,r,5,'기준값'); r+=1
_orng=lambda c: f'{DO}!${c}${DO_R[0]}:${c}${DO_R[1]}'
_nrng=lambda c: f'{DN}!${c}${DN_R[0]}:${c}${DN_R[1]}'
_o_ex='+'.join(f'SUMIFS({_orng("F")},{_orng("D")},"{OLD_ITEMS[o]}")' for o in OLD_MAJ)
_o_wt='+'.join(f'SUMIFS({_orng("E")},{_orng("D")},"{OLD_ITEMS[o]}")' for o in OLD_MAJ)
_n_ex='+'.join(f'SUMIFS({_nrng("F")},{_nrng("D")},"{NEW_ITEMS[o]}")' for o in NEW_MAJ)
_n_im='+'.join(f'SUMIFS({_nrng("H")},{_nrng("D")},"{NEW_ITEMS[o]}")' for o in NEW_MAJ)
_n_wt='+'.join(f'SUMIFS({_nrng("E")},{_nrng("D")},"{NEW_ITEMS[o]}")' for o in NEW_MAJ)
put(ws,r,1,'성질별(수출만)',bold=True); put(ws,r,2,'='+_o_ex,NUM)
put(ws,r,3,'-',align='center'); put(ws,r,4,'='+_o_wt,NUM1)
put(ws,r,5,'대분류 4개 합계'); r+=1
put(ws,r,1,'신성질별',bold=True); put(ws,r,2,'='+_n_ex,NUM)
put(ws,r,3,'='+_n_im,NUM); put(ws,r,4,'='+_n_wt,NUM1)
put(ws,r,5,'대분류 3개 합계'); r+=1
put(ws,r,1,'품목별(HS)',bold=True); put(ws,r,2,f'=SUM({DH}!F{DH_R[0]}:F{DH_R[1]})',NUM,color=RED)
put(ws,r,3,f'=SUM({DH}!G{DH_R[0]}:G{DH_R[1]})',NUM,color=RED); put(ws,r,4,'-',align='center')
put(ws,r,5,'HS 01~97류만 → 과소',fill=WARN); r+=1
put(ws,r,1,'국가별',bold=True); put(ws,r,2,f'=SUM({DC}!E{DC_R[0]}:E{DC_R[1]})',NUM)
put(ws,r,3,f'=SUM({DC}!F{DC_R[0]}:F{DC_R[1]})',NUM); put(ws,r,4,'-',align='center')
put(ws,r,5,'반올림 누적(4,475행)'); r+=1
put(ws,r,1,'총괄 대비 차이(최대)',bold=True)
put(ws,r,2,f'=MAX(ABS(B{X+1}-B{X}),ABS(B{X+2}-B{X}),ABS(B{X+4}-B{X}))',NUM,bold=True)
put(ws,r,3,f'=MAX(ABS(C{X+2}-C{X}),ABS(C{X+4}-C{X}))',NUM,bold=True)
put(ws,r,4,f'=MAX(ABS(D{X+1}-D{X}),ABS(D{X+2}-D{X}))',NUM1,bold=True)
put(ws,r,5,'품목별 제외 기준',align='center'); r+=1
put(ws,r,1,'판정',bold=True)
put(ws,r,2,f'=IF(B{r-1}<={TB},"총괄·성질별·신성질별·국가별 일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,3,f'=IF(C{r-1}<={TB},"일치","불일치")',align='center',bold=True,color=GREEN)
put(ws,r,4,f'=IF(D{r-1}<={TW},"일치","불일치")',align='center',bold=True,color=GREEN)
r+=2
sec(ws,r,'검증 총평'); r+=1
_c='+'.join(f'COUNTIF({cl}1:{cl}{r-1},"불일치")' for cl in ['E','I','G'])
c=ws.cell(r,1,f'=IF({_c}=0,"검증 결과 : 품목별 시트의 체계적 차이(검증 7)를 제외하면 모든 항목이 허용오차 이내에서 일치합니다.",'
              f'"검증 결과 : 불일치 항목이 있습니다. 각 표의 판정열을 확인하세요.")')
c.font=Font(name=FONT,sz=11,b=True,color=GREEN); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=10)
ws.freeze_panes='A6'

# ══════════════════════════════════════════════════════════════════
# 분석2_월별추이 (19개월)
# ══════════════════════════════════════════════════════════════════
ws2=wb.create_sheet('분석2_월별추이'); ws2.sheet_view.showGridLines=False
widths(ws2,[12,9,12,14,15,15,15,10,10,10,10,14,14,10,12,12,12,15,16])
title(ws2,1,'② 월별 추이 (2025.01 ~ 2026.07, 19개월)',19)
note(ws2,2,1,'「데이터_총괄」을 참조합니다. 일평균 = 금액÷조업일수(영업일 보정), 단가 = 금액÷중량, 순상품교역조건 = 수출단가÷수입단가.',19)
note(ws2,3,1,'전년동월비는 12개월 전 값이 있는 2026.01부터 계산됩니다. 12개월 이동합은 최근 12개월 누계로, 계절성을 제거한 추세선입니다.',19)
H2=['기간','조업일수','수출 건수','수출 중량(톤)','수출 금액(천$)','수입 금액(천$)','무역수지(천$)',
    '수출 전월비','수입 전월비','수출 전년동월비','수입 전년동월비','수지 전년동월증감','일평균 수출(천$)',
    '일평균 전월비','수출단가(천$/톤)','수입단가(천$/톤)','순상품교역조건','수출 12개월 이동합','수지 12개월 이동합']
HR2=5
for i,h in enumerate(H2): head(ws2,HR2,i+1,h)
ws2.row_dimensions[HR2].height=34
D2=HR2+1
for k in range(NM):
    r=D2+k; dr=DT_R[0]+k
    put(ws2,r,1,f'={DT}!A{dr}',align='center',bold=True)
    put(ws2,r,2,f'={DT}!D{dr}',NUM1); put(ws2,r,3,f'={DT}!E{dr}',NUM)
    put(ws2,r,4,f'={DT}!F{dr}',NUM1); put(ws2,r,5,f'={DT}!G{dr}',NUM)
    put(ws2,r,6,f'={DT}!J{dr}',NUM); put(ws2,r,7,f'={DT}!K{dr}',NUM)
    put(ws2,r,8,'' if k==0 else f'=E{r}/E{r-1}-1',PCT)
    put(ws2,r,9,'' if k==0 else f'=F{r}/F{r-1}-1',PCT)
    put(ws2,r,10,'' if k<12 else f'=E{r}/E{r-12}-1',PCT)
    put(ws2,r,11,'' if k<12 else f'=F{r}/F{r-12}-1',PCT)
    put(ws2,r,12,'' if k<12 else f'=G{r}-G{r-12}',NUM)
    put(ws2,r,13,f'=E{r}/B{r}',NUM)
    put(ws2,r,14,'' if k==0 else f'=M{r}/M{r-1}-1',PCT)
    put(ws2,r,15,f'=E{r}/D{r}',UNIT); put(ws2,r,16,f'=F{r}/{DT}!I{dr}',UNIT)
    put(ws2,r,17,f'=O{r}/P{r}','0.000')
    put(ws2,r,18,'' if k<11 else f'=SUM(E{r-11}:E{r})',NUM)
    put(ws2,r,19,'' if k<11 else f'=SUM(G{r-11}:G{r})',NUM)
E2=D2+NM-1
r=E2+2
SUMROWS={}
for lbl,rows in [('2025년 연간(12개월)',(D2,D2+11)),('2025.01~07',(D2,D2+6)),('2026.01~07',(D2+12,E2))]:
    a,b=rows
    put(ws2,r,1,lbl,bold=True,align='center'); SUMROWS[lbl]=r
    for col,fmt in [(2,NUM1),(3,NUM),(4,NUM1),(5,NUM),(6,NUM),(7,NUM)]:
        put(ws2,r,col,f'=SUM({L(col)}{a}:{L(col)}{b})',fmt,bold=True)
    put(ws2,r,13,f'=E{r}/B{r}',NUM,bold=True)
    put(ws2,r,15,f'=E{r}/D{r}',UNIT,bold=True)
    put(ws2,r,16,f'=F{r}/SUM({DT}!I{DT_R[0]+(a-D2)}:I{DT_R[0]+(b-D2)})',UNIT,bold=True)
    put(ws2,r,17,f'=O{r}/P{r}','0.000',bold=True)
    r+=1
put(ws2,r,1,'동기비(26.1~7 vs 25.1~7)',bold=True,align='center')
for col in [2,3,4,5,6,13,15,16,17]:
    put(ws2,r,col,f'={L(col)}{SUMROWS["2026.01~07"]}/{L(col)}{SUMROWS["2025.01~07"]}-1',PCT,bold=True)
put(ws2,r,7,f'=G{SUMROWS["2026.01~07"]}-G{SUMROWS["2025.01~07"]}',NUM,bold=True)
YOY_ROW=r
note(ws2,r+1,1,'맨 아래 행의 무역수지(G열)는 비율이 아니라 증감액(천달러)입니다.',19)
ws2.freeze_panes=f'B{D2}'

# ══════════════════════════════════════════════════════════════════
# 분석3_전년동기비교 (성질별·신성질별)
# ══════════════════════════════════════════════════════════════════
ws3=wb.create_sheet('분석3_전년동기비교'); ws3.sheet_view.showGridLines=False
widths(ws3,[8,26,16,16,16,11,11,11,16,16,11,16,16,16])
title(ws3,1,'③ 전년 동기 대비 (2026.01~07 vs 2025.01~07)',14)
note(ws3,2,1,'모든 값은 정제 시트에 대한 SUMIFS입니다(연도·월 조건). 기여도 = 해당 항목 증감액 ÷ 전체 수출 증감액.',14)
note(ws3,3,1,'성질별 시트에는 수출만 있으므로 [B]는 수출만, 신성질별에는 수입도 있어 [C]는 수출·수입·무역수지를 모두 비교합니다.',14)
r=5
sec(ws3,r,'[A] 총괄 지표',14); r+=1
for i,h in enumerate(['','지표','2025.01~07','2026.01~07','증감','증감률','2025년 연간','2026년 연환산*']): head(ws3,r,i+1,h)
r+=1
A_ROW=r
tot_metrics=[('수출 금액(천$)','E',NUM),('수입 금액(천$)','F',NUM),('무역수지(천$)','G',NUM),
             ('수출 중량(톤)','D',NUM1),('수출 건수','C',NUM),('조업일수','B',NUM1),
             ('일평균 수출(천$)','M',NUM),('수출단가(천$/톤)','O',UNIT),('수입단가(천$/톤)','P',UNIT),
             ('순상품교역조건','Q','0.000')]
for nm,col,fmt in tot_metrics:
    put(ws3,r,1,'총괄',align='center')
    put(ws3,r,2,nm,bold=True)
    put(ws3,r,3,f"='분석2_월별추이'!{col}{SUMROWS['2025.01~07']}",fmt)
    put(ws3,r,4,f"='분석2_월별추이'!{col}{SUMROWS['2026.01~07']}",fmt)
    put(ws3,r,5,f'=D{r}-C{r}',fmt)
    put(ws3,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(ws3,r,7,f"='분석2_월별추이'!{col}{SUMROWS['2025년 연간(12개월)']}",fmt)
    put(ws3,r,8,f'=IF(OR($B{r}="순상품교역조건",LEFT($B{r},3)="수출단",LEFT($B{r},3)="수입단",LEFT($B{r},3)="일평균"),D{r},D{r}/7*12)',fmt)
    r+=1
note(ws3,r,1,'* 2026년 연환산 = 1~7월 실적 ÷ 7 × 12 (단순 환산이며 계절성·조업일수는 반영하지 않은 참고치입니다). 단가·일평균·교역조건은 환산 없이 실적값입니다.',14)
r+=2

def cmp_block(ws, r, label, sheet, keycol, valcol, items, majors, name_prefix=''):
    """SUMIFS 기반 동기 비교 블록 (수출 금액)"""
    sec(ws,r,label,14); r+=1
    for i,h in enumerate(['구분','항목','2025.01~07','2026.01~07','증감액','증감률','기여도','26년 비중',
                          '2025년 연간','2026년 7월','전월비','전년동월비']): head(ws,r,i+1,h)
    r+=1
    r0=r
    tot_ref_25=f'{sheet}!{valcol}'
    for k,nm in enumerate(items):
        isM = k in majors
        put(ws,r,1,'대분류' if isM else '중분류',align='center',bold=isM, fill=SUB if isM else None)
        put(ws,r,2,nm,bold=isM, fill=SUB if isM else None)
        base=f'{sheet}!${valcol}${DO_R[0] if sheet==DO else DN_R[0]}:${valcol}${DO_R[1] if sheet==DO else DN_R[1]}'
        rng=lambda c: f'{sheet}!${c}${DO_R[0] if sheet==DO else DN_R[0]}:${c}${DO_R[1] if sheet==DO else DN_R[1]}'
        put(ws,r,3,f'=SUMIFS({base},{rng("D")},$B{r},{rng("B")},2025,{rng("C")},"<=7")',NUM,bold=isM)
        put(ws,r,4,f'=SUMIFS({base},{rng("D")},$B{r},{rng("B")},2026,{rng("C")},"<=7")',NUM,bold=isM)
        put(ws,r,5,f'=D{r}-C{r}',NUM); put(ws,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT,bold=isM)
        _e26 = "'분석2_월별추이'!E%d" % SUMROWS['2026.01~07']
        _e25 = "'분석2_월별추이'!E%d" % SUMROWS['2025.01~07']
        put(ws,r,7,'=IFERROR(E%d/(%s-%s),"-")' % (r,_e26,_e25),PCT2)
        put(ws,r,8,'=IFERROR(D%d/%s,"-")' % (r,_e26),PCT2)
        put(ws,r,9,f'=SUMIFS({base},{rng("D")},$B{r},{rng("B")},2025)',NUM)
        put(ws,r,10,f'=SUMIFS({base},{rng("D")},$B{r},{rng("A")},"2026.07")',NUM)
        put(ws,r,11,f'=IFERROR(J{r}/SUMIFS({base},{rng("D")},$B{r},{rng("A")},"2026.06")-1,"-")',PCT)
        put(ws,r,12,f'=IFERROR(J{r}/SUMIFS({base},{rng("D")},$B{r},{rng("A")},"2025.07")-1,"-")',PCT)
        r+=1
    return r0, r

B0,r = cmp_block(ws3, r, '[B] 성질별 (수출 금액, 천달러)', DO, 'D', 'F', OLD_ITEMS, OLD_MAJ)
r+=1
C0,r = cmp_block(ws3, r, '[C] 신성질별 (수출 금액, 천달러)', DN, 'D', 'F', NEW_ITEMS, NEW_MAJ)
r+=1
sec(ws3,r,'[D] 신성질별 — 수입 및 무역수지 (천달러)',14); r+=1
for i,h in enumerate(['구분','항목','수입 25.1~7','수입 26.1~7','증감액','증감률','수출 26.1~7','무역수지 25.1~7','무역수지 26.1~7','수지 증감']): head(ws3,r,i+1,h)
r+=1
D0=r
nb=f'{DN}!$H${DN_R[0]}:$H${DN_R[1]}'; ne=f'{DN}!$F${DN_R[0]}:$F${DN_R[1]}'
nrng=lambda c: f'{DN}!${c}${DN_R[0]}:${c}${DN_R[1]}'
for k,nm in enumerate(NEW_ITEMS):
    isM = k in NEW_MAJ
    put(ws3,r,1,'대분류' if isM else '중분류',align='center',bold=isM, fill=SUB if isM else None)
    put(ws3,r,2,nm,bold=isM, fill=SUB if isM else None)
    put(ws3,r,3,f'=SUMIFS({nb},{nrng("D")},$B{r},{nrng("B")},2025,{nrng("C")},"<=7")',NUM,bold=isM)
    put(ws3,r,4,f'=SUMIFS({nb},{nrng("D")},$B{r},{nrng("B")},2026,{nrng("C")},"<=7")',NUM,bold=isM)
    put(ws3,r,5,f'=D{r}-C{r}',NUM); put(ws3,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(ws3,r,7,f'=SUMIFS({ne},{nrng("D")},$B{r},{nrng("B")},2026,{nrng("C")},"<=7")',NUM)
    put(ws3,r,8,f'=SUMIFS({ne},{nrng("D")},$B{r},{nrng("B")},2025,{nrng("C")},"<=7")-C{r}',NUM)
    put(ws3,r,9,f'=G{r}-D{r}',NUM,bold=isM)
    put(ws3,r,10,f'=I{r}-H{r}',NUM)
    r+=1
ws3.freeze_panes='C6'

# ── 빌드 시점 정렬 순서 계산 ────────────────────────────────────────
M25_7=[m for m in M if m.startswith('2025')][:7]
def agg(d,i,ms): return {k: sum(v[m][i] for m in ms if m in v) for k,v in d.items()}
hs26=agg(HS,1,M26); cty26=agg(CTY,1,M26)
HS_ORDER=[k for k,_ in sorted(hs26.items(), key=lambda x:-x[1])]
CTY_ORDER=[k for k,_ in sorted(cty26.items(), key=lambda x:-x[1])]
TOP_HS=HS_ORDER[:8]; TOP_CTY=CTY_ORDER[:8]; TOPN_CTY=40

# ══════════════════════════════════════════════════════════════════
# 분석4_품목별(HS)
# ══════════════════════════════════════════════════════════════════
ws4=wb.create_sheet('분석4_품목별(HS)'); ws4.sheet_view.showGridLines=False
widths(ws4,[8,42,15,15,15,10,10,10,15,15,10,15,15,15])
title(ws4,1,'④ 품목별(HS 2단위) — 전년 동기 대비',14)
note(ws4,2,1,'원본에 들어 있는 HS류 전체(01~97류 중 96개 + 99류)입니다. 정렬은 2026.01~07 수출금액 기준(작성 시점 고정, 수식 자동 재정렬 아님). 품목명은 INDEX/MATCH로 원본에서 가져옵니다.',14)
note(ws4,3,1,'★ 이 시트의 합계는 총괄보다 수출 -0.04% 작습니다(검증 7 참조). 98류(특수분류)가 원본에 없기 때문이며, 품목 간 비교·순위에는 문제가 없습니다.',14,color=RED)
r=5
for i,h in enumerate(['HS','품목명','수출 25.1~7','수출 26.1~7','증감액','증감률','기여도','26년 비중',
                      '수입 25.1~7','수입 26.1~7','수입 증감률','무역수지 25.1~7','무역수지 26.1~7','수지 증감']): head(ws4,r,i+1,h)
ws4.row_dimensions[r].height=30
r+=1
H4=r
hb=f'{DH}!$F${DH_R[0]}:$F${DH_R[1]}'; hi=f'{DH}!$G${DH_R[0]}:$G${DH_R[1]}'
hrng=lambda c: f'{DH}!${c}${DH_R[0]}:${c}${DH_R[1]}'
EX26=f"'분석2_월별추이'!E{SUMROWS['2026.01~07']}"; EX25=f"'분석2_월별추이'!E{SUMROWS['2025.01~07']}"
for code in HS_ORDER:
    put(ws4,r,1,code,align='center').number_format='@'
    put(ws4,r,2,f'=IFERROR(INDEX({hrng("E")},MATCH($A{r},{hrng("D")},0)),"")')
    put(ws4,r,3,f'=SUMIFS({hb},{hrng("D")},$A{r},{hrng("B")},2025,{hrng("C")},"<=7")',NUM)
    put(ws4,r,4,f'=SUMIFS({hb},{hrng("D")},$A{r},{hrng("B")},2026,{hrng("C")},"<=7")',NUM)
    put(ws4,r,5,f'=D{r}-C{r}',NUM); put(ws4,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(ws4,r,7,f'=IFERROR(E{r}/({EX26}-{EX25}),"-")',PCT2)
    put(ws4,r,8,f'=IFERROR(D{r}/{EX26},"-")',PCT2)
    put(ws4,r,9,f'=SUMIFS({hi},{hrng("D")},$A{r},{hrng("B")},2025,{hrng("C")},"<=7")',NUM)
    put(ws4,r,10,f'=SUMIFS({hi},{hrng("D")},$A{r},{hrng("B")},2026,{hrng("C")},"<=7")',NUM)
    put(ws4,r,11,f'=IFERROR(J{r}/I{r}-1,"-")',PCT)
    put(ws4,r,12,f'=C{r}-I{r}',NUM); put(ws4,r,13,f'=D{r}-J{r}',NUM); put(ws4,r,14,f'=M{r}-L{r}',NUM)
    r+=1
E4=r-1
put(ws4,r,2,'합계(HS 01~97)',bold=True)
for col in [3,4,5,9,10,12,13,14]: put(ws4,r,col,f'=SUM({L(col)}{H4}:{L(col)}{E4})',NUM,bold=True)
put(ws4,r,6,f'=D{r}/C{r}-1',PCT,bold=True); put(ws4,r,11,f'=J{r}/I{r}-1',PCT,bold=True)
put(ws4,r,7,f'=E{r}/({EX26}-{EX25})',PCT2,bold=True); put(ws4,r,8,f'=D{r}/{EX26}',PCT2,bold=True)
HS_SUM=r; r+=2
sec(ws4,r,'[참고] 상위 8개 품목 월별 수출금액 (그래프용, 천달러)',14); r+=1
HSM_HDR=r
head(ws4,r,1,'HS'); head(ws4,r,2,'품목명')
for k in range(NM): head(ws4,r,3+k,M[k])
r+=1
HSM=r
for code in TOP_HS:
    put(ws4,r,1,code,align='center').number_format='@'
    put(ws4,r,2,f'=IFERROR(INDEX({hrng("E")},MATCH($A{r},{hrng("D")},0)),"")')
    for k in range(NM):
        put(ws4,r,3+k,f'=SUMIFS({hb},{hrng("D")},$A{r},{hrng("A")},"{M[k]}")',NUM)
    r+=1
HSM_END=r-1
ws4.freeze_panes='C6'

# ══════════════════════════════════════════════════════════════════
# 분석5_국가별
# ══════════════════════════════════════════════════════════════════
ws5=wb.create_sheet('분석5_국가별'); ws5.sheet_view.showGridLines=False
widths(ws5,[6,20,15,15,15,10,10,11,11,10,15,15,10,15,15,15])
title(ws5,1,'⑤ 국가별 — 전년 동기 대비 (상위 40개국 + 기타)',16)
note(ws5,2,1,'정렬은 2026.01~07 수출금액 기준(작성 시점 고정). 「기타」 행은 총괄 합계에서 상위 40개국을 뺀 값이므로 248개국 전체가 빠짐없이 집계됩니다.',16)
r=4
for i,h in enumerate(['순위','국가','수출 25.1~7','수출 26.1~7','증감액','증감률','기여도','점유율 25','점유율 26','점유율 변화',
                      '수입 25.1~7','수입 26.1~7','수입 증감률','수지 25.1~7','수지 26.1~7','수지 증감']): head(ws5,r,i+1,h)
ws5.row_dimensions[r].height=30
r+=1
C5=r
cb=f'{DC}!$E${DC_R[0]}:$E${DC_R[1]}'; ci=f'{DC}!$F${DC_R[0]}:$F${DC_R[1]}'
crng=lambda c: f'{DC}!${c}${DC_R[0]}:${c}${DC_R[1]}'
IM26=f"'분석2_월별추이'!F{SUMROWS['2026.01~07']}"; IM25=f"'분석2_월별추이'!F{SUMROWS['2025.01~07']}"
for i,name in enumerate(CTY_ORDER[:TOPN_CTY]):
    put(ws5,r,1,i+1,align='center')
    put(ws5,r,2,name,bold=(i<10))
    put(ws5,r,3,f'=SUMIFS({cb},{crng("D")},$B{r},{crng("B")},2025,{crng("C")},"<=7")',NUM)
    put(ws5,r,4,f'=SUMIFS({cb},{crng("D")},$B{r},{crng("B")},2026,{crng("C")},"<=7")',NUM)
    put(ws5,r,5,f'=D{r}-C{r}',NUM); put(ws5,r,6,f'=IFERROR(D{r}/C{r}-1,"-")',PCT)
    put(ws5,r,7,f'=IFERROR(E{r}/({EX26}-{EX25}),"-")',PCT2)
    put(ws5,r,8,f'=C{r}/{EX25}',PCT2); put(ws5,r,9,f'=D{r}/{EX26}',PCT2)
    put(ws5,r,10,f'=(I{r}-H{r})*100',PP)
    put(ws5,r,11,f'=SUMIFS({ci},{crng("D")},$B{r},{crng("B")},2025,{crng("C")},"<=7")',NUM)
    put(ws5,r,12,f'=SUMIFS({ci},{crng("D")},$B{r},{crng("B")},2026,{crng("C")},"<=7")',NUM)
    put(ws5,r,13,f'=IFERROR(L{r}/K{r}-1,"-")',PCT)
    put(ws5,r,14,f'=C{r}-K{r}',NUM); put(ws5,r,15,f'=D{r}-L{r}',NUM); put(ws5,r,16,f'=O{r}-N{r}',NUM)
    r+=1
E5=r-1
put(ws5,r,1,'-',align='center'); put(ws5,r,2,'기타(그 외 국가)',bold=True)
put(ws5,r,3,f'={EX25}-SUM(C{C5}:C{E5})',NUM,bold=True)
put(ws5,r,4,f'={EX26}-SUM(D{C5}:D{E5})',NUM,bold=True)
put(ws5,r,5,f'=D{r}-C{r}',NUM); put(ws5,r,6,f'=D{r}/C{r}-1',PCT)
put(ws5,r,7,f'=E{r}/({EX26}-{EX25})',PCT2); put(ws5,r,8,f'=C{r}/{EX25}',PCT2); put(ws5,r,9,f'=D{r}/{EX26}',PCT2)
put(ws5,r,10,f'=(I{r}-H{r})*100',PP)
put(ws5,r,11,f'={IM25}-SUM(K{C5}:K{E5})',NUM); put(ws5,r,12,f'={IM26}-SUM(L{C5}:L{E5})',NUM)
put(ws5,r,13,f'=L{r}/K{r}-1',PCT)
put(ws5,r,14,f'=C{r}-K{r}',NUM); put(ws5,r,15,f'=D{r}-L{r}',NUM); put(ws5,r,16,f'=O{r}-N{r}',NUM)
OTHER=r; r+=1
put(ws5,r,2,'합계(전체 국가)',bold=True)
for col in [3,4,5,11,12,14,15,16]: put(ws5,r,col,f'=SUM({L(col)}{C5}:{L(col)}{OTHER})',NUM,bold=True)
put(ws5,r,6,f'=D{r}/C{r}-1',PCT,bold=True); put(ws5,r,13,f'=L{r}/K{r}-1',PCT,bold=True)
put(ws5,r,8,f'=C{r}/{EX25}',PCT2,bold=True); put(ws5,r,9,f'=D{r}/{EX26}',PCT2,bold=True)
CTY_SUM=r; r+=1
note(ws5,r,1,'합계 행이 총괄과 100.00%로 맞으면 국가 집계가 빠짐없이 이루어진 것입니다.',16)
r+=2
sec(ws5,r,'[참고] 상위 8개국 월별 수출금액 (그래프용, 천달러)',16); r+=1
CM_HDR=r; head(ws5,r,2,'국가')
for k in range(NM): head(ws5,r,3+k,M[k])
r+=1
CM=r
for name in TOP_CTY:
    put(ws5,r,2,name,bold=True)
    for k in range(NM):
        put(ws5,r,3+k,f'=SUMIFS({cb},{crng("D")},$B{r},{crng("A")},"{M[k]}")',NUM)
    r+=1
CM_END=r-1
r+=1
sec(ws5,r,'[참고] 수출 집중도 추이 (그래프용)',16); r+=1
CC_HDR=r; head(ws5,r,2,'지표')
for k in range(NM): head(ws5,r,3+k,M[k])
r+=1
CC=r
put(ws5,r,2,'중국+홍콩+대만 비중',bold=True)
for k in range(NM):
    tot=f"'분석2_월별추이'!E{D2+k}"
    s='+'.join(f'SUMIFS({cb},{crng("D")},"{c}",{crng("A")},"{M[k]}")' for c in ['중국','홍콩','대만'])
    put(ws5,r,3+k,f'=({s})/{tot}',PCT2)
r+=1
put(ws5,r,2,'상위 5개국 비중',bold=True)
for k in range(NM):
    tot=f"'분석2_월별추이'!E{D2+k}"
    s='+'.join(f'SUMIFS({cb},{crng("D")},"{c}",{crng("A")},"{M[k]}")' for c in CTY_ORDER[:5])
    put(ws5,r,3+k,f'=({s})/{tot}',PCT2)
r+=1
put(ws5,r,2,'미국 비중',bold=True)
for k in range(NM):
    tot=f"'분석2_월별추이'!E{D2+k}"
    put(ws5,r,3+k,f'=SUMIFS({cb},{crng("D")},"미국",{crng("A")},"{M[k]}")/{tot}',PCT2)
CC_END=r
note(ws5,r+1,2,'상위 5개국 = 2026.01~07 수출 기준 ' + ', '.join(CTY_ORDER[:5]),15)
ws5.freeze_panes='C5'

# ══════════════════════════════════════════════════════════════════
# 분석6_그래프
# ══════════════════════════════════════════════════════════════════
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
ws6=wb.create_sheet('분석6_그래프'); ws6.sheet_view.showGridLines=False
widths(ws6,[24,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13,13])
title(ws6,1,'⑥ 그래프 — 분석2~분석5 데이터를 그대로 참조',21)
note(ws6,2,1,'아래 보조표는 그래프 전용이며 모두 수식입니다. 원본이 바뀌면 표와 그래프가 함께 갱신됩니다.',21)
A2='분석2_월별추이'; A3='분석3_전년동기비교'; A4="'분석4_품목별(HS)'"; A5='분석5_국가별'
r=4
sec(ws6,r,'보조표 1. 신성질별 대분류 월별 수출·수입 (천달러)',21); r+=1
S1H=r; head(ws6,r,1,'구분')
for k in range(NM): head(ws6,r,2+k,M[k])
r+=1; S1=r
nrng=lambda c: f'{DN}!${c}${DN_R[0]}:${c}${DN_R[1]}'
for kind,col in [('수출','F'),('수입','H')]:
    for o in NEW_MAJ:
        put(ws6,r,1,f'{NEW_ITEMS[o]} {kind}',bold=True)
        for k in range(NM):
            put(ws6,r,2+k,f'=SUMIFS({nrng(col)},{nrng("D")},"{NEW_ITEMS[o]}",{nrng("A")},"{M[k]}")',NUM)
        r+=1
S1E=r-1; r+=1
sec(ws6,r,'보조표 2. 국가별 1~7월 수출 증감액 (2026 vs 2025, 천달러) — 상위 15 / 하위 5',21); r+=1
S2H=r; head(ws6,r,1,'국가'); head(ws6,r,2,'증감액'); head(ws6,r,3,'증감률')
r+=1; S2=r
cty_delta=sorted(((sum(CTY[c][m][1] for m in M26 if m in CTY[c])-sum(CTY[c][m][1] for m in M25_7 if m in CTY[c]), c) for c in CTY), reverse=True)
pick=[c for _,c in cty_delta[:15]]+[c for _,c in cty_delta[-5:]]
cty_row={c:C5+i for i,c in enumerate(CTY_ORDER[:TOPN_CTY])}
for c in pick:
    put(ws6,r,1,c)
    if c in cty_row:
        put(ws6,r,2,f"='{A5}'!E{cty_row[c]}",NUM); put(ws6,r,3,f"='{A5}'!F{cty_row[c]}",PCT)
    else:
        crng=lambda x: f'{DC}!${x}${DC_R[0]}:${x}${DC_R[1]}'
        e26=f'SUMIFS({DC}!$E${DC_R[0]}:$E${DC_R[1]},{crng("D")},$A{r},{crng("B")},2026,{crng("C")},"<=7")'
        e25=f'SUMIFS({DC}!$E${DC_R[0]}:$E${DC_R[1]},{crng("D")},$A{r},{crng("B")},2025,{crng("C")},"<=7")'
        put(ws6,r,2,f'={e26}-{e25}',NUM); put(ws6,r,3,f'=IFERROR({e26}/{e25}-1,"-")',PCT)
    r+=1
S2E=r-1; r+=1
sec(ws6,r,'보조표 3. 품목별(HS) 1~7월 수출 증감액 상위 15 (천달러)',21); r+=1
S3H=r; head(ws6,r,1,'품목'); head(ws6,r,2,'증감액'); head(ws6,r,3,'증감률')
r+=1; S3=r
hs_delta=sorted(((sum(HS[c][m][1] for m in M26 if m in HS[c])-sum(HS[c][m][1] for m in M25_7 if m in HS[c]), c) for c in HS), reverse=True)[:15]
hs_row={c:H4+i for i,c in enumerate(HS_ORDER)}
for _,c in hs_delta:
    put(ws6,r,1,f'{c}. {HSNAME[c][:22]}')
    put(ws6,r,2,f'={A4}!E{hs_row[c]}',NUM); put(ws6,r,3,f'={A4}!F{hs_row[c]}',PCT)
    r+=1
S3E=r-1; r+=1
sec(ws6,r,'보조표 4. 신성질별 중분류 무역수지 (2026.01~07, 천달러)',21); r+=1
S4H=r; head(ws6,r,1,'품목'); head(ws6,r,2,'무역수지')
r+=1; S4=r
for i,nm in enumerate(NEW_ITEMS):
    if i in NEW_MAJ: continue
    put(ws6,r,1,nm); put(ws6,r,2,f"='{A3}'!I{D0+i}",NUM)
    r+=1
S4E=r-1
CT=r+2

PALETTE=['4472C4','ED7D31','A5A5A5','FFC000','5B9BD5','70AD47','264478','9E480E',
         '636363','997300','255E91','43682B','7CAFDD','F1975A','B7B7B7','FFCD33']
def flat(ch, off=0):
    isbar=isinstance(ch,BarChart)
    for i,s_ in enumerate(ch.series):
        col=PALETTE[(i+off)%len(PALETTE)]
        if isbar:
            s_.graphicalProperties.solidFill=col
            s_.graphicalProperties.line.noFill=True
        else:
            s_.smooth=False
            s_.graphicalProperties.line.solidFill=col
            s_.graphicalProperties.line.width=22000
def sub(ch, off=0):
    ch.x_axis.delete=True; flat(ch, off)
def add(ch,anchor,t,w=30,h=13,ylab=None,xlab=None,legend=True):
    flat(ch); ch.title=t; ch.width=w; ch.height=h; ch.style=2
    if ylab: ch.y_axis.title=ylab
    if xlab: ch.x_axis.title=xlab
    if not legend: ch.legend=None
    ch.y_axis.majorGridlines=ChartLines()
    ws6.add_chart(ch,anchor)

ws2n='분석2_월별추이'
cats2=Reference(ws2,min_col=1,min_row=D2,max_row=E2)
# 1. 수출·수입 + 무역수지
b=BarChart(); b.type='col'; b.grouping='clustered'
b.add_data(Reference(ws2,min_col=5,max_col=6,min_row=HR2,max_row=E2),titles_from_data=True)
b.set_categories(cats2)
l=LineChart(); l.add_data(Reference(ws2,min_col=7,min_row=HR2,max_row=E2),titles_from_data=True)
l.y_axis.axId=200; l.y_axis.title='무역수지(천$)'; l.y_axis.crosses='max'; sub(l,2); b+=l
add(b,f'A{CT}','1. 월별 수출·수입 금액과 무역수지 (2025.01~2026.07)',ylab='금액(천달러)')
# 2. 전년동월비
l2=LineChart()
l2.add_data(Reference(ws2,min_col=10,max_col=11,min_row=HR2,max_row=E2),titles_from_data=True)
l2.set_categories(cats2)
add(l2,f'A{CT+26}','2. 전년 동월 대비 증감률 (2026.01~07)',ylab='전년동월비')
# 3. 12개월 이동합
l3=LineChart()
l3.add_data(Reference(ws2,min_col=18,max_col=19,min_row=HR2,max_row=E2),titles_from_data=True)
l3.set_categories(cats2)
add(l3,f'A{CT+52}','3. 12개월 이동합 — 계절성을 제거한 추세',ylab='금액(천달러)')
# 4. 일평균 수출과 조업일수
l4=LineChart(); l4.add_data(Reference(ws2,min_col=13,min_row=HR2,max_row=E2),titles_from_data=True)
l4.set_categories(cats2)
b4=BarChart(); b4.type='col'; b4.add_data(Reference(ws2,min_col=2,min_row=HR2,max_row=E2),titles_from_data=True)
b4.y_axis.axId=200; b4.y_axis.title='조업일수'; b4.y_axis.crosses='max'; b4.gapWidth=200; sub(b4,1); l4+=b4
add(l4,f'A{CT+78}','4. 일평균 수출금액과 조업일수',ylab='일평균 수출(천달러)')
# 5. 단가와 교역조건
l5=LineChart(); l5.add_data(Reference(ws2,min_col=15,max_col=16,min_row=HR2,max_row=E2),titles_from_data=True)
l5.set_categories(cats2)
l5b=LineChart(); l5b.add_data(Reference(ws2,min_col=17,min_row=HR2,max_row=E2),titles_from_data=True)
l5b.y_axis.axId=200; l5b.y_axis.title='교역조건(배)'; l5b.y_axis.crosses='max'; sub(l5b,2); l5+=l5b
add(l5,f'A{CT+104}','5. 수출·수입 단가와 순상품교역조건',ylab='천달러/톤')
# 6·7. 신성질별 대분류 수출/수입
cats6=Reference(ws6,min_col=2,max_col=1+NM,min_row=S1H)
l6=LineChart()
l6.add_data(Reference(ws6,min_col=1,max_col=1+NM,min_row=S1,max_row=S1+2),from_rows=True,titles_from_data=True)
l6.set_categories(cats6)
add(l6,f'A{CT+130}','6. 신성질별 대분류 월별 수출',ylab='금액(천달러)')
l7=LineChart()
l7.add_data(Reference(ws6,min_col=1,max_col=1+NM,min_row=S1+3,max_row=S1E),from_rows=True,titles_from_data=True)
l7.set_categories(cats6)
add(l7,f'A{CT+156}','7. 신성질별 대분류 월별 수입',ylab='금액(천달러)')
# 8. HS 상위 8 월별
l8=LineChart()
l8.add_data(Reference(ws4,min_col=2,max_col=2+NM,min_row=HSM,max_row=HSM_END),from_rows=True,titles_from_data=True)
l8.set_categories(Reference(ws4,min_col=3,max_col=2+NM,min_row=HSM_HDR))
add(l8,f'A{CT+182}','8. 품목별(HS) 상위 8개 월별 수출',ylab='금액(천달러)')
# 9. 국가 상위 8 월별
l9=LineChart()
l9.add_data(Reference(ws5,min_col=2,max_col=2+NM,min_row=CM,max_row=CM_END),from_rows=True,titles_from_data=True)
l9.set_categories(Reference(ws5,min_col=3,max_col=2+NM,min_row=CM_HDR))
add(l9,f'A{CT+208}','9. 국가별 상위 8개국 월별 수출',ylab='금액(천달러)')
# 10. 국가 증감액
b10=BarChart(); b10.type='bar'
b10.add_data(Reference(ws6,min_col=2,min_row=S2H,max_row=S2E),titles_from_data=True)
b10.set_categories(Reference(ws6,min_col=1,min_row=S2,max_row=S2E))
add(b10,f'A{CT+234}','10. 국가별 1~7월 수출 증감액 (상위 15 / 하위 5)',h=16,ylab='증감액(천달러)',legend=False)
# 11. HS 증감액
b11=BarChart(); b11.type='bar'
b11.add_data(Reference(ws6,min_col=2,min_row=S3H,max_row=S3E),titles_from_data=True)
b11.set_categories(Reference(ws6,min_col=1,min_row=S3,max_row=S3E))
add(b11,f'A{CT+264}','11. 품목별(HS) 1~7월 수출 증감액 상위 15',h=14,ylab='증감액(천달러)',legend=False)
# 12. 집중도
l12=LineChart()
l12.add_data(Reference(ws5,min_col=2,max_col=2+NM,min_row=CC,max_row=CC_END),from_rows=True,titles_from_data=True)
l12.set_categories(Reference(ws5,min_col=3,max_col=2+NM,min_row=CC_HDR))
add(l12,f'A{CT+292}','12. 수출 집중도 추이 (전체 수출 대비 비중)',ylab='비중')
# 13. 신성질별 중분류 무역수지
b13=BarChart(); b13.type='bar'
b13.add_data(Reference(ws6,min_col=2,min_row=S4H,max_row=S4E),titles_from_data=True)
b13.set_categories(Reference(ws6,min_col=1,min_row=S4,max_row=S4E))
add(b13,f'A{CT+318}','13. 신성질별 중분류 무역수지 (2026.01~07)',h=14,ylab='무역수지(천달러)',legend=False)

# ══════════════════════════════════════════════════════════════════
# 분석7_해설
# ══════════════════════════════════════════════════════════════════
ws7=wb.create_sheet('분석7_해설'); ws7.sheet_view.showGridLines=False
widths(ws7,[32,17,17,17,17,17,15,15,15])
title(ws7,1,'⑦ 분석 결과 해설 — 숫자가 말하는 것',9)
note(ws7,2,1,'[핵심 수치] 표는 모두 수식이며 원본이 바뀌면 자동 갱신됩니다. 본문 서술의 숫자는 2025.01~2026.07 자료 기준입니다.',9)
r=4
sec(ws7,r,'[핵심 수치] 2026.01~07 vs 2025.01~07 (단위: 천달러, 톤)',9); r+=1
for i,h in enumerate(['지표','2025.01~07','2026.01~07','증감','증감률','2025년 연간']): head(ws7,r,i+1,h)
r+=1
K0=r
for i,(nm,col,fmt) in enumerate([('수출 금액','E',NUM),('수입 금액','F',NUM),('무역수지','G',NUM),
        ('수출 중량(톤)','D',NUM1),('조업일수','B',NUM1),('일평균 수출','M',NUM),
        ('수출단가(천$/톤)','O',UNIT),('수입단가(천$/톤)','P',UNIT),('순상품교역조건','Q','0.000')]):
    put(ws7,r,1,nm,bold=True)
    put(ws7,r,2,f"='{A2}'!{col}{SUMROWS['2025.01~07']}",fmt)
    put(ws7,r,3,f"='{A2}'!{col}{SUMROWS['2026.01~07']}",fmt)
    put(ws7,r,4,f'=C{r}-B{r}',fmt); put(ws7,r,5,f'=IFERROR(C{r}/B{r}-1,"-")',PCT)
    put(ws7,r,6,f"='{A2}'!{col}{SUMROWS['2025년 연간(12개월)']}",fmt)
    r+=1
r+=1
sec(ws7,r,'[수출 증가 331 → 2,000억 달러를 누가 만들었나] 1~7월 증감액 상위',9); r+=1
for i,h in enumerate(['구분','항목','2025.01~07','2026.01~07','증감액','증감률','기여도']): head(ws7,r,i+1,h)
r+=1
drivers=[('신성질별','라.IT부품',C0+16),('신성질별','다.IT제품',C0+15),('신성질별','다.광산물',C0+8),
         ('신성질별','가.수송장비',C0+13),('신성질별','나.내구소비재',C0+2)]
for kind,nm,src in drivers:
    put(ws7,r,1,kind,align='center'); put(ws7,r,2,nm,bold=True)
    for c_,sc in [(3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G')]:
        put(ws7,r,c_,f"='{A3}'!{sc}{src}", NUM if c_<6 else (PCT if c_==6 else PCT2))
    r+=1
for code in HS_ORDER[:2]:
    put(ws7,r,1,'품목(HS)',align='center'); put(ws7,r,2,f'{code}류')
    for c_,sc in [(3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G')]:
        put(ws7,r,c_,f"={A4}!{sc}{hs_row[code]}", NUM if c_<6 else (PCT if c_==6 else PCT2))
    r+=1
for c in CTY_ORDER[:3]:
    put(ws7,r,1,'국가',align='center'); put(ws7,r,2,c)
    for c_,sc in [(3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G')]:
        put(ws7,r,c_,f"='{A5}'!{sc}{cty_row[c]}", NUM if c_<6 else (PCT if c_==6 else PCT2))
    r+=1
r+=1

def para(ws,row,text,span=9,sz=10,color='000000',bold=False):
    c=ws.cell(row,1,text); c.font=Font(name=FONT,sz=sz,b=bold,color=color)
    c.alignment=Alignment(vertical='top',wrap_text=True)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    ws.row_dimensions[row].height=max(16,15*(len(text)//68+1))
    return c

SECT=[
 ('1. 자료 구성과 이번 분석의 범위',[
  '· 5개 원본 시트(총괄 / 품목별 / 성질별 / 신성질별 / 국가별)는 모두 2025.01~2026.07 19개월, 수리일 기준, 단위는 건·톤(TON)·천달러입니다.',
  '· 원본은 숫자가 텍스트(쉼표·공백 포함)로 저장되어 있어 계산이 불가능한 상태였습니다. 그래서 「데이터_○○」 정제 시트를 만들어 VALUE/SUBSTITUTE 수식으로 숫자화하고, 모든 분석은 그 정제 시트를 참조합니다. 원본 5개 시트는 값 하나도 수정하지 않았습니다.',
  '· 시트별 담고 있는 범위가 다릅니다. 성질별은 수출만, 신성질별·품목별·국가별은 수출과 수입을 모두 담습니다. 국가별에는 중량이 없어 국가별 단가는 계산할 수 없습니다.',
  '· 2026년은 7월까지만 있으므로 연간 비교는 성립하지 않습니다. 이 분석의 기본 비교 단위는 2026.01~07 vs 2025.01~07(동기 7개월)이며, 분석3의 연환산 열은 단순 참고치입니다.',
 ]),
 ('2. 합계 정합성 — 네 개 시트는 일치, 품목별 시트만 체계적으로 작습니다',[
  '· 9종 검증(총괄 총계행 vs 19개월 합 / 무역수지 항등식 / 성질별 대분류·중분류 / 신성질별 대분류·중분류(수출·수입) / 품목별 / 국가별 / 5개 시트 교차)을 수식으로 수행했습니다.',
  '· 총괄·성질별·신성질별·국가별은 19개월 총계가 수출 1조 3,043억 9천만 천달러 선에서 최대 27천달러(0.000002%) 차이로 일치합니다. 국가별은 4,475행, 신성질별은 323행의 반올림이 누적된 결과로 데이터 오류가 아닙니다.',
  '· ★ 품목별(HS) 시트만 다릅니다. 19개월 합이 수출 −4억 9,946만 천달러(−0.038%), 수입 −3억 8,429만 천달러(−0.036%)로 총괄보다 작고, 이 차이가 19개월 내내 매월 나타납니다(수출 월 2~3천만 천달러). 반올림으로는 설명되지 않는 체계적 차이입니다.',
  '· 원인을 원본에서 직접 확인했습니다. 품목별 시트에 실제로 존재하는 HS류는 01~97류 중 96개(77류는 HS 자체 결번)와 값이 전부 0인 99류뿐이고, 98류가 통째로 없습니다. 98류는 여행자 휴대품·별송품·소액물품 등 특수분류로, 총괄에는 포함되지만 품목별 조회에서는 빠진 것으로 보입니다. 실무상 의미: 품목별 시트로 전체 수출입 총액을 대체하면 약 0.04% 과소집계됩니다. 품목 간 비교·순위 분석에는 영향이 없습니다.',
  '· 2026.06~07에는 품목별 수입 차이가 −5,795만·−4,745만 천달러로 평소(−1~2천만)보다 커집니다. 특수분류 수입이 최근 늘었을 가능성이 있어, 관세청 원자료에서 98류를 따로 조회해 확인해 보시길 권합니다.',
 ]),
 ('3. 총괄 추이 — 2026년의 수출 급증은 달력 효과가 아닙니다',[
  '· 2026.01~07 수출은 5,950.6억 달러로 전년 동기 3,953.9억 달러 대비 +50.5%(+1,996.7억 달러) 늘었습니다. 수입은 +18.2%에 그쳐, 무역수지는 339.3억 달러 → 1,677.6억 달러로 1,338.3억 달러 확대되었습니다.',
  '· 전년동월비는 2026.01 +33.9%에서 매달 높아져 2026.06 +70.4%로 정점을 찍고 7월 +63.0%입니다. 12개월 이동합(분석2 R열)은 7,093억 → 9,090억 달러로 계단식 상승이며, 계절성을 제거해도 추세가 꺾인 신호는 아직 없습니다.',
  '· 조업일수는 155.5일 → 156.5일로 +0.6%뿐입니다. 일평균 수출이 254만 → 380만 천달러로 +49.5% 늘었으니, 증가분은 영업일 증가가 아닌 실질 증가입니다.',
  '· 다만 2026.07은 전월비 −2.9%, 일평균 기준으로는 −9.0%입니다(7월 조업일수 24일 > 6월 22.5일). 월 합계가 보여주는 것보다 실제 둔화 폭이 큽니다.',
 ]),
 ('4. 성장의 정체 — 물량은 줄었는데 금액은 50% 늘었습니다',[
  '· 수출 중량은 1억 1,474만 톤 → 1억 987만 톤으로 오히려 −4.2% 감소했습니다. 그런데 금액은 +50.5%입니다. 수출 단가가 3.446 → 5.416 천달러/톤(+57.2%)으로 뛰었기 때문입니다.',
  '· 수입 단가도 1.166 → 1.362(+16.8%) 올랐지만 상승 폭이 훨씬 작아, 순상품교역조건은 2.955 → 3.978배(+34.6%)로 개선되었습니다. 무역수지 1,338억 달러 증가의 실체는 물량 확대가 아니라 이 단가 격차입니다.',
  '· 이 구조는 되돌림에 취약합니다. 실제로 2026.07 수출단가는 6.499 → 5.603(−13.8%)로 떨어졌고, 같은 달 무역수지도 359억 → 304억 달러로 줄었습니다. 단가가 정상화되면 물량이 그대로여도 흑자는 빠르게 축소됩니다.',
 ]),
 ('5. 품목 — 반도체·IT 한 축이 증가분의 3분의 2를 만들었습니다',[
  '· HS 2단위 기준으로 85류(전기기기·반도체 등)가 1,231.3억 → 2,568.7억 달러(+108.6%)로, 전체 수출 증가분의 67.0%를 혼자 만들었습니다. 84류(기계류)가 +36.2억 달러(기여도 18.1%)로 뒤를 잇습니다. 두 품목만으로 증가분의 85%입니다.',
  '· 신성질별로 보면 같은 현상이 IT부품 981.2억 → 2,443.4억 달러(+149.0%, 기여도 73.2%), IT제품 +98.1%로 나타납니다. 자본재 전체는 +77.8%입니다.',
  '· 반대로 소비재는 −0.4%로 사실상 역성장했고, 그중 내구소비재는 −7.2%(−28.1억 달러)입니다. 87류(자동차 등 차량)도 −1.0%로 줄었습니다. 즉 이번 호황은 전 품목 확산형이 아니라 IT 편중형입니다.',
  '· 71류(귀금속·보석)는 +154.9%로 증가율이 가장 높지만 비중은 1.7%로 작고, 월별 등락이 매우 큽니다. 규모가 작은 품목의 높은 증가율에 과도한 의미를 두지 않는 편이 좋습니다.',
  '· 무역수지 관점에서는 IT부품이 +1,745.4억 달러 흑자를 내는 반면, 광산물이 −709.1억 달러, 기계류가 −82.2억 달러 적자입니다. 27류(광물성 연료)는 −557.4억 달러로 단일 최대 적자 품목입니다. 수출 호황에도 에너지·원자재 수입 구조는 그대로라는 뜻입니다.',
 ]),
 ('6. 국가 — 증가는 중화권과 미국에 집중되어 있습니다',[
  '· 증가액 상위는 중국 +496.7억(기여도 24.9%), 미국 +384.8억(19.3%), 홍콩 +277.8억(13.9%), 베트남 +197.0억(9.9%), 대만 +159.7억(8.0%) 달러입니다. 이 5개국이 증가분의 76%입니다.',
  '· 특히 홍콩은 +161.9%, 대만 +63.6%, 말레이시아 +76.1%, 필리핀 +78.4%로 반도체 공급망에 걸친 국가들이 함께 뛰었습니다. 홍콩·대만은 최종 소비지라기보다 중계·후공정 성격이 강해, 실수요와 재고 이동을 구분해서 볼 필요가 있습니다.',
  '· 대중 무역수지는 −82.8억 달러 적자에서 +188.5억 달러 흑자로 271.2억 달러 전환되었습니다. 대미 흑자도 301.8억 → 584.6억 달러로 커졌습니다. 반면 사우디아라비아(−144.6억), 일본(−129.4억), 아랍에미리트(−73.4억)는 적자가 유지·확대되었습니다. 에너지 수입국에 대한 적자 구조는 그대로입니다.',
  '· 집중도는 뚜렷하게 높아졌습니다. 상위 5개국 비중은 2025.01 56.2%에서 2026.06 64.8%까지 올랐고(7월 62.1%), 중국·홍콩·대만 합계 비중은 27.1% → 34.4%가 되었습니다. 품목(IT)과 지역(중화권)의 이중 집중입니다.',
  '· 감소국도 있습니다. 덴마크 −65.0%, 카자흐스탄 −41.2%, 라이베리아 −22.8%입니다. 라이베리아·마샬군도는 선박 편의치적국이므로 선박 인도 일정에 따라 크게 흔들리는 항목이며, 추세로 읽지 않는 편이 좋습니다.',
 ]),
 ('7. 2026년 7월에 무슨 일이 있었나',[
  '· 7월 수출은 989.6억 달러로 전월비 −2.9%(일평균 −9.0%)이지만 전년동월비로는 여전히 +63.0%입니다. 즉 추세 이탈이라기보다 높은 수준에서의 조정입니다.',
  '· 감소는 특정 지역에 몰려 있습니다. 대만 −31.8%, 홍콩 −26.8%, 싱가포르 −22.5%, 미국 −12.6%인 반면 중국은 +8.4%, 베트남 +15.4%, 인도 +22.8%로 늘었습니다.',
  '· 품목에서는 84류(기계류) −16.6%, 71류(귀금속) −21.8%, 85류(전기기기) −1.9%입니다. 85류가 소폭 감소에 그친 반면 84류 감소가 컸다는 점은, 반도체 자체보다 장비·기계 쪽 주문이 먼저 쉬어간 모습으로 읽힙니다.',
  '· 한편 7월 수입 중량은 5,114만 톤으로 19개월 중 최대인데 수입 금액은 +3.8%만 늘었습니다(수입 단가 −11.3%). 원자재를 더 싸게 더 많이 들여왔다는 뜻으로, 향후 수출 마진에는 우호적인 신호입니다.',
 ]),
 ('8. 이 분석이 갖는 의미',[
  '· ① 이중 집중 리스크. 수출 증가의 67%가 HS 85류, 76%가 상위 5개국에서 나왔습니다. 반도체 사이클과 중화권 수요 중 하나만 꺾여도 전체 수출이 같이 꺾이는 구조입니다. 분산 효과를 기대할 만한 완충 품목(소비재·자동차)은 오히려 감소했습니다.',
  '· ② 흑자의 질. 지금의 흑자는 물량(−4.2%)이 아니라 단가(+57.2%)에서 나왔습니다. 단가는 사이클성이 강하므로, 흑자 1,678억 달러를 항상성 있는 체력으로 보기는 어렵습니다. 7월의 단가 −13.8%와 흑자 축소가 그 민감도를 보여줍니다.',
  '· ③ 수입도 같이 늘고 있습니다. IT부품 수입이 +40.6%(496.4억 → 698.0억 달러)입니다. 중간재를 들여와 완제품·부품을 내보내는 구조라, 수출이 꺾이면 수입도 함께 줄어 무역수지는 생각보다 완만하게 조정될 수 있습니다.',
  '· ④ 에너지·원자재 적자는 상수입니다. 27류 −557.4억, 26류 −147.8억 달러 적자는 수출 호황과 무관하게 유지됩니다. 유가·광물가가 반등하면 흑자는 수출과 무관하게 줄어듭니다.',
  '· ⑤ 매월 볼 지표 다섯 개 : HS 85류 전년동월비(방향), 수출 단가(흑자의 지속성), 조업일수 보정 일평균(달력 착시 제거), 상위 5개국 비중(집중도), 순상품교역조건(채산성). 모두 분석2·4·5에서 매월 값만 갱신하면 그대로 추적됩니다.',
 ]),
 ('9. 한계와 확인이 필요한 부분',[
  '· 2026년은 7개월치뿐이므로 연간 비교는 불가능합니다. 분석3의 연환산 열은 1~7월을 단순히 12개월로 늘린 참고치이며, 계절성(연말 밀어내기 등)을 반영하지 않습니다. 2025년 실적을 보면 12월이 연중 최고(695.1억 달러)였습니다.',
  '· 품목별 시트의 −0.04% 과소집계(검증 7)는 98류 부재를 원본에서 확인한 결과이지만, 차이 금액이 정확히 98류 규모와 일치하는지는 별도 조회로만 검증할 수 있습니다. 총액이 중요한 용도라면 총괄 시트를 기준으로 삼으십시오.',
  '· 국가별 시트에는 중량이 없어 국가별 단가·물량 분석이 불가능하고, 성질별 시트에는 수입이 없어 성질별 무역수지를 만들 수 없습니다(신성질별로는 가능해 분석3-D에 넣었습니다).',
  '· 금액은 명목 달러 기준이며 환율·물가 조정이 되어 있지 않습니다. 단가 상승분에는 실제 가격 상승과 품목 구성 변화가 섞여 있어, 완전히 분리하려면 품목별 수량(개수) 자료가 필요합니다.',
  '· 홍콩·대만·마샬군도·라이베리아처럼 중계무역·편의치적 성격이 강한 상대국은 최종 수요지와 다를 수 있습니다. 최종 소비지 기준 분석이 필요하면 별도 자료가 있어야 합니다.',
 ]),
]
for t,lines in SECT:
    sec(ws7,r,t,9); r+=1
    for ln in lines:
        para(ws7,r,ln); r+=1
    r+=1
note(ws7,r,1,'작성 : 업로드된 5개 원본 파일만을 근거로 산출했으며 외부 자료를 참조하지 않았습니다.',9)

exec(open('build_e.py').read())

# ══════════════════════════════════════════════════════════════════
# 0_안내 (목차)
# ══════════════════════════════════════════════════════════════════
ws0=wb.create_sheet('0_안내', 0); ws0.sheet_view.showGridLines=False
widths(ws0,[26,66,22])
title(ws0,1,'수출입 실적 통합 분석 (2025.01 ~ 2026.07)',3)
note(ws0,2,1,'업로드하신 5개 원본 파일(총괄·품목별·성질별·신성질별·국가별)을 한 파일로 모으고, 원본은 수정하지 않은 채 분석 시트를 추가했습니다.',3)
r=4
sec(ws0,r,'시트 구성',3); r+=1
for i,h in enumerate(['시트','내용','비고']): head(ws0,r,i+1,h)
r+=1
rows=[('수출입 총괄','2025.01~2026.07 월별 총괄 (원본 그대로)','원본 · 수정 없음'),
      ('수출입 실적(품목별)','HS 2단위 품목별 수출입 (원본 그대로)','원본 · 수정 없음'),
      ('수출입 실적(성질별)','성질별 수출 (원본 그대로)','원본 · 수정 없음'),
      ('수출입 실적(신성질별)','신성질별 수출입 (원본 그대로)','원본 · 수정 없음'),
      ('수출입 실적(국가별)','국가별 수출입 (원본 그대로)','원본 · 수정 없음'),
      ('데이터_총괄 외 4개','원본의 텍스트 숫자를 VALUE 수식으로 숫자화한 정제 시트','전부 수식 · 직접 입력 금지'),
      ('분석1_정합성검증','개별항목 합 vs 총합 9종 검증','품목별 시트의 체계적 차이 확인'),
      ('분석2_월별추이','19개월 추이 · 전월비 · 전년동월비 · 일평균 · 단가 · 12개월 이동합',''),
      ('분석3_전년동기비교','2026.1~7 vs 2025.1~7 (총괄 · 성질별 · 신성질별 수출입)',''),
      ('분석4_품목별(HS)','HS류별 동기 비교 · 기여도 · 무역수지','상위 8개 월별 표 포함'),
      ('분석5_국가별','상위 40개국 + 기타 · 점유율 · 무역수지','집중도 추이 포함'),
      ('분석6_그래프','그래프 13종 + 보조표(총량·품목·국가)',''),
      ('분석7_해설','핵심 수치표와 해석 · 한계','먼저 읽으시면 좋습니다'),
      ('분석8_물량·가격분해','금액 증감을 물량효과·가격효과·교차효과로 분해 + 피셔지수','수출 심층'),
      ('분석9_품목단가분해(HS)','HS류별 단가·물량 분해와 성장 유형 분류','수출 심층'),
      ('분석10_집중도·확산도','HHI · 유효 품목수 · 확산지수','수출 심층'),
      ('분석11_모멘텀·계절성','3개월이평 YoY · 계절조정 · IT/비IT · 잔여기간 시나리오','수출 심층'),
      ('분석12_국가심층','HS85 동조성 · 건당금액 · 순위변동','수출 심층'),
      ('분석13_민감도','IT 단가 충격 시뮬레이션과 임계점','입력 셀 있음'),
      ('분석14_원인분석','관측 사실 ↔ 원인 가설 ↔ 검증 방법','해석 포함'),
      ('분석15_그래프2','심층 분석 그래프 10종','')]
for a,b,c in rows:
    put(ws0,r,1,a,bold=True); put(ws0,r,2,b); put(ws0,r,3,c,align='center')
    r+=1
r+=1
sec(ws0,r,'읽는 순서 제안',3); r+=1
for t in ['① 분석7_해설 → 분석14_원인분석 — 무엇이 어떻게 변했고 왜 그런지 먼저 확인',
          '② 분석1_정합성검증 — 숫자를 믿어도 되는지 확인 (품목별 시트 주의사항 포함)',
          '③ 분석8·9 — 수출 증가가 물량인지 가격인지 분해해서 확인 (이 분석의 핵심)',
          '④ 분석10·11·12 — 집중도 · 모멘텀 · 국가 동조성으로 구조와 속도 확인',
          '⑤ 분석13_민감도 — 노란색 셀에 가정을 넣어 충격 크기를 직접 확인',
          '⑥ 분석6·15_그래프 — 그림으로 확인']:
    put(ws0,r,1,t,border=False); ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
r+=1
note(ws0,r,1,'※ 모든 분석 셀은 원본을 참조하는 수식입니다. 원본 시트에 다음 달 데이터를 같은 형식으로 추가하고 정제 시트의 참조 범위를 늘리면 분석 전체가 그대로 갱신됩니다.',3)
wb.save('final.xlsx'); print('final 저장')
