# -*- coding: utf-8 -*-
"""원본 시트(수출입 총괄 / 성질별 / 신성질별)를 건드리지 않고 분석 시트를 추가한다."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = 'work.xlsx'
OUT = 'analysis.xlsx'

S_TOT = "'수출입 총괄'"
S_OLD = "'수출입 실적(성질별)'"
S_NEW = "'수출입 실적(신성질별)'"

MONTHS = ['2026.01','2026.02','2026.03','2026.04','2026.05','2026.06','2026.07']
TOT_ROWS = list(range(7, 14))          # 총괄 월별 행
OLD_START = [7 + 21*i for i in range(7)]
NEW_START = [7 + 17*i for i in range(7)]

OLD_ITEMS = [  # (offset, level, name)
    (0,1,'1.식료및직접소비재'), (1,1,'2.원료및연료'), (2,1,'3.경공업품'),
    (3,2,'가.섬유원료'), (4,2,'나.섬유사'), (5,2,'다.직물'), (6,2,'라.기타섬유제품'),
    (7,2,'마.의류'), (8,2,'바.목제품'), (9,2,'사.가죽,고무및신발류'),
    (10,2,'아.귀금속및보석류'), (11,2,'자.기타비금속광물제품'),
    (12,2,'차.완구,운동용구및악기'), (13,2,'카.기타'),
    (14,1,'4.중화학공업품'), (15,2,'가.화공품'), (16,2,'나.철강제품'),
    (17,2,'다.기계류와정밀기기'), (18,2,'라.전기,전자제품'), (19,2,'마.수송장비'), (20,2,'바.기타'),
]
NEW_ITEMS = [
    (0,1,'1.소비재'), (1,2,'가.직접소비재'), (2,2,'나.내구소비재'), (3,2,'다.비내구소비재'),
    (4,2,'라.간이세율적용분'),
    (5,1,'2.원자재'), (6,2,'가.동식물성 연·원료'), (7,2,'나.섬유류'), (8,2,'다.광산물'),
    (9,2,'라.철강 및 금속제품'), (10,2,'마.화학공업제품'), (11,2,'바.기타 원자재'),
    (12,1,'3.자본재'), (13,2,'가.수송장비'), (14,2,'나.기계류'), (15,2,'다.IT제품'), (16,2,'라.IT부품'),
]
OLD_MAJ = [0,1,2,14]
NEW_MAJ = [0,5,12]

FONT = '맑은 고딕'
NAVY = '1F3864'; BLUE = '0000FF'; GREEN = '008000'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
SUB_FILL = PatternFill('solid', fgColor='D9E2F3')
IN_FILL  = PatternFill('solid', fgColor='FFFF00')
NOTE_FILL= PatternFill('solid', fgColor='F2F2F2')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0'; NUM1 = '#,##0.0'; PCT = '0.0%;[RED]-0.0%;-'; PCT2='0.00%'
UNIT = '#,##0.000'

wb = openpyxl.load_workbook(SRC)

# ── 원본 구조 검증(assert) ────────────────────────────────────────────────
w_old = wb['수출입 실적(성질별)']; w_new = wb['수출입 실적(신성질별)']
for m, st in enumerate(OLD_START):
    for off, lv, nm in OLD_ITEMS:
        assert w_old.cell(st+off, 2).value == nm, (m, nm, w_old.cell(st+off,2).value)
for m, st in enumerate(NEW_START):
    for off, lv, nm in NEW_ITEMS:
        assert w_new.cell(st+off, 2).value == nm, (m, nm, w_new.cell(st+off,2).value)

def style_title(ws, row, text, span):
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, sz=14, b=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22

def style_head(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font = Font(name=FONT, sz=10, b=True, color='FFFFFF')
    c.fill = HDR_FILL; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BOX
    return c

def note(ws, row, col, text, span=8):
    c = ws.cell(row, col, text)
    c.font = Font(name=FONT, sz=9, i=True, color='595959')
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    return c

def sec(ws, row, text, span=9):
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, sz=11, b=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c.fill = SUB_FILL
    return c

def put(ws, row, col, value, fmt=None, bold=False, color='000000', align=None, border=True, sz=10):
    c = ws.cell(row, col, value)
    c.font = Font(name=FONT, sz=sz, b=bold, color=color)
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical='center')
    if border: c.border = BOX
    return c

# ══════════════════════════════════════════════════════════════════════════
# 시트 1 : 분석1_정합성검증
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('분석1_정합성검증')
ws.sheet_view.showGridLines = False
for col, wdt in zip('ABCDEFGHIJ', [22,20,20,16,14,20,20,16,14,10]):
    ws.column_dimensions[col].width = wdt

style_title(ws, 1, '① 합계 정합성 검증 — 개별항목의 합 vs 총합', 9)
note(ws, 2, 1, '원본 3개 시트(수출입 총괄 / 성질별 / 신성질별)의 모든 값을 수식으로 직접 참조하여 검증합니다. 원본 시트는 수정하지 않았습니다.', 9)
note(ws, 3, 1, '허용오차: 원본이 소수 첫째자리(중량)·정수(금액)로 반올림되어 제공되므로, 아래 허용오차 이내의 차이는 반올림 오차로 판정합니다. (노란색 셀 = 사용자 입력값)', 9)

put(ws, 4, 1, '허용오차(중량, 톤)', bold=True); c=put(ws,4,2,1.0,NUM1,color=BLUE); c.fill=IN_FILL
put(ws, 4, 3, '허용오차(금액, 천달러)', bold=True); c=put(ws,4,4,5,NUM,color=BLUE); c.fill=IN_FILL
TOL_W, TOL_A = '$B$4', '$D$4'

r = 6
sec(ws, r, '[검증 1] 「수출입 총괄」 총계행 vs 2026.01~07 월별 합계'); r += 1
heads = ['항목','총계행(원본)','월별 합계','차이','판정']
for i,h in enumerate(heads): style_head(ws, r, i+1, h)
r += 1
metrics = [('조업일수','B',NUM1,TOL_W), ('수출 건수','C',NUM,TOL_A), ('수출 중량(톤)','D',NUM1,TOL_W),
           ('수출 금액(천$)','E',NUM,TOL_A), ('수입 건수','F',NUM,TOL_A), ('수입 중량(톤)','G',NUM1,TOL_W),
           ('수입 금액(천$)','H',NUM,TOL_A), ('무역수지(천$)','I',NUM,TOL_A)]
v1_start = r
for name, col, fmt, tol in metrics:
    put(ws, r, 1, name, bold=True)
    put(ws, r, 2, f'={S_TOT}!{col}6', fmt)
    put(ws, r, 3, f'=SUM({S_TOT}!{col}7:{col}13)', fmt)
    put(ws, r, 4, f'=B{r}-C{r}', NUM1)
    put(ws, r, 5, f'=IF(ABS(D{r})<={tol},"일치(반올림오차 이내)","불일치")', align='center')
    r += 1
r += 1

sec(ws, r, '[검증 2] 무역수지 항등식 : 수출금액 − 수입금액 = 무역수지'); r += 1
for i,h in enumerate(['기간','수출금액','수입금액','계산값(수출−수입)','원본 무역수지','차이','판정']):
    style_head(ws, r, i+1, h)
r += 1
for idx, src in enumerate(['6'] + [str(x) for x in TOT_ROWS]):
    put(ws, r, 1, '총계' if idx == 0 else MONTHS[idx-1], bold=(idx==0))
    put(ws, r, 2, f'={S_TOT}!E{src}', NUM)
    put(ws, r, 3, f'={S_TOT}!H{src}', NUM)
    put(ws, r, 4, f'=B{r}-C{r}', NUM)
    put(ws, r, 5, f'={S_TOT}!I{src}', NUM)
    put(ws, r, 6, f'=D{r}-E{r}', NUM)
    put(ws, r, 7, f'=IF(ABS(F{r})<={TOL_A},"일치(반올림오차 이내)","불일치")', align='center')
    r += 1
r += 1

sec(ws, r, '[검증 3] 「성질별」 대분류 4개 합 vs 「총괄」 수출 실적'); r += 1
for i,h in enumerate(['기간','성질별 금액 합','총괄 수출금액','차이','판정','성질별 중량 합','총괄 수출중량','차이','판정']):
    style_head(ws, r, i+1, h)
r += 1
for m in range(7):
    st = OLD_START[m]
    amt = '+'.join(f'{S_OLD}!D{st+o}' for o in OLD_MAJ)
    wgt = '+'.join(f'{S_OLD}!C{st+o}' for o in OLD_MAJ)
    put(ws, r, 1, MONTHS[m], bold=True)
    put(ws, r, 2, '='+amt, NUM); put(ws, r, 3, f'={S_TOT}!E{TOT_ROWS[m]}', NUM)
    put(ws, r, 4, f'=B{r}-C{r}', NUM1)
    put(ws, r, 5, f'=IF(ABS(D{r})<={TOL_A},"일치","불일치")', align='center')
    put(ws, r, 6, '='+wgt, NUM1); put(ws, r, 7, f'={S_TOT}!D{TOT_ROWS[m]}', NUM1)
    put(ws, r, 8, f'=F{r}-G{r}', NUM1)
    put(ws, r, 9, f'=IF(ABS(H{r})<={TOL_W},"일치","불일치")', align='center')
    r += 1
put(ws, r, 1, '7개월 누계', bold=True)
put(ws, r, 2, f'=SUM(B{r-7}:B{r-1})', NUM, bold=True); put(ws, r, 3, f'={S_OLD}!D6', NUM, bold=True)
put(ws, r, 4, f'=B{r}-C{r}', NUM1); put(ws, r, 5, f'=IF(ABS(D{r})<={TOL_A}*7,"일치(누계 반올림오차)","불일치")', align='center')
put(ws, r, 6, f'=SUM(F{r-7}:F{r-1})', NUM1, bold=True); put(ws, r, 7, f'={S_OLD}!C6', NUM1, bold=True)
put(ws, r, 8, f'=F{r}-G{r}', NUM1); put(ws, r, 9, f'=IF(ABS(H{r})<={TOL_W}*7,"일치(누계 반올림오차)","불일치")', align='center')
r += 1
note(ws, r, 1, 'C열·G열의 총계 비교값은 「성질별」 시트 6행(총계행)입니다. 즉 대분류 4개 × 7개월의 합이 시트 자체 총계와 일치하는지를 봅니다.', 9)
r += 2

sec(ws, r, '[검증 4] 「성질별」 중분류 합 vs 대분류 (경공업품 / 중화학공업품)'); r += 1
for i,h in enumerate(['기간','3.경공업품(대)','중분류 가~카 합','차이','판정','4.중화학공업품(대)','중분류 가~바 합','차이','판정']):
    style_head(ws, r, i+1, h)
r += 1
for m in range(7):
    st = OLD_START[m]
    put(ws, r, 1, MONTHS[m], bold=True)
    put(ws, r, 2, f'={S_OLD}!D{st+2}', NUM)
    put(ws, r, 3, f'=SUM({S_OLD}!D{st+3}:D{st+13})', NUM)
    put(ws, r, 4, f'=B{r}-C{r}', NUM1)
    put(ws, r, 5, f'=IF(ABS(D{r})<={TOL_A},"일치","불일치")', align='center')
    put(ws, r, 6, f'={S_OLD}!D{st+14}', NUM)
    put(ws, r, 7, f'=SUM({S_OLD}!D{st+15}:D{st+20})', NUM)
    put(ws, r, 8, f'=F{r}-G{r}', NUM1)
    put(ws, r, 9, f'=IF(ABS(H{r})<={TOL_A},"일치","불일치")', align='center')
    r += 1
r += 1

sec(ws, r, '[검증 5] 「신성질별」 대분류 3개 합 vs 「총괄」 수출 실적'); r += 1
for i,h in enumerate(['기간','신성질별 금액 합','총괄 수출금액','차이','판정','신성질별 중량 합','총괄 수출중량','차이','판정']):
    style_head(ws, r, i+1, h)
r += 1
for m in range(7):
    st = NEW_START[m]
    amt = '+'.join(f'{S_NEW}!D{st+o}' for o in NEW_MAJ)
    wgt = '+'.join(f'{S_NEW}!C{st+o}' for o in NEW_MAJ)
    put(ws, r, 1, MONTHS[m], bold=True)
    put(ws, r, 2, '='+amt, NUM); put(ws, r, 3, f'={S_TOT}!E{TOT_ROWS[m]}', NUM)
    put(ws, r, 4, f'=B{r}-C{r}', NUM1)
    put(ws, r, 5, f'=IF(ABS(D{r})<={TOL_A},"일치","불일치")', align='center')
    put(ws, r, 6, '='+wgt, NUM1); put(ws, r, 7, f'={S_TOT}!D{TOT_ROWS[m]}', NUM1)
    put(ws, r, 8, f'=F{r}-G{r}', NUM1)
    put(ws, r, 9, f'=IF(ABS(H{r})<={TOL_W},"일치","불일치")', align='center')
    r += 1
put(ws, r, 1, '7개월 누계', bold=True)
put(ws, r, 2, f'=SUM(B{r-7}:B{r-1})', NUM, bold=True); put(ws, r, 3, f'={S_NEW}!D6', NUM, bold=True)
put(ws, r, 4, f'=B{r}-C{r}', NUM1); put(ws, r, 5, f'=IF(ABS(D{r})<={TOL_A}*7,"일치(누계 반올림오차)","불일치")', align='center')
put(ws, r, 6, f'=SUM(F{r-7}:F{r-1})', NUM1, bold=True); put(ws, r, 7, f'={S_NEW}!C6', NUM1, bold=True)
put(ws, r, 8, f'=F{r}-G{r}', NUM1); put(ws, r, 9, f'=IF(ABS(H{r})<={TOL_W}*7,"일치(누계 반올림오차)","불일치")', align='center')
r += 2

sec(ws, r, '[검증 6] 「신성질별」 중분류 합 vs 대분류'); r += 1
for i,h in enumerate(['기간','1.소비재 차이','2.원자재 차이','3.자본재 차이','판정(금액)']):
    style_head(ws, r, i+1, h)
r += 1
for m in range(7):
    st = NEW_START[m]
    put(ws, r, 1, MONTHS[m], bold=True)
    put(ws, r, 2, f'={S_NEW}!D{st+0}-SUM({S_NEW}!D{st+1}:D{st+4})', NUM1)
    put(ws, r, 3, f'={S_NEW}!D{st+5}-SUM({S_NEW}!D{st+6}:D{st+11})', NUM1)
    put(ws, r, 4, f'={S_NEW}!D{st+12}-SUM({S_NEW}!D{st+13}:D{st+16})', NUM1)
    put(ws, r, 5, f'=IF(MAX(ABS(B{r}),ABS(C{r}),ABS(D{r}))<={TOL_A},"일치","불일치")', align='center')
    r += 1
r += 1

sec(ws, r, '[검증 7] 시트 간 교차검증 — 동일한 수출 실적을 3개 시트가 같은 값으로 말하는가'); r += 1
for i,h in enumerate(['구분','수출 금액(천$)','수출 중량(톤)']):
    style_head(ws, r, i+1, h)
r += 1
for label, ref in [('수출입 총괄 (총계행)', f'{S_TOT}!E6|{S_TOT}!D6'),
                   ('성질별 (총계행)', f'{S_OLD}!D6|{S_OLD}!C6'),
                   ('신성질별 (총계행)', f'{S_NEW}!D6|{S_NEW}!C6')]:
    a, b = ref.split('|')
    put(ws, r, 1, label, bold=True); put(ws, r, 2, '='+a, NUM); put(ws, r, 3, '='+b, NUM1)
    r += 1
put(ws, r, 1, '최대 편차', bold=True)
put(ws, r, 2, f'=MAX(B{r-3}:B{r-1})-MIN(B{r-3}:B{r-1})', NUM1, bold=True)
put(ws, r, 3, f'=MAX(C{r-3}:C{r-1})-MIN(C{r-3}:C{r-1})', NUM1, bold=True)
r += 1
put(ws, r, 1, '판정', bold=True)
put(ws, r, 2, f'=IF(B{r-1}<={TOL_A},"3개 시트 일치","불일치")', align='center')
put(ws, r, 3, f'=IF(C{r-1}<={TOL_W},"3개 시트 일치","불일치")', align='center')
r += 2

sec(ws, r, '검증 총평'); r += 1
_cnt = '+'.join(f'COUNTIF({cl}1:{cl}{r-1},"불일치")' for cl in ['B','C','D','E','G','I'])
c = ws.cell(r, 1, f'=IF({_cnt}=0,"검증 결과: 모든 항목이 허용오차(반올림) 이내에서 일치합니다.","검증 결과: 불일치 항목이 있습니다. 위 표의 판정열을 확인하세요.")')
c.font = Font(name=FONT, sz=11, b=True, color=GREEN)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
note(ws, r, 1, '주) 원본 데이터는 소수점 반올림 상태로 제공되므로 개별 항목 합계와 총계 간에 ±수 단위(천달러) / ±0.x톤의 차이가 발생할 수 있습니다. '
               '이는 데이터 오류가 아니라 반올림 누적 오차입니다.', 9)
CHK_LAST = r

# ══════════════════════════════════════════════════════════════════════════
# 시트 2 : 분석2_월별추이
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('분석2_월별추이')
ws2.sheet_view.showGridLines = False
widths = [12,10,12,15,15,15,15,11,11,14,15,11,13,13,13,16,16,10]
for i,wd in enumerate(widths): ws2.column_dimensions[get_column_letter(i+1)].width = wd

style_title(ws2, 1, '② 월별 추이 — 총괄 지표 · 조업일수 보정 · 단가', 17)
note(ws2, 2, 1, '「수출입 총괄」 시트를 수식으로 참조합니다. 단가 = 금액(천달러) ÷ 중량(톤), 일평균 = 금액 ÷ 조업일수(월별 영업일 차이 보정).', 17)
note(ws2, 3, 1, '순상품교역조건 = 수출단가 ÷ 수입단가. 값이 높을수록 같은 물량을 팔아 더 많은 수입 물량을 살 수 있음을 뜻합니다.', 17)

H2 = ['기간','조업일수','수출 건수','수출 중량(톤)','수출 금액(천$)','수입 금액(천$)','무역수지(천$)',
      '수출 전월비','수입 전월비','수지 전월증감(천$)','일평균 수출(천$)','일평균 수출 전월비',
      '수출단가(천$/톤)','수입단가(천$/톤)','순상품교역조건','누적 수출(천$)','누적 수지(천$)']
HR2 = 5
for i,h in enumerate(H2): style_head(ws2, HR2, i+1, h)
ws2.row_dimensions[HR2].height = 32
D2_START = HR2 + 1
for m in range(7):
    r = D2_START + m; src = TOT_ROWS[m]
    put(ws2, r, 1, MONTHS[m], bold=True, align='center')
    put(ws2, r, 2, f'={S_TOT}!B{src}', NUM1)
    put(ws2, r, 3, f'={S_TOT}!C{src}', NUM)
    put(ws2, r, 4, f'={S_TOT}!D{src}', NUM1)
    put(ws2, r, 5, f'={S_TOT}!E{src}', NUM)
    put(ws2, r, 6, f'={S_TOT}!H{src}', NUM)
    put(ws2, r, 7, f'={S_TOT}!I{src}', NUM)
    put(ws2, r, 8, '' if m == 0 else f'=E{r}/E{r-1}-1', PCT)
    put(ws2, r, 9, '' if m == 0 else f'=F{r}/F{r-1}-1', PCT)
    put(ws2, r,10, '' if m == 0 else f'=G{r}-G{r-1}', NUM)
    put(ws2, r,11, f'=E{r}/B{r}', NUM)
    put(ws2, r,12, '' if m == 0 else f'=K{r}/K{r-1}-1', PCT)
    put(ws2, r,13, f'=E{r}/D{r}', UNIT)
    put(ws2, r,14, f'=F{r}/{S_TOT}!G{src}', UNIT)
    put(ws2, r,15, f'=M{r}/N{r}', '0.000')
    put(ws2, r,16, f'=SUM($E${D2_START}:E{r})', NUM)
    put(ws2, r,17, f'=SUM($G${D2_START}:G{r})', NUM)
R2_END = D2_START + 6
r = R2_END + 1
put(ws2, r, 1, '7개월 누계/평균', bold=True, align='center')
for col, fmt in [(2,NUM1),(3,NUM),(4,NUM1),(5,NUM),(6,NUM),(7,NUM)]:
    cl = get_column_letter(col)
    put(ws2, r, col, f'=SUM({cl}{D2_START}:{cl}{R2_END})', fmt, bold=True)
put(ws2, r, 8, f'=E{R2_END}/E{D2_START}-1', PCT, bold=True)
put(ws2, r, 9, f'=F{R2_END}/F{D2_START}-1', PCT, bold=True)
put(ws2, r,10, f'=G{R2_END}-G{D2_START}', NUM, bold=True)
put(ws2, r,11, f'=E{r}/B{r}', NUM, bold=True)
put(ws2, r,12, f'=K{R2_END}/K{D2_START}-1', PCT, bold=True)
put(ws2, r,13, f'=E{r}/D{r}', UNIT, bold=True)
put(ws2, r,14, f'=F{r}/SUM({S_TOT}!G7:G13)', UNIT, bold=True)
put(ws2, r,15, f'=M{r}/N{r}', '0.000', bold=True)
put(ws2, r,16, f'=E{r}', NUM, bold=True)
put(ws2, r,17, f'=G{r}', NUM, bold=True)
SUM2_ROW = r
note(ws2, r+1, 1, '누계/평균 행의 전월비(H·I·L열)는 1월 대비 7월 증감률(=기간 전체 변화)입니다.', 17)

# ══════════════════════════════════════════════════════════════════════════
# 시트 3 : 분석3_품목별증감
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('분석3_품목별증감')
ws3.sheet_view.showGridLines = False
w3 = [8,26] + [14]*7 + [12,15,12,15,16,10,11,11]
for i,wd in enumerate(w3): ws3.column_dimensions[get_column_letter(i+1)].width = wd

style_title(ws3, 1, '③ 품목별 증감 분석 — 전월비 · 기간내 증감 · 기여도', 18)
note(ws3, 2, 1, '모든 수치는 수출 금액(천달러) 기준이며 원본 시트를 수식으로 참조합니다. '
                '기여도 = 해당 품목의 1월→7월 증감액 ÷ 전체 수출의 1월→7월 증감액.', 18)
note(ws3, 3, 1, '"최대/최소 월증감률"은 2~7월 6개 구간의 전월비 중 최대·최소값으로, 값이 클수록 월별 변동이 심한(불규칙한) 품목입니다.', 18)

H3 = ['구분','품목'] + MONTHS + ['7월 전월비','7월 전월증감액','1월→7월 증감률','1월→7월 증감액',
                                '7개월 누계','누계 비중','기여도','최대 월증감률','최소 월증감률']

def item_block(ws, start_row, title, sheet_ref, starts, items, total_ref):
    sec(ws, start_row, title, 18)
    hr = start_row + 1
    for i,h in enumerate(H3): style_head(ws, hr, i+1, h)
    ws.row_dimensions[hr].height = 30
    r0 = hr + 1
    for k,(off, lv, nm) in enumerate(items):
        r = r0 + k
        put(ws, r, 1, '대분류' if lv == 1 else '중분류', align='center', bold=(lv==1))
        put(ws, r, 2, nm, bold=(lv==1))
        if lv == 1:
            ws.cell(r,1).fill = SUB_FILL; ws.cell(r,2).fill = SUB_FILL
        for m in range(7):
            c = put(ws, r, 3+m, f'={sheet_ref}!D{starts[m]+off}', NUM, bold=(lv==1))
            if lv == 1: c.fill = SUB_FILL
        put(ws, r,10, f'=I{r}/H{r}-1', PCT, bold=(lv==1))
        put(ws, r,11, f'=I{r}-H{r}', NUM)
        put(ws, r,12, f'=I{r}/C{r}-1', PCT, bold=(lv==1))
        put(ws, r,13, f'=I{r}-C{r}', NUM)
        put(ws, r,14, f'=SUM(C{r}:I{r})', NUM)
        put(ws, r,15, f'=N{r}/{total_ref}', PCT2)
        put(ws, r,16, f'=IFERROR(M{r}/({S_TOT}!E13-{S_TOT}!E7),"-")', PCT2)
        put(ws, r,17, f'=MAX(D{r}/C{r}-1,E{r}/D{r}-1,F{r}/E{r}-1,G{r}/F{r}-1,H{r}/G{r}-1,I{r}/H{r}-1)', PCT)
        put(ws, r,18, f'=MIN(D{r}/C{r}-1,E{r}/D{r}-1,F{r}/E{r}-1,G{r}/F{r}-1,H{r}/G{r}-1,I{r}/H{r}-1)', PCT)
    rN = r0 + len(items)
    put(ws, rN, 2, '합계(대분류 계)', bold=True)
    majs = [i for i,(o,lv,n) in enumerate(items) if lv == 1]
    for m in range(7):
        cl = get_column_letter(3+m)
        put(ws, rN, 3+m, '='+'+'.join(f'{cl}{r0+i}' for i in majs), NUM, bold=True)
    put(ws, rN,10, f'=I{rN}/H{rN}-1', PCT, bold=True)
    put(ws, rN,11, f'=I{rN}-H{rN}', NUM, bold=True)
    put(ws, rN,12, f'=I{rN}/C{rN}-1', PCT, bold=True)
    put(ws, rN,13, f'=I{rN}-C{rN}', NUM, bold=True)
    put(ws, rN,14, f'=SUM(C{rN}:I{rN})', NUM, bold=True)
    put(ws, rN,15, f'=N{rN}/{total_ref}', PCT2, bold=True)
    put(ws, rN,16, f'=M{rN}/({S_TOT}!E13-{S_TOT}!E7)', PCT2, bold=True)
    return r0, rN

A_START, A_TOTROW = item_block(ws3, 5, '[A] 성질별 (수출 금액, 천달러)', S_OLD, OLD_START, OLD_ITEMS, f'{S_OLD}!$D$6')
B_HEAD = A_TOTROW + 3
B_START, B_TOTROW = item_block(ws3, B_HEAD, '[B] 신성질별 (수출 금액, 천달러)', S_NEW, NEW_START, NEW_ITEMS, f'{S_NEW}!$D$6')
note(ws3, B_TOTROW + 1, 1, '누계 비중은 각 시트 총계행(7개월 수출 총액) 대비 비중입니다. 한 표에 대분류와 중분류가 함께 들어 있으므로 전체 행의 비중을 단순 합산하면 중복 집계됩니다(대분류 행만 더해야 100%가 됩니다).', 18)

# 중분류 정렬 순서(1월→7월 증감액 기준, 빌드 시점 원본값으로 정렬)
_src = wb['수출입 실적(신성질별)']
_moves = []
for off, lv, nm in NEW_ITEMS:
    if lv != 2: continue
    v1 = _src.cell(NEW_START[0]+off, 4).value
    v7 = _src.cell(NEW_START[6]+off, 4).value
    _moves.append((v7 - v1, nm, off))
_moves.sort(reverse=True)

# ══════════════════════════════════════════════════════════════════════════
# 시트 4 : 분석4_그래프
# ══════════════════════════════════════════════════════════════════════════
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList

ws4 = wb.create_sheet('분석4_그래프')
ws4.sheet_view.showGridLines = False
for i,wd in enumerate([26]+[13]*8): ws4.column_dimensions[get_column_letter(i+1)].width = wd

style_title(ws4, 1, '④ 그래프 — 분석2·분석3 데이터를 그대로 참조', 9)
note(ws4, 2, 1, '아래 보조 데이터표는 그래프 전용이며 모두 수식(분석2·분석3 참조)입니다. 원본 값이 바뀌면 표와 그래프가 함께 갱신됩니다.', 9)

# 보조표 1 : 지수화(2026.01 = 100)
r = 4
sec(ws4, r, '보조표 1. 2026.01 = 100 지수 (물량·금액·단가의 방향 비교)', 9); r += 1
IDX_HDR = r
style_head(ws4, r, 1, '지표')
for m in range(7): style_head(ws4, r, 2+m, MONTHS[m])
r += 1
IDX_START = r
idx_rows = [('수출 금액 지수', 'E'), ('수출 중량 지수', 'D'), ('수출 단가 지수', 'M'), ('수입 금액 지수', 'F')]
for nm, col in idx_rows:
    put(ws4, r, 1, nm, bold=True)
    for m in range(7):
        put(ws4, r, 2+m, f"=분석2_월별추이!{col}{D2_START+m}/분석2_월별추이!${col}${D2_START}*100", '0.0')
    r += 1
IDX_END = r - 1
r += 1

# 보조표 2 : 신성질별 중분류 1월→7월 증감액 (증감액 순 정렬)
sec(ws4, r, '보조표 2. 신성질별 중분류 — 1월 대비 7월 증감액 (천달러, 증감액 순)', 9); r += 1
MOV_HDR = r
style_head(ws4, r, 1, '품목'); style_head(ws4, r, 2, '1월→7월 증감액'); style_head(ws4, r, 3, '1월→7월 증감률')
r += 1
MOV_START = r
row_of = {nm: B_START + i for i,(off, lv, nm) in enumerate(NEW_ITEMS)}
for _d, nm, off in _moves:
    put(ws4, r, 1, nm)
    put(ws4, r, 2, f'=분석3_품목별증감!M{row_of[nm]}', NUM)
    put(ws4, r, 3, f'=분석3_품목별증감!L{row_of[nm]}', PCT)
    r += 1
MOV_END = r - 1
note(ws4, r, 1, '정렬 순서는 파일 작성 시점의 값 기준으로 고정되어 있습니다(수식 자동 재정렬 아님).', 9)
r += 2
CHART_TOP = r + 1

def flatten(ch):
    for ser in ch.series:
        ser.smooth = False

def sub_axis(ch):
    ch.x_axis.delete = True
    ch.x_axis.title = None
    flatten(ch)

def add_chart(ch, anchor, title, w=26, h=12, ylab=None, xlab='기간'):
    flatten(ch)
    ch.title = title; ch.width = w; ch.height = h
    ch.style = 2
    if ylab: ch.y_axis.title = ylab
    if xlab: ch.x_axis.title = xlab
    ch.y_axis.majorGridlines = ChartLines()
    ws4.add_chart(ch, anchor)

MC = [get_column_letter(3+i) for i in range(7)]  # 분석3 월 컬럼 C~I

# 그래프 1 : 수출·수입 금액(막대) + 무역수지(선, 보조축)
b1 = BarChart(); b1.type = 'col'; b1.grouping = 'clustered'
data = Reference(ws2, min_col=5, max_col=6, min_row=HR2, max_row=R2_END)
b1.add_data(data, titles_from_data=True)
b1.set_categories(Reference(ws2, min_col=1, min_row=D2_START, max_row=R2_END))
l1 = LineChart()
l1.add_data(Reference(ws2, min_col=7, min_row=HR2, max_row=R2_END), titles_from_data=True)
l1.y_axis.axId = 200; l1.y_axis.title = '무역수지(천$)'
l1.y_axis.crosses = 'max'
sub_axis(l1)
b1 += l1
add_chart(b1, f'A{CHART_TOP}', '1. 월별 수출·수입 금액과 무역수지 (2026.01~07)', ylab='금액(천달러)', xlab=None)

# 그래프 2 : 조업일수 보정 — 월 수출금액 vs 일평균 수출금액
l2 = LineChart()
l2.add_data(Reference(ws2, min_col=5, min_row=HR2, max_row=R2_END), titles_from_data=True)
l2b = LineChart()
l2b.add_data(Reference(ws2, min_col=11, min_row=HR2, max_row=R2_END), titles_from_data=True)
l2b.y_axis.axId = 200; l2b.y_axis.title = '일평균(천$)'; l2b.y_axis.crosses = 'max'
sub_axis(l2b)
l2 += l2b
l2.set_categories(Reference(ws2, min_col=1, min_row=D2_START, max_row=R2_END))
add_chart(l2, f'A{CHART_TOP+25}', '2. 조업일수 보정 — 월 수출금액 vs 일평균 수출금액', ylab='월 수출금액(천달러)', xlab=None)

# 그래프 3 : 성질별 4대분류 누적 막대
b3 = BarChart(); b3.type = 'col'; b3.grouping = 'percentStacked'; b3.overlap = 100
maj_rows_old = [A_START + i for i,(o,lv,n) in enumerate(OLD_ITEMS) if lv == 1]
for rr in maj_rows_old:
    ref = Reference(ws3, min_col=2, max_col=9, min_row=rr, max_row=rr)
    b3.add_data(ref, from_rows=True, titles_from_data=True)
b3.set_categories(Reference(ws3, min_col=3, max_col=9, min_row=A_START-1))
add_chart(b3, f'A{CHART_TOP+50}', '3. 성질별 대분류 구성비 추이 (월별 수출금액 100% 기준)', ylab='구성비')

# 그래프 4 : 신성질별 3대분류 선
l4 = LineChart()
maj_rows_new = [B_START + i for i,(o,lv,n) in enumerate(NEW_ITEMS) if lv == 1]
for rr in maj_rows_new:
    l4.add_data(Reference(ws3, min_col=2, max_col=9, min_row=rr, max_row=rr), from_rows=True, titles_from_data=True)
l4.set_categories(Reference(ws3, min_col=3, max_col=9, min_row=B_START-1))
add_chart(l4, f'A{CHART_TOP+75}', '4. 신성질별 대분류(소비재·원자재·자본재) 월별 수출금액', ylab='금액(천달러)')

# 그래프 5 : 자본재 세부 4개
l5 = LineChart()
for nm in ['가.수송장비','나.기계류','다.IT제품','라.IT부품']:
    rr = row_of[nm]
    l5.add_data(Reference(ws3, min_col=2, max_col=9, min_row=rr, max_row=rr), from_rows=True, titles_from_data=True)
l5.set_categories(Reference(ws3, min_col=3, max_col=9, min_row=B_START-1))
add_chart(l5, f'A{CHART_TOP+100}', '5. 자본재 세부품목 월별 수출금액 — IT부품이 견인', ylab='금액(천달러)')

# 그래프 6 : 전기전자 vs 나머지 중화학 (성질별)
l6 = LineChart()
row_of_old = {nm: A_START + i for i,(off, lv, nm) in enumerate(OLD_ITEMS)}
for nm in ['라.전기,전자제품','마.수송장비','다.기계류와정밀기기','가.화공품','나.철강제품']:
    rr = row_of_old[nm]
    l6.add_data(Reference(ws3, min_col=2, max_col=9, min_row=rr, max_row=rr), from_rows=True, titles_from_data=True)
l6.set_categories(Reference(ws3, min_col=3, max_col=9, min_row=A_START-1))
add_chart(l6, f'A{CHART_TOP+125}', '6. 중화학공업품 세부품목 월별 수출금액', ylab='금액(천달러)')

# 그래프 7 : 1월→7월 증감액 (가로 막대)
b7 = BarChart(); b7.type = 'bar'; b7.grouping = 'clustered'
b7.add_data(Reference(ws4, min_col=2, min_row=MOV_HDR, max_row=MOV_END), titles_from_data=True)
b7.set_categories(Reference(ws4, min_col=1, min_row=MOV_START, max_row=MOV_END))
b7.legend = None
add_chart(b7, f'A{CHART_TOP+150}', '7. 신성질별 중분류 1월 대비 7월 수출금액 증감액', h=14, ylab=None, xlab='증감액(천달러)')

# 그래프 8 : 지수 비교
l8 = LineChart()
l8.add_data(Reference(ws4, min_col=1, max_col=8, min_row=IDX_START, max_row=IDX_END), from_rows=True, titles_from_data=True)
l8.set_categories(Reference(ws4, min_col=2, max_col=8, min_row=IDX_HDR))
add_chart(l8, f'A{CHART_TOP+175}', '8. 물량·금액·단가 지수 비교 (2026.01 = 100)', ylab='지수')

# 그래프 9 : 수출·수입 단가와 교역조건
l9 = LineChart()
l9.add_data(Reference(ws2, min_col=13, max_col=14, min_row=HR2, max_row=R2_END), titles_from_data=True)
l9.set_categories(Reference(ws2, min_col=1, min_row=D2_START, max_row=R2_END))
l9b = LineChart()
l9b.add_data(Reference(ws2, min_col=15, min_row=HR2, max_row=R2_END), titles_from_data=True)
l9b.y_axis.axId = 200; l9b.y_axis.title = '교역조건(배)'; l9b.y_axis.crosses = 'max'
sub_axis(l9b)
l9 += l9b
add_chart(l9, f'A{CHART_TOP+200}', '9. 수출·수입 단가와 순상품교역조건', ylab='천달러/톤', xlab=None)

# ══════════════════════════════════════════════════════════════════════════
# 시트 5 : 분석5_전년대비(입력)
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('분석5_전년대비(입력)')
ws5.sheet_view.showGridLines = False
for i,wd in enumerate([16,18,18,14,18,18,14,18,18,14]): ws5.column_dimensions[get_column_letter(i+1)].width = wd

style_title(ws5, 1, '⑤ 전년 동월/동기 대비 분석 (전년도 값 입력 시 자동 계산)', 10)
note(ws5, 2, 1, '★ 중요 : 업로드하신 원본 파일에는 2026.01~07 자료만 들어 있고 2025년 자료가 없습니다. '
                '따라서 전년 대비(YoY) 수치는 파일만으로는 산출할 수 없어, 값을 임의로 만들어 넣지 않았습니다.', 10)
note(ws5, 3, 1, '노란색 셀에 관세청 수출입무역통계의 2025년 동월 값을 붙여넣으면 전년동월비가 자동으로 계산되고, 아래 요약도 자동 갱신됩니다.', 10)

r = 5
sec(ws5, r, '[1] 월별 총괄 — 전년 동월 대비 (노란색 = 입력 셀, 단위: 천달러)', 10); r += 1
for i,h in enumerate(['기간','2026년 수출','2025년 수출(입력)','전년동월비','2026년 수입','2025년 수입(입력)','전년동월비',
                      '2026년 수지','2025년 수지(자동)','수지 증감']):
    style_head(ws5, r, i+1, h)
ws5.row_dimensions[r].height = 30
r += 1
Y_START = r
for m in range(7):
    rr = Y_START + m
    put(ws5, rr, 1, MONTHS[m], bold=True, align='center')
    put(ws5, rr, 2, f'=분석2_월별추이!E{D2_START+m}', NUM)
    c = put(ws5, rr, 3, None, NUM, color=BLUE); c.fill = IN_FILL
    put(ws5, rr, 4, f'=IF(C{rr}="","",B{rr}/C{rr}-1)', PCT)
    put(ws5, rr, 5, f'=분석2_월별추이!F{D2_START+m}', NUM)
    c = put(ws5, rr, 6, None, NUM, color=BLUE); c.fill = IN_FILL
    put(ws5, rr, 7, f'=IF(F{rr}="","",E{rr}/F{rr}-1)', PCT)
    put(ws5, rr, 8, f'=B{rr}-E{rr}', NUM)
    put(ws5, rr, 9, f'=IF(OR(C{rr}="",F{rr}=""),"",C{rr}-F{rr})', NUM)
    put(ws5, rr,10, f'=IF(I{rr}="","",H{rr}-I{rr})', NUM)
Y_END = Y_START + 6
r = Y_END + 1
put(ws5, r, 1, '7개월 누계', bold=True, align='center')
for col in [2,3,5,6,8,9]:
    cl = get_column_letter(col)
    put(ws5, r, col, f'=SUM({cl}{Y_START}:{cl}{Y_END})', NUM, bold=True)
put(ws5, r, 4, f'=IF(C{r}=0,"",B{r}/C{r}-1)', PCT, bold=True)
put(ws5, r, 7, f'=IF(F{r}=0,"",E{r}/F{r}-1)', PCT, bold=True)
put(ws5, r,10, f'=IF(I{r}=0,"",H{r}-I{r})', NUM, bold=True)
Y_SUM = r
r += 2

sec(ws5, r, '[2] 대분류 7개월 누계 — 전년 동기 대비 (노란색 = 입력 셀, 단위: 천달러)', 10); r += 1
for i,h in enumerate(['구분','품목','2026년 누계','2025년 누계(입력)','전년동기비','증감액']):
    style_head(ws5, r, i+1, h)
r += 1
for lbl, sheet_nm, items, start in [('성질별', '분석3_품목별증감', OLD_ITEMS, A_START),
                                    ('신성질별', '분석3_품목별증감', NEW_ITEMS, B_START)]:
    for i,(off, lv, nm) in enumerate(items):
        if lv != 1: continue
        put(ws5, r, 1, lbl, align='center')
        put(ws5, r, 2, nm, bold=True)
        put(ws5, r, 3, f'={sheet_nm}!N{start+i}', NUM)
        c = put(ws5, r, 4, None, NUM, color=BLUE); c.fill = IN_FILL
        put(ws5, r, 5, f'=IF(D{r}="","",C{r}/D{r}-1)', PCT)
        put(ws5, r, 6, f'=IF(D{r}="","",C{r}-D{r})', NUM)
        r += 1
note(ws5, r, 1, '출처 표기 예시 : 관세청 수출입무역통계(unipass.customs.go.kr) — 성질별/신성질별 수출입실적, 2025.01~07, 수리일 기준. '
                '값을 입력할 때 통계 기준(수리일)과 중량단위(톤)를 원본과 동일하게 맞추셔야 비교가 성립합니다.', 10)
r += 2
sec(ws5, r, '[참고] 전년 자료가 없어도 확인 가능한 대안 지표', 10); r += 1
alts = [('연초 대비(1월→7월) 증감률', "분석3_품목별증감 L열 — 계절성은 걷어내지 못하지만 기간 내 추세는 볼 수 있습니다."),
        ('조업일수 보정 일평균', "분석2_월별추이 K·L열 — 2월(19일)처럼 영업일이 짧은 달의 착시를 제거합니다."),
        ('단가(금액/중량)', "분석2_월별추이 M·N열 — 물량 효과와 가격 효과를 분리합니다."),
        ('구성비 변화', "분석3_품목별증감 O열 — 어떤 품목이 포트폴리오에서 커졌는지 봅니다.")]
for a, b in alts:
    put(ws5, r, 1, a, bold=True); ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    put(ws5, r, 2, b, border=False)
    r += 1

# ══════════════════════════════════════════════════════════════════════════
# 시트 6 : 분석6_해설
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('분석6_해설')
ws6.sheet_view.showGridLines = False
for i,wd in enumerate([30,16,16,16,16,16,16,16,16]): ws6.column_dimensions[get_column_letter(i+1)].width = wd

style_title(ws6, 1, '⑥ 분석 결과 해설 — 숫자가 말하는 것', 9)
note(ws6, 2, 1, '아래 [핵심 수치] 표는 모두 수식이며 원본이 바뀌면 자동 갱신됩니다. 본문 서술의 숫자는 2026.01~07 자료 기준입니다.', 9)

A2 = '분석2_월별추이'; A3 = '분석3_품목별증감'
J1, J7 = D2_START, R2_END   # 분석2의 1월/7월 행
r = 4
sec(ws6, r, '[핵심 수치] (단위: 천달러, 톤)', 9); r += 1
for i,h in enumerate(['지표','2026.01','2026.07','기간 최고','7개월 누계','1월→7월 변화']):
    style_head(ws6, r, i+1, h)
r += 1
KEY = [
    ('수출 금액', f'{A2}!E{J1}', f'{A2}!E{J7}', f'MAX({A2}!E{J1}:E{J7})', f'{A2}!E{SUM2_ROW}', NUM, True),
    ('수입 금액', f'{A2}!F{J1}', f'{A2}!F{J7}', f'MAX({A2}!F{J1}:F{J7})', f'{A2}!F{SUM2_ROW}', NUM, True),
    ('무역수지', f'{A2}!G{J1}', f'{A2}!G{J7}', f'MAX({A2}!G{J1}:G{J7})', f'{A2}!G{SUM2_ROW}', NUM, True),
    ('수출 중량(톤)', f'{A2}!D{J1}', f'{A2}!D{J7}', f'MAX({A2}!D{J1}:D{J7})', f'{A2}!D{SUM2_ROW}', NUM1, True),
    ('수출 단가(천$/톤)', f'{A2}!M{J1}', f'{A2}!M{J7}', f'MAX({A2}!M{J1}:M{J7})', f'{A2}!M{SUM2_ROW}', UNIT, True),
    ('수입 단가(천$/톤)', f'{A2}!N{J1}', f'{A2}!N{J7}', f'MAX({A2}!N{J1}:N{J7})', f'{A2}!N{SUM2_ROW}', UNIT, True),
    ('순상품교역조건', f'{A2}!O{J1}', f'{A2}!O{J7}', f'MAX({A2}!O{J1}:O{J7})', f'{A2}!O{SUM2_ROW}', '0.000', True),
    ('일평균 수출', f'{A2}!K{J1}', f'{A2}!K{J7}', f'MAX({A2}!K{J1}:K{J7})', f'{A2}!K{SUM2_ROW}', NUM, True),
]
KEY_START = r
for nm, c1, c7, cmax, ctot, fmt, pct in KEY:
    put(ws6, r, 1, nm, bold=True)
    put(ws6, r, 2, '='+c1, fmt); put(ws6, r, 3, '='+c7, fmt)
    put(ws6, r, 4, '='+cmax, fmt); put(ws6, r, 5, '='+ctot, fmt)
    put(ws6, r, 6, f'=C{r}/B{r}-1', PCT)
    r += 1
KEY_END = r - 1
r += 1

sec(ws6, r, '[증가 기여도 상위] 1월 대비 7월 수출 증가분을 누가 만들었나', 9); r += 1
for i,h in enumerate(['품목(신성질별 중분류)','1월','7월','증감액','증감률','전체 증가분 대비 기여도']):
    style_head(ws6, r, i+1, h)
r += 1
for _d, nm, off in _moves[:6]:
    rr = row_of[nm]
    put(ws6, r, 1, nm, bold=True)
    put(ws6, r, 2, f'={A3}!C{rr}', NUM); put(ws6, r, 3, f'={A3}!I{rr}', NUM)
    put(ws6, r, 4, f'={A3}!M{rr}', NUM); put(ws6, r, 5, f'={A3}!L{rr}', PCT)
    put(ws6, r, 6, f'={A3}!P{rr}', PCT2)
    r += 1
r += 1

def para(ws, row, text, span=9, bold=False, sz=10, color='000000', height=None, formula=False):
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, sz=sz, b=bold, color=color)
    c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    if height: ws.row_dimensions[row].height = height
    return c

SECT = [
 ('1. 무엇을 어떻게 분석했나', [
  '· 원본 3개 시트(수출입 총괄 / 성질별 / 신성질별)는 2026.01~07, 수리일 기준, 단위는 건·톤(TON)·천달러입니다.',
  '· 성질별·신성질별 시트는 검색조건이 [성질명: 수출]이므로 수출 실적만 담고 있습니다. 수입은 총괄 시트에만 총액으로 존재하며, 그래서 품목별 수입·품목별 무역수지는 이 파일만으로는 계산할 수 없습니다.',
  '· 원본 3개 시트는 값 하나도 수정하지 않았고, 분석1~분석6 시트를 새로 추가해 모두 수식으로 원본을 참조하도록 만들었습니다.',
 ]),
 ('2. 합계 정합성 — 결론: 데이터는 신뢰할 수 있습니다', [
  '· 검증 항목(분석1 시트): ① 총괄 총계행 vs 월별 합계(8개 지표) ② 무역수지 항등식(수출−수입=수지) ③ 성질별 대분류 4개 합 vs 총괄 수출 ④ 성질별 중분류 합 vs 대분류 ⑤ 신성질별 대분류 3개 합 vs 총괄 수출 ⑥ 신성질별 중분류 합 vs 대분류 ⑦ 3개 시트 총계 교차검증.',
  '· 모든 항목이 일치했습니다. 관측된 최대 편차는 금액 ±2천달러, 중량 ±0.2톤입니다. 7개월 수출 총액이 5,950.6억 달러이므로 총액의 1억분의 1에도 못 미치는 차이이며, 원본이 소수 첫째자리(중량)·정수(금액)로 반올림되어 제공되기 때문에 생기는 반올림 누적 오차입니다. 데이터 오류가 아닙니다.',
  '· 의미: 성질별과 신성질별은 같은 원자료를 다른 기준으로 나눈 두 개의 뷰이고, 어느 쪽으로 집계해도 총괄과 같은 총액이 나옵니다. 따라서 두 분류를 교차해 해석해도 무방합니다(예: 신성질별 IT부품 급증 = 성질별 전기·전자제품 급증).',
  '· 유일한 눈에 띄는 점: 2026.07의 무역수지는 수출−수입 계산값과 원본 표기값이 1천달러 차이 납니다. 역시 반올림 차이입니다.',
 ]),
 ('3. 총괄 추이 — 수출 확대와 흑자 폭 확대, 그러나 7월은 실질 둔화', [
  '· 수출은 1월 658.4억 달러에서 6월 1,019.6억 달러까지 올라간 뒤 7월 989.6억 달러로 꺾였습니다(전월비 −2.9%). 7개월 누계는 5,950.6억 달러입니다.',
  '· 무역수지는 7개월 내내 흑자였고 1월 87.0억 달러에서 6월 359.0억 달러로 4배 이상 커졌다가 7월 303.9억 달러로 줄었습니다. 누계 흑자는 1,677.6억 달러입니다.',
  '· 조업일수 보정이 중요합니다. 2월은 조업일수가 19일로 가장 짧아 월 합계가 작아 보이지만 일평균 수출은 1월(23.5일)보다 오히려 27% 높습니다. 즉 2월의 부진은 착시입니다.',
  '· 반대로 7월은 조업일수가 24일로 6월(22.5일)보다 많았습니다. 그래서 월 합계로는 −2.9%인 감소가 일평균 기준으로는 −9.0%가 됩니다. 7월 둔화는 달력 효과가 아니라 실질적인 둔화라는 뜻입니다. 이 차이를 보려면 분석2 시트의 K·L열(일평균)을 보셔야 합니다.',
 ]),
 ('4. 성장의 정체 — 물량이 아니라 단가가 끌어올렸습니다', [
  '· 수출 중량은 1월 1,674만 톤에서 7월 1,766만 톤으로 +5.5% 늘었을 뿐인데, 수출 금액은 +50.3% 늘었습니다.',
  '· 그 차이는 전부 단가입니다. 수출 단가(금액÷중량)는 3.933에서 5.603 천달러/톤으로 +42.5% 올랐습니다(6월 6.499가 최고). 같은 무게를 팔아 훨씬 비싼 값을 받고 있다는 뜻이고, 이는 물량 증가가 아니라 고부가가치 품목(반도체·IT부품 등)으로의 구성 변화가 성장을 만들었음을 보여줍니다.',
  '· 수입 단가도 1.143에서 1.341로 올랐지만(+17.3%) 상승 폭이 작아, 순상품교역조건(수출단가÷수입단가)은 3.441에서 4.178배로 개선되었습니다. 같은 물량을 팔아 더 많은 수입 물량을 감당할 수 있게 되었다는 뜻이며, 이것이 흑자 확대의 구조적 배경입니다.',
  '· 다만 7월에는 수출 단가가 6.499에서 5.603으로 −13.8% 떨어졌고 교역조건도 4.298에서 4.178로 낮아졌습니다. 단가로 벌어온 흑자는 단가가 되돌아설 때 같은 속도로 줄어듭니다.',
 ]),
 ('5. 품목별 — 사실상 IT 한 축이 전체를 움직였습니다', [
  '· 1월→7월 수출 증가분 331.2억 달러 가운데 IT부품 한 품목이 206.4억 달러(62.3%)를 만들었습니다. IT제품(32.4억, 9.8%)까지 더하면 IT 두 품목이 증가분의 72%입니다. 성질별로 보면 같은 현상이 전기·전자제품 +244.7억 달러(증가분의 73.9%)로 나타납니다.',
  '· 그 결과 수출에서 IT부품+IT제품이 차지하는 비중은 1월 40.9%에서 7월 51.3%로 올랐고, 자본재 전체 비중은 63.5%에서 70.0%가 되었습니다. 수출 포트폴리오가 눈에 띄게 IT 쪽으로 쏠렸습니다.',
  '· 7월 감소도 같은 품목이 만들었습니다. 7월 전월비 감소액의 대부분은 IT부품(−37.1억, −8.0%)과 전기·전자제품(−38.2억, −6.7%)이고, 여기에 귀금속및보석류(−26.9%)가 더해졌습니다. 반대로 목제품(+17.3%), 철강제품(+7.1%), 기계류와정밀기기(+4.8%), 직물(+5.7%)은 7월에도 늘었습니다. 즉 7월 조정은 전방위 부진이 아니라 IT·귀금속에 국한된 조정입니다.',
  '· 변동성이 큰 품목: 귀금속및보석류는 월별 증감률이 −26.9%~+54.6%로 요동칩니다(3월 15.4억 달러로 급등 후 등락). 기타 원자재는 −61%~+274%로 더 심하지만 금액 규모가 작아 총액에는 영향이 미미합니다. 광산물은 3월에 +50.9% 급등했습니다.',
  '· 꾸준한 품목: 가죽·고무및신발류(1→7월 +20.8%, 6개월 중 5개월 상승), 기계류(+27.4%), 화학공업제품(+21.5%)은 큰 등락 없이 우상향했습니다. 소비재는 +14.0%로 완만해, IT 붐이 소비재 수출까지 끌어올리지는 않았습니다.',
 ]),
 ('6. 이 분석이 갖는 의미', [
  '· ① 집중 리스크가 실체화되었습니다. 수출의 절반이 IT 두 품목에서 나옵니다. 이 구조에서는 반도체·IT 사이클이 곧 국가 수출 사이클이며, 7월의 −8% 조정이 일시적 재고조정인지 사이클 전환인지가 하반기 최대 변수입니다. 8월 이후 IT부품 증감률 하나만 봐도 전체 방향을 상당 부분 예측할 수 있습니다.',
  '· ② 흑자의 질을 구분해야 합니다. 지금의 흑자는 수출 물량이 늘어서가 아니라 단가가 올라서 생겼습니다(물량 +5.5% vs 금액 +50.3%). 가격이 정상화되면 물량이 그대로여도 흑자는 빠르게 줄어듭니다. 7월에 이미 단가 −13.8%와 함께 흑자가 359억→304억 달러로 줄어든 것이 그 예고편입니다.',
  '· ③ 수입 쪽에는 우호적인 신호가 있습니다. 7월 수입 중량은 5,114만 톤으로 7개월 중 최대(+17.1%)였는데 수입 금액은 +3.8%만 늘었습니다. 수입 단가가 −11.3% 떨어졌다는 뜻이고, 원자재를 더 싸게 더 많이 들여왔다는 의미입니다. 원자재 투입 비용 하락은 통상 1~2분기 뒤 수출 마진에 우호적으로 작용합니다.',
  '· ④ 실무적으로 볼 지표는 네 개입니다. IT부품 월간 증감률(방향), 수출 단가(흑자의 지속성), 조업일수 보정 일평균(달력 착시 제거), 순상품교역조건(교역 채산성). 분석2·분석3 시트에서 매월 값만 갱신하면 그대로 추적됩니다.',
 ]),
 ('7. 이 분석의 한계와 확인이 필요한 부분', [
  '· 전년 대비(YoY)는 계산하지 않았습니다. 업로드된 파일에 2025년 자료가 없기 때문이며, 없는 값을 임의로 만들지 않았습니다. 분석5 시트의 노란색 칸에 2025년 값을 붙여넣으면 전년동월비가 자동 계산됩니다.',
  '· 7개월(2026.01~07)만으로는 계절성을 분리할 수 없습니다. 예컨대 2월 부진은 조업일수로 설명되지만, 3월 급등(광산물·귀금속)이 계절 요인인지 일회성인지는 전년 자료가 있어야 판별됩니다.',
  '· 품목별 수입 자료가 없어 품목별 무역수지와 IT부품의 수입 의존도(중간재 수입 → 완제품 수출 구조)는 확인할 수 없습니다. 필요하시면 관세청 통계에서 성질별 수입 실적을 같은 조건(수리일 기준, 톤)으로 추가로 받아 주시면 같은 틀로 확장해 드릴 수 있습니다.',
  '· 금액은 명목 달러 기준이며 환율·물가 조정이 되어 있지 않습니다. 단가 상승분에는 실제 제품 가격 상승과 품목 구성 변화가 섞여 있어, 이를 완전히 분리하려면 품목별 수량(개수) 자료가 필요합니다.',
 ]),
]
for title, lines in SECT:
    sec(ws6, r, title, 9); r += 1
    for ln in lines:
        h = max(16, 15 * (len(ln)//60 + 1))
        para(ws6, r, ln, height=h)
        r += 1
    r += 1
note(ws6, r, 1, '작성: 원본 파일(2026.01~07 수출입 실적)만을 근거로 산출. 외부 자료를 추가로 참조하지 않았으며, 전년 비교치는 자료 부재로 산출하지 않았습니다.', 9)

ws2.freeze_panes = f'B{D2_START}'
ws3.freeze_panes = f'C{A_START}'
ws5.freeze_panes = f'B{Y_START}'
ws.freeze_panes = 'A6'
wb.save(OUT)
print('saved', OUT)
